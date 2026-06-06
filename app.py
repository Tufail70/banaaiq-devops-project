# Flask-Migrate commands:
# flask db init
# flask db migrate
# flask db upgrade

from datetime import date, datetime, timedelta
from copy import deepcopy
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError, as_completed
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from difflib import SequenceMatcher
from functools import wraps
from io import BytesIO
import base64
import hashlib
import html
import json
import logging
import os
import re
import secrets
import tempfile
import traceback
import uuid
from urllib.parse import urljoin, urlparse

from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_mail import Mail, Message
from flask_migrate import Migrate
from flask_login import LoginManager, current_user, login_required, login_user, logout_user
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFError, CSRFProtect, generate_csrf
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Flowable, HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
import requests
from sqlalchemy import inspect, text
from sqlalchemy.exc import SQLAlchemyError
from dotenv import load_dotenv
from werkzeug.exceptions import HTTPException
from itsdangerous import URLSafeTimedSerializer as _URLSafeTimedSerializer, SignatureExpired, BadSignature
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from wtforms import BooleanField, PasswordField, SelectField, StringField, SubmitField, TelField, TextAreaField
from wtforms.validators import DataRequired, Length

from config import Config, normalize_database_url, resolve_database_url
from models import (
    AIQueryLog,
    BOQ,
    BOQActual,
    BOQPackage,
    DPR,
    EngineerPackage,
    FeatureProject,
    Invoice,
    InventoryAssignment,
    InventoryItem,
    Notification,
    PaymentMethod,
    Plan,
    Project,
    ProjectAssignment,
    ProjectCounter,
    ProjectMilestone,
    ROLE_PROJECT_MANAGER,
    ROLE_SITE_ENGINEER,
    StockRequest,
    Subscription,
    Task,
    TaskActivity,
    TaskLog,
    USER_ROLE_OPTIONS,
    UsageLog,
    User,
    db,
    password_hash_looks_valid,
)
from decorators import role_required
from payments import cancel_subscription, charge_invoice, create_customer, create_subscription, get_publishable_key, update_payment_method, get_payment_provider
from utils import calculate_health_score, calculate_vat, can_use_ai, generate_invoice_number, generate_project_code, get_cached_health_score, get_monthly_ai_usage, get_subscription_status, invalidate_health_cache, log_ai_query, project_health_payload, send_cancellation_email, send_payment_failed_final_email, send_payment_retry_email, send_trial_expiry_warning_email, send_trial_welcome_email

try:
    import arabic_reshaper
except Exception:
    arabic_reshaper = None

try:
    from bidi.algorithm import get_display
except Exception:
    get_display = None

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# ── Sentry — init before Flask app creation ────────────────────────────────
def _sentry_scrub_pii(event, hint):
    """Strip passwords, tokens, and payment details from Sentry events."""
    _SENSITIVE = {
        "password", "password_confirm", "current_password", "new_password",
        "token", "csrf_token", "access_token", "refresh_token",
        "api_key", "secret", "card_number", "cvv", "cvc", "card", "iban",
        "bank_account",
    }

    def _scrub(d):
        if isinstance(d, dict):
            return {k: "[SCRUBBED]" if k.lower() in _SENSITIVE else _scrub(v) for k, v in d.items()}
        if isinstance(d, list):
            return [_scrub(i) for i in d]
        return d

    if "request" in event:
        if "data" in event["request"]:
            event["request"]["data"] = _scrub(event["request"]["data"])
        if "cookies" in event["request"]:
            event["request"]["cookies"] = "[SCRUBBED]"
        hdrs = event["request"].get("headers", {})
        for h in ("Authorization", "Cookie"):
            if h in hdrs:
                hdrs[h] = "[SCRUBBED]"
        event["request"]["headers"] = hdrs
    return event


try:
    import sentry_sdk as _sentry_sdk
    from sentry_sdk.integrations.flask import FlaskIntegration as _FlaskInteg
    from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration as _SAInteg
    _SENTRY_DSN = os.getenv("SENTRY_DSN", "").strip()
    if _SENTRY_DSN:
        _sentry_sdk.init(
            dsn=_SENTRY_DSN,
            integrations=[
                _FlaskInteg(transaction_style="endpoint"),
                _SAInteg(),
            ],
            environment=os.getenv("SENTRY_ENVIRONMENT", "development"),
            traces_sample_rate=float(os.getenv("SENTRY_TRACES_SAMPLE_RATE", "0.1")),
            send_default_pii=False,
            release=os.getenv("GIT_COMMIT_SHA", "unknown"),
            before_send=_sentry_scrub_pii,
        )
        print(f"[Sentry] Initialized — env={os.getenv('SENTRY_ENVIRONMENT', 'development')}")
    else:
        print("[Sentry] DSN not set — error tracking disabled (set SENTRY_DSN in production)")
except ImportError:
    _sentry_sdk = None
    print("[Sentry] sentry-sdk not installed — pip install sentry-sdk[flask]")

DECIMAL_CURRENCY = Decimal("0.01")
DECIMAL_QUANTITY = Decimal("0.001")
CURRENCY_RATES = {
    "SAR": Decimal("1"),
    "INR": Decimal("22.27"),
    "USD": Decimal("0.27"),
}
CURRENCY_METADATA = {
    "SAR": {
        "code": "SAR",
        "symbol": "\u0631\u064a\u0627\u0644",
        "display_name": "Saudi Riyal",
        "selector_label": "SAR",
    },
    "INR": {
        "code": "INR",
        "symbol": "\u20b9",
        "display_name": "Indian Rupee",
        "selector_label": "INR",
    },
    "USD": {
        "code": "USD",
        "symbol": "$",
        "display_name": "US Dollar",
        "selector_label": "USD",
    },
}
GCC_COUNTRY_CODES = {"SA", "AE", "QA", "KW", "BH", "OM"}
LOCALHOST_IPS = {"127.0.0.1", "::1", "localhost"}
load_dotenv(os.path.join(BASE_DIR, ".env"))
_is_dev = os.getenv("FLASK_ENV", "production").lower() == "development"
logging.basicConfig(level=logging.DEBUG if _is_dev else logging.WARNING)
for _noisy_logger in ("openai", "httpx", "httpcore"):
    logging.getLogger(_noisy_logger).setLevel(logging.WARNING)


def mask_database_url(database_url):
    if not database_url:
        return "<missing>"
    return re.sub(r":[^:@/]+@", ":***@", database_url, count=1)


configured_database_url = normalize_database_url(resolve_database_url())

app = Flask(__name__, static_folder="static", template_folder="templates")
gunicorn_error_logger = logging.getLogger("gunicorn.error")
if gunicorn_error_logger.handlers:
    app.logger.handlers = gunicorn_error_logger.handlers
app.logger.setLevel(logging.DEBUG if _is_dev else logging.WARNING)
app.config.from_object(Config)
# SECRET_KEY loaded from Config; no fallback override here (config.py raises if missing in prod)
app.config["SQLALCHEMY_DATABASE_URI"] = configured_database_url
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}
app.config["SQLALCHEMY_ECHO"] = False
app.config["WTF_CSRF_ENABLED"] = True
app.config["WTF_CSRF_TIME_LIMIT"] = 3600
app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USE_SSL"] = False
app.config["MAIL_USERNAME"] = os.getenv("MAIL_USERNAME") or "iqbaana@gmail.com"
app.config["MAIL_PASSWORD"] = os.getenv("MAIL_PASSWORD") or ""
app.config["MAIL_DEFAULT_SENDER"] = ("BanaaIQ Contact Form", "iqbaana@gmail.com")
app.config["GOOGLE_CLIENT_ID"] = os.getenv("GOOGLE_CLIENT_ID", "").strip()
app.config["LINKEDIN_CLIENT_ID"] = os.getenv("LINKEDIN_CLIENT_ID", "").strip()
if not os.environ.get("SECRET_KEY"):
    app.logger.warning("SECRET_KEY is not set. Using fallback-dev-key.")
if not os.environ.get("DATABASE_URL"):
    app.logger.warning(
        "DATABASE_URL is not set. Using a local development SQLite database."
    )
app.logger.debug(
    "SQLALCHEMY_DATABASE_URI resolved to %s",
    mask_database_url(app.config["SQLALCHEMY_DATABASE_URI"]),
)
client = OpenAI(api_key=app.config["OPENAI_API_KEY"]) if app.config["OPENAI_API_KEY"] else None
mail = Mail(app)
csrf = CSRFProtect(app)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

db.init_app(app)
migrate = Migrate(app, db, render_as_batch=True)

login_manager = LoginManager()
login_manager.login_view = "auth_login"
login_manager.login_message = "Please log in to continue. / يرجى تسجيل الدخول للمتابعة."
login_manager.login_message_category = "info"
login_manager.init_app(app)

# ── Sentry user context — set/clear on login/logout ───────────────────────
from flask_login import user_logged_in, user_logged_out

@user_logged_in.connect_via(app)
def _sentry_set_user(sender, user, **kw):
    if _sentry_sdk:
        _sentry_sdk.set_user({
            "id": str(user.id),
            "email": user.email,
            "username": getattr(user, "username", None) or user.email,
            "role": getattr(user, "role", None),
        })

@user_logged_out.connect_via(app)
def _sentry_clear_user(sender, user, **kw):
    if _sentry_sdk:
        _sentry_sdk.set_user(None)


def wants_json_response():
    if request.path.startswith("/api/"):
        return True
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return True
    return request.accept_mimetypes.best == "application/json"


@login_manager.unauthorized_handler
def handle_unauthorized():
    if wants_json_response():
        return jsonify(success=False, error="Authentication required."), 401
    return redirect(url_for("auth_login", next=request.full_path if request.query_string else request.path))


def is_protected_path(path):
    protected_prefixes = (
        "/api/",
        "/billing",
        "/dashboard",
        "/my-workspace",
        "/projects",
        "/workspace",
    )
    auth_paths = ("/auth/login", "/login", "/auth/register", "/register", "/auth/forgot", "/auth/reset")
    return path.startswith(protected_prefixes) or path.startswith(auth_paths)


def apply_no_store_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

# ── Session token validation — detects "sign out everywhere" on other devices ─
@app.before_request
def _validate_session_token():
    """If session has a token but it doesn't match the user's current token,
    the user signed out everywhere from another device — kick this session."""
    if not current_user.is_authenticated or session.get("is_guest"):
        return
    stored_token = session.get("session_token")
    # Only enforce if a token was recorded in this session (new logins always set it)
    if stored_token and stored_token != getattr(current_user, "session_token", None):
        logout_user()
        session.clear()
        if request.endpoint not in ("static", "auth_login"):
            flash("Signed out for security. / تم تسجيل الخروج لأمانك.", "info")
        return redirect(url_for("auth_login"))


# ── Request timing middleware ────────────────────────────────────────────────
import time as _time
from flask import g as _g

@app.before_request
def _start_timer():
    _g.request_start = _time.time()

@app.after_request
def _log_slow_request(response):
    elapsed = _time.time() - getattr(_g, 'request_start', _time.time())
    if elapsed > 0.5:
        app.logger.warning(
            "SLOW %.2fs %s %s status=%s",
            elapsed, request.method, request.path, response.status_code
        )
    return response
# ────────────────────────────────────────────────────────────────────────────


def rollback_db_session():
    try:
        db.session.rollback()
    except Exception as rollback_error:
        app.logger.error(f"Database rollback error: {rollback_error}")


def auth_error_message(action, error):
    error_text = str(error).lower()
    if "no such table" in error_text or ("relation" in error_text and "does not exist" in error_text):
        return "The account database is still being initialized. Please try again in a moment."
    return f"We couldn't {action} right now. Please try again in a moment."


def seed_plans():
    if Plan.query.count() > 0:
        return []

    plans = [
        Plan(
            name="starter",
            display_name="Starter",
            monthly_price=199,
            annual_price=1499,
            max_users=1,
            max_ai_queries=50,
            features='["DPR","BOQ","Inventory","Tasks"]',
        ),
        Plan(
            name="professional",
            display_name="Professional",
            monthly_price=499,
            annual_price=4990,
            max_users=5,
            max_ai_queries=None,
            features='["DPR","BOQ","Inventory","Tasks","Translator","Priority Support"]',
        ),
        Plan(
            name="enterprise",
            display_name="Enterprise",
            monthly_price=None,
            annual_price=None,
            max_users=None,
            max_ai_queries=None,
            features='["All Features","SLA","Dedicated Onboarding","Custom Integrations"]',
        ),
    ]
    for plan in plans:
        db.session.add(plan)
    db.session.commit()
    return [plan.name for plan in plans]


def ensure_users_phone_column():
    try:
        inspector = inspect(db.engine)
        user_columns = {column["name"] for column in inspector.get_columns("users")}
        if "phone" in user_columns:
            return False
        db.session.execute(text("ALTER TABLE users ADD COLUMN phone VARCHAR(30)"))
        db.session.commit()
        app.logger.info("Added users.phone column.")
        return True
    except Exception as error:
        rollback_db_session()
        traceback.print_exc()
        app.logger.error("Unable to add users.phone column automatically: %s", error)
        return False


def ensure_users_lockout_columns():
    """Add failed_login_attempts and locked_until columns for account lockout."""
    try:
        inspector = inspect(db.engine)
        user_columns = {column["name"] for column in inspector.get_columns("users")}
        added = []
        if "failed_login_attempts" not in user_columns:
            db.session.execute(text("ALTER TABLE users ADD COLUMN failed_login_attempts INTEGER DEFAULT 0"))
            added.append("failed_login_attempts")
        if "locked_until" not in user_columns:
            db.session.execute(text("ALTER TABLE users ADD COLUMN locked_until TIMESTAMP"))
            added.append("locked_until")
        if added:
            db.session.commit()
            app.logger.info("Added account lockout columns: %s", ", ".join(added))
        return bool(added)
    except Exception as error:
        rollback_db_session()
        app.logger.error("Unable to add lockout columns: %s", error)
        return False


def ensure_boq_intelligence_columns():
    column_defs = {
        "parent_boq_id": "INTEGER REFERENCES boqs(id)",
        "generation_mode": "VARCHAR(30)",
        "project_description": "TEXT",
        "audit_score": "INTEGER",
        "audit_results_json": "TEXT",
        "audit_generated_at": "TIMESTAMP",
        "project_type": "VARCHAR(80)",
        "floor_area_sqm": "FLOAT",
        "is_master": "BOOLEAN DEFAULT FALSE",
        "distributed_at": "TIMESTAMP",
    }
    try:
        inspector = inspect(db.engine)
        if not inspector.has_table("boqs"):
            return False
        existing_columns = {column["name"] for column in inspector.get_columns("boqs")}
        added = []
        for column_name, sql_type in column_defs.items():
            if column_name in existing_columns:
                continue
            db.session.execute(text(f"ALTER TABLE boqs ADD COLUMN {column_name} {sql_type}"))
            added.append(column_name)
        if added:
            db.session.commit()
            app.logger.info("Added BOQ intelligence columns: %s", ", ".join(added))
            return True
        return False
    except Exception as error:
        rollback_db_session()
        traceback.print_exc()
        app.logger.error("Unable to add BOQ intelligence columns automatically: %s", error)
        return False


def ensure_project_command_center_columns():
    table_columns = {
        "projects": {
            "reference_number": "VARCHAR(100)",
            "project_code": "VARCHAR(20)",
            "client_name": "VARCHAR(200)",
            "client_contact": "VARCHAR(200)",
            "contract_value": "NUMERIC(15, 2)",
            "currency": "VARCHAR(5) DEFAULT 'SAR'",
            "start_date": "DATE",
            "planned_completion": "DATE",
            "actual_completion": "DATE",
            "location_city": "VARCHAR(100)",
            "location_zone": "VARCHAR(100)",
            "project_type": "VARCHAR(50)",
            "lead_engineer": "VARCHAR(200)",
            "project_manager": "VARCHAR(200)",
            "health_score": "INTEGER DEFAULT 100",
            "health_status": "VARCHAR(10) DEFAULT 'green'",
            "health_summary": "TEXT",
            "health_components_json": "TEXT DEFAULT '{}'",
            "health_last_calculated": "TIMESTAMP",
            "completion_summary": "TEXT",
            "completion_summary_ar": "TEXT",
            "updated_at": "TIMESTAMP",
        },
        "dprs": {
            "project_id": "INTEGER",
        },
        "boqs": {
            "project_id": "INTEGER",
        },
        "tasks": {
            "project_id": "INTEGER",
            "is_private_to_engineer": "BOOLEAN DEFAULT FALSE",
            "engineer_notes": "TEXT",
        },
        "inventory_items": {
            "project_id": "INTEGER",
        },
        "engineer_package": {
            "item_statuses_json": "TEXT DEFAULT '{}'",
        },
    }
    try:
        inspector = inspect(db.engine)
        added = []
        for table_name, column_defs in table_columns.items():
            if not inspector.has_table(table_name):
                continue
            existing_columns = {column["name"] for column in inspector.get_columns(table_name)}
            for column_name, sql_type in column_defs.items():
                if column_name in existing_columns:
                    continue
                db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {sql_type}"))
                added.append(f"{table_name}.{column_name}")
        if added:
            db.session.commit()
            app.logger.info("Added Project Command Center columns: %s", ", ".join(added))
            return True
        return False
    except Exception as error:
        rollback_db_session()
        traceback.print_exc()
        app.logger.error("Unable to add Project Command Center columns automatically: %s", error)
        return False


_AUTH_LOOP_PATHS = (
    "/auth/login", "/login",
    "/auth/register", "/register",
    "/auth/forgot",
    "/auth/reset",
    "/logout", "/auth/logout",
)


def validate_password_strength(password):
    """Return (True, None) if password meets complexity requirements, else (False, error_msg)."""
    if not password or len(password) < 8:
        return False, "Password must be at least 8 characters. / يجب أن تكون كلمة المرور 8 أحرف على الأقل."
    if not any(c.isupper() for c in password):
        return False, "Password must contain at least one uppercase letter. / يجب أن تحتوي على حرف كبير واحد على الأقل."
    if not any(c.islower() for c in password):
        return False, "Password must contain at least one lowercase letter. / يجب أن تحتوي على حرف صغير واحد على الأقل."
    if not any(c.isdigit() for c in password):
        return False, "Password must contain at least one number. / يجب أن تحتوي على رقم واحد على الأقل."
    return True, None


def get_safe_next_url(next_url):
    if not next_url:
        return None

    next_url = next_url.strip()
    # Reject absolute URLs (open-redirect)
    if "://" in next_url or next_url.startswith("//"):
        return None

    host_url = request.host_url
    ref_url = urlparse(host_url)
    test_url = urlparse(urljoin(host_url, next_url))
    if test_url.scheme not in ("http", "https"):
        return None
    if ref_url.netloc != test_url.netloc:
        return None

    # Don't loop back to auth pages
    path = test_url.path or "/"
    if any(path.startswith(p) for p in _AUTH_LOOP_PATHS):
        return None

    safe_path = path
    if test_url.query:
        safe_path = f"{safe_path}?{test_url.query}"
    return safe_path


def get_plan_amount(plan, billing_cycle):
    if not plan:
        return None
    amount = plan.annual_price if billing_cycle == "annual" else plan.monthly_price
    return to_currency_decimal(amount or 0) if amount is not None else None


def get_billing_period_end(start_at, billing_cycle):
    return start_at + timedelta(days=365 if billing_cycle == "annual" else 30)


def get_default_payment_method(user):
    if not getattr(user, "is_authenticated", False):
        return None
    return PaymentMethod.query.filter_by(user_id=user.id, is_default=True).order_by(PaymentMethod.created_at.desc()).first()


def build_invoice_for_subscription(user, subscription_obj, amount, period_start, period_end, status="pending"):
    # Determine currency from user preference (INR for India, SAR for GCC)
    user_currency = (getattr(user, "preferred_currency", None) or "INR").upper()
    vat_amount, total_amount = calculate_vat(amount)
    invoice = Invoice(
        user_id=user.id,
        subscription_id=subscription_obj.id,
        invoice_number=generate_invoice_number(),
        amount=to_currency_decimal(amount),
        vat_amount=to_currency_decimal(vat_amount),
        total_amount=to_currency_decimal(total_amount),
        currency=user_currency,
        status=status,
        due_date=period_start,
        period_start=period_start,
        period_end=period_end,
    )
    db.session.add(invoice)
    db.session.flush()
    return invoice


def attempt_invoice_charge(invoice, subscription_obj, payment_method_record):
    if not payment_method_record:
        invoice.status = "pending"
        subscription_obj.status = "past_due"
        invoice.retry_count = (invoice.retry_count or 0) + 1
        invoice.next_retry_at = datetime.utcnow() + timedelta(days=3)
        return False

    payment_result = charge_invoice(invoice, payment_method_record.gateway_token or "mock_token")
    if payment_result.get("status") == "succeeded":
        invoice.status = "paid"
        invoice.paid_at = datetime.utcnow()
        invoice.next_retry_at = None
        subscription_obj.status = "active"
        return True

    invoice.status = "failed"
    invoice.retry_count = (invoice.retry_count or 0) + 1
    invoice.next_retry_at = datetime.utcnow() + timedelta(days=3)
    subscription_obj.status = "past_due"
    return False


def normalize_subscription_status(status_value):
    if not status_value:
        return "none"
    return status_value.split(":", 1)[0]


def get_subscription_days_left(status_value):
    if not status_value or not status_value.startswith("trialing:"):
        return None
    try:
        return max(int(status_value.split(":", 1)[1]), 0)
    except (TypeError, ValueError):
        return None


def get_ai_quota_snapshot(user):
    if not getattr(user, "is_authenticated", False):
        return {
            "used": 0,
            "remaining": 0,
            "limit": 0,
            "unlimited": False,
        }

    subscription_obj = getattr(user, "subscription", None)
    plan = getattr(subscription_obj, "plan", None)
    if not subscription_obj or not plan:
        return {
            "used": 0,
            "remaining": 0,
            "limit": 0,
            "unlimited": False,
        }

    used = get_monthly_ai_usage(user)
    if plan.max_ai_queries is None:
        return {
            "used": used,
            "remaining": None,
            "limit": None,
            "unlimited": True,
        }

    remaining = max(plan.max_ai_queries - used, 0)
    return {
        "used": used,
        "remaining": remaining,
        "limit": plan.max_ai_queries,
        "unlimited": False,
    }


def build_ai_response_meta(user):
    quota = get_ai_quota_snapshot(user)
    return {
        "calls_used": quota["used"],
        "calls_remaining": quota["remaining"],
        "ai_limit": quota["limit"],
        "ai_unlimited": quota["unlimited"],
    }


def detect_currency_for_user(user):
    """Return preferred currency for a logged-in user.
    Priority: user.preferred_currency → user.country → default INR (India launch market).
    """
    if not user or not getattr(user, "is_authenticated", False):
        return "INR"
    pref = getattr(user, "preferred_currency", None)
    if pref and pref.upper() in CURRENCY_METADATA:
        return pref.upper()
    country = getattr(user, "country", None)
    if country:
        return map_country_code_to_currency(country)
    return "INR"


def build_plan_view_model(plan, currency_code="INR"):
    """Build display model for a plan in the given currency.
    Uses native INR/USD prices when available; falls back to SAR conversion."""
    currency_code = (currency_code or "INR").upper()

    # Native prices (preferred — set by migration 016 backfill)
    native_monthly = plan.get_price(currency_code, "monthly") if hasattr(plan, "get_price") else None
    native_annual = plan.get_price(currency_code, "annual") if hasattr(plan, "get_price") else None

    # Fallback: use SAR prices + rate conversion
    sar_monthly = float(plan.monthly_price) if plan.monthly_price is not None else None
    sar_annual = float(plan.annual_price) if plan.annual_price is not None else None

    if native_monthly and native_monthly > 0:
        monthly_price_num = native_monthly
        annual_price_num = native_annual if (native_annual and native_annual > 0) else None
    else:
        monthly_price_num = sar_monthly
        annual_price_num = sar_annual

    save_percent = 0
    if monthly_price_num and annual_price_num:
        save_percent = int(round((1 - (annual_price_num / (monthly_price_num * 12))) * 100))

    # Format display strings
    if monthly_price_num is not None and native_monthly and native_monthly > 0:
        # Use native formatted price directly (no SAR conversion needed)
        symbols = {"INR": "₹", "SAR": "SAR ", "USD": "$"}
        sym = symbols.get(currency_code, "")
        monthly_display = f"{int(monthly_price_num):,}"
        annual_display = f"{int(annual_price_num):,}" if annual_price_num else None
    else:
        monthly_display = format_currency_number_from_sar(monthly_price_num, currency_code, decimals=0)
        annual_display = format_currency_number_from_sar(annual_price_num, currency_code, decimals=0)

    return {
        "id": plan.id,
        "name": plan.name,
        "display_name": plan.display_name,
        "monthly_price": monthly_price_num,
        "annual_price": annual_price_num,
        "monthly_price_display": monthly_display,
        "annual_price_display": annual_display,
        "max_users": plan.max_users,
        "max_ai_queries": plan.max_ai_queries,
        "features": plan.features_list,
        "save_percent": save_percent,
    }


def ensure_ai_access(feature_name):
    allowed, error_msg = can_use_ai(current_user)
    if allowed:
        return None
    return (
        jsonify(
            error=error_msg,
            upgrade_required=True,
            pricing_url=url_for("pricing"),
            feature=feature_name,
            success=False,
        ),
        403,
    )


def record_ai_usage(feature_name):
    if not current_user.is_authenticated:
        return
    try:
        log_ai_query(current_user, feature_name)
    except Exception as error:
        rollback_db_session()
        app.logger.error("AI query log failed: %s", error)


def verify_no_plain_passwords():
    users = User.query.all()
    issues_found = False
    for user in users:
        ph = user.password_hash or ""
        if not password_hash_looks_valid(ph):
            app.logger.error(
                "PLAIN TEXT PASSWORD FOUND: user id %s email %s",
                user.id,
                user.email,
            )
            issues_found = True
    if not issues_found:
        app.logger.info("Password audit passed — all passwords hashed correctly.")


def _check_migration_state():
    """Warn if the DB revision is behind the Alembic head. Does not alter schema."""
    try:
        from alembic.config import Config
        from alembic.script import ScriptDirectory
        from alembic.runtime.migration import MigrationContext

        migrations_dir = os.path.join(app.root_path, "migrations")
        alembic_cfg = Config(os.path.join(migrations_dir, "alembic.ini"))
        alembic_cfg.set_main_option("script_location", migrations_dir)
        script = ScriptDirectory.from_config(alembic_cfg)
        head_rev = script.get_current_head()

        with db.engine.connect() as conn:
            ctx = MigrationContext.configure(conn)
            current_rev = ctx.get_current_revision()

        if current_rev != head_rev:
            app.logger.warning(
                "[DB] Migration mismatch — DB at %s, code expects %s. "
                "Run `flask db upgrade` before starting the server.",
                current_rev,
                head_rev,
            )
        else:
            app.logger.info("[DB] Schema at %s ✓", head_rev)
    except Exception as e:
        app.logger.error("[DB] Could not verify migration state: %s", e)


def initialize_database():
    with app.app_context():
        try:
            inspector = inspect(db.engine)
            if not inspector.has_table("users"):
                app.logger.info("[DB] Fresh database detected. Creating initial schema...")
                db.create_all()
                import os
                os.system("flask db stamp head")
            
            # Schema is managed exclusively by Alembic migrations.
            # Run `flask db upgrade` before starting the server.
            _check_migration_state()
            ensure_users_phone_column()
            ensure_users_lockout_columns()
            ensure_boq_intelligence_columns()
            ensure_project_command_center_columns()
            seeded_plans = seed_plans()
            if seeded_plans:
                app.logger.info("Seeded plans: %s", ", ".join(seeded_plans))
            verify_no_plain_passwords()
        except Exception as error:
            rollback_db_session()
            traceback.print_exc()
            app.logger.error(str(error))
            app.logger.error(
                "Database initialization failed for %s",
                mask_database_url(app.config["SQLALCHEMY_DATABASE_URI"]),
            )


initialize_database()

BLOG_ARTICLES = {
    "vision-2030-ai-construction": "blog/article1.html",
    "multilingual-construction-teams-best-practices": "blog/article2.html",
    "predictive-inventory-management-ksa": "blog/article3.html",
}

# Job-title choices (professional designations — stored in User.job_title)
ROLE_CHOICES = list(USER_ROLE_OPTIONS)
REGISTER_JOB_TITLE_CHOICES = [("", "Select your job title...")] + [(t, t) for t in ROLE_CHOICES]

# Account-type choices (access-control level — stored in User.role)
REGISTER_ACCOUNT_TYPE_CHOICES = [
    ("", "Select account type..."),
    (ROLE_PROJECT_MANAGER, "Project Manager"),
    (ROLE_SITE_ENGINEER, "Site Engineer"),
]

BOQ_ENGINEER_ROLES = [
    "Site Engineer",
    "QS / Quantity Surveyor",
    "Project Manager",
    "Procurement Engineer",
    "MEP Engineer",
    "Civil Engineer",
    "Planning Engineer",
    "HSE Officer",
    "Document Controller",
    "Other",
]

ENGINEER_ROLES = {
    "mep_engineer": {
        "label": "MEP Engineer",
        "focus": [
            "HVAC equipment",
            "piping",
            "electrical",
            "plumbing",
            "ventilation",
            "mechanical systems",
            "VRF",
            "chiller",
            "AHU",
            "FCU",
            "conduit",
            "cable tray",
            "insulation",
        ],
        "questions": [
            "Are all equipment models SASO certified?",
            "Is piping sized correctly per load?",
            "Are accessories complete per system?",
            "Is insulation specified for all pipes?",
            "Are control systems included?",
        ],
        "bill_items": [
            "Equipment supply and installation",
            "Piping materials and labour",
            "Accessories and fittings",
            "Insulation supply and installation",
            "Controls and commissioning",
            "Testing and balancing",
        ],
    },
    "civil_engineer": {
        "label": "Civil / Structural Engineer",
        "focus": [
            "concrete",
            "rebar",
            "formwork",
            "excavation",
            "foundation",
            "structural steel",
            "masonry",
            "waterproofing",
            "earthworks",
        ],
        "questions": [
            "Are concrete grades specified?",
            "Is rebar reinforcement detailed?",
            "Are foundation depths adequate?",
            "Is waterproofing included?",
            "Are structural connections detailed?",
        ],
        "bill_items": [
            "Excavation and earthworks",
            "Concrete supply and placement",
            "Rebar supply and fixing",
            "Formwork supply, fix, strip",
            "Waterproofing",
            "Backfilling and compaction",
        ],
    },
    "quantity_surveyor": {
        "label": "Quantity Surveyor (QS)",
        "focus": [
            "quantities",
            "rates",
            "costs",
            "provisional sums",
            "pc sums",
            "contingency",
            "preliminaries",
            "subcontractors",
            "materials",
        ],
        "questions": [
            "Are quantities measured correctly?",
            "Are unit rates reasonable for Saudi market?",
            "Are provisional sums included?",
            "Is contingency adequate (10-15%)?",
            "Are all trades covered?",
            "Is VAT 15% included?",
        ],
        "bill_items": [
            "Measured works by trade",
            "Provisional sums",
            "PC sums for specialist works",
            "Preliminaries and site overheads",
            "Contingency allowance",
            "VAT 15%",
        ],
    },
    "project_manager": {
        "label": "Project Manager",
        "focus": [
            "scope",
            "programme",
            "milestones",
            "deliverables",
            "risks",
            "packages",
            "subcontractors",
            "procurement",
        ],
        "questions": [
            "Is scope clearly defined?",
            "Are package boundaries clear?",
            "Are long-lead items identified?",
            "Are subcontract scopes complete?",
            "Are risks priced?",
            "Is programme realistic?",
        ],
        "bill_items": [
            "Main contractor preliminaries",
            "Specialist subcontract packages",
            "Long-lead procurement items",
            "Testing and commissioning",
            "As-built documentation",
            "Training and handover",
        ],
    },
    "procurement_officer": {
        "label": "Procurement Officer",
        "focus": [
            "materials",
            "suppliers",
            "vendors",
            "delivery",
            "specifications",
            "approved makes",
            "alternatives",
        ],
        "questions": [
            "Are approved makes specified?",
            "Are delivery lead times considered?",
            "Are alternatives/equivalents allowed?",
            "Are Saudi standards (SASO) specified?",
            "Are local suppliers available?",
            "Are quantities confirmed for POs?",
        ],
        "bill_items": [
            "Equipment purchase orders",
            "Materials supply contracts",
            "Import duties and freight",
            "Storage and handling",
            "Quality inspection costs",
        ],
    },
    "hse_officer": {
        "label": "HSE Officer",
        "focus": [
            "safety",
            "PPE",
            "scaffolding",
            "temporary works",
            "fire protection",
            "emergency",
            "signage",
            "training",
        ],
        "questions": [
            "Is HSE plan included in BOQ?",
            "Are scaffolding costs included?",
            "Is PPE provision costed?",
            "Are safety signs included?",
            "Is fire protection adequate?",
            "Are safety inductions budgeted?",
        ],
        "bill_items": [
            "HSE plan preparation",
            "Scaffolding and access",
            "PPE for all workers",
            "Safety signage and barriers",
            "Fire extinguishers on site",
            "Safety training budget",
        ],
    },
}

DPR_PROJECTS = [
    "Riyadh Logistics Hub Phase 2",
    "NEOM Access Road Package 4",
    "Dammam Industrial Zone B2",
]

DPR_RECORDS = [
    {"id": 1, "date": "20 Jan 2025", "project": "Riyadh Logistics Hub Phase 2", "zone": "Zone A - Foundation", "workers": 24, "weather": "Clear", "status": "Completed", "has_ai_summary": True},
    {"id": 2, "date": "21 Jan 2025", "project": "Riyadh Logistics Hub Phase 2", "zone": "Zone B - Columns", "workers": 31, "weather": "Hot", "status": "Completed", "has_ai_summary": True},
    {"id": 3, "date": "22 Jan 2025", "project": "NEOM Access Road Package 4", "zone": "Section 3 - Earthworks", "workers": 18, "weather": "Dusty", "status": "Completed", "has_ai_summary": False},
    {"id": 4, "date": "23 Jan 2025", "project": "Riyadh Logistics Hub Phase 2", "zone": "Zone A - Slab Pour", "workers": 28, "weather": "Clear", "status": "Completed", "has_ai_summary": True},
    {"id": 5, "date": "24 Jan 2025", "project": "Dammam Industrial Zone B2", "zone": "Block 4 - MEP Rough-in", "workers": 15, "weather": "Hot", "status": "Draft", "has_ai_summary": False},
]

BOQ_TEMPLATES = {
    "foundation": [
        {"desc": "Site Clearance and Grubbing", "unit": "m2", "qty": 2500, "rate": 15},
        {"desc": "Bulk Earthworks - Cut to Spoil", "unit": "m3", "qty": 1800, "rate": 45},
        {"desc": "Hardcore Filling and Compaction", "unit": "m3", "qty": 650, "rate": 85},
        {"desc": "Blinding Concrete (75mm)", "unit": "m2", "qty": 2200, "rate": 35},
        {"desc": "Reinforced Concrete Foundations", "unit": "m3", "qty": 480, "rate": 850},
        {"desc": "Rebar Supply and Fix", "unit": "ton", "qty": 68, "rate": 4200},
        {"desc": "Formwork to Foundations", "unit": "m2", "qty": 1200, "rate": 55},
        {"desc": "Waterproofing Membrane", "unit": "m2", "qty": 2500, "rate": 45},
    ],
    "mep": [
        {"desc": "Electrical Main Distribution Board", "unit": "pcs", "qty": 2, "rate": 45000},
        {"desc": "Sub Distribution Boards", "unit": "pcs", "qty": 8, "rate": 12000},
        {"desc": "Cable 16mm2 - Supply and Install", "unit": "m", "qty": 2400, "rate": 85},
        {"desc": "Conduit PVC 25mm", "unit": "m", "qty": 3600, "rate": 18},
        {"desc": "Chilled Water Pipe 4 inch", "unit": "m", "qty": 450, "rate": 380},
        {"desc": "AHU Unit 10TR", "unit": "pcs", "qty": 6, "rate": 28000},
        {"desc": "Fire Alarm System - Complete", "unit": "ls", "qty": 1, "rate": 185000},
        {"desc": "CCTV System - 32 Camera", "unit": "ls", "qty": 1, "rate": 95000},
    ],
    "fitout": [
        {"desc": "Gypsum Board Partitions", "unit": "m2", "qty": 3200, "rate": 95},
        {"desc": "Ceramic Floor Tiling", "unit": "m2", "qty": 2800, "rate": 85},
        {"desc": "Suspended Ceiling Works", "unit": "m2", "qty": 2600, "rate": 110},
        {"desc": "Wooden Doors Complete", "unit": "pcs", "qty": 140, "rate": 1800},
        {"desc": "Painting Works - Internal", "unit": "m2", "qty": 9500, "rate": 22},
        {"desc": "Aluminum Glazing", "unit": "m2", "qty": 420, "rate": 680},
    ],
}

BOQ_SAVED = [
    {
        "id": 1,
        "title": "Foundation Works BOQ",
        "project": "Riyadh Hub Ph2",
        "items": 12,
        "total": 2450000,
        "last_modified": "20 Jan 2025",
        "status": "Approved",
        "template_key": "foundation",
    },
    {
        "id": 2,
        "title": "MEP Package BOQ",
        "project": "NEOM Package 4",
        "items": 18,
        "total": 1890000,
        "last_modified": "18 Jan 2025",
        "status": "Draft",
        "template_key": "mep",
    },
    {
        "id": 3,
        "title": "Fit-Out BOQ",
        "project": "Dammam Zone B2",
        "items": 15,
        "total": 980000,
        "last_modified": "15 Jan 2025",
        "status": "Under Review",
        "template_key": "fitout",
    },
]

INVENTORY_ITEMS = [
    {"id": 1, "name": "Rebar 10mm", "category": "Steel", "unit": "Ton", "stock": 2.5, "threshold": 5, "status": "critical", "updated": "24 Jan 2025", "value_sar": 10500, "supplier": "Al Rajhi Steel", "location": "Yard A", "notes": "Emergency reorder required"},
    {"id": 2, "name": "Rebar 16mm", "category": "Steel", "unit": "Ton", "stock": 8.0, "threshold": 5, "status": "ok", "updated": "24 Jan 2025", "value_sar": 33600, "supplier": "Al Rajhi Steel", "location": "Yard A", "notes": ""},
    {"id": 3, "name": "OPC Cement", "category": "Concrete", "unit": "Bag", "stock": 145, "threshold": 200, "status": "low", "updated": "23 Jan 2025", "value_sar": 7250, "supplier": "Yamama Cement", "location": "Store 1", "notes": "Reorder placed 22 Jan"},
    {"id": 4, "name": "Ready Mix Concrete", "category": "Concrete", "unit": "m3", "stock": 0, "threshold": 20, "status": "critical", "updated": "22 Jan 2025", "value_sar": 0, "supplier": "Saudi Readymix", "location": "On demand", "notes": "Order on call basis"},
    {"id": 5, "name": "Concrete Blocks", "category": "Masonry", "unit": "Pcs", "stock": 3200, "threshold": 500, "status": "ok", "updated": "20 Jan 2025", "value_sar": 16000, "supplier": "Local Supplier", "location": "Yard B", "notes": ""},
    {"id": 6, "name": "Electrical Conduit 25mm", "category": "Electrical", "unit": "m", "stock": 85, "threshold": 100, "status": "low", "updated": "24 Jan 2025", "value_sar": 1530, "supplier": "Saudi Cable Co.", "location": "Store 2", "notes": ""},
    {"id": 7, "name": "Cable 16mm2", "category": "Electrical", "unit": "m", "stock": 55, "threshold": 100, "status": "low", "updated": "23 Jan 2025", "value_sar": 4675, "supplier": "Saudi Cable Co.", "location": "Store 2", "notes": "Check with MEP team"},
    {"id": 8, "name": "PVC Pipe 4in", "category": "Plumbing", "unit": "m", "stock": 220, "threshold": 50, "status": "ok", "updated": "21 Jan 2025", "value_sar": 8800, "supplier": "Gulf Pipes", "location": "Store 3", "notes": ""},
    {"id": 9, "name": "Ceramic Tiles 60x60", "category": "Finishing", "unit": "m2", "stock": 320, "threshold": 100, "status": "ok", "updated": "19 Jan 2025", "value_sar": 22400, "supplier": "RAK Ceramics", "location": "Store 4", "notes": ""},
    {"id": 10, "name": "Waterproofing Membrane", "category": "Finishing", "unit": "Roll", "stock": 12, "threshold": 10, "status": "ok", "updated": "18 Jan 2025", "value_sar": 9600, "supplier": "BASF", "location": "Store 4", "notes": ""},
    {"id": 11, "name": "Safety Helmets", "category": "Safety", "unit": "Pcs", "stock": 18, "threshold": 30, "status": "low", "updated": "24 Jan 2025", "value_sar": 1800, "supplier": "3M Saudi", "location": "Safety Store", "notes": "HSE requirement minimum 30"},
    {"id": 12, "name": "Formwork Panels", "category": "Shuttering", "unit": "m2", "stock": 95, "threshold": 50, "status": "ok", "updated": "22 Jan 2025", "value_sar": 28500, "supplier": "PERI Saudi", "location": "Yard C", "notes": ""},
]

USAGE_LOG = [
    {"date": "24 Jan 25", "item": "Rebar 10mm", "used": "0.5T", "by": "Ahmed", "zone": "Zone A"},
    {"date": "23 Jan 25", "item": "OPC Cement", "used": "25", "by": "Khalid", "zone": "Zone B"},
    {"date": "22 Jan 25", "item": "Safety Helmets", "used": "5", "by": "Sara", "zone": "Site"},
    {"date": "21 Jan 25", "item": "Cable 16mm2", "used": "20m", "by": "Tariq", "zone": "L3"},
    {"date": "20 Jan 25", "item": "Formwork Panels", "used": "15m2", "by": "Ali", "zone": "Zone C"},
]

TASKS_DATA = [
    {
        "id": 1,
        "title": "Survey Foundation Zone C",
        "assignee": "Ali Hassan",
        "initials": "AH",
        "avatar_color": "#0a0a0a",
        "due": "25 Jan 2025",
        "priority": "high",
        "category": "Civil",
        "status": "backlog",
        "description": "Complete topographic survey of Zone C foundation area before concrete pour begins. Coordinate with site engineer for benchmark reference points.",
        "subtasks": [{"text": "Set up survey equipment", "done": False}, {"text": "Take 15 reference points", "done": False}, {"text": "Generate survey report", "done": False}],
        "comments": [
            {"author": "Ali Hassan", "initials": "AH", "text": "Survey equipment booked for Monday", "time": "2 hours ago"},
            {"author": "Mohammed Al-Rashidi", "initials": "MR", "text": "Please coordinate with the concrete team", "time": "1 hour ago"},
        ],
        "created": "20 Jan 2025",
    },
    {
        "id": 2,
        "title": "Submit HSE Weekly Report",
        "assignee": "Sara Al-Ghamdi",
        "initials": "SG",
        "avatar_color": "#e8c547",
        "due": "26 Jan 2025",
        "priority": "medium",
        "category": "Safety",
        "status": "backlog",
        "description": "Compile and submit the weekly HSE report. Include incident log, toolbox talks, and PPE compliance rates.",
        "subtasks": [{"text": "Collect incident reports", "done": False}, {"text": "Compile toolbox attendance", "done": False}, {"text": "Calculate PPE compliance", "done": False}],
        "comments": [],
        "created": "20 Jan 2025",
    },
    {
        "id": 3,
        "title": "Order Rebar 10mm - Emergency",
        "assignee": "Mohammed Al-Rashidi",
        "initials": "MR",
        "avatar_color": "#dc3545",
        "due": "22 Jan 2025",
        "priority": "high",
        "category": "Procurement",
        "status": "backlog",
        "description": "Critical stock shortage. Place emergency order for minimum 15 tons Rebar 10mm. Obtain 3 quotes and submit PO.",
        "subtasks": [{"text": "Get 3 supplier quotes", "done": True}, {"text": "Submit PO for approval", "done": False}, {"text": "Confirm delivery date", "done": False}],
        "comments": [{"author": "Inventory System", "initials": "IS", "text": "Stock at critical level (2.5 tons remaining)", "time": "Yesterday"}],
        "created": "21 Jan 2025",
    },
    {
        "id": 4,
        "title": "Pour Concrete Slab B2",
        "assignee": "Khalid Nasser",
        "initials": "KN",
        "avatar_color": "#28a745",
        "due": "24 Jan 2025",
        "priority": "high",
        "category": "Civil",
        "status": "inprogress",
        "description": "Concrete pour for Slab B2. Ready-mix ordered. Ensure formwork inspection complete. Target 480m3 in single pour.",
        "subtasks": [{"text": "Formwork final inspection", "done": True}, {"text": "Rebar inspection sign-off", "done": True}, {"text": "Concrete pour execution", "done": False}, {"text": "Curing compound application", "done": False}],
        "comments": [
            {"author": "Khalid Nasser", "initials": "KN", "text": "Formwork inspection passed", "time": "3 hours ago"},
            {"author": "Ahmed Al-Otaibi", "initials": "AO", "text": "Rebar signed off by consultant", "time": "2 hours ago"},
        ],
        "created": "19 Jan 2025",
    },
    {
        "id": 5,
        "title": "Install Electrical Conduit Level 3",
        "assignee": "Tariq Saad",
        "initials": "TS",
        "avatar_color": "#6f42c1",
        "due": "28 Jan 2025",
        "priority": "medium",
        "category": "MEP",
        "status": "inprogress",
        "description": "Install PVC conduit runs for Level 3 electrical works as per approved drawings. Coordinate with civil for slab penetrations.",
        "subtasks": [{"text": "Mark conduit routes", "done": True}, {"text": "Install main runs", "done": False}, {"text": "Slab penetration sleeves", "done": False}, {"text": "Inspection request", "done": False}],
        "comments": [],
        "created": "20 Jan 2025",
    },
    {
        "id": 6,
        "title": "Update BOQ Package 3",
        "assignee": "Ahmed Al-Otaibi",
        "initials": "AO",
        "avatar_color": "#17a2b8",
        "due": "30 Jan 2025",
        "priority": "low",
        "category": "Admin",
        "status": "inprogress",
        "description": "Update BOQ Package 3 with variation orders VO-012 and VO-013. Recalculate totals and resubmit.",
        "subtasks": [{"text": "Incorporate VO-012", "done": True}, {"text": "Incorporate VO-013", "done": False}, {"text": "Recalculate totals", "done": False}, {"text": "Submit for approval", "done": False}],
        "comments": [{"author": "Ahmed Al-Otaibi", "initials": "AO", "text": "VOs received from client on 21 Jan", "time": "Yesterday"}],
        "created": "21 Jan 2025",
    },
    {
        "id": 7,
        "title": "Waterproofing Inspection - Roof",
        "assignee": "Sara Al-Ghamdi",
        "initials": "SG",
        "avatar_color": "#e8c547",
        "due": "20 Jan 2025",
        "priority": "high",
        "category": "QA/QC",
        "status": "review",
        "description": "Third-party waterproofing inspection on roof slab. Witness flood test and document results. Issue NCR if needed.",
        "subtasks": [{"text": "Witness flood test", "done": True}, {"text": "Check membrane laps", "done": True}, {"text": "Document findings", "done": True}, {"text": "Issue inspection report", "done": False}],
        "comments": [{"author": "Sara Al-Ghamdi", "initials": "SG", "text": "Flood test passed. Minor lap defect at grid F4", "time": "Yesterday"}],
        "created": "18 Jan 2025",
    },
    {
        "id": 8,
        "title": "Subcontractor Invoice Review Q4",
        "assignee": "Mohammed Al-Rashidi",
        "initials": "MR",
        "avatar_color": "#dc3545",
        "due": "21 Jan 2025",
        "priority": "medium",
        "category": "Finance",
        "status": "review",
        "description": "Review and certify Q4 subcontractor invoices for 3 packages. Total value approx SAR 2.3M.",
        "subtasks": [{"text": "Review civil invoices", "done": True}, {"text": "Review MEP invoices", "done": True}, {"text": "Review finishing invoices", "done": False}, {"text": "Submit certificate", "done": False}],
        "comments": [],
        "created": "17 Jan 2025",
    },
    {
        "id": 9,
        "title": "Site Clearance Zone A",
        "assignee": "Ali Hassan",
        "initials": "AH",
        "avatar_color": "#0a0a0a",
        "due": "18 Jan 2025",
        "priority": "medium",
        "category": "Civil",
        "status": "done",
        "description": "Complete site clearance, grubbing, and removal of existing structures in Zone A. Area: 2500m2.",
        "subtasks": [{"text": "Vegetation removal", "done": True}, {"text": "Structure demolition", "done": True}, {"text": "Debris removal", "done": True}, {"text": "Area handover", "done": True}],
        "comments": [{"author": "Ali Hassan", "initials": "AH", "text": "Completed on time. Area handed over to civil team.", "time": "6 days ago"}],
        "created": "14 Jan 2025",
    },
    {
        "id": 10,
        "title": "DPR Week 3 Submission",
        "assignee": "Khalid Nasser",
        "initials": "KN",
        "avatar_color": "#28a745",
        "due": "19 Jan 2025",
        "priority": "low",
        "category": "Admin",
        "status": "done",
        "description": "Compile and submit Week 3 Daily Progress Reports to client portal.",
        "subtasks": [{"text": "Compile all 6 DPRs", "done": True}, {"text": "Add photo log", "done": True}, {"text": "Submit to client portal", "done": True}],
        "comments": [{"author": "Khalid Nasser", "initials": "KN", "text": "Submitted on 19 Jan. Client acknowledged.", "time": "5 days ago"}],
        "created": "16 Jan 2025",
    },
]

TEAM_MEMBERS = [
    {"name": "Mohammed Al-Rashidi", "initials": "MR", "role": "Project Manager", "tasks": 3, "availability": "available", "color": "#dc3545"},
    {"name": "Ahmed Al-Otaibi", "initials": "AO", "role": "Site Engineer", "tasks": 2, "availability": "busy", "color": "#17a2b8"},
    {"name": "Sara Al-Ghamdi", "initials": "SG", "role": "Safety Officer", "tasks": 2, "availability": "available", "color": "#e8c547"},
    {"name": "Khalid Nasser", "initials": "KN", "role": "Civil Engineer", "tasks": 2, "availability": "onsite", "color": "#28a745"},
    {"name": "Tariq Saad", "initials": "TS", "role": "MEP Engineer", "tasks": 1, "availability": "available", "color": "#6f42c1"},
]

NOTIFICATIONS = [
    {
        "id": 1,
        "type": "alert",
        "icon": "fa-triangle-exclamation",
        "color": "danger",
        "title": "Low Stock Alert",
        "title_ar": "تنبيه انخفاض المخزون",
        "message": "Rebar 10mm is critically low (2.5 tons remaining)",
        "message_ar": "حديد التسليح 10مم منخفض بشكل حرج (المتبقي 2.5 طن)",
        "time": "2 hours ago",
        "time_ar": "منذ ساعتين",
        "read": False,
        "link": "/dashboard/inventory",
    },
    {
        "id": 2,
        "type": "task",
        "icon": "fa-tasks",
        "color": "warning",
        "title": "Task Due Today",
        "title_ar": "مهمة مستحقة اليوم",
        "message": "Survey Foundation Zone C is due today",
        "message_ar": "مهمة مسح أساسات المنطقة C مستحقة اليوم",
        "time": "3 hours ago",
        "time_ar": "منذ 3 ساعات",
        "read": False,
        "link": "/dashboard/tasks",
    },
    {
        "id": 3,
        "type": "report",
        "icon": "fa-clipboard-list",
        "color": "primary",
        "title": "DPR Pending",
        "title_ar": "تقرير DPR معلّق",
        "message": "Daily Progress Report not yet submitted for today",
        "message_ar": "لم يتم إرسال تقرير التقدم اليومي لليوم بعد",
        "time": "5 hours ago",
        "time_ar": "منذ 5 ساعات",
        "read": False,
        "link": "/dashboard/dpr/new",
    },
    {
        "id": 4,
        "type": "success",
        "icon": "fa-circle-check",
        "color": "success",
        "title": "Report Generated",
        "title_ar": "تم إنشاء التقرير",
        "message": "BOQ Package 3 was successfully exported to PDF",
        "message_ar": "تم تصدير حزمة BOQ رقم 3 إلى PDF بنجاح",
        "time": "Yesterday, 4:30 PM",
        "time_ar": "أمس، 4:30 مساءً",
        "read": True,
        "link": "/dashboard/boq",
    },
    {
        "id": 5,
        "type": "info",
        "icon": "fa-user-plus",
        "color": "info",
        "title": "Welcome to BanaaIQ",
        "title_ar": "مرحباً بك في BanaaIQ",
        "message": "Your account is set up. Start by creating your first project.",
        "message_ar": "تم إعداد حسابك. ابدأ بإنشاء مشروعك الأول.",
        "time": "Yesterday, 9:00 AM",
        "time_ar": "أمس، 9:00 صباحاً",
        "read": True,
        "link": "/dashboard",
    },
]

DEMO_DPR_DETAILS = {
    1: {
        "notes": "Foundation reinforcement works were completed and inspected. Concrete placement in Zone A reached planned volume.",
        "issues": "Minor delay in rebar delivery, recovered by end of day.",
        "workers": [
            {"name": "Mohammed Al-Rashidi", "role": "Site Engineer", "hours": 9, "present": True},
            {"name": "Ahmed Al-Otaibi", "role": "Civil Foreman", "hours": 9, "present": True},
            {"name": "Rami Nasser", "role": "Safety Officer", "hours": 8, "present": True},
        ],
        "ai_summary": "ENGLISH SUMMARY:\nFoundation works progressed as planned with stable labor productivity.\n\nKEY INSIGHT:\nCrew allocation is aligned with current scope.\n\nARABIC SUMMARY:\nتقدمت أعمال الأساسات وفق الخطة مع إنتاجية مستقرة.",
    },
    2: {
        "notes": "Column shuttering and steel fixing advanced across Zone B.",
        "issues": "High temperature required additional hydration breaks.",
        "workers": [
            {"name": "Faisal Al-Qahtani", "role": "Site Engineer", "hours": 9, "present": True},
            {"name": "Salem Al-Dossari", "role": "Civil Foreman", "hours": 9, "present": True},
            {"name": "Imran Malik", "role": "Bar Bender", "hours": 8, "present": True},
        ],
        "ai_summary": "Column activities closed planned segments with controlled heat-risk measures.",
    },
    3: {
        "notes": "Earthworks in Section 3 focused on grading and compaction.",
        "issues": "Dust reduced visibility and slowed truck cycles.",
        "workers": [
            {"name": "Nawaf Al-Harbi", "role": "Earthworks Supervisor", "hours": 8, "present": True},
            {"name": "Ali Khan", "role": "Machine Operator", "hours": 8, "present": True},
            {"name": "Yousef Al-Anazi", "role": "Surveyor", "hours": 7, "present": True},
        ],
        "ai_summary": "",
    },
    4: {
        "notes": "Slab pour preparation and formwork checks were finalized.",
        "issues": "No major issues reported.",
        "workers": [
            {"name": "Mohammed Al-Rashidi", "role": "Site Engineer", "hours": 9, "present": True},
            {"name": "Ahmed Al-Otaibi", "role": "Civil Foreman", "hours": 9, "present": True},
            {"name": "Ravi Kumar", "role": "Concrete Supervisor", "hours": 8, "present": True},
        ],
        "ai_summary": "Slab pour activities were delivered per sequence with stable performance.",
    },
    5: {
        "notes": "MEP rough-in layout checks initiated in Block 4.",
        "issues": "Pending vendor confirmation for one cable tray batch.",
        "workers": [
            {"name": "Hassan Al-Shehri", "role": "MEP Engineer", "hours": 8, "present": True},
            {"name": "Rodel Santos", "role": "Electrical Foreman", "hours": 8, "present": True},
            {"name": "Ibrahim Al-Ghamdi", "role": "Store Keeper", "hours": 7, "present": True},
        ],
        "ai_summary": "",
    },
}


class ContactForm(FlaskForm):
    name = StringField("Name", validators=[DataRequired(), Length(max=100)])
    company = StringField("Company", validators=[DataRequired(), Length(max=120)])
    role = SelectField("Role", choices=[(r, r) for r in ROLE_CHOICES], validators=[DataRequired()])
    email = StringField("Email", validators=[DataRequired(), Length(max=120)])
    phone = TelField("Phone", validators=[DataRequired(), Length(max=30)], default="+966")
    message = TextAreaField("Message", validators=[DataRequired(), Length(min=20, max=1500)])
    interest = SelectField("Interest", choices=[("Demo", "Demo"), ("Partnership", "Partnership"), ("Acquisition", "Acquisition"), ("Hiring", "Hiring")], validators=[DataRequired()])
    submit = SubmitField("Send Inquiry")


class LoginForm(FlaskForm):
    email = StringField("Email", validators=[DataRequired(), Length(max=120)])
    password = PasswordField("Password", validators=[DataRequired(), Length(min=8)])
    submit = SubmitField("Login")


REGISTER_COUNTRY_CHOICES = [
    ("IN",    "India"),
    ("SA",    "Saudi Arabia"),
    ("AE",    "United Arab Emirates"),
    ("KW",    "Kuwait"),
    ("QA",    "Qatar"),
    ("BH",    "Bahrain"),
    ("OM",    "Oman"),
    ("OTHER", "Other"),
]


class RegisterForm(FlaskForm):
    full_name = StringField("Full Name", validators=[DataRequired(), Length(max=100)])
    company = StringField("Company Name", validators=[DataRequired(), Length(max=100)])
    # account_type: access-control level stored in User.role
    account_type = SelectField(
        "Account Type",
        choices=REGISTER_ACCOUNT_TYPE_CHOICES,
        validators=[DataRequired()],
    )
    # job_title: professional designation stored in User.job_title (required for site_engineer)
    job_title = SelectField(
        "Job Title",
        choices=REGISTER_JOB_TITLE_CHOICES,
        validators=[],
    )
    country = SelectField(
        "Country",
        choices=REGISTER_COUNTRY_CHOICES,
        default="IN",
        validators=[DataRequired()],
    )
    email = StringField("Email", validators=[DataRequired(), Length(max=120)])
    password = PasswordField("Password", validators=[DataRequired(), Length(min=8)])
    confirm_password = PasswordField("Confirm Password", validators=[DataRequired(), Length(min=8)])
    terms = BooleanField("I agree to the Terms of Use and Privacy Policy")
    submit = SubmitField("Create Account")


class ForgotForm(FlaskForm):
    email = StringField("Email", validators=[DataRequired(), Length(max=120)])
    submit = SubmitField("Send Reset Link")


def gregorian_to_hijri(dt):
    y, m, d = dt.year, dt.month, dt.day
    a = (14 - m) // 12
    y2 = y + 4800 - a
    m2 = m + 12 * a - 3
    jd = d + ((153 * m2 + 2) // 5) + 365 * y2 + (y2 // 4) - (y2 // 100) + (y2 // 400) - 32045
    l = jd - 1948440 + 10632
    n = (l - 1) // 10631
    l = l - 10631 * n + 354
    j = ((10985 - l) // 5316) * ((50 * l) // 17719) + (l // 5670) * ((43 * l) // 15238)
    l = l - ((30 - j) // 15) * ((17719 * j) // 50) - (j // 16) * ((15238 * j) // 43) + 29
    hm = (24 * l) // 709
    hd = l - (709 * hm) // 24
    hy = 30 * n + j - 30
    return hy, hm, hd


def parse_demo_date(text):
    if isinstance(text, datetime):
        return text.date()
    if isinstance(text, date):
        return text
    if not text:
        return None
    for fmt in ("%d %b %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(text), fmt).date()
        except (TypeError, ValueError):
            continue
    return None


def format_display_date(value):
    parsed = parse_demo_date(value)
    return parsed.strftime("%d %b %Y") if parsed else ""


def format_html_date(value):
    parsed = parse_demo_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else ""


def to_decimal(value, quantum):
    raw_value = 0 if value in (None, "") else value
    try:
        return Decimal(str(raw_value)).quantize(quantum, rounding=ROUND_HALF_UP)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0").quantize(quantum, rounding=ROUND_HALF_UP)


def to_currency_decimal(value):
    return to_decimal(value, DECIMAL_CURRENCY)


def to_quantity_decimal(value):
    return to_decimal(value, DECIMAL_QUANTITY)


def map_country_code_to_currency(country_code):
    country_code = str(country_code or "").upper()
    if country_code == "IN":
        return "INR"
    if country_code in GCC_COUNTRY_CODES:
        return "SAR"
    return "USD"


def build_currency_context(currency_code, source="detected", country_code=None, remote_addr=None):
    metadata = CURRENCY_METADATA.get(currency_code, CURRENCY_METADATA["SAR"]).copy()
    metadata.update(
        source=source,
        country_code=country_code,
        ip=remote_addr,
        options=[{"code": item["code"], "label": item["selector_label"]} for item in CURRENCY_METADATA.values()],
    )
    return metadata


def detect_user_currency():
    remote_addr = (request.remote_addr or "").strip()
    cached = session.get("detected_currency")

    if isinstance(cached, dict) and cached.get("ip") == remote_addr and cached.get("code") in CURRENCY_METADATA:
        return cached

    if not remote_addr or remote_addr in LOCALHOST_IPS:
        detected = {
            "code": "SAR",
            "source": "localhost",
            "country_code": "SA",
            "ip": remote_addr or "127.0.0.1",
        }
        session["detected_currency"] = detected
        session.modified = True
        return detected

    # Default to INR — India is the primary launch market
    detected = {
        "code": "INR",
        "source": "fallback",
        "country_code": None,
        "ip": remote_addr,
    }
    try:
        response = requests.get(f"http://ip-api.com/json/{remote_addr}", timeout=2)
        if response.ok:
            payload = response.json() or {}
            if str(payload.get("status") or "").lower() == "success":
                country_code = str(payload.get("countryCode") or "").upper()
                detected["country_code"] = country_code
                detected["code"] = map_country_code_to_currency(country_code)
                detected["source"] = "detected"
    except requests.RequestException as error:
        app.logger.warning("Currency detection failed for %s: %s", remote_addr, error)

    session["detected_currency"] = detected
    session.modified = True
    return detected


def get_active_currency():
    selected_currency = str(session.get("selected_currency") or "").upper()
    if selected_currency in CURRENCY_METADATA:
        return build_currency_context(selected_currency, source="manual")
    # If logged-in user has a preferred currency, honour it
    if current_user and getattr(current_user, "is_authenticated", False) and not getattr(current_user, "is_guest", False):
        user_pref = getattr(current_user, "preferred_currency", None)
        if user_pref and user_pref.upper() in CURRENCY_METADATA:
            return build_currency_context(user_pref.upper(), source="user_preference")
    detected = detect_user_currency()
    return build_currency_context(
        detected.get("code", "INR"),
        source=detected.get("source", "detected"),
        country_code=detected.get("country_code"),
        remote_addr=detected.get("ip"),
    )


def convert_sar_amount(value, currency_code):
    if value is None:
        return None

    amount = to_currency_decimal(value)
    currency_code = str(currency_code or "SAR").upper()
    rate = CURRENCY_RATES.get(currency_code, CURRENCY_RATES["SAR"])
    converted = amount * rate

    if currency_code == "INR":
        return (converted / Decimal("10")).quantize(Decimal("1"), rounding=ROUND_HALF_UP) * Decimal("10")
    if currency_code == "USD":
        return converted.quantize(Decimal("1"), rounding=ROUND_HALF_UP)
    return converted.quantize(DECIMAL_CURRENCY, rounding=ROUND_HALF_UP)


def format_currency_number_from_sar(value, currency_code, decimals=None):
    converted = convert_sar_amount(value, currency_code)
    if converted is None:
        return None

    if decimals is None:
        decimals = 2 if str(currency_code or "SAR").upper() == "SAR" else 0
    return f"{float(converted):,.{decimals}f}"


def format_currency_from_sar(value, currency_code, decimals=None):
    formatted_value = format_currency_number_from_sar(value, currency_code, decimals=decimals)
    if formatted_value is None:
        return "Custom"
    currency = CURRENCY_METADATA.get(str(currency_code or "SAR").upper(), CURRENCY_METADATA["SAR"])
    return f"{currency['symbol']} {formatted_value}"


@app.context_processor
def inject_currency_helpers():
    return {"format_currency_from_sar": format_currency_from_sar}


def get_dpr_records():
    if is_demo_mode():
        return list(DPR_RECORDS)
    if current_user.is_authenticated:
        return [d.to_dict() for d in DPR.query.filter_by(user_id=current_user.id).order_by(DPR.created_at.desc()).all()]
    return []


def get_dpr_record(record_id):
    if is_demo_mode():
        for record in DPR_RECORDS:
            if record["id"] == record_id:
                return record
        return None
    if current_user.is_authenticated:
        dpr = DPR.query.filter_by(id=record_id, user_id=current_user.id).first()
        return dpr.to_dict() if dpr else None
    return None


def get_dpr_detail(record_id):
    if is_demo_mode():
        return DEMO_DPR_DETAILS.get(record_id, {"notes": "", "issues": "", "workers": [], "ai_summary": ""})
    if current_user.is_authenticated:
        dpr = DPR.query.filter_by(id=record_id, user_id=current_user.id).first()
        if dpr:
            return {
                "notes": dpr.progress_notes or "",
                "progress_notes": dpr.progress_notes or "",
                "issues": dpr.issues or "",
                "workers": dpr.workers,
                "ai_summary": dpr.ai_summary or "",
            }
    return {"notes": "", "issues": "", "workers": [], "ai_summary": ""}


def is_demo_mode():
    return bool(session.get("is_guest"))


def get_user_projects():
    if current_user.is_authenticated and not is_demo_mode():
        if getattr(current_user, "role", None) == ROLE_PROJECT_MANAGER:
            return Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()
        assigned_ids = {
            row.project_id
            for row in ProjectAssignment.query.filter_by(user_id=current_user.id).all()
        }
        package_project_ids = {
            row.project_id
            for row in EngineerPackage.query.filter_by(assigned_user_id=current_user.id).all()
            if row.project_id
        }
        project_ids = sorted(assigned_ids | package_project_ids)
        if not project_ids:
            return []
        return Project.query.filter(Project.id.in_(project_ids)).order_by(Project.name).all()
    return []


def get_feature_projects(feature):
    if session.get("is_guest"):
        return []
    return FeatureProject.query.filter_by(user_id=current_user.id, feature=feature).order_by(FeatureProject.name).all()


def get_selected_feature_project(feature, selected_project_id):
    if is_demo_mode():
        return None
    if not selected_project_id or str(selected_project_id) == "all":
        return None
    try:
        project_id = int(selected_project_id)
    except (TypeError, ValueError):
        return None
    project = FeatureProject.query.filter_by(id=project_id, user_id=current_user.id, feature=feature).first()
    return project


def resolve_feature_project(feature, selected_project_id):
    project = get_selected_feature_project(feature, selected_project_id)
    if project:
        return str(project.id), project
    return "all", None


def get_selected_feature_project_name(feature, selected_project_id, fallback="All Projects"):
    project = get_selected_feature_project(feature, selected_project_id)
    return project.name if project else fallback


def get_selected_project(selected_project_id):
    if is_demo_mode():
        return None
    if not selected_project_id or str(selected_project_id) == "all":
        return None
    try:
        project_id = int(selected_project_id)
    except (TypeError, ValueError):
        return None
    project = db.session.get(Project, project_id)
    if not project:
        return None
    if getattr(current_user, "role", None) == ROLE_PROJECT_MANAGER:
        return project if project.user_id == current_user.id else None
    if getattr(current_user, "role", None) == ROLE_SITE_ENGINEER:
        assigned = ProjectAssignment.query.filter_by(
            project_id=project.id,
            user_id=current_user.id,
        ).first()
        packaged = EngineerPackage.query.filter_by(
            project_id=project.id,
            assigned_user_id=current_user.id,
        ).first()
        return project if assigned or packaged else None
    if project.user_id != current_user.id:
        return None
    return project


def resolve_selected_project(selected_project_id):
    project = get_selected_project(selected_project_id)
    if project:
        return str(project.id), project
    return "all", None


def get_selected_project_name(selected_project_id, fallback="All Projects"):
    project = get_selected_project(selected_project_id)
    return project.name if project else fallback


def get_inventory_items_for_user():
    if is_demo_mode():
        return [dict(item) for item in INVENTORY_ITEMS]
    if current_user.is_authenticated:
        return [item.to_dict() for item in InventoryItem.query.filter_by(user_id=current_user.id).order_by(InventoryItem.created_at.desc()).all()]
    return []


def save_inventory_items_for_user(items):
    # No-op now; inventory is persisted in database.
    return items


def compute_inventory_stats(items):
    total_items = len(items)
    low_stock = len([i for i in items if i.get("status") == "low"])
    critical = len([i for i in items if i.get("status") == "critical"])
    total_value = sum(float(i.get("value_sar", 0) or 0) for i in items)
    return total_items, low_stock, critical, total_value


def get_tasks_for_user(feature_project_id=None):
    # Legacy stub — tasks are now project-scoped; use task index route instead
    return []


def get_or_create_task_project(project_name):
    safe_name = sanitize_input(project_name or "General", 200).strip() or "General"
    if safe_name == "General":
        return None, safe_name
    project = FeatureProject.query.filter_by(
        user_id=current_user.id,
        name=safe_name,
        feature="tasks",
    ).first()
    if not project:
        project = FeatureProject(
            user_id=current_user.id,
            name=safe_name,
            color="#0a0a0a",
            feature="tasks",
        )
        db.session.add(project)
        db.session.flush()
    return project, safe_name


def create_task_log(task, action, details, field_changed="", old_value="", new_value=""):
    """Write a TaskActivity entry (formerly TaskLog — retained for call-site compat)."""
    if not current_user.is_authenticated or session.get("is_guest"):
        return
    task_id = task.id if hasattr(task, "id") else int(task)
    detail_str = details or ""
    if field_changed and (old_value or new_value):
        detail_str = f"{details} [{field_changed}: {old_value} → {new_value}]"
    db.session.add(
        TaskActivity(
            task_id=task_id,
            actor_id=current_user.id,
            action=action[:40],
            details=detail_str[:2000] if detail_str else None,
        )
    )


def task_language_label(code):
    labels = {
        "ar": "Arabic",
        "ur": "Urdu",
        "hi": "Hindi",
        "en": "English",
        "fr": "French",
        "zh": "Chinese",
    }
    return labels.get((code or "en").lower(), (code or "English").upper())


def save_tasks_for_user(tasks):
    # No-op now; tasks are persisted in database.
    return tasks


def build_kanban(tasks_list):
    return {
        "backlog": [t for t in tasks_list if t.get("status") == "backlog"],
        "inprogress": [t for t in tasks_list if t.get("status") == "inprogress"],
        "review": [t for t in tasks_list if t.get("status") == "review"],
        "done": [t for t in tasks_list if t.get("status") == "done"],
    }


PROJECT_TYPE_OPTIONS = [
    ("villa", "Villa"),
    ("commercial", "Commercial"),
    ("infrastructure", "Infrastructure"),
    ("industrial", "Industrial"),
    ("residential", "Residential"),
    ("mixed-use", "Mixed-Use"),
]

PROJECT_STATUS_OPTIONS = [
    ("active", "Active"),
    ("on_hold", "On Hold"),
    ("completed", "Completed"),
    ("cancelled", "Cancelled"),
]

TRADE_OPTIONS = ["Civil", "MEP", "Finishing", "Structure", "Other"]

TASK_STATUS_LABELS = {
    "not_started": "Not Started",
    "in_progress": "In Progress",
    "review": "Review",
    "done": "Done",
}

ENGINEER_ITEM_STATUS_LABELS = {
    "not_started": "Not Started",
    "in_progress": "In Progress",
    "complete": "Complete",
}

TRADE_MATERIAL_KEYWORDS = {
    "civil": ["concrete", "rebar", "steel", "cement", "block", "masonry", "formwork", "excavation", "aggregate", "sand"],
    "mep": ["electrical", "cable", "conduit", "pipe", "plumbing", "hvac", "chilled", "fire", "mechanical", "duct"],
    "finishing": ["tile", "paint", "ceiling", "gypsum", "door", "glazing", "aluminum", "waterproof", "floor", "finish"],
    "structure": ["concrete", "rebar", "steel", "formwork", "column", "beam", "slab", "foundation", "shuttering"],
}


def parse_date_field(field_name):
    raw_value = sanitize_input(request.form.get(field_name, ""), 20).strip()
    if not raw_value:
        return None
    try:
        return datetime.strptime(raw_value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_decimal_field(field_name):
    raw_value = sanitize_input(request.form.get(field_name, ""), 40).replace(",", "").strip()
    if not raw_value:
        return None
    return to_currency_decimal(raw_value)


def project_status_label(status):
    lookup = dict(PROJECT_STATUS_OPTIONS)
    return lookup.get(status or "active", "Active")


def project_type_label(project_type):
    lookup = dict(PROJECT_TYPE_OPTIONS)
    return lookup.get(project_type or "", (project_type or "Project").replace("_", " ").title())


def get_project_or_403(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    # PM: must be the project creator.
    if current_user.role == ROLE_PROJECT_MANAGER:
        if project.user_id != current_user.id:
            abort(403)
        return project
    # SE: must have an EngineerPackage assignment OR a ProjectAssignment on this project.
    has_package = EngineerPackage.query.filter_by(
        project_id=project.id,
        assigned_user_id=current_user.id,
    ).first()
    if has_package:
        return project
    has_assignment = ProjectAssignment.query.filter_by(
        project_id=project.id,
        user_id=current_user.id,
    ).first()
    if not has_assignment:
        abort(403)
    return project


def get_package_or_403(package_id, allow_owner=True):
    package = db.session.get(EngineerPackage, package_id)
    if not package:
        abort(404)
    is_assigned_user = package.assigned_user_id == current_user.id
    is_owner = allow_owner and package.project and package.project.user_id == current_user.id
    if not (is_assigned_user or is_owner):
        abort(403)
    return package


def project_timeline_progress(project):
    if not project.start_date or not project.planned_completion:
        return 0
    total_days = max((project.planned_completion - project.start_date).days, 1)
    elapsed_days = (date.today() - project.start_date).days
    return max(0, min(100, round((elapsed_days / total_days) * 100)))


def milestone_date_at(project, fraction):
    if not project.start_date or not project.planned_completion:
        return project.start_date or project.planned_completion
    duration_days = max((project.planned_completion - project.start_date).days, 0)
    return project.start_date + timedelta(days=round(duration_days * fraction))


def create_default_project_milestones(project):
    defaults = [
        ("Project Kickoff", project.start_date),
        ("Design Approval", milestone_date_at(project, 0.20)),
        ("Structure Complete", milestone_date_at(project, 0.50)),
        ("MEP Complete", milestone_date_at(project, 0.75)),
        ("Handover", project.planned_completion),
    ]
    for title, planned_date in defaults:
        db.session.add(
            ProjectMilestone(
                project=project,
                title=title,
                planned_date=planned_date,
                status="pending",
            )
        )


def sync_milestone_statuses(project):
    today = date.today()
    for milestone in project.milestones:
        if milestone.actual_date:
            milestone.status = "completed"
        elif milestone.planned_date and milestone.planned_date < today and project.status != "completed":
            milestone.status = "overdue"
        elif milestone.status != "completed":
            milestone.status = "pending"


def project_form_data(project=None):
    return {
        "name": sanitize_input(request.form.get("name", ""), 200).strip(),
        "reference_number": sanitize_input(request.form.get("reference_number", ""), 100).strip(),
        "project_type": sanitize_input(request.form.get("project_type", ""), 50).strip(),
        "location_city": sanitize_input(request.form.get("location_city", ""), 100).strip(),
        "location_zone": sanitize_input(request.form.get("location_zone", ""), 100).strip(),
        "client_name": sanitize_input(request.form.get("client_name", ""), 200).strip(),
        "client_contact": sanitize_input(request.form.get("client_contact", ""), 200).strip(),
        "contract_value": parse_decimal_field("contract_value"),
        "start_date": parse_date_field("start_date"),
        "planned_completion": parse_date_field("planned_completion"),
        "lead_engineer": sanitize_input(request.form.get("lead_engineer", ""), 200).strip(),
        "project_manager": sanitize_input(request.form.get("project_manager", ""), 200).strip(),
        "status": sanitize_input(request.form.get("status", getattr(project, "status", "active")), 20).strip() or "active",
    }


def apply_project_form(project, form_data):
    project.name = form_data["name"]
    project.reference_number = form_data["reference_number"] or None
    project.project_type = form_data["project_type"] or None
    project.location_city = form_data["location_city"] or None
    project.location_zone = form_data["location_zone"] or None
    project.client_name = form_data["client_name"] or None
    project.client_contact = form_data["client_contact"] or None
    project.contract_value = form_data["contract_value"]
    project.currency = "SAR"
    project.start_date = form_data["start_date"]
    project.planned_completion = form_data["planned_completion"]
    project.lead_engineer = form_data["lead_engineer"] or None
    project.project_manager = form_data["project_manager"] or None
    if form_data["status"] in dict(PROJECT_STATUS_OPTIONS):
        project.status = form_data["status"]
    project.health_score = project.health_score if project.health_score is not None else 100
    project.health_status = project.health_status or "green"
    project.health_summary = project.health_summary or "Project workspace created. Link DPRs, BOQs, tasks, and inventory to begin live health tracking."


def add_custom_project_milestones(project):
    titles = request.form.getlist("custom_milestone_title")
    dates = request.form.getlist("custom_milestone_date")
    for index, title in enumerate(titles):
        safe_title = sanitize_input(title, 200).strip()
        if not safe_title:
            continue
        planned_date = None
        if index < len(dates) and dates[index]:
            try:
                planned_date = datetime.strptime(dates[index], "%Y-%m-%d").date()
            except ValueError:
                planned_date = None
        db.session.add(ProjectMilestone(project=project, title=safe_title, planned_date=planned_date, status="pending"))


def project_workspace_data(project):
    sync_milestone_statuses(project)
    dprs = DPR.query.filter_by(user_id=current_user.id, project_id=project.id).order_by(DPR.date.desc(), DPR.created_at.desc()).all()
    boqs = BOQ.query.filter_by(user_id=current_user.id, project_id=project.id).order_by(BOQ.created_at.desc()).all()
    tasks = Task.query.filter_by(project_id=project.id).order_by(Task.created_at.desc()).all()
    inventory_items = InventoryItem.query.filter_by(user_id=current_user.id, project_id=project.id).order_by(InventoryItem.updated_at.desc()).all()
    packages = EngineerPackage.query.filter_by(project_id=project.id).order_by(EngineerPackage.created_at.desc()).all()
    total_boq_value = sum(float(boq.grand_total or 0) for boq in boqs)
    total_inventory_value = sum(float(item.value_sar or 0) for item in inventory_items)
    # Build workspace kanban using new status values
    _ws_kanban = {"not_started": [], "in_progress": [], "review": [], "done": []}
    for _t in tasks:
        _col = _t.status if _t.status in _ws_kanban else "not_started"
        _ws_kanban[_col].append(_t)
    kanban = _ws_kanban
    # Build master inventory batch summary for the workspace inventory tab
    batch_map = {}
    for item in inventory_items:
        bid = item.master_inventory_batch_id
        if not bid:
            continue
        if bid not in batch_map:
            batch_map[bid] = {"batch_id": bid, "item_count": 0, "total_sar": 0.0, "created_at": item.created_at}
        batch_map[bid]["item_count"] += 1
        batch_map[bid]["total_sar"] += float(item.value_sar or 0) * float(item.stock or 0)
    inventory_batches = sorted(batch_map.values(), key=lambda b: b["created_at"] or "", reverse=True)

    # ── Rich overview stats ────────────────────────────────────────────────────
    _today = date.today()
    _week_ago = _today - timedelta(days=7)

    # DPR stats
    _dpr_dates = [d.date for d in dprs if d.date]
    _dpr_this_week = sum(1 for d in _dpr_dates if d >= _week_ago)
    _dpr_last_date = max(_dpr_dates) if _dpr_dates else None

    # BOQ stats
    _boq_count = len(boqs)
    _task_done = sum(1 for t in tasks if t.status == "done")
    _task_total = len(tasks)
    _boq_completion_pct = round((_task_done / _task_total * 100) if _task_total else 0)

    # Inventory stats
    _inv_alert = sum(1 for i in inventory_items if getattr(i, "status", None) in ("low", "critical"))
    _inv_total_sar = sum(float(i.value_sar or 0) * float(i.stock or 0) for i in inventory_items)

    # Task stats
    _task_status_counts = {s: sum(1 for t in tasks if t.status == s) for s in ("not_started", "in_progress", "review", "done")}
    _task_overdue = sum(1 for t in tasks if t.status != "done" and t.due_date and t.due_date < _today)

    # Engineers progress (tasks per assignment)
    _engineer_rows = []
    for asgn in project.assignments:
        _eng_tasks = [t for t in tasks if t.assigned_to_id == asgn.user_id]
        _eng_done = sum(1 for t in _eng_tasks if t.status == "done")
        _eng_total = len(_eng_tasks)
        _eng_pct = round(_eng_done / _eng_total * 100) if _eng_total else 0
        _engineer_rows.append({
            "name": asgn.user.full_name,
            "role_on_project": asgn.role_on_project or "",
            "done": _eng_done,
            "total": _eng_total,
            "pct": _eng_pct,
        })

    # Timeline
    _tl_progress = project_timeline_progress(project)
    _tl_elapsed = ((_today - project.start_date).days) if project.start_date else None
    if project.planned_completion and _today <= project.planned_completion:
        _tl_remaining = (project.planned_completion - _today).days
    elif project.planned_completion:
        _tl_remaining = 0
    else:
        _tl_remaining = None

    overview_stats = {
        "dpr_count": len(dprs),
        "dpr_this_week": _dpr_this_week,
        "dpr_last_date": _dpr_last_date,
        "boq_count": _boq_count,
        "boq_total_sar": total_boq_value,
        "boq_completion_pct": _boq_completion_pct,
        "inv_count": len(inventory_items),
        "inv_total_sar": _inv_total_sar,
        "inv_alert_count": _inv_alert,
        "task_counts": _task_status_counts,
        "task_total": _task_total,
        "task_overdue": _task_overdue,
        "engineers": _engineer_rows,
        "timeline_pct": _tl_progress,
        "timeline_elapsed_days": _tl_elapsed,
        "timeline_remaining_days": _tl_remaining,
    }

    return {
        "dprs": dprs,
        "boqs": boqs,
        "tasks": tasks,
        "inventory_items": inventory_items,
        "inventory_batches": inventory_batches,
        "packages": packages,
        "total_boq_value": total_boq_value,
        "total_inventory_value": total_inventory_value,
        "kanban": kanban,
        "stats": {
            "dprs": len(dprs),
            "boq_value": total_boq_value,
            "tasks": len(tasks),
            "inventory": len(inventory_items),
        },
        "overview_stats": overview_stats,
        "timeline_progress": _tl_progress,
    }


def _pdf_escape(value):
    return html.escape(str(value or "-")).replace("\n", "<br/>")


def _fmt_report_date(value):
    if not value:
        return "-"
    if isinstance(value, datetime):
        value = value.date()
    try:
        return value.strftime("%d %b %Y")
    except Exception:
        return str(value)


def _fmt_sar(value):
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        amount = 0.0
    return f"SAR {amount:,.2f}"


def _working_days_between(start, end):
    if not start or not end:
        return 0
    if end < start:
        start, end = end, start
    days = 0
    cursor = start
    while cursor <= end:
        if cursor.weekday() < 5:
            days += 1
        cursor += timedelta(days=1)
    return days


def _project_boqs(project):
    return BOQ.query.filter_by(user_id=project.user_id, project_id=project.id).order_by(BOQ.updated_at.desc(), BOQ.created_at.desc()).all()


def _project_tasks(project):
    return Task.query.filter_by(project_id=project.id).order_by(Task.created_at.desc()).all()


def _project_inventory(project):
    return InventoryItem.query.filter_by(user_id=project.user_id, project_id=project.id).order_by(InventoryItem.updated_at.desc()).all()


def _project_dprs(project):
    return DPR.query.filter_by(user_id=project.user_id, project_id=project.id).order_by(DPR.date.desc(), DPR.created_at.desc()).all()


def _project_boq_actuals(boqs):
    boq_ids = [boq.id for boq in boqs]
    if not boq_ids:
        return []
    return BOQActual.query.filter(BOQActual.boq_id.in_(boq_ids)).all()


def _actual_total_for_project(project):
    boqs = _project_boqs(project)
    actuals_total = sum(float(actual.actual_total_sar or 0) for actual in _project_boq_actuals(boqs))
    if actuals_total > 0:
        return actuals_total
    return sum(float(item.value_sar or 0) for item in _project_inventory(project))


def _project_financial_snapshot(project):
    boqs = _project_boqs(project)
    inventory_items = _project_inventory(project)
    boq_total = sum(float(boq.grand_total or 0) for boq in boqs)
    inventory_total = sum(float(item.value_sar or 0) for item in inventory_items)
    actual_estimate = _actual_total_for_project(project)
    variance = actual_estimate - boq_total
    variance_pct = (variance / boq_total * 100) if boq_total else 0.0
    boq_status = "No BOQ linked"
    if boqs:
        statuses = sorted({boq.status or "Draft" for boq in boqs})
        boq_status = ", ".join(statuses)
    return {
        "boq_total": boq_total,
        "inventory_total": inventory_total,
        "actual_estimate": actual_estimate,
        "variance": variance,
        "variance_pct": variance_pct,
        "boq_status": boq_status,
        "boqs": boqs,
        "inventory_items": inventory_items,
    }


def _major_boq_section(row):
    trade = classify_boq_trade(_boq_item_desc(row), row.get("section", ""), row.get("sheet_name", ""))
    trade_lower = trade.lower()
    if any(token in trade_lower for token in ["mechanical", "electrical", "plumbing", "fire", "elv", "hvac", "ict"]):
        return "MEP"
    if any(token in trade_lower for token in ["architectural", "finish"]):
        return "Finishing"
    return "Civil"


def _boq_vs_actual_sections(project):
    boqs = _project_boqs(project)
    actuals = _project_boq_actuals(boqs)
    actual_map = {}
    for actual in actuals:
        actual_map[(actual.boq_id, str(actual.item_no))] = actual_map.get((actual.boq_id, str(actual.item_no)), 0.0) + float(actual.actual_total_sar or 0)

    sections = {
        "Civil": {"budget": 0.0, "actual": 0.0},
        "MEP": {"budget": 0.0, "actual": 0.0},
        "Finishing": {"budget": 0.0, "actual": 0.0},
    }
    for boq in boqs:
        for idx, row in enumerate(boq.items or [], start=1):
            section = _major_boq_section(row)
            item_no = _boq_item_no(row, idx)
            sections[section]["budget"] += _boq_item_total(row)
            sections[section]["actual"] += actual_map.get((boq.id, str(item_no)), 0.0)

    rows = []
    for section, values in sections.items():
        budget = values["budget"]
        actual = values["actual"]
        variance = actual - budget
        if budget <= 0 and actual <= 0:
            status = "No linked items"
        elif variance <= budget * 0.05:
            status = "On plan"
        elif variance <= budget * 0.15:
            status = "Watch"
        else:
            status = "Over plan"
        rows.append(
            {
                "description": section,
                "budget": budget,
                "actual": actual,
                "variance": variance,
                "status": status,
            }
        )
    return rows


class TimelineBar(Flowable):
    def __init__(self, start_date, planned_date, actual_date, width=16 * cm, height=1.4 * cm):
        super().__init__()
        self.start_date = start_date
        self.planned_date = planned_date
        self.actual_date = actual_date
        self.width = width
        self.height = height

    def draw(self):
        navy = colors.HexColor("#0a0a0a")
        gold = colors.HexColor("#e8c547")
        soft = colors.HexColor("#efebe1")
        y = self.height / 2
        self.canv.setStrokeColor(soft)
        self.canv.setLineWidth(10)
        self.canv.line(0, y, self.width, y)

        total_days = max(((self.planned_date or self.start_date) - (self.start_date or self.planned_date)).days, 1) if self.start_date and self.planned_date else 1
        if self.start_date and self.planned_date and self.actual_date:
            actual_ratio = (self.actual_date - self.start_date).days / total_days
        else:
            actual_ratio = 1
        marker_x = max(0, min(self.width, self.width * actual_ratio))

        self.canv.setStrokeColor(navy)
        self.canv.setLineWidth(10)
        self.canv.line(0, y, self.width, y)
        self.canv.setFillColor(gold)
        self.canv.circle(marker_x, y, 6, fill=1, stroke=0)

        self.canv.setFont("Helvetica", 7)
        self.canv.setFillColor(colors.HexColor("#6b6b6b"))
        self.canv.drawString(0, 2, _fmt_report_date(self.start_date))
        self.canv.drawCentredString(self.width, 2, _fmt_report_date(self.planned_date))
        self.canv.setFillColor(gold)
        self.canv.drawCentredString(marker_x, self.height - 10, f"Actual: {_fmt_report_date(self.actual_date)}")


def _report_styles():
    styles = getSampleStyleSheet()
    arabic_font = "Cairo" if setup_arabic_font() else "Helvetica"
    return {
        "normal": ParagraphStyle("CCNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#1a1a1a")),
        "small": ParagraphStyle("CCSmall", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, leading=9, textColor=colors.HexColor("#6b6b6b")),
        "bold": ParagraphStyle("CCBold", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=8.5, leading=11, textColor=colors.HexColor("#0a0a0a")),
        "title": ParagraphStyle("CCTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, textColor=colors.HexColor("#0a0a0a"), spaceAfter=8),
        "cover_logo": ParagraphStyle("CCCoverLogo", parent=styles["Title"], fontName=arabic_font, fontSize=22, textColor=colors.HexColor("#e8c547"), alignment=TA_CENTER),
        "cover_subtitle": ParagraphStyle("CCCoverSubtitle", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=13, textColor=colors.white, alignment=TA_CENTER),
        "cover_ar": ParagraphStyle("CCCoverAr", parent=styles["BodyText"], fontName=arabic_font, fontSize=13, textColor=colors.HexColor("#e8c547"), alignment=TA_CENTER),
        "cover_project": ParagraphStyle("CCCoverProject", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=28, textColor=colors.white, alignment=TA_CENTER, leading=33),
        "section_en": ParagraphStyle("CCSectionEn", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=12, textColor=colors.white),
        "section_ar": ParagraphStyle("CCSectionAr", parent=styles["BodyText"], fontName=arabic_font, fontSize=11, textColor=colors.HexColor("#e8c547"), alignment=TA_RIGHT),
        "arabic": ParagraphStyle("CCArabic", parent=styles["BodyText"], fontName=arabic_font, fontSize=10, leading=15, alignment=TA_RIGHT, textColor=colors.HexColor("#1a1a1a")),
        "exec": ParagraphStyle("CCExec", parent=styles["BodyText"], fontName="Helvetica", fontSize=10, leading=15, textColor=colors.HexColor("#1a1a1a")),
        "sign": ParagraphStyle("CCSign", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=10, leading=14, textColor=colors.HexColor("#0a0a0a")),
    }


def _pdf_para(value, style):
    return Paragraph(_pdf_escape(value), style)


def _pdf_arabic(value, style):
    return Paragraph(_pdf_escape(process_arabic_text(value)), style)


def _section_header(title, arabic_title, styles):
    table = Table(
        [[Paragraph(_pdf_escape(title), styles["section_en"]), _pdf_arabic(arabic_title, styles["section_ar"])]],
        colWidths=[9 * cm, 9 * cm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0a0a0a")),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor("#e8c547")),
                ("PADDING", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return table


def _styled_pdf_table(rows, col_widths, styles, header=True, extra_styles=None, small=False):
    body_style = styles["small"] if small else styles["normal"]
    bold_style = styles["bold"]
    header_style = ParagraphStyle("CCTableHeader", parent=bold_style, textColor=colors.white, fontSize=7.5 if small else 8.5)
    data = []
    for row_index, row in enumerate(rows):
        row_style = header_style if row_index == 0 and header else body_style
        data.append([cell if hasattr(cell, "wrap") else _pdf_para(cell, row_style) for cell in row])
    table = Table(data, colWidths=col_widths, repeatRows=1 if header else 0)
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e4ddd0")),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1 if header else 0), (-1, -1), [colors.white, colors.HexColor("#fbf9f4")]),
    ]
    if header:
        commands.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0a0a0a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ]
        )
    if extra_styles:
        commands.extend(extra_styles)
    table.setStyle(TableStyle(commands))
    return table


def _summary_box(label, value, styles, accent="#e8c547"):
    table = Table(
        [[_pdf_para(label, styles["small"])], [_pdf_para(value, ParagraphStyle(f"Box{label[:8]}", parent=styles["bold"], fontSize=12, textColor=colors.HexColor("#0a0a0a")))]],
        colWidths=[5.6 * cm],
    )
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fbf9f4")),
                ("BOX", (0, 0), (-1, -1), 0.75, colors.HexColor(accent)),
                ("LINEABOVE", (0, 0), (-1, 0), 3, colors.HexColor(accent)),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    return table


def _completion_report_data(project):
    dprs = _project_dprs(project)
    tasks = _project_tasks(project)
    inventory_items = _project_inventory(project)
    packages = EngineerPackage.query.filter_by(project_id=project.id).order_by(EngineerPackage.created_at.desc()).all()
    financial = _project_financial_snapshot(project)
    completed_tasks = [task for task in tasks if task.status == "done"]
    actual_end = project.actual_completion or date.today()
    working_days = _working_days_between(project.start_date, actual_end)
    dpr_coverage = round((len(dprs) / working_days * 100), 1) if working_days else 0
    days_delta = (actual_end - project.planned_completion).days if project.planned_completion else 0
    return {
        "dprs": dprs,
        "tasks": tasks,
        "completed_tasks": completed_tasks,
        "inventory_items": inventory_items,
        "packages": packages,
        "financial": financial,
        "boq_sections": _boq_vs_actual_sections(project),
        "actual_end": actual_end,
        "working_days": working_days,
        "dpr_coverage": dpr_coverage,
        "days_delta": days_delta,
    }


def _completion_ai_summary(project, data):
    """Return (english, arabic, ai_generated). Caches result in Project.completion_summary."""
    # ── Use cached summary if already set ─────────────────────────────────
    cached_en = getattr(project, "completion_summary", None)
    cached_ar = getattr(project, "completion_summary_ar", None)
    if cached_en and cached_en.strip():
        return cached_en.strip(), (cached_ar or "").strip() or cached_en.strip(), True

    # ── Try AI ────────────────────────────────────────────────────────────
    ai_ok, ai_msg = can_use_ai(current_user)
    if ai_ok:
        start = _fmt_report_date(project.start_date)
        actual_completion = _fmt_report_date(data["actual_end"])
        total_tasks = len(data["tasks"])
        completed_tasks = len(data["completed_tasks"])
        dpr_count = len(data["dprs"])
        prompt = f"""
Write a professional project completion executive summary for a Saudi construction project with these details:
Project: {project.name}
Client: {project.client_name}
Contract Value: SAR {project.contract_value}
Duration: {start} to {actual_completion}
Final Health Score: {project.health_score}/100
DPRs submitted: {dpr_count}
Tasks completed: {completed_tasks}/{total_tasks}

Write 3 sentences in English, then 3 sentences in Arabic. Professional tone.
Suitable for client handover presentation.
Format: English paragraph, then Arabic paragraph.
"""
        success, text = call_openai(prompt, max_tokens=420, temperature=0.35)
        if success:
            record_ai_usage("projects")
            english_lines = []
            arabic_lines = []
            for line in [line.strip() for line in text.splitlines() if line.strip()]:
                cleaned = line.replace("English:", "").replace("Arabic:", "").strip()
                if not cleaned:
                    continue
                if is_arabic_text(cleaned):
                    arabic_lines.append(cleaned)
                else:
                    english_lines.append(cleaned)
            english = " ".join(english_lines).strip()
            arabic = " ".join(arabic_lines).strip()
            if english or arabic:
                en_final = english or f"{project.name} has reached completion with a final health score of {project.health_score}/100."
                ar_final = arabic or "تم إعداد هذا الملخص التنفيذي بواسطة منصة بناء IQ لدعم تسليم المشروع."
                # Persist for future calls
                try:
                    project.completion_summary = en_final
                    project.completion_summary_ar = ar_final
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                return en_final, ar_final, True

    # ── Fallback ─────────────────────────────────────────────────────────
    en_fallback = (
        "Project completed. See detailed sections below."
    )
    ar_fallback = "اكتمل المشروع. راجع الأقسام التالية."
    try:
        project.completion_summary = en_fallback
        project.completion_summary_ar = ar_fallback
        db.session.commit()
    except Exception:
        db.session.rollback()
    return en_fallback, ar_fallback, False


def build_completion_report_pdf(project):
    data = _completion_report_data(project)
    styles = _report_styles()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.45 * cm, leftMargin=1.45 * cm, topMargin=1.35 * cm, bottomMargin=1.2 * cm)
    navy = colors.HexColor("#0a0a0a")
    gold = colors.HexColor("#e8c547")
    today_text = date.today().strftime("%d %B %Y")

    def cover_page(canvas_obj, _doc):
        width, height = A4
        canvas_obj.saveState()
        canvas_obj.setFillColor(navy)
        canvas_obj.rect(0, 0, width, height, fill=1, stroke=0)
        canvas_obj.setFillColor(gold)
        canvas_obj.rect(0, 0, width, 1.15 * cm, fill=1, stroke=0)
        canvas_obj.setFillColor(colors.white)
        canvas_obj.setFont("Helvetica-Bold", 8)
        canvas_obj.drawString(1.4 * cm, 0.42 * cm, f"Prepared by BanaaIQ AI - {today_text}")
        canvas_obj.drawRightString(width - 1.4 * cm, 0.42 * cm, "Confidential - For Client Presentation")
        canvas_obj.restoreState()

    def later_page(canvas_obj, _doc):
        width, height = A4
        canvas_obj.saveState()
        canvas_obj.setFillColor(navy)
        canvas_obj.rect(0, height - 0.55 * cm, width, 0.55 * cm, fill=1, stroke=0)
        canvas_obj.setFillColor(gold)
        canvas_obj.setFont("Helvetica-Bold", 7)
        canvas_obj.drawString(1.45 * cm, height - 0.35 * cm, "BanaaIQ | Project Command Center")
        canvas_obj.drawRightString(width - 1.45 * cm, height - 0.35 * cm, project.project_code or "BIQ-PENDING")
        canvas_obj.setFillColor(colors.HexColor("#6b6b6b"))
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.drawCentredString(width / 2, 0.55 * cm, f"BanaaIQ Completion Report | Page {_doc.page}")
        canvas_obj.restoreState()

    story = []

    story.append(Spacer(1, 2.1 * cm))
    story.append(_pdf_arabic("بناء IQ | BanaaIQ", styles["cover_logo"]))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph("Project Completion Report", styles["cover_subtitle"]))
    story.append(_pdf_arabic("تقرير إتمام المشروع", styles["cover_ar"]))
    story.append(Spacer(1, 2.2 * cm))
    cover_card = Table(
        [
            [_pdf_para(project.name, styles["cover_project"])],
            [_pdf_para(f"{project.project_code or 'BIQ-PENDING'}", ParagraphStyle("CodeBadge", parent=styles["bold"], alignment=TA_CENTER, textColor=gold, fontSize=11))],
            [_pdf_para(f"Client: {project.client_name or '-'}", ParagraphStyle("CoverDetail1", parent=styles["normal"], alignment=TA_CENTER, fontSize=10))],
            [_pdf_para(f"Location: {project.location_city or '-'}", ParagraphStyle("CoverDetail2", parent=styles["normal"], alignment=TA_CENTER, fontSize=10))],
        ],
        colWidths=[15.5 * cm],
    )
    cover_card.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, 0), navy),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("BOX", (0, 0), (-1, -1), 1.2, gold),
                ("PADDING", (0, 0), (-1, -1), 12),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    story.append(cover_card)
    story.append(PageBreak())

    overview_rows = [
        ["Project Name", project.name],
        ["Project Code", project.project_code or "BIQ-PENDING"],
        ["Client", project.client_name or "-"],
        ["Location", project.location_city or "-"],
        ["Project Type", project_type_label(project.project_type)],
        ["Lead Engineer", project.lead_engineer or "-"],
        ["Project Manager", project.project_manager or "-"],
        ["Contract Value", _fmt_sar(project.contract_value)],
        ["Start Date", _fmt_report_date(project.start_date)],
        ["Planned Completion", _fmt_report_date(project.planned_completion)],
        ["Actual Completion", _fmt_report_date(data["actual_end"])],
        ["Days Ahead/Behind", f"{data['days_delta']:+d} days" if project.planned_completion else "-"],
        ["Final Health Score", f"{project.health_score if project.health_score is not None else 100}/100 ({project.health_status or 'green'})"],
    ]
    story.append(_section_header("Project Overview", "نظرة عامة", styles))
    story.append(Spacer(1, 0.3 * cm))
    story.append(_styled_pdf_table(overview_rows, [6 * cm, 12 * cm], styles, header=False))
    story.append(PageBreak())

    financial = data["financial"]
    variance_text = f"{_fmt_sar(financial['variance'])} ({financial['variance_pct']:+.1f}%)"
    story.append(_section_header("Financial Summary", "الملخص المالي", styles))
    story.append(Spacer(1, 0.25 * cm))
    story.append(Table(
        [[
            _summary_box("BOQ Contract Value", _fmt_sar(financial["boq_total"]), styles),
            _summary_box("Estimated Actual Cost", _fmt_sar(financial["inventory_total"]), styles, "#15325f"),
            _summary_box("Variance", variance_text, styles, "#b68100" if financial["variance"] > 0 else "#28a745"),
        ]],
        colWidths=[6 * cm, 6 * cm, 6 * cm],
    ))
    story.append(Spacer(1, 0.35 * cm))
    financial_rows = [["Description", "Budgeted SAR", "Actual SAR", "Variance", "Status"]]
    for row in data["boq_sections"]:
        financial_rows.append([row["description"], _fmt_sar(row["budget"]), _fmt_sar(row["actual"]), _fmt_sar(row["variance"]), row["status"]])
    total_budget = sum(row["budget"] for row in data["boq_sections"])
    total_actual = sum(row["actual"] for row in data["boq_sections"])
    financial_rows.append(["TOTAL", _fmt_sar(total_budget), _fmt_sar(total_actual), _fmt_sar(total_actual - total_budget), "Final"])
    story.append(_styled_pdf_table(financial_rows, [5.4 * cm, 3.3 * cm, 3.3 * cm, 3 * cm, 3 * cm], styles, extra_styles=[("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f8eea8"))]))
    story.append(Spacer(1, 0.2 * cm))
    story.append(_pdf_para("Actual costs derived from inventory consumption records in BanaaIQ.", styles["small"]))
    story.append(PageBreak())

    story.append(_section_header("Timeline Performance", "الأداء الزمني", styles))
    story.append(Spacer(1, 0.25 * cm))
    milestone_rows = [["Milestone", "Planned Date", "Actual Date", "Status"]]
    for milestone in project.milestones:
        milestone_rows.append([milestone.title or "-", _fmt_report_date(milestone.planned_date), _fmt_report_date(milestone.actual_date) if milestone.actual_date else "Not recorded", (milestone.status or "pending").replace("_", " ").title()])
    story.append(_styled_pdf_table(milestone_rows, [6.2 * cm, 3.8 * cm, 3.8 * cm, 4.2 * cm], styles))
    story.append(Spacer(1, 0.35 * cm))
    if data["days_delta"] < 0:
        schedule_phrase = f"{abs(data['days_delta'])} days ahead of schedule"
    elif data["days_delta"] > 0:
        schedule_phrase = f"{data['days_delta']} days behind schedule"
    else:
        schedule_phrase = "on the planned completion date"
    story.append(_pdf_para(f"Project {project.name} was planned to start on {_fmt_report_date(project.start_date)}. The planned completion was {_fmt_report_date(project.planned_completion)}. Actual completion was {schedule_phrase}.", styles["normal"]))
    story.append(Spacer(1, 0.45 * cm))
    story.append(TimelineBar(project.start_date, project.planned_completion, data["actual_end"]))
    story.append(PageBreak())

    story.append(_section_header("Site Activity Summary", "ملخص النشاط", styles))
    story.append(Spacer(1, 0.25 * cm))
    story.append(Table([[
        _summary_box("Total DPRs Submitted", str(len(data["dprs"])), styles),
        _summary_box("Total Working Days", str(data["working_days"]), styles, "#15325f"),
        _summary_box("DPR Coverage", f"{data['dpr_coverage']}%", styles, "#28a745" if data["dpr_coverage"] >= 70 else "#b68100"),
    ]], colWidths=[6 * cm, 6 * cm, 6 * cm]))
    story.append(Spacer(1, 0.3 * cm))
    dpr_rows = [["Date", "Zone", "Weather", "Key Activity", "Issues"]]
    for dpr in data["dprs"][:10]:
        dpr_rows.append([
            _fmt_report_date(dpr.date),
            dpr.zone or "-",
            dpr.weather or "-",
            (dpr.progress_notes or "-")[:100],
            "Y" if (dpr.issues or "").strip() else "N",
        ])
    if len(data["dprs"]) > 10:
        dpr_rows.append([f"+ {len(data['dprs']) - 10} more DPRs", "", "", "Full archive available in BanaaIQ.", ""])
    story.append(_styled_pdf_table(dpr_rows, [2.5 * cm, 3 * cm, 2.4 * cm, 8 * cm, 2.1 * cm], styles, small=True))
    story.append(Spacer(1, 0.2 * cm))
    story.append(_pdf_para("Full DPR archive available in BanaaIQ platform.", styles["small"]))
    story.append(PageBreak())

    total_tasks = len(data["tasks"])
    completed_tasks = len(data["completed_tasks"])
    completion_rate = round((completed_tasks / total_tasks * 100), 1) if total_tasks else 0
    status_counts = {status: 0 for status in TASK_STATUS_LABELS}
    for task in data["tasks"]:
        status_counts[task.status] = status_counts.get(task.status, 0) + 1
    story.append(_section_header("Tasks Completed", "المهام المنجزة", styles))
    story.append(Spacer(1, 0.25 * cm))
    story.append(Table([[
        _summary_box("Total Tasks", str(total_tasks), styles),
        _summary_box("Completed", str(completed_tasks), styles, "#28a745"),
        _summary_box("Completion Rate", f"{completion_rate}%", styles, "#15325f"),
    ]], colWidths=[6 * cm, 6 * cm, 6 * cm]))
    story.append(Spacer(1, 0.3 * cm))
    status_rows = [["Status", "Count", "% of Total"]]
    for status_key in ["done", "inprogress", "review", "backlog"]:
        count = status_counts.get(status_key, 0)
        pct = round((count / total_tasks * 100), 1) if total_tasks else 0
        status_rows.append([TASK_STATUS_LABELS.get(status_key, status_key.title()), str(count), f"{pct}%"])
    story.append(_styled_pdf_table(status_rows, [8 * cm, 4 * cm, 6 * cm], styles))
    story.append(Spacer(1, 0.3 * cm))
    completed_rows = [["Task Title", "Assignee", "Job Role", "Completed Date", "Priority"]]
    for task in sorted(data["completed_tasks"], key=lambda t: t.updated_at or t.created_at or datetime.min, reverse=True)[:10]:
        _aname = task.assigned_to.full_name if task.assigned_to else "-"
        completed_rows.append([task.name, _aname, "-", _fmt_report_date(task.updated_at), (task.priority or "-").title()])
    story.append(_styled_pdf_table(completed_rows, [5.5 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 2.9 * cm], styles, small=True))
    story.append(PageBreak())

    story.append(_section_header("Materials Summary", "ملخص المواد", styles))
    story.append(Spacer(1, 0.25 * cm))
    inventory_ids = [item.id for item in data["inventory_items"]]
    source_actual_rows = (
        BOQActual.query.filter(BOQActual.source == "inventory", BOQActual.source_id.in_(inventory_ids)).all()
        if inventory_ids
        else []
    )
    consumed_by_item = {}
    for actual in source_actual_rows:
        consumed_by_item[actual.source_id] = consumed_by_item.get(actual.source_id, 0.0) + float(actual.actual_qty_used or 0)
    material_rows = [["Material", "Category", "Unit", "Opening", "Current", "Consumed", "Unit Cost", "Total Value"]]
    material_extra_styles = []
    for row_index, item in enumerate(data["inventory_items"], start=1):
        consumed = consumed_by_item.get(item.id, 0.0)
        current_stock = float(item.stock or 0)
        opening = current_stock + consumed
        unit_cost = (float(item.value_sar or 0) / current_stock) if current_stock else 0.0
        material_rows.append([item.name, item.category or "-", item.unit or "-", f"{opening:,.2f}", f"{current_stock:,.2f}", f"{consumed:,.2f}", _fmt_sar(unit_cost), _fmt_sar(item.value_sar)])
        if item.status == "critical":
            material_extra_styles.append(("TEXTCOLOR", (0, row_index), (-1, row_index), colors.HexColor("#dc3545")))
        elif item.status == "low":
            material_extra_styles.append(("TEXTCOLOR", (0, row_index), (-1, row_index), colors.HexColor("#c8a730")))
    material_rows.append(["TOTAL", "", "", "", "", "", "", _fmt_sar(sum(float(item.value_sar or 0) for item in data["inventory_items"]))])
    material_extra_styles.append(("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f8eea8")))
    story.append(_styled_pdf_table(material_rows, [3.3 * cm, 2 * cm, 1.2 * cm, 2 * cm, 2 * cm, 1.8 * cm, 2.3 * cm, 3.4 * cm], styles, small=True, extra_styles=material_extra_styles))
    story.append(PageBreak())

    story.append(_section_header("Team Performance", "أداء الفريق", styles))
    story.append(Spacer(1, 0.25 * cm))
    engineer_rows = [["Engineer", "Trade / Package", "Scope SAR", "Completion", "DPRs", "Tasks Done"]]
    for package in data["packages"]:
        engineer_key = _normalize_match_text(package.assigned_engineer_name)
        dpr_count = sum(1 for dpr in data["dprs"] if any(_normalize_match_text(worker.get("name")) == engineer_key for worker in dpr.workers))
        task_count = sum(1 for task in data["completed_tasks"] if _normalize_match_text((task.assigned_to.full_name if task.assigned_to else "") or "") == engineer_key)
        engineer_rows.append([
            package.assigned_engineer_name,
            package.trade or package.scope_description or "-",
            _fmt_sar(package.package_value),
            f"{float(package.completion_percentage or 0):.0f}%",
            str(dpr_count),
            str(task_count),
        ])
    story.append(_styled_pdf_table(engineer_rows, [4.3 * cm, 4.4 * cm, 3 * cm, 2.4 * cm, 1.8 * cm, 2.1 * cm], styles, small=True))
    story.append(PageBreak())

    english_summary, arabic_summary, ai_generated = _completion_ai_summary(project, data)
    story.append(_section_header("Executive Summary", "الملخص التنفيذي", styles))
    story.append(Spacer(1, 0.45 * cm))
    story.append(Paragraph(_pdf_escape(english_summary), styles["exec"]))
    story.append(Spacer(1, 0.35 * cm))
    story.append(HRFlowable(width="100%", thickness=1.4, color=gold))
    story.append(Spacer(1, 0.35 * cm))
    story.append(_pdf_arabic(arabic_summary, styles["arabic"]))
    story.append(Spacer(1, 2.7 * cm))
    source_text = "Generated by BanaaIQ AI" if ai_generated else "Prepared by BanaaIQ AI template"
    story.append(_pdf_para(f"{source_text} - {project.project_code or 'BIQ-PENDING'} - {today_text}", styles["small"]))
    story.append(PageBreak())

    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph("This report was automatically generated by BanaaIQ Intelligent Construction Management Platform.", styles["title"]))
    story.append(Spacer(1, 1.2 * cm))
    sign_rows = [
        ["Project Manager:", "____________________________", "Date:", "____________"],
        ["Client Representative:", "____________________________", "Date:", "____________"],
        [
            "BanaaIQ Verification:",
            Paragraph('<font name="ZapfDingbats">3</font> Auto-verified', styles["normal"]),
            "Date:",
            today_text,
        ],
    ]
    story.append(_styled_pdf_table(sign_rows, [4.5 * cm, 6.5 * cm, 2 * cm, 5 * cm], styles, header=False))
    story.append(Spacer(1, 3 * cm))
    footer_bar = Table([[_pdf_arabic("بناء IQ | BanaaIQ", styles["cover_logo"])]], colWidths=[18 * cm])
    footer_bar.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), navy), ("PADDING", (0, 0), (-1, -1), 14), ("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    story.append(footer_bar)

    doc.build(story, onFirstPage=cover_page, onLaterPages=later_page)
    buffer.seek(0)
    return buffer


def _refresh_stale_health(projects):
    cutoff = datetime.utcnow() - timedelta(hours=1)
    for project in projects:
        if not project.health_last_calculated or project.health_last_calculated < cutoff:
            calculate_health_score(project)


def _project_alert_lines(project):
    today = date.today()
    lines = []
    overdue = Task.query.filter(Task.project_id == project.id, Task.due_date < today, Task.status != "done").all()
    safety_count = 0
    other_count = len(overdue) - safety_count
    if safety_count:
        lines.append(f"{project.name}: {safety_count} overdue safety task(s)")
    elif other_count:
        lines.append(f"{project.name}: {other_count} overdue task(s)")

    critical_items = InventoryItem.query.filter(
        InventoryItem.project_id == project.id,
        InventoryItem.status == "critical",
    ).all()
    for item in critical_items[:2]:
        lines.append(f"{project.name}: {item.name} stock critical")

    financial = _project_financial_snapshot(project)
    if financial["boq_total"] and financial["variance_pct"] > 10:
        lines.append(f"{project.name}: budget {financial['variance_pct']:.0f}% over plan")
    return lines



def boq_row_value(row, *keys):
    for key in keys:
        if key in row and row.get(key) not in (None, ""):
            return row.get(key)
    return ""


def package_boq_items(package):
    if not package or not package.boq:
        return []
    statuses = package.item_statuses
    rows = []
    for index, item in enumerate(package.boq.items or []):
        quantity = float(boq_row_value(item, "quantity", "qty") or 0)
        rate = float(boq_row_value(item, "rate", "unit_rate") or 0)
        total = float(boq_row_value(item, "total") or (quantity * rate))
        status_key = statuses.get(str(index), "not_started")
        rows.append(
            {
                "index": index,
                "description": boq_row_value(item, "desc", "description", "item_description"),
                "unit": boq_row_value(item, "unit", "uom"),
                "quantity": quantity,
                "rate": rate,
                "total": total,
                "status": status_key,
                "status_label": ENGINEER_ITEM_STATUS_LABELS.get(status_key, "Not Started"),
            }
        )
    return rows


def refresh_package_completion(package):
    items = package_boq_items(package)
    if not items:
        return
    completed = sum(1 for item in items if item["status"] == "complete")
    package.completion_percentage = round((completed / len(items)) * 100, 1)
    if completed == len(items):
        package.status = "completed"
    elif package.status == "completed":
        package.status = "active"


def trade_filtered_inventory_items(package):
    if not package:
        return []
    query = InventoryItem.query.filter_by(project_id=package.project_id)
    items = query.order_by(InventoryItem.updated_at.desc()).all()
    trade_key = str(package.trade or "other").lower()
    keywords = TRADE_MATERIAL_KEYWORDS.get(trade_key)
    if not keywords:
        return items
    filtered = []
    for item in items:
        haystack = f"{item.name or ''} {item.category or ''} {item.notes or ''}".lower()
        if any(keyword in haystack for keyword in keywords):
            filtered.append(item)
    return filtered


def engineer_workspace_data(package):
    if not package:
        return {}
    items = package_boq_items(package)
    current_name = _normalize_match_text(current_user.full_name)
    tasks = Task.query.filter(
        Task.project_id == package.project_id,
        Task.assigned_to_id == current_user.id,
    ).order_by(Task.created_at.desc()).all()
    materials = trade_filtered_inventory_items(package)
    dprs = DPR.query.filter_by(project_id=package.project_id, user_id=current_user.id).order_by(DPR.date.desc(), DPR.created_at.desc()).all()
    completed_items = sum(1 for item in items if item["status"] == "complete")
    days_remaining = None
    if package.project and package.project.planned_completion:
        days_remaining = max((package.project.planned_completion - date.today()).days, 0)
    overdue_tasks = sum(1 for task in tasks if task.due_date and task.due_date < date.today() and task.status != "done")
    critical_items = sum(1 for item in materials if item.status == "critical")
    _eng_kanban = {"not_started": [], "in_progress": [], "review": [], "done": []}
    for _t in tasks:
        _col = _t.status if _t.status in _eng_kanban else "not_started"
        _eng_kanban[_col].append(_t)
    return {
        "items": items,
        "tasks": tasks,
        "kanban": _eng_kanban,
        "materials": materials,
        "dprs": dprs,
        "summary": {
            "completed_items": completed_items,
            "total_items": len(items),
            "days_remaining": days_remaining,
            "overdue_tasks": overdue_tasks,
            "critical_items": critical_items,
        },
    }


def get_boq_by_id(boq_id):
    if is_demo_mode():
        for boq in BOQ_SAVED:
            if boq["id"] == boq_id:
                return boq
        return None
    if current_user.is_authenticated:
        boq = BOQ.query.filter_by(id=boq_id, user_id=current_user.id).first()
        return boq.to_dict() if boq else None
    return None


def build_live_notifications():
    if is_demo_mode():
        items = [dict(item) for item in INVENTORY_ITEMS]
        notifications = []
        for item in items:
            status = str(item.get("status", "")).lower()
            if status not in {"critical", "low"}:
                continue
            notifications.append(
                {
                    "id": f"inventory-{item['id']}",
                    "type": "inventory",
                    "icon": "fa-triangle-exclamation",
                    "color": "danger" if status == "critical" else "warning",
                    "title": "Critical Inventory Alert" if status == "critical" else "Low Inventory Alert",
                    "title_ar": "تنبيه مخزون حرج" if status == "critical" else "تنبيه انخفاض المخزون",
                    "message": f"{item['name']} is at {item['stock']} {item['unit']} against a threshold of {item['threshold']} {item['unit']}.",
                    "message_ar": f"{item['name']} عند مستوى {item['stock']} {item['unit']} مقابل حد أدنى {item['threshold']} {item['unit']}.",
                    "time": "Live",
                    "time_ar": "مباشر",
                    "link": "/dashboard/inventory",
                }
            )
        return notifications

    if not current_user.is_authenticated:
        return []

    notifications = []
    db_notifications = (
        Notification.query.filter_by(user_id=current_user.id)
        .order_by(Notification.created_at.desc())
        .limit(5)
        .all()
    )
    notifications.extend(notification.to_dict() for notification in db_notifications)
    low_items = (
        InventoryItem.query.filter_by(user_id=current_user.id)
        .order_by(InventoryItem.updated_at.desc())
        .all()
    )
    for item in low_items:
        status = item.status
        if status not in {"critical", "low"}:
            continue
        feature_project = FeatureProject.query.filter_by(id=item.feature_project_id, user_id=current_user.id, feature="inventory").first() if item.feature_project_id else None
        project_name = feature_project.name if feature_project else "Unassigned"
        project_link = url_for("inventory_index") if feature_project else url_for("inventory_index")
        notifications.append(
            {
                "id": f"inventory-{item.id}",
                "type": "inventory",
                "icon": "fa-triangle-exclamation",
                "color": "danger" if status == "critical" else "warning",
                "title": "Critical Inventory Alert" if status == "critical" else "Low Inventory Alert",
                "title_ar": "تنبيه مخزون حرج" if status == "critical" else "تنبيه انخفاض المخزون",
                "message": f"{item.name} on {project_name} is at {item.stock} {item.unit or ''} against a threshold of {item.threshold} {item.unit or ''}.",
                "message_ar": f"{item.name} في مشروع {project_name} عند مستوى {item.stock} {item.unit or ''} مقابل حد أدنى {item.threshold} {item.unit or ''}.",
                "time": "Live",
                "time_ar": "مباشر",
                "link": project_link,
            }
        )
    return notifications


def get_localized_notifications(lang):
    read_ids = set(session.get("read_notifications", []))
    localized = []
    for n in build_live_notifications():
        item = dict(n)
        item["read"] = bool(item.get("read")) or item["id"] in read_ids
        if lang == "ar":
            item["title"] = item.get("title_ar", item.get("title", ""))
            item["message"] = item.get("message_ar", item.get("message", ""))
            item["time"] = item.get("time_ar", item.get("time", ""))
        localized.append(item)
    return localized


def default_boq_rows():
    return [{"desc": "", "unit": "m2", "qty": "", "rate": "", "notes": ""} for _ in range(5)]


def boq_template_rows(template_key):
    rows = []
    for item in BOQ_TEMPLATES.get(template_key, []):
        rows.append(
            {
                "desc": item.get("desc", ""),
                "unit": item.get("unit", "m2"),
                "qty": item.get("qty", ""),
                "rate": item.get("rate", ""),
                "notes": "",
            }
        )
    return rows


def normalize_boq_role(role_name):
    raw = sanitize_input(role_name or "", 120).strip()
    lookup = re.sub(r"[^a-z0-9]+", " ", raw.lower()).strip()
    aliases = {
        "site engineer": ("site_engineer", "Site Engineer"),
        "site eng": ("site_engineer", "Site Engineer"),
        "quantity surveyor": ("quantity_surveyor", "QS / Quantity Surveyor"),
        "qs": ("quantity_surveyor", "QS / Quantity Surveyor"),
        "qs quantity surveyor": ("quantity_surveyor", "QS / Quantity Surveyor"),
        "project manager": ("project_manager", "Project Manager"),
        "pm": ("project_manager", "Project Manager"),
        "procurement engineer": ("procurement_engineer", "Procurement Engineer"),
        "procurement": ("procurement_engineer", "Procurement Engineer"),
        "mep engineer": ("mep_engineer", "MEP Engineer"),
        "mechanical engineer": ("mep_engineer", "MEP Engineer"),
        "electrical engineer": ("mep_engineer", "MEP Engineer"),
        "civil engineer": ("civil_engineer", "Civil Engineer"),
        "planning engineer": ("planning_engineer", "Planning Engineer"),
        "planner": ("planning_engineer", "Planning Engineer"),
        "hse officer": ("hse_officer", "HSE Officer"),
        "safety officer": ("hse_officer", "HSE Officer"),
        "document controller": ("document_controller", "Document Controller"),
    }
    if lookup in aliases:
        return aliases[lookup]
    return "general", (raw or "Engineer")


def classify_boq_trade(description="", section="", sheet_name=""):
    haystack = f"{description} {section} {sheet_name}".lower()
    trade_keywords = [
        ("Mechanical / HVAC", ["hvac", "vrf", "ahu", "fcu", "duct", "chilled water", "ventilation", "diffuser", "grille", "heat pump", "air handling"]),
        ("Electrical", ["electrical", "cable", "lighting", "panel", "db", "transformer", "switchgear", "conduit", "power", "socket", "earthing"]),
        ("Plumbing / Drainage", ["plumbing", "drainage", "sanitary", "water supply", "ppr", "upvc", "sewer", "manhole", "valve", "pump"]),
        ("Fire Fighting / Alarm", ["fire", "sprinkler", "hydrant", "alarm", "smoke detector", "hose reel", "firefighting"]),
        ("ELV / ICT", ["elv", "cctv", "data", "fiber", "telecom", "access control", "pa system", "bms"]),
        ("Structural / Concrete", ["rebar", "concrete", "foundation", "slab", "column", "beam", "formwork", "pile", "excavation", "structural"]),
        ("Architectural / Finishes", ["tile", "paint", "gypsum", "ceiling", "partition", "joinery", "door", "window", "waterproofing", "finishing"]),
        ("External Works", ["road", "kerb", "interlock", "landscape", "fence", "paving", "asphalt", "external"]),
        ("Procurement / General", ["preliminaries", "temporary", "provisional", "mobilization", "logistics", "general requirement"]),
    ]
    for trade, keywords in trade_keywords:
        if any(keyword in haystack for keyword in keywords):
            return trade
    return "General Construction"


def _boq_cell_is_empty(value):
    if value is None:
        return True
    try:
        if value != value:
            return True
    except Exception:
        pass
    return isinstance(value, str) and not value.replace("\xa0", " ").strip()


def _clean_boq_cell(value):
    if _boq_cell_is_empty(value):
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()
        return cleaned or None
    return value


def _stringify_boq_cell(value):
    if _boq_cell_is_empty(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _boq_to_float(value):
    if _boq_cell_is_empty(value):
        return 0.0
    try:
        cleaned = (
            str(value)
            .replace(",", "")
            .replace("SAR", "")
            .replace("sar", "")
            .replace("%", "")
            .strip()
        )
        if not cleaned:
            return 0.0
        return float(cleaned)
    except Exception:
        return 0.0


def _normalize_boq_text_token(value):
    return re.sub(r"\s+", " ", _stringify_boq_cell(value).lower()).strip(" .:-_/")


def _is_probable_boq_unit(value):
    token = _normalize_boq_text_token(value)
    if not token:
        return False
    known_units = {
        "no",
        "nos",
        "nr",
        "ea",
        "each",
        "pcs",
        "pc",
        "set",
        "sets",
        "lot",
        "ls",
        "item",
        "m",
        "m2",
        "m3",
        "mtr",
        "meter",
        "metre",
        "sqm",
        "sq m",
        "sqmt",
        "cum",
        "kg",
        "ton",
        "tons",
        "tr",
        "kw",
        "mh",
    }
    if token in known_units:
        return True
    return bool(re.fullmatch(r"[a-z]{1,4}\d*", token))


def _is_boq_placeholder_text(value):
    token = _normalize_boq_text_token(value)
    if not token:
        return True
    placeholder_tokens = {
        "description",
        "item",
        "item description",
        "particulars",
        "scope",
        "scope of work",
        "type",
        "unit",
        "uom",
        "qty",
        "quantity",
        "rate",
        "unit rate",
        "amount",
        "total",
        "value",
        "model",
        "code",
        "ref",
        "reference",
        "remarks",
        "specification",
        "section",
        "trade",
        "sheet",
        "sheet1",
        "construction",
        "general construction",
    }
    if token in placeholder_tokens:
        return True
    if _is_probable_boq_unit(token):
        return True
    return bool(re.fullmatch(r"[\d./()+%-]+", token))


def _detect_boq_header(rows):
    header_keywords = [
        "description",
        "model",
        "item",
        "qty",
        "quantity",
        "unit",
        "rate",
        "total",
        "no",
        "type",
        "particulars",
        "scope",
        "code",
    ]
    header_row_idx = 0
    header_map = {}

    for idx, row in enumerate(rows[:20]):
        row_lower = [_stringify_boq_cell(cell).lower() for cell in row]
        matches = sum(1 for cell in row_lower if any(kw in cell for kw in header_keywords))
        if matches >= 2:
            header_row_idx = idx
            for col_idx, cell in enumerate(row_lower):
                if cell in ("no", "#", "sr", "s.no", "s/no", "item no"):
                    header_map["no"] = col_idx
                elif any(k in cell for k in ["model", "code", "ref"]):
                    header_map["model"] = col_idx
                elif any(k in cell for k in ["desc", "item", "particular", "scope", "work"]):
                    header_map["description"] = col_idx
                elif cell == "unit" or "uom" in cell:
                    header_map["unit"] = col_idx
                elif any(k in cell for k in ["qty", "quant", "total quantity"]):
                    header_map["quantity"] = col_idx
                elif any(k in cell for k in ["rate", "price", "cost", "unit price"]):
                    header_map["rate"] = col_idx
                elif any(k in cell for k in ["amount", "total", "value"]):
                    header_map["total"] = col_idx
            break

    if not header_map:
        for row in rows[1:6]:
            non_empty = [(i, v) for i, v in enumerate(row) if not _boq_cell_is_empty(v)]
            if len(non_empty) >= 3:
                if len(row) >= 4:
                    header_map = {"no": 0, "model": 1, "description": 2, "quantity": 3}
                else:
                    header_map = {"description": 0, "quantity": len(row) - 1}
                break
    return header_row_idx, header_map


def _build_flat_boq_headers(row, column_count):
    headers = []
    seen = set()
    for idx in range(column_count):
        raw = _stringify_boq_cell(row[idx] if idx < len(row) else "")
        base = re.sub(r"\s+", " ", raw).strip(" :-_/") or f"Column {idx + 1}"
        candidate = base
        suffix = 2
        while candidate.lower() in seen:
            candidate = f"{base} {suffix}"
            suffix += 1
        seen.add(candidate.lower())
        headers.append(candidate)
    return headers


def _expand_worksheet_merged_rows(worksheet):
    merged_values = {}
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        anchor_value = worksheet.cell(row=min_row, column=min_col).value
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                merged_values[(row_idx, col_idx)] = anchor_value

    rows = []
    for row_idx in range(1, worksheet.max_row + 1):
        current_row = []
        for col_idx in range(1, worksheet.max_column + 1):
            value = merged_values.get((row_idx, col_idx), worksheet.cell(row=row_idx, column=col_idx).value)
            current_row.append(_clean_boq_cell(value))
        rows.append(current_row)

    if not rows:
        return []
    max_columns = max((len(row) for row in rows), default=0)
    normalized = []
    for row in rows:
        padded = list(row) + [None] * (max_columns - len(row))
        if any(not _boq_cell_is_empty(cell) for cell in padded):
            normalized.append(padded)
    if not normalized:
        return []

    keep_indexes = []
    for col_idx in range(max_columns):
        if any(not _boq_cell_is_empty(row[col_idx]) for row in normalized):
            keep_indexes.append(col_idx)

    cleaned_rows = []
    for row in normalized:
        trimmed = [row[idx] for idx in keep_indexes]
        if any(not _boq_cell_is_empty(cell) for cell in trimmed):
            cleaned_rows.append(trimmed)
    return cleaned_rows


def _rows_from_boq_source(source):
    if source is None:
        return []
    if isinstance(source, list):
        return [list(row) for row in source if isinstance(row, (list, tuple))]
    if hasattr(source, "where") and hasattr(source, "values"):
        return source.where(source.notna(), None).values.tolist()
    return []


def _flatten_boq_dataframe(sheet_name, dataframe):
    rows = _rows_from_boq_source(dataframe)
    if not rows:
        return []

    header_row_idx, header_map = _detect_boq_header(rows)
    column_count = max(len(row) for row in rows)
    header_source = rows[header_row_idx] if header_row_idx < len(rows) else []
    header_labels = _build_flat_boq_headers(header_source, column_count)
    flat_rows = []
    current_section = sheet_name or "General"

    def get_val(row, key):
        col_idx = header_map.get(key)
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    for row_number, raw_row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        row = list(raw_row) + [None] * (column_count - len(raw_row))
        if all(_boq_cell_is_empty(cell) for cell in row):
            continue

        non_empty_text = [_stringify_boq_cell(cell) for cell in row if not _boq_cell_is_empty(cell)]
        numeric_values = [_boq_to_float(cell) for cell in row if not _boq_cell_is_empty(cell)]
        if len(non_empty_text) <= 2 and not any(value > 0 for value in numeric_values):
            section_candidate = " ".join(non_empty_text).strip()
            if section_candidate and len(section_candidate) > 3 and not section_candidate.replace(".", "").isdigit():
                current_section = section_candidate[:120]
                continue

        no_value = get_val(row, "no")
        model_value = get_val(row, "model")
        desc_value = get_val(row, "description")
        unit_value = get_val(row, "unit")
        qty_value = get_val(row, "quantity")
        rate_value = get_val(row, "rate")
        total_value = get_val(row, "total")

        text_candidates = []
        for idx, cell in enumerate(row):
            text = _stringify_boq_cell(cell)
            if not text:
                continue
            if idx == header_map.get("no"):
                continue
            if re.fullmatch(r"[\d.,()+/%-]+", text):
                continue
            if _is_probable_boq_unit(text) or _is_boq_placeholder_text(text):
                continue
            text_candidates.append(text)

        description = _stringify_boq_cell(desc_value)
        model = _stringify_boq_cell(model_value)
        if _is_boq_placeholder_text(description):
            description = ""
        if _is_boq_placeholder_text(model) or _normalize_boq_text_token(model) == _normalize_boq_text_token(description):
            model = ""
        if not description:
            description = max((text for text in text_candidates if len(text) > 2), key=len, default="")

        quantity = _boq_to_float(qty_value)
        unit_rate = _boq_to_float(rate_value)
        amount = _boq_to_float(total_value)
        if quantity <= 0 and unit_rate > 0 and amount > 0:
            quantity = round(amount / unit_rate, 3)
        if amount <= 0 and quantity > 0 and unit_rate > 0:
            amount = round(quantity * unit_rate, 2)

        if not description and not model and quantity <= 0 and amount <= 0:
            continue

        trade_category = classify_boq_trade(
            description=description or model,
            section=current_section,
            sheet_name=sheet_name,
        )
        flat_rows.append(
            {
                "source_sheet": sheet_name,
                "row_number": row_number,
                "section": current_section or "General",
                "trade_category": trade_category,
                "item_no": _stringify_boq_cell(no_value) or str(len(flat_rows) + 1),
                "model": model,
                "description": description,
                "unit": _stringify_boq_cell(unit_value) or "No.",
                "quantity": round(quantity, 3) if quantity else 0,
                "unit_rate": round(unit_rate, 2) if unit_rate else 0,
                "amount": round(amount, 2) if amount else 0,
                "is_data_row": bool(quantity > 0 or unit_rate > 0 or amount > 0),
                "raw_values": {
                    header_labels[idx]: _stringify_boq_cell(row[idx])
                    for idx in range(column_count)
                    if idx < len(row) and not _boq_cell_is_empty(row[idx])
                },
            }
        )
    return flat_rows


def _build_boq_items_from_flat_rows(flat_rows):
    items = []
    pending_context = {}

    for row in flat_rows:
        sheet_name = str(row.get("source_sheet", "") or "")
        section = str(row.get("section", "General") or "General")
        context_key = (sheet_name, section)
        description = _stringify_boq_cell(row.get("description"))
        model = _stringify_boq_cell(row.get("model"))
        quantity = float(row.get("quantity", 0) or 0)
        unit_rate = float(row.get("unit_rate", 0) or 0)
        amount = float(row.get("amount", 0) or 0)

        if _is_boq_placeholder_text(description):
            description = ""
        if _is_boq_placeholder_text(model) or _normalize_boq_text_token(model) == _normalize_boq_text_token(description):
            model = ""

        if not row.get("is_data_row"):
            if description:
                pending_context[context_key] = description
            continue

        if not description:
            description = pending_context.get(context_key, "")

        if not description:
            raw_text_values = []
            for value in (row.get("raw_values") or {}).values():
                text = _stringify_boq_cell(value)
                if not text or _is_boq_placeholder_text(text):
                    continue
                if _normalize_boq_text_token(text) == _normalize_boq_text_token(model):
                    continue
                raw_text_values.append(text)
            description = max(raw_text_values, key=len, default="")

        if not description and model:
            description = model

        if not description:
            continue
        if quantity <= 0 and amount <= 0:
            continue
        if unit_rate <= 0 and quantity > 0 and amount > 0:
            unit_rate = round(amount / quantity, 2)

        items.append(
            {
                "no": str(row.get("item_no", len(items) + 1) or len(items) + 1),
                "model": model,
                "description": description,
                "unit": _stringify_boq_cell(row.get("unit")) or "No.",
                "quantity": round(quantity, 3),
                "rate": round(unit_rate, 2) if unit_rate else 0,
                "total": round(amount, 2) if amount else round(quantity * unit_rate, 2),
                "section": section,
                "source_sheet": sheet_name,
                "trade_category": row.get("trade_category", ""),
            }
        )
    return items


def extract_boq_universal(file_bytes, filename):
    """
    Universal BOQ extractor.
    Handles: HVAC schedules, civil BOQs, MEP takeoffs, multi-section Excel,
    mixed column structures, and Arabic/English content.
    Returns a structured dict with sections, all items, raw text, and method.
    """
    import io

    result = {
        "project_title": "",
        "file_name": filename,
        "sections": [],
        "all_items": [],
        "raw_text": "",
        "extraction_method": "",
    }

    section_patterns = [
        r"^\d+\.\d+",
        r"^section\b",
        r"^part\s+\d",
        r"^chapter\b",
        r"^[A-Z0-9\s/&().-]{5,}$",
    ]

    def is_section_header(row):
        non_empty = [_stringify_boq_cell(c) for c in row if not _boq_cell_is_empty(c)]
        if not non_empty:
            return False, ""
        if len(non_empty) <= 2:
            value = non_empty[0].strip()
            if len(value) > 3:
                for pattern in section_patterns:
                    if re.match(pattern, value, re.IGNORECASE):
                        return True, value
                if len(value) > 5 and value.upper() == value:
                    return True, value
                if re.match(r"^\d+\.\d+\s+", value):
                    return True, value
        return False, ""

    def normalize_section_title(title):
        clean = re.sub(r"^\d+\.\d+\s*", "", (title or "")).strip(" -")
        lower = clean.lower()
        if "accessor" in lower:
            return "Accessories"
        if "pipe" in lower or "piping" in lower:
            return "Piping"
        if "insulation" in lower:
            return "Insulation"
        if lower == "other" or "other" in lower:
            return "Other"
        if "equipment" in lower:
            return "Equipment"
        return clean or title or "General"

    def header_map_from_row(row):
        mapping = {}
        for idx, cell in enumerate(row):
            token = _normalize_boq_text_token(cell)
            if token in {"no", "no.", "#", "s/no", "item no"}:
                mapping["item_no"] = idx
            elif any(k in token for k in ["type", "item", "particular", "desc", "description"]):
                mapping.setdefault("description", idx)
            elif any(k in token for k in ["model", "code", "ref"]):
                mapping["model"] = idx
            elif token in {"unit", "uom", "unit type"}:
                mapping["unit"] = idx
            elif any(k in token for k in ["qty", "quantity", "vol", "nos"]):
                mapping["quantity"] = idx
            elif any(k in token for k in ["rate", "price"]):
                mapping["rate"] = idx
            elif any(k in token for k in ["amount", "total", "value"]):
                mapping["amount"] = idx
            elif any(k in token for k in ["size", "diameter", "thickness"]):
                mapping.setdefault("description_extra", []).append(idx)
        return mapping

    def is_header_row(row):
        keywords = ["type", "model", "desc", "qty", "quantity", "unit", "no.", "item", "diameter", "size", "thickness"]
        lower = [_normalize_boq_text_token(c) for c in row]
        return sum(1 for cell in lower if any(k in cell for k in keywords)) >= 2

    def infer_default_section(row_strings, current):
        combined = " ".join(row_strings).lower()
        if current and current != "General":
            return current
        if any(term in combined for term in ["odu", "idu", "ahu", "fcu", "vrf", "gmv", "outdoor unit", "indoor unit"]):
            return "Equipment"
        if any(term in combined for term in ["controller", "branch", "accessories", "connector"]):
            return "Accessories"
        if any(term in combined for term in ["pipe", "piping", "diameter", "length", "copper tube"]):
            return "Piping"
        if any(term in combined for term in ["insulation", "thickness"]):
            return "Insulation"
        if any(term in combined for term in ["concrete", "rebar", "excavation", "formwork"]):
            return "Civil"
        if any(term in combined for term in ["cable", "panel", "db", "conduit", "tray"]):
            return "Electrical"
        return "General"

    def parse_row_item(row, current_section, header_map, row_text):
        row_values = list(row)
        row_strings = [_stringify_boq_cell(cell) for cell in row_values]
        non_empty_strings = [text for text in row_strings if text]
        if not non_empty_strings:
            return None

        quantity = None
        if "quantity" in header_map:
            quantity = _boq_to_float(row_values[header_map["quantity"]])
        if quantity is None or quantity <= 0:
            for cell in reversed(row_values):
                if isinstance(cell, (int, float)) and cell > 0:
                    quantity = float(cell)
                    break
        quantity = quantity or 0

        unit = ""
        if "unit" in header_map:
            unit = _stringify_boq_cell(row_values[header_map["unit"]])
        if not unit:
            known_units = {"no", "nos", "nr", "ea", "each", "pcs", "pc", "set", "sets", "lot", "ls", "m", "m2", "m3", "kg", "ton", "tr", "sqm", "cum", "hr"}
            for cell in row_values:
                token = _normalize_boq_text_token(cell)
                if token in known_units:
                    unit = _stringify_boq_cell(cell)
                    break

        rate = _boq_to_float(row_values[header_map["rate"]]) if "rate" in header_map else 0
        amount = _boq_to_float(row_values[header_map["amount"]]) if "amount" in header_map else 0
        item_no = _stringify_boq_cell(row_values[header_map["item_no"]]) if "item_no" in header_map else ""
        model = _stringify_boq_cell(row_values[header_map["model"]]) if "model" in header_map else ""

        desc_parts = []
        used_idx = {
            header_map.get("item_no"),
            header_map.get("model"),
            header_map.get("unit"),
            header_map.get("quantity"),
            header_map.get("rate"),
            header_map.get("amount"),
        }
        if "description" in header_map:
            desc_value = _stringify_boq_cell(row_values[header_map["description"]])
            if desc_value and not _is_boq_placeholder_text(desc_value):
                desc_parts.append(desc_value)
        for extra_idx in header_map.get("description_extra", []):
            extra_value = _stringify_boq_cell(row_values[extra_idx])
            if extra_value and not _is_boq_placeholder_text(extra_value):
                desc_parts.append(extra_value)
                used_idx.add(extra_idx)
        for idx, text in enumerate(row_strings):
            if idx in used_idx or not text:
                continue
            if _is_boq_placeholder_text(text):
                continue
            if text == model:
                continue
            desc_parts.append(text)

        if not desc_parts and model:
            desc_parts.append(model)
        description = " | ".join(dict.fromkeys(part.strip() for part in desc_parts if part.strip()))
        if not description and quantity <= 0 and rate <= 0 and amount <= 0:
            return None

        section = current_section or infer_default_section(non_empty_strings, current_section)
        if not section or section == "General":
            section = infer_default_section(non_empty_strings, current_section)

        return {
            "section": section or "General",
            "item_no": item_no,
            "model": model,
            "description": description,
            "quantity": quantity,
            "unit": unit,
            "rate": rate,
            "amount": amount,
            "raw_row": row_text,
        }

    if filename.lower().endswith((".xlsx", ".xls")):
        try:
            workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
            all_text_rows = []
            all_items_flat = []
            sections_found = []

            for sheet_name in workbook.sheetnames[:5]:
                ws = workbook[sheet_name]
                rows = _expand_worksheet_merged_rows(ws)
                if not rows:
                    continue

                if not result["project_title"]:
                    for row in rows[:5]:
                        texts = [_stringify_boq_cell(cell) for cell in row if not _boq_cell_is_empty(cell)]
                        for text in texts:
                            token = _normalize_boq_text_token(text)
                            if len(text) > 5 and not is_header_row([text]) and "sheet" not in token:
                                result["project_title"] = text.strip()
                                break
                        if result["project_title"]:
                            break

                current_section = "General"
                current_header_map = {}
                current_section_items = []

                def flush_section():
                    nonlocal current_section_items
                    if current_section_items:
                        sections_found.append({"title": current_section, "items": current_section_items})
                        current_section_items = []

                for row_index, row in enumerate(rows, start=1):
                    if not any(not _boq_cell_is_empty(c) for c in row):
                        continue

                    row_text = " | ".join(_stringify_boq_cell(c) for c in row if not _boq_cell_is_empty(c)).strip(" |")
                    if result["project_title"] and row_index <= 3 and row_text == result["project_title"]:
                        continue

                    is_sec, sec_title = is_section_header(row)
                    if row_text:
                        all_text_rows.append(row_text)

                    if is_sec:
                        flush_section()
                        current_section = normalize_section_title(sec_title)
                        current_header_map = {}
                        all_text_rows.append(f"[SECTION: {current_section}]")
                        continue

                    if is_header_row(row):
                        current_header_map = header_map_from_row(row)
                        if current_section == "General" and any("model" in _normalize_boq_text_token(c) or "type" in _normalize_boq_text_token(c) for c in row):
                            current_section = "Equipment"
                        header_text = " | ".join(_stringify_boq_cell(c) for c in row if not _boq_cell_is_empty(c))
                        all_text_rows.append(f"HEADERS: {header_text}")
                        continue

                    item = parse_row_item(row, current_section, current_header_map, row_text)
                    if not item:
                        if current_section_items and row_text and len(row_text) > 3:
                            current_section_items[-1]["description"] = (
                                current_section_items[-1]["description"] + " " + row_text
                            ).strip()
                            current_section_items[-1]["raw_row"] = (
                                current_section_items[-1]["raw_row"] + " || " + row_text
                            ).strip()
                        continue

                    if not item["description"] and current_section_items:
                        item["description"] = current_section_items[-1]["description"]
                    if not item["description"]:
                        continue

                    current_section_items.append(item)
                    all_items_flat.append(
                        {
                            "section": item["section"],
                            "description": item["description"],
                            "quantity": item["quantity"],
                            "unit": item["unit"],
                            "model": item["model"],
                            "item_no": item["item_no"],
                            "raw_row": item["raw_row"],
                            "source_sheet": sheet_name,
                        }
                    )

                flush_section()

            result["sections"] = sections_found
            result["all_items"] = all_items_flat
            result["raw_text"] = "\n".join(all_text_rows)[:6000]
            result["extraction_method"] = "excel_direct"
        except Exception as e:
            result["error"] = str(e)
            result["extraction_method"] = "failed"
    elif filename.lower().endswith(".pdf"):
        try:
            import pdfplumber

            pages_text = []
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages[:20]:
                    text = page.extract_text()
                    if text:
                        pages_text.append(text)
                    tables = page.extract_tables() or []
                    for table in tables:
                        for row in table or []:
                            if row:
                                pages_text.append(" | ".join(str(c) if c else "" for c in row))

            full_text = "\n".join(pages_text)
            result["raw_text"] = full_text[:6000]
            result["extraction_method"] = "pdf"
            lines = [line.strip() for line in full_text.split("\n") if line.strip()]
            for line in lines[:5]:
                if len(line) > 5:
                    result["project_title"] = line
                    break
        except ImportError:
            result["error"] = "Install pdfplumber: pip install pdfplumber"
        except Exception as e:
            result["error"] = str(e)

    return result


def build_boq_flat_preview(rows, limit=60):
    preview_rows = []
    for row in (rows or [])[:limit]:
        preview_rows.append(
            {
                "source_sheet": row.get("source_sheet", ""),
                "section": row.get("section", ""),
                "trade_category": row.get("trade_category", ""),
                "item_no": row.get("item_no", row.get("no", "")),
                "model": row.get("model", ""),
                "description": row.get("description", ""),
                "unit": row.get("unit", ""),
                "quantity": row.get("quantity", 0),
                "unit_rate": row.get("unit_rate", row.get("rate", row.get("unit_rate_sar", 0))),
                "amount": row.get("amount", row.get("total", row.get("total_sar", 0))),
            }
        )
    return preview_rows


def parse_boq_excel_with_pandas(file_bytes, filename=""):
    try:
        import pandas as pd
    except ModuleNotFoundError:
        pd = None

    ext = os.path.splitext((filename or "").lower())[1]
    is_zip_excel = file_bytes[:2] == b"PK"
    sheet_names = []
    workbook_data = {}
    last_error = None

    if ext != ".xls" or is_zip_excel:
        try:
            workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
            for worksheet in workbook.worksheets[:15]:
                sheet_names.append(worksheet.title)
                rows = _expand_worksheet_merged_rows(worksheet)
                if rows:
                    workbook_data[worksheet.title] = rows
        except Exception as exc:
            last_error = exc

    if not workbook_data:
        if pd is None:
            if ext == ".xls" and not is_zip_excel:
                raise ValueError("Legacy .xls uploads require pandas + xlrd in the active Python environment.")
            if last_error is not None:
                raise ValueError(f"Workbook open failed and pandas is unavailable: {last_error}") from last_error
            raise ValueError("pandas is unavailable in the active Python environment.")
        read_kwargs = {
            "sheet_name": None,
            "header": None,
            "dtype": object,
        }
        if ext == ".xls" and not is_zip_excel:
            read_kwargs["engine"] = "xlrd"
        elif ext in {".xlsx", ".xlsm", ".xltx", ".xltm"} or is_zip_excel:
            read_kwargs["engine"] = "openpyxl"
        try:
            workbook_data = pd.read_excel(BytesIO(file_bytes), **read_kwargs)
        except Exception as exc:
            if last_error is not None:
                raise ValueError(f"{last_error}; fallback read failed: {exc}") from exc
            raise

        sheet_names = list(workbook_data.keys())[:15]
        cleaned_data = {}
        for sheet_name, df in list(workbook_data.items())[:15]:
            if df is None or df.empty:
                continue
            df = df.apply(lambda col: col.map(_clean_boq_cell))
            df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)
            if not df.empty:
                cleaned_data[sheet_name] = df
        workbook_data = cleaned_data

    extracted_items = []
    flat_rows = []

    for sheet_name, df in list(workbook_data.items())[:15]:
        rows = _rows_from_boq_source(df)
        if not rows:
            continue
        sheet_flat_rows = _flatten_boq_dataframe(sheet_name, rows)
        flat_rows.extend(sheet_flat_rows)

        sheet_items = _build_boq_items_from_flat_rows(sheet_flat_rows)
        if sheet_items:
            extracted_items.extend(sheet_items)
            continue

        tmp_wb = Workbook()
        tmp_ws = tmp_wb.active
        tmp_ws.title = "tmp"
        for row in rows:
            tmp_ws.append([None if _boq_cell_is_empty(v) else v for v in row])

        fallback_items = smart_extract_boq(tmp_ws)
        for item in fallback_items:
            enriched = dict(item)
            enriched["source_sheet"] = sheet_name
            enriched["trade_category"] = classify_boq_trade(
                description=enriched.get("description", ""),
                section=enriched.get("section", ""),
                sheet_name=sheet_name,
            )
            extracted_items.append(enriched)

    return sheet_names, extracted_items, flat_rows


def summarize_boq_trades(items):
    grouped = {}
    for item in items:
        trade = item.get("trade_category") or classify_boq_trade(
            description=item.get("description", ""),
            section=item.get("section", ""),
            sheet_name=item.get("source_sheet", ""),
        )
        bucket = grouped.setdefault(
            trade,
            {
                "trade": trade,
                "item_count": 0,
                "quantity_total": 0.0,
                "sheets": set(),
                "sample_items": [],
            },
        )
        bucket["item_count"] += 1
        try:
            bucket["quantity_total"] += float(item.get("quantity", 0) or 0)
        except Exception:
            pass
        if item.get("source_sheet"):
            bucket["sheets"].add(item["source_sheet"])
        sample = item.get("description") or item.get("model") or ""
        if sample and len(bucket["sample_items"]) < 3:
            bucket["sample_items"].append(sample[:80])

    summary = []
    for trade, info in grouped.items():
        summary.append(
            {
                "trade": trade,
                "item_count": info["item_count"],
                "quantity_total": round(info["quantity_total"], 2),
                "sheet_count": len(info["sheets"]),
                "sample_items": info["sample_items"],
            }
        )
    summary.sort(key=lambda item: item["item_count"], reverse=True)
    return summary


def build_boq_guidance_context(user_name, role_name, project_name, client_name, sheet_names, extracted_items, analyzed):
    role_slug, role_display = normalize_boq_role(role_name)
    trade_breakdown = summarize_boq_trades(extracted_items)
    sections = analyzed.get("sections", []) or []
    financials = analyzed.get("financial_summary", {}) or {}
    primary_trades = [item["trade"] for item in trade_breakdown[:3]]
    flat_table_total_rows = int(analyzed.get("flat_table_total_rows") or 0)
    data_issues = analyzed.get("data_issues", []) or []
    billing_guidance = analyzed.get("billing_guidance", {}) or {}
    extraction_summary = analyzed.get("extraction_summary", {}) or {}

    role_focus_map = {
        "site_engineer": [
            "Confirm line items against current site scope and latest drawings before approving quantities.",
            "Walk down the top trades on site and validate missing logistics, access, and installation constraints.",
            "Flag construction-sequence risks early so procurement and QS teams can price them properly.",
        ],
        "quantity_surveyor": [
            "Validate quantity take-off logic trade by trade and reconcile unusual spikes before pricing sign-off.",
            "Check provisional sums, exclusions, and unit consistency before issuing commercial review notes.",
            "Benchmark high-value lines against supplier quotations or previous Saudi project data.",
        ],
        "project_manager": [
            "Review the dominant trades first and align them with budget, schedule milestones, and procurement lead times.",
            "Use the guidance notes to assign quantity validation, supplier checks, and approval ownership by team.",
            "Issue a coordinated review cycle so engineering, commercial, and procurement comments close together.",
        ],
        "procurement_engineer": [
            "Prioritize supplier RFQs for the highest-value trades and long-lead packages detected in the workbook.",
            "Turn the categorized trade summary into enquiry packages and assign target quotation dates immediately.",
            "Check whether transport, testing, or supplier compliance costs are missing from commercial lines.",
        ],
        "mep_engineer": [
            "Validate MEP quantities against latest coordinated drawings, risers, equipment schedules, and service routes.",
            "Check for scope gaps between HVAC, electrical, plumbing, fire alarm, and ELV interfaces.",
            "Highlight any testing, commissioning, and specialist accessory items missing from the workbook.",
        ],
        "civil_engineer": [
            "Review structural and civil line items against drawings, excavation limits, and method statements.",
            "Verify concrete, rebar, formwork, and finishing interfaces for sequencing and temporary works coverage.",
            "Escalate unclear measurement assumptions before commercial submission.",
        ],
        "planning_engineer": [
            "Map the top trades to near-term milestones and long-lead procurement activities before baseline review.",
            "Check whether phased execution, night works, or logistics constraints should drive separate BOQ lines.",
            "Use the guidance to align package release dates with procurement and installation readiness.",
        ],
        "hse_officer": [
            "Review the trade breakdown for temporary works, access systems, testing, and safety-critical packages.",
            "Flag risk-prone trades early so HSE controls are priced into temporary works and logistics scope.",
            "Check whether permits, protection, fire stopping, and safe access provisions appear in the BOQ.",
        ],
        "document_controller": [
            "Verify the uploaded workbook aligns with the latest revision history, sheet naming, and controlled document list.",
            "Check that all priced trades map back to approved drawing packages or marked tender clarifications.",
            "Record any assumptions or manual revisions before the BOQ moves to the next approval stage.",
        ],
        "general": [
            "Review the categorized trades and verify the BOQ scope before issuing the next commercial action.",
            "Assign each major trade to the right reviewer and capture missing assumptions early.",
            "Use the step-by-step notes below to move from file validation to commercial sign-off cleanly.",
        ],
    }

    return {
        "user_name": user_name or "Engineer",
        "role_name": role_name or role_display,
        "role_display": role_display,
        "role_slug": role_slug,
        "project_name": project_name or "Construction Project",
        "client_name": client_name or "Internal Review",
        "sheet_names": sheet_names[:8],
        "trade_breakdown": trade_breakdown,
        "primary_trades": primary_trades,
        "boq_sections": sections[:8],
        "total_items": analyzed.get("total_items_found") or len(extracted_items),
        "flat_table_total_rows": flat_table_total_rows,
        "financial_summary": financials,
        "executive_summary": analyzed.get("executive_summary", ""),
        "recommendations": analyzed.get("recommendations", [])[:5],
        "risk_flags": analyzed.get("risk_flags", [])[:5],
        "data_issues": data_issues[:6],
        "billing_guidance": billing_guidance,
        "extraction_summary": extraction_summary,
        "role_focus_steps": role_focus_map.get(role_slug, role_focus_map["general"]),
    }


def render_boq_role_guidance(context):
    role_slug = context.get("role_slug", "general")
    template_name = f"dashboard/boq/guidance/{role_slug}.html"
    try:
        app.jinja_env.get_template(template_name)
    except Exception:
        template_name = "dashboard/boq/guidance/general.html"
    html = render_template(template_name, **context)
    plain = re.sub(r"<\s*br\s*/?>", "\n", html)
    plain = re.sub(r"</(li|p|div|h3|h4|h5|tr)>", "\n", plain)
    plain = re.sub(r"<[^>]+>", "", plain)
    plain = re.sub(r"\n{3,}", "\n\n", plain)
    return {
        "template_name": template_name,
        "html": html,
        "text": plain.strip(),
    }


def sanitize_input(text, max_length=2000):
    if not text:
        return ""
    dangerous_phrases = ["ignore previous instructions", "disregard all", "new instruction", "system:", "assistant:", "###", "jailbreak"]
    text = str(text)
    for phrase in dangerous_phrases:
        text = text.replace(phrase, "[removed]").replace(phrase.title(), "[removed]").replace(phrase.upper(), "[removed]")
    return text[:max_length]


def sanitize(text, max_length=2000):
    return sanitize_input(text, max_length=max_length)


# Excel formula-injection prefix characters (=, +, -, @, TAB, CR)
_EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def safe_excel_cell(value):
    """Prevent Excel formula injection.

    If a string cell value starts with a formula-trigger character,
    prepend a single quote so Excel treats it as literal text rather
    than executing it as a formula or DDE link.
    """
    if isinstance(value, str) and value and value[0] in _EXCEL_FORMULA_PREFIXES:
        return "'" + value
    return value


def setup_arabic_font():
    """Register Cairo Arabic font with ReportLab."""
    if getattr(setup_arabic_font, "_registered", False):
        return True

    candidates = [
        os.path.join(app.root_path, "static", "fonts", "Cairo.ttf"),
        os.path.join(app.root_path, "static", "fonts", "Cairo-Regular.ttf"),
    ]

    for font_path in candidates:
        if not os.path.exists(font_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont("Cairo", font_path))
            setup_arabic_font._registered = True
            return True
        except Exception as e:
            app.logger.error(f"Font registration failed: {e}")
            return False

    app.logger.warning(f"Cairo font not found. Checked: {candidates}")
    return False


def process_arabic_text(text):
    """
    Convert Arabic text so ReportLab renders it correctly
    (reshaping + bidi algorithm).
    """
    if not text:
        return text
    if arabic_reshaper is None or get_display is None:
        return text
    try:
        reshaped = arabic_reshaper.reshape(text)
        return get_display(reshaped)
    except Exception as e:
        app.logger.error(f"Arabic processing error: {e}")
        return text


def is_arabic_text(text):
    """Check if text contains Arabic characters."""
    if not text:
        return False
    chars = [c for c in str(text) if not c.isspace()]
    if not chars:
        return False
    arabic_chars = sum(1 for c in chars if "\u0600" <= c <= "\u06FF")
    return arabic_chars > len(chars) * 0.3


def setup_hindi_font():
    """Register Noto Sans Devanagari font with ReportLab (for Hindi text)."""
    if getattr(setup_hindi_font, "_registered", False):
        return True

    candidates = [
        os.path.join(app.root_path, "static", "fonts", "NotoSansDevanagari-Regular.ttf"),
        os.path.join(app.root_path, "static", "fonts", "NotoSansDevanagari.ttf"),
    ]
    for font_path in candidates:
        if not os.path.exists(font_path):
            continue
        try:
            pdfmetrics.registerFont(TTFont("NotoDevanagari", font_path))
            setup_hindi_font._registered = True
            return True
        except Exception as e:
            app.logger.error(f"Hindi font registration failed: {e}")
            return False

    app.logger.warning(f"NotoSansDevanagari font not found — Hindi PDF columns will use fallback.")
    return False


def is_devanagari_text(text):
    """Check if text contains Devanagari (Hindi) characters."""
    if not text:
        return False
    chars = [c for c in str(text) if not c.isspace()]
    if not chars:
        return False
    deva_chars = sum(1 for c in chars if "\u0900" <= c <= "\u097F")
    return deva_chars > len(chars) * 0.2


def smart_extract_boq(worksheet):
    """
    Intelligently extracts BOQ items from ANY Excel worksheet structure.
    Handles standard BOQs, HVAC schedules, material lists, and equipment BOQs.
    """
    items = []

    all_rows = []
    for row in worksheet.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() for cell in row):
            all_rows.append(list(row))

    if not all_rows:
        return items

    header_keywords = [
        "description",
        "model",
        "item",
        "qty",
        "quantity",
        "unit",
        "rate",
        "total",
        "no",
        "type",
        "particulars",
        "scope",
        "code",
    ]

    header_row_idx = 0
    header_map = {}

    for idx, row in enumerate(all_rows[:15]):
        row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
        matches = sum(1 for cell in row_lower if any(kw in cell for kw in header_keywords))
        if matches >= 2:
            header_row_idx = idx
            for col_idx, cell in enumerate(row_lower):
                if cell in ("no", "#", "sr", "s.no", "s/no"):
                    header_map["no"] = col_idx
                elif any(k in cell for k in ["model", "code", "ref"]):
                    header_map["model"] = col_idx
                elif any(k in cell for k in ["desc", "item", "particular", "scope", "work"]):
                    header_map["description"] = col_idx
                elif cell == "unit" or "uom" in cell:
                    header_map["unit"] = col_idx
                elif any(k in cell for k in ["qty", "quant", "total quantity"]):
                    header_map["quantity"] = col_idx
                elif any(k in cell for k in ["rate", "price", "sar", "cost", "unit price"]):
                    header_map["rate"] = col_idx
            break

    if not header_map:
        for row in all_rows[1:6]:
            non_empty = [(i, v) for i, v in enumerate(row) if v is not None and str(v).strip()]
            if len(non_empty) >= 3:
                if len(row) >= 4:
                    header_map = {"no": 0, "model": 1, "description": 2, "quantity": 3}
                else:
                    header_map = {"description": 0, "quantity": len(row) - 1}
                break

    current_section = "General"

    def get_val(row, key):
        col_idx = header_map.get(key)
        if col_idx is None or col_idx >= len(row):
            return None
        return row[col_idx]

    def to_float(val):
        if val is None:
            return 0.0
        try:
            cleaned = (
                str(val)
                .replace(",", "")
                .replace("SAR", "")
                .replace("sar", "")
                .replace(" ", "")
                .strip()
            )
            return float(cleaned) if cleaned else 0.0
        except Exception:
            return 0.0

    for row in all_rows[header_row_idx + 1 :]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) <= 2:
            section_candidate = str(non_empty[0]).strip() if non_empty else ""
            if section_candidate and len(section_candidate) > 3 and not section_candidate.replace(".", "").isdigit():
                current_section = section_candidate
                continue

        no = get_val(row, "no")
        model = get_val(row, "model")
        description = get_val(row, "description")
        unit = get_val(row, "unit")
        quantity = get_val(row, "quantity")
        rate = get_val(row, "rate")

        desc_val = description
        if _is_boq_placeholder_text(desc_val):
            desc_val = ""
        if not desc_val and not _is_boq_placeholder_text(model):
            desc_val = str(model).strip()
        if not desc_val or not str(desc_val).strip():
            continue

        qty_clean = to_float(quantity)
        if qty_clean <= 0:
            continue

        rate_clean = to_float(rate)

        items.append(
            {
                "no": str(no).strip() if no is not None and str(no).strip() else str(len(items) + 1),
                "model": "" if _is_boq_placeholder_text(model) else (str(model).strip() if model is not None else ""),
                "description": str(desc_val).strip(),
                "unit": str(unit).strip() if unit is not None and str(unit).strip() else "No.",
                "quantity": qty_clean,
                "rate": rate_clean,
                "total": qty_clean * rate_clean,
                "section": current_section,
            }
        )

    return items


def _normalize_json_response(result_text):
    clean = (result_text or "").strip()
    if "```" in clean:
        parts = clean.split("```")
        for part in parts:
            part = part.strip()
            if "{" in part and "}" in part:
                clean = part
                if clean.lower().startswith("json"):
                    clean = clean[4:].strip()
                break
    return clean


def _parse_json_like_response(result_text):
    clean = _normalize_json_response(result_text)
    if "{" in clean and "}" in clean:
        clean = clean[clean.find("{") : clean.rfind("}") + 1]
    try:
        return json.loads(clean)
    except Exception:
        try:
            import ast

            parsed = ast.literal_eval(clean)
            if isinstance(parsed, (dict, list)):
                return parsed
        except Exception:
            pass
    raise ValueError("AI response was not valid JSON.")


def _friendly_ai_error(error):
    message = str(error or "").lower()
    if "timeout" in message or "timed out" in message:
        return "AI analysis is taking longer than expected. Please try again in a moment."
    return "AI service temporarily unavailable. Please try again."


def call_openai_json(system_prompt, user_prompt, max_tokens=2500, temperature=0.1):
    if client is None:
        return False, "AI service is not configured. Please set OPENAI_API_KEY.", ""
    last_response = ""
    for attempt in range(2):
        prompt = user_prompt
        if attempt:
            prompt = (
                f"{user_prompt}\n\n"
                "The previous response could not be parsed as valid JSON. "
                "Return valid JSON only, with double-quoted property names and no markdown.\n\n"
                f"Previous response:\n{last_response[:3000]}"
            )
        try:
            kwargs = {
                "model": app.config.get("OPENAI_MODEL", "gpt-4o-mini"),
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt},
                ],
                "max_tokens": max_tokens,
                "temperature": temperature,
            }
            try:
                response = client.chat.completions.create(
                    **kwargs,
                    response_format={"type": "json_object"},
                )
            except TypeError:
                response = client.chat.completions.create(**kwargs)
            session["ai_calls"] = session.get("ai_calls", 0) + 1
            session.modified = True
            last_response = (response.choices[0].message.content or "").strip()
            try:
                parsed = _parse_json_like_response(last_response)
                if isinstance(parsed, dict):
                    return True, parsed, last_response
            except Exception as parse_error:
                app.logger.warning(
                    "OpenAI JSON parse failed on attempt %s: %s",
                    attempt + 1,
                    parse_error,
                )
        except Exception as error:
            app.logger.exception("OpenAI JSON call failed")
            return False, _friendly_ai_error(error), last_response
    return False, "AI returned an unreadable response. Please try again.", last_response


def _boq_float(value):
    return _boq_to_float(value)


def _normalize_boq_editor_item(item, index=1):
    qty = _boq_float(item.get("quantity", item.get("qty", 0)))
    rate_raw = item.get("rate")
    if rate_raw in (None, "", 0, "0"):
        rate_raw = item.get("unit_rate_sar", item.get("unit_rate", 0))
    total_raw = item.get("total")
    if total_raw in (None, "", 0, "0"):
        total_raw = item.get("total_sar", item.get("amount", 0))
    rate = _boq_float(rate_raw)
    total = _boq_float(total_raw)
    if rate <= 0 and qty > 0 and total > 0:
        rate = round(total / qty, 2)
    if total <= 0 and qty > 0 and rate > 0:
        total = round(qty * rate, 2)
    desc = sanitize_input(item.get("desc") or item.get("description", ""), 600).strip()
    return {
        "no": str(item.get("no", item.get("item_no", index)) or index),
        "desc": desc,
        "description": desc,
        "unit": sanitize_input(item.get("unit", "No."), 30) or "No.",
        "qty": float(qty),
        "quantity": float(qty),
        "rate": float(rate),
        "total": float(total),
        "notes": sanitize_input(item.get("notes", item.get("section", "")), 500),
    }


def normalize_boq_editor_items(items):
    normalized = []
    for idx, item in enumerate(items or [], start=1):
        if not isinstance(item, dict):
            continue
        row = _normalize_boq_editor_item(item, idx)
        if not row["description"] and row["quantity"] <= 0 and row["rate"] <= 0:
            continue
        normalized.append(row)
    return normalized


def _boq_totals_from_rows(rows):
    subtotal = Decimal("0.00")
    for row in rows or []:
        total = row.get("total")
        if total in (None, ""):
            total = Decimal(str(row.get("quantity", row.get("qty", 0)) or 0)) * Decimal(str(row.get("rate", 0) or 0))
        subtotal += to_currency_decimal(total)
    subtotal = to_currency_decimal(subtotal)
    vat = to_currency_decimal(subtotal * Decimal("0.15"))
    grand_total = to_currency_decimal(subtotal + vat)
    return subtotal, vat, grand_total


def _rows_from_generated_boq(parsed):
    rows = []
    for section in parsed.get("sections", []) or []:
        section_name = sanitize_input(section.get("section_name", "General"), 120) or "General"
        for item in section.get("items", []) or []:
            if not isinstance(item, dict):
                continue
            qty = _boq_float(item.get("quantity", 0))
            rate = _boq_float(item.get("unit_rate_sar", item.get("rate", 0)))
            total = _boq_float(item.get("total_sar", item.get("total", qty * rate)))
            if total <= 0 and qty > 0 and rate > 0:
                total = round(qty * rate, 2)
            notes = item.get("notes") or section_name
            rows.append(
                _normalize_boq_editor_item(
                    {
                        "item_no": item.get("item_no", len(rows) + 1),
                        "description": item.get("description", ""),
                        "unit": item.get("unit", "No."),
                        "quantity": qty,
                        "rate": rate,
                        "total": total,
                        "notes": notes,
                    },
                    len(rows) + 1,
                )
            )
    return rows


def _boq_item_no(row, index=1):
    return str(row.get("no") or row.get("item_no") or index)


def _boq_item_desc(row):
    return row.get("description") or row.get("desc") or ""


def _boq_item_qty(row):
    return _boq_float(row.get("quantity", row.get("qty", 0)))


def _boq_item_rate(row):
    return _boq_float(row.get("rate", row.get("unit_rate_sar", row.get("unit_rate", 0))))


def _boq_item_total(row):
    qty = _boq_item_qty(row)
    rate = _boq_item_rate(row)
    return _boq_float(row.get("total", row.get("total_sar", qty * rate))) or round(qty * rate, 2)


def _boq_prompt_items(rows):
    items = []
    for idx, row in enumerate(rows or [], start=1):
        items.append(
            {
                "item_no": _boq_item_no(row, idx),
                "description": _boq_item_desc(row),
                "unit": row.get("unit", "No."),
                "quantity": _boq_item_qty(row),
                "unit_rate_sar": _boq_item_rate(row),
                "total_sar": _boq_item_total(row),
                "notes": row.get("notes", ""),
            }
        )
    return items


def _load_boq_json_meta(boq_obj):
    try:
        parsed = json.loads(boq_obj.ai_suggestions or "{}")
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def build_boq_diff_rows(original_rows, modified_rows):
    original_map = {_boq_item_no(row, idx): row for idx, row in enumerate(original_rows or [], start=1)}
    modified_map = {_boq_item_no(row, idx): row for idx, row in enumerate(modified_rows or [], start=1)}
    ordered_keys = []
    for idx, row in enumerate(original_rows or [], start=1):
        key = _boq_item_no(row, idx)
        if key not in ordered_keys:
            ordered_keys.append(key)
    for idx, row in enumerate(modified_rows or [], start=1):
        key = _boq_item_no(row, idx)
        if key not in ordered_keys:
            ordered_keys.append(key)

    diff_rows = []
    for key in ordered_keys:
        old = original_map.get(key)
        new = modified_map.get(key)
        if old and new:
            changed = any(
                [
                    _boq_item_desc(old).strip() != _boq_item_desc(new).strip(),
                    (old.get("unit") or "") != (new.get("unit") or ""),
                    round(_boq_item_qty(old), 3) != round(_boq_item_qty(new), 3),
                    round(_boq_item_rate(old), 2) != round(_boq_item_rate(new), 2),
                    round(_boq_item_total(old), 2) != round(_boq_item_total(new), 2),
                ]
            )
            status = "changed" if changed else "unchanged"
        elif new:
            status = "new"
        else:
            status = "removed"
        diff_rows.append({"item_no": key, "original": old, "modified": new, "status": status})
    return diff_rows


def build_boq_version_context(boq_obj):
    if not boq_obj:
        return {}
    parent = BOQ.query.filter_by(id=boq_obj.parent_boq_id, user_id=boq_obj.user_id).first() if boq_obj.parent_boq_id else None
    children = BOQ.query.filter_by(parent_boq_id=boq_obj.id, user_id=boq_obj.user_id).order_by(BOQ.created_at.desc()).all()
    meta = _load_boq_json_meta(boq_obj)
    original_total = float(parent.grand_total or 0) if parent else 0.0
    new_total = float(boq_obj.grand_total or 0)
    change = new_total - original_total
    change_pct = (change / original_total * 100) if original_total else 0.0
    return {
        "parent": parent,
        "children": children,
        "modification": meta.get("phase2_modification", {}) if isinstance(meta.get("phase2_modification"), dict) else {},
        "original_total": original_total,
        "new_total": new_total,
        "change": change,
        "change_pct": change_pct,
        "diff_rows": build_boq_diff_rows(parent.items, boq_obj.items) if parent else [],
    }


def _normalize_match_text(value):
    return re.sub(r"\s+", " ", str(value or "").lower()).strip()


def _best_boq_item_match(rows, text, min_score=0.62):
    needle = _normalize_match_text(text)
    if not needle:
        return None, 0.0
    best_row = None
    best_score = 0.0
    for idx, row in enumerate(rows or [], start=1):
        desc = _normalize_match_text(_boq_item_desc(row))
        if not desc:
            continue
        score = SequenceMatcher(None, needle, desc).ratio()
        if needle in desc or desc in needle:
            score = max(score, 0.82)
        if score > best_score:
            best_row = row
            best_score = score
    if best_score < min_score:
        return None, best_score
    return best_row, best_score


def _latest_boq_for_project(user_id, project_name):
    project_key = _normalize_match_text(project_name)
    query = BOQ.query.filter_by(user_id=user_id).order_by(BOQ.updated_at.desc(), BOQ.created_at.desc())
    candidates = query.all()
    if project_key:
        exact = [boq for boq in candidates if _normalize_match_text(boq.project) == project_key and boq.items]
        if exact:
            return exact[0]
        fuzzy = [
            (SequenceMatcher(None, project_key, _normalize_match_text(boq.project)).ratio(), boq)
            for boq in candidates
            if boq.items
        ]
        fuzzy = sorted(fuzzy, key=lambda item: item[0], reverse=True)
        if fuzzy and fuzzy[0][0] >= 0.72:
            return fuzzy[0][1]
    return next((boq for boq in candidates if boq.items), None)


def _actuals_grouped_by_item(boq_id, package_id=None):
    query = BOQActual.query.filter_by(boq_id=boq_id)
    if package_id is not None:
        query = query.filter_by(package_id=package_id)
    grouped = {}
    for actual in query.order_by(BOQActual.recorded_at.desc()).all():
        grouped.setdefault(str(actual.item_no), []).append(actual)
    return grouped


def _actual_aggregate(actuals):
    qty = sum(float(actual.actual_qty_used or 0) for actual in actuals)
    total = sum(float(actual.actual_total_sar or 0) for actual in actuals)
    return qty, total


def _refresh_package_completion(package):
    if not package:
        return
    grouped = _actuals_grouped_by_item(package.boq_id, package.id)
    weighted_done = 0.0
    weighted_budget = 0.0
    for idx, item in enumerate(package.items, start=1):
        item_no = _boq_item_no(item, idx)
        budget_qty = _boq_item_qty(item)
        budget_total = _boq_item_total(item)
        actual_qty, _actual_total = _actual_aggregate(grouped.get(item_no, []))
        ratio = 0.0 if budget_qty <= 0 else min(actual_qty / budget_qty, 1.0)
        weighted_done += budget_total * ratio
        weighted_budget += budget_total
    package.completion_percentage = round((weighted_done / weighted_budget * 100) if weighted_budget else 0.0, 2)
    if package.completion_percentage >= 99.5:
        package.status = "completed"
    elif package.status == "completed":
        package.status = "active"


def _refresh_boq_actual_snapshots(boq_obj):
    grouped = _actuals_grouped_by_item(boq_obj.id)
    rows = []
    for idx, row in enumerate(boq_obj.items, start=1):
        updated = dict(row)
        item_no = _boq_item_no(row, idx)
        actual_qty, actual_total = _actual_aggregate(grouped.get(item_no, []))
        budget_qty = _boq_item_qty(row)
        budget_rate = _boq_item_rate(row)
        expected_for_actual_qty = actual_qty * budget_rate
        updated["actual_qty_used"] = round(actual_qty, 3)
        updated["actual_total_sar"] = round(actual_total, 2)
        updated["remaining_qty"] = round(max(budget_qty - actual_qty, 0), 3)
        updated["actual_variance_sar"] = round(actual_total - expected_for_actual_qty, 2)
        updated["actual_variance_percentage"] = round(((actual_total - expected_for_actual_qty) / expected_for_actual_qty * 100) if expected_for_actual_qty else 0, 2)
        rows.append(updated)
    boq_obj.items = rows
    for package in boq_obj.packages:
        _refresh_package_completion(package)


def _create_boq_actual(boq_obj, item, qty, source, source_id=None, recorded_by="", package_id=None, actual_rate=None, review_required=False, recorded_at=None):
    budget_qty = _boq_item_qty(item)
    budget_rate = _boq_item_rate(item)
    budget_total = _boq_item_total(item)
    entry_qty = max(_boq_float(qty), 0.0)
    applied_qty = 0.0 if review_required else entry_qty
    rate = _boq_float(actual_rate) if actual_rate not in (None, "") else budget_rate
    actual_total = applied_qty * rate
    expected_total = applied_qty * budget_rate
    variance = actual_total - expected_total
    variance_pct = (variance / expected_total * 100) if expected_total else 0.0
    recorder = sanitize_input(recorded_by or getattr(current_user, "full_name", "") or "Engineer", 120)
    if review_required and "(needs manual review)" not in recorder:
        recorder = f"{recorder} (needs manual review)"
    actual = BOQActual(
        boq_id=boq_obj.id,
        package_id=package_id,
        item_no=_boq_item_no(item),
        item_description=sanitize_input(_boq_item_desc(item), 600),
        budgeted_qty=budget_qty,
        actual_qty_used=applied_qty,
        budgeted_rate_sar=budget_rate,
        actual_rate_sar=rate if actual_rate not in (None, "") else None,
        budgeted_total_sar=budget_total,
        actual_total_sar=actual_total,
        variance_sar=variance,
        variance_percentage=variance_pct,
        source=sanitize_input(source, 30),
        source_id=source_id,
        recorded_at=recorded_at or datetime.utcnow(),
        recorded_by=recorder,
    )
    db.session.add(actual)
    db.session.flush()
    _refresh_boq_actual_snapshots(boq_obj)
    return actual


def record_actuals_from_dpr(dpr_obj):
    """Create BOQ actual records from explicit DPR progress notation."""
    if not dpr_obj or not getattr(dpr_obj, "project_id", None):
        return 0

    notes = sanitize_input(getattr(dpr_obj, "progress_notes", "") or "", 2500)
    if not notes.strip():
        return 0

    boqs = (
        BOQ.query.filter_by(project_id=dpr_obj.project_id)
        .order_by(BOQ.updated_at.desc(), BOQ.created_at.desc())
        .all()
    )
    if not boqs:
        return 0

    patterns = [
        re.compile(r"\b(?:boq|item)\s*#?\s*([A-Za-z0-9.\-_/]+)\s*(?:qty|quantity|completed|done|used|=|:)\s*([0-9]+(?:\.[0-9]+)?)", re.IGNORECASE),
        re.compile(r"\[BOQ\s*([A-Za-z0-9.\-_/]+)\]\s*([0-9]+(?:\.[0-9]+)?)", re.IGNORECASE),
    ]
    matches = []
    for pattern in patterns:
        matches.extend(pattern.findall(notes))
    if not matches:
        return 0

    BOQActual.query.filter_by(source="dpr", source_id=dpr_obj.id).delete(synchronize_session=False)

    created = 0
    recorded_at = datetime.combine(dpr_obj.date, datetime.min.time()) if dpr_obj.date else datetime.utcnow()
    for item_no, raw_qty in matches[:25]:
        qty = _boq_float(raw_qty)
        if qty <= 0:
            continue
        item_no_text = str(item_no).strip()
        matched_boq = None
        matched_item = None
        for boq_obj in boqs:
            for item in _boq_prompt_items(boq_obj.items):
                if _boq_item_no(item).lower() == item_no_text.lower():
                    matched_boq = boq_obj
                    matched_item = item
                    break
            if matched_item:
                break
        if not matched_boq or not matched_item:
            continue
        _create_boq_actual(
            matched_boq,
            matched_item,
            qty,
            source="dpr",
            source_id=dpr_obj.id,
            recorded_by=getattr(current_user, "full_name", "") or "DPR",
            recorded_at=recorded_at,
        )
        created += 1

    if created:
        db.session.commit()
    else:
        db.session.rollback()
    return created


def _tracker_item_status(budget_qty, budget_total, actual_qty, actual_total):
    qty_ratio = (actual_qty / budget_qty) if budget_qty else 0.0
    cost_ratio = (actual_total / budget_total) if budget_total else 0.0
    if cost_ratio >= 1.15 or qty_ratio >= 1.0:
        return "critical"
    if cost_ratio >= 1.05 or qty_ratio >= 0.85:
        return "at_risk"
    return "on_track"


def compute_boq_tracker_summary(boq_obj, package_id=None):
    rows = boq_obj.items
    if package_id:
        package = BOQPackage.query.filter_by(id=package_id, boq_id=boq_obj.id).first()
        rows = package.items if package else []
    grouped = _actuals_grouped_by_item(boq_obj.id, package_id)
    item_rows = []
    spent_total = 0.0
    budget_total = 0.0
    weighted_done = 0.0
    alerts = []
    worst = "on_track"
    for idx, row in enumerate(rows or [], start=1):
        item_no = _boq_item_no(row, idx)
        actuals = grouped.get(item_no, [])
        actual_qty, actual_total = _actual_aggregate(actuals)
        budget_qty = _boq_item_qty(row)
        line_budget = _boq_item_total(row)
        remaining_qty = max(budget_qty - actual_qty, 0)
        status = _tracker_item_status(budget_qty, line_budget, actual_qty, actual_total)
        if status == "critical":
            worst = "critical"
        elif status == "at_risk" and worst != "critical":
            worst = "at_risk"
        completion_ratio = min(actual_qty / budget_qty, 1.0) if budget_qty else 0.0
        weighted_done += line_budget * completion_ratio
        budget_total += line_budget
        spent_total += actual_total
        if budget_qty and actual_qty / budget_qty >= 0.85:
            alerts.append(f"{_boq_item_desc(row)} is {actual_qty / budget_qty * 100:.0f}% consumed - reorder recommended")
        if line_budget and actual_total / line_budget >= 1.12:
            alerts.append(f"{_boq_item_desc(row)} cost is {(actual_total / line_budget - 1) * 100:.0f}% over budget")
        item_rows.append(
            {
                "item_no": item_no,
                "description": _boq_item_desc(row),
                "unit": row.get("unit", ""),
                "budgeted_qty": budget_qty,
                "actual_qty": actual_qty,
                "remaining_qty": remaining_qty,
                "budget_sar": line_budget,
                "actual_sar": actual_total,
                "variance_sar": actual_total - (actual_qty * _boq_item_rate(row)),
                "status": status,
                "actuals": [actual.to_dict() for actual in actuals],
            }
        )
    completion = (weighted_done / budget_total * 100) if budget_total else 0.0
    contract_value = float(boq_obj.grand_total or 0) if not package_id else round(budget_total * 1.15, 2)
    remaining = max(contract_value - spent_total, 0.0)
    forecast = spent_total / (completion / 100) if completion > 0 else contract_value
    if forecast <= 0:
        forecast = contract_value
    if not alerts and item_rows:
        latest_actual = BOQActual.query.filter_by(boq_id=boq_obj.id).order_by(BOQActual.recorded_at.desc()).first()
        if latest_actual and latest_actual.recorded_at and (datetime.utcnow() - latest_actual.recorded_at).days >= 7:
            alerts.append("BOQ items have not been updated in 7 days")
    return {
        "contract_value": contract_value,
        "budget_total": budget_total,
        "spent_total": spent_total,
        "remaining_total": remaining,
        "completion_percentage": round(completion, 2),
        "forecast_final_cost": round(forecast, 2),
        "forecast_delta_pct": round(((forecast - contract_value) / contract_value * 100) if contract_value else 0.0, 2),
        "items": item_rows,
        "alerts": alerts[:8],
        "status": worst if spent_total > 0 else "none",
        "chart": build_boq_burn_chart(boq_obj, contract_value),
    }


def build_boq_burn_chart(boq_obj, contract_value):
    start = boq_obj.created_at or datetime.utcnow()
    actuals = BOQActual.query.filter_by(boq_id=boq_obj.id).order_by(BOQActual.recorded_at.asc()).all()
    if actuals:
        end = max([actual.recorded_at for actual in actuals if actual.recorded_at] or [datetime.utcnow()])
    else:
        end = datetime.utcnow()
    weeks = max(6, min(16, int(((end - start).days // 7) + 2)))
    labels = [f"W{k + 1}" for k in range(weeks)]
    planned = [round(contract_value * ((idx + 1) / weeks), 2) for idx in range(weeks)]
    actual_by_week = [0.0 for _ in range(weeks)]
    for actual in actuals:
        week_idx = min(max(int(((actual.recorded_at or start) - start).days // 7), 0), weeks - 1)
        actual_by_week[week_idx] += float(actual.actual_total_sar or 0)
    running = 0.0
    actual_line = []
    for value in actual_by_week:
        running += value
        actual_line.append(round(running, 2))
    return {"labels": labels, "planned": planned, "actual": actual_line}


def tracker_status_for_boq(boq_obj):
    if not boq_obj or not BOQActual.query.filter_by(boq_id=boq_obj.id).first():
        return "none"
    return compute_boq_tracker_summary(boq_obj).get("status", "none")


def _package_type_for_item(item):
    trade = classify_boq_trade(_boq_item_desc(item), item.get("notes", ""))
    trade = (trade or "").lower()
    if "mep" in trade or "elect" in trade or "plumb" in trade or "mechanical" in trade:
        return "mep"
    if "finish" in trade or "tile" in trade or "paint" in trade or "joinery" in trade:
        return "finishing"
    if "structure" in trade or "concrete" in trade or "rebar" in trade or "formwork" in trade:
        return "structure"
    if "external" in trade or "landscape" in trade:
        return "external"
    if "provisional" in trade:
        return "provisional"
    return "civil"


def fallback_package_suggestions(rows):
    labels = {
        "civil": ("Civil Works", "Civil Engineer"),
        "mep": ("MEP Package", "MEP Engineer"),
        "finishing": ("Finishing Works", "Finishing Engineer"),
        "structure": ("Structure Package", "Structural Engineer"),
        "external": ("External Works", "Site Engineer"),
        "provisional": ("Provisional Sums", "Quantity Surveyor"),
    }
    grouped = {}
    for idx, row in enumerate(rows or [], start=1):
        ptype = _package_type_for_item(row)
        grouped.setdefault(ptype, []).append(_boq_item_no(row, idx))
    suggestions = []
    row_map = {_boq_item_no(row, idx): row for idx, row in enumerate(rows or [], start=1)}
    for ptype, numbers in grouped.items():
        name, role = labels.get(ptype, ("Other Package", "Site Engineer"))
        subtotal = sum(_boq_item_total(row_map.get(no, {})) for no in numbers)
        suggestions.append(
            {
                "package_name": name,
                "package_type": ptype,
                "suggested_engineer_role": role,
                "item_numbers": numbers,
                "package_subtotal_sar": round(subtotal, 2),
                "reasoning": "Grouped by trade type and construction sequence.",
            }
        )
    return suggestions


def package_access_allowed(package):
    if not package or not getattr(current_user, "is_authenticated", False):
        return False
    if package.boq and package.boq.user_id == current_user.id:
        return True
    engineer_name = _normalize_match_text(package.assigned_engineer_name)
    engineer_email = _normalize_match_text(package.assigned_engineer_email)
    return engineer_name == _normalize_match_text(current_user.full_name) or engineer_email == _normalize_match_text(current_user.email)


BOQ_SMART_HEADER_ALIASES = {
    "item_no": ["item no", "item", "no", "s/no", "serial", "#", "رقم", "م", "مسلسل", "البند"],
    "description": [
        "description",
        "item description",
        "particulars",
        "scope",
        "work description",
        "name",
        "بيان",
        "الوصف",
        "وصف",
        "البند",
        "الأعمال",
        "الاعمال",
        "تفاصيل",
        "مواصفات",
    ],
    "unit": ["unit", "uom", "u/m", "unit of measure", "وحدة", "الوحدة", "وحدة القياس", "وحده"],
    "quantity": ["qty", "quantity", "quantities", "qnty", "الكمية", "كمية", "عدد"],
    "unit_rate": ["rate", "unit rate", "unit price", "price", "cost", "سعر الوحدة", "السعر", "سعر", "تكلفة", "معدل"],
    "total": ["total", "amount", "value", "line total", "الإجمالي", "اجمالي", "المجموع", "القيمة", "المبلغ"],
}


def _boq_header_score(header, aliases):
    import difflib

    token = _normalize_boq_text_token(header)
    if not token:
        return 0.0
    best = 0.0
    for alias in aliases:
        alias_token = _normalize_boq_text_token(alias)
        if not alias_token:
            continue
        if token == alias_token:
            best = max(best, 1.0)
        elif alias_token in token or token in alias_token:
            best = max(best, 0.86)
        else:
            best = max(best, difflib.SequenceMatcher(None, token, alias_token).ratio())
    return best


def _guess_boq_column_mapping(headers):
    mapping = {}
    used_indexes = set()
    for field, aliases in BOQ_SMART_HEADER_ALIASES.items():
        best_idx = None
        best_score = 0
        for idx, header in enumerate(headers or []):
            if idx in used_indexes:
                continue
            score = _boq_header_score(header, aliases)
            if score > best_score:
                best_idx = idx
                best_score = score
        if best_idx is not None and best_score >= 0.72:
            mapping[field] = best_idx
            used_indexes.add(best_idx)
    return mapping


def _boq_sheet_score(rows):
    if not rows:
        return 0
    score = 0
    for row in rows[:120]:
        numeric_cells = sum(1 for cell in row if _boq_to_float(cell) > 0)
        text_cells = sum(1 for cell in row if _stringify_boq_cell(cell) and not _stringify_boq_cell(cell).replace(".", "").isdigit())
        score += min(numeric_cells, 5) * 2 + min(text_cells, 4)
    for row in rows[:20]:
        headers = [_stringify_boq_cell(c) for c in row]
        mapping = _guess_boq_column_mapping(headers)
        if "description" in mapping:
            score += 20
        if "quantity" in mapping:
            score += 15
        if "unit_rate" in mapping or "total" in mapping:
            score += 12
    return score


def _detect_smart_boq_header(rows):
    best = {"idx": 0, "mapping": {}, "score": 0}
    for idx, row in enumerate((rows or [])[:25]):
        headers = [_stringify_boq_cell(c) or f"Column {i + 1}" for i, c in enumerate(row)]
        mapping = _guess_boq_column_mapping(headers)
        score = len(mapping) * 10
        if "description" in mapping:
            score += 12
        if "quantity" in mapping:
            score += 8
        if "unit_rate" in mapping or "total" in mapping:
            score += 8
        if score > best["score"]:
            best = {"idx": idx, "mapping": mapping, "score": score}
    return best["idx"], best["mapping"], best["score"]


def _headers_from_row(row):
    headers = []
    seen = set()
    for idx, cell in enumerate(row or []):
        base = _stringify_boq_cell(cell) or f"Column {idx + 1}"
        candidate = base
        suffix = 2
        while candidate.lower() in seen:
            candidate = f"{base} {suffix}"
            suffix += 1
        seen.add(candidate.lower())
        headers.append(candidate)
    return headers


def _ai_map_boq_columns(headers, sample_rows):
    system_prompt = "You map BOQ spreadsheet columns for construction quantity surveyors. Respond in JSON only."
    user_prompt = f"""These are column headers from a BOQ Excel file: {headers}
Map them to these fields: description, unit, quantity, unit_rate, total. Some columns may not exist.
Use these first rows as context: {sample_rows}
Return JSON: {{"description": "col_name", "unit": "col_name", "quantity": "col_name", "unit_rate": "col_name", "total": "col_name", "unmapped": ["col1","col2"]}}"""
    success, parsed, _raw = call_openai_json(system_prompt, user_prompt, max_tokens=700, temperature=0.0)
    if not success:
        app.logger.warning("BOQ AI column mapping failed: %s", parsed)
        return {}
    index_by_name = {str(header).strip().lower(): idx for idx, header in enumerate(headers)}
    mapping = {}
    for field in ("description", "unit", "quantity", "unit_rate", "total"):
        value = str(parsed.get(field, "") or "").strip().lower()
        if value in index_by_name:
            mapping[field] = index_by_name[value]
    return mapping


def _extract_smart_rows_with_mapping(sheet_name, rows, header_idx, headers, mapping, mapping_source="fuzzy"):
    preview_rows = []
    warnings = []
    if not mapping or "description" not in mapping:
        warnings.append("Could not confidently identify the description column.")
        return preview_rows, warnings

    def get(row, field):
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return ""
        return row[idx]

    for offset, raw_row in enumerate((rows or [])[header_idx + 1 :], start=header_idx + 2):
        row = list(raw_row) + [None] * max(0, len(headers) - len(raw_row))
        if not any(not _boq_cell_is_empty(cell) for cell in row):
            continue
        description = _stringify_boq_cell(get(row, "description"))
        if _is_boq_placeholder_text(description):
            continue
        if description.lower().strip() in {"total", "subtotal", "grand total"} or description.strip() in {"الإجمالي", "المجموع"}:
            continue
        unit = _stringify_boq_cell(get(row, "unit")) or "No."
        quantity = _boq_float(get(row, "quantity"))
        unit_rate = _boq_float(get(row, "unit_rate"))
        total = _boq_float(get(row, "total"))
        if unit_rate <= 0 and total > 0 and quantity > 0:
            unit_rate = round(total / quantity, 2)
        if total <= 0 and quantity > 0 and unit_rate > 0:
            total = round(quantity * unit_rate, 2)

        confidence = 90 if quantity > 0 and (unit_rate > 0 or total > 0) else 65
        row_warnings = []
        if quantity <= 0:
            confidence -= 25
            row_warnings.append("quantity missing or zero")
        if unit_rate <= 0 and total <= 0:
            confidence -= 15
            row_warnings.append("rate/total missing")
        if not description:
            confidence -= 40
            row_warnings.append("description missing")
        if mapping_source == "ai":
            confidence -= 5

        if not description and quantity <= 0 and total <= 0:
            continue
        item = {
            "id": len(preview_rows) + 1,
            "selected": confidence >= 55,
            "source_sheet": sheet_name,
            "row_number": offset,
            "description": sanitize_input(description, 600),
            "unit": sanitize_input(unit, 30) or "No.",
            "quantity": round(quantity, 3),
            "unit_rate": round(unit_rate, 2),
            "total": round(total, 2),
            "rate": round(unit_rate, 2),
            "notes": f"{sheet_name} row {offset}",
            "confidence": max(0, min(100, confidence)),
            "warning": "; ".join(row_warnings),
        }
        if row_warnings:
            warnings.append(f"Row {offset}: {item['warning']}")
        preview_rows.append(item)
    return preview_rows, warnings[:30]


def _fallback_items_to_preview(items, source_name="BOQ Upload"):
    rows = []
    warnings = []
    for idx, item in enumerate(items or [], start=1):
        normalized = _normalize_boq_editor_item(item, idx)
        confidence = 85
        row_warnings = []
        if normalized["quantity"] <= 0:
            confidence -= 25
            row_warnings.append("quantity missing or zero")
        if normalized["rate"] <= 0 and normalized["total"] <= 0:
            confidence -= 15
            row_warnings.append("rate/total missing")
        rows.append(
            {
                "id": idx,
                "selected": confidence >= 55,
                "source_sheet": item.get("source_sheet", source_name),
                "row_number": item.get("row_number", idx),
                "description": normalized["description"],
                "unit": normalized["unit"],
                "quantity": normalized["quantity"],
                "unit_rate": normalized["rate"],
                "rate": normalized["rate"],
                "total": normalized["total"],
                "notes": normalized["notes"] or item.get("section", ""),
                "confidence": confidence,
                "warning": "; ".join(row_warnings),
            }
        )
        if row_warnings:
            warnings.append(f"Row {idx}: {'; '.join(row_warnings)}")
    return rows, warnings[:30]


def parse_boq_excel_smart(file_bytes, filename):
    sheet_names = []
    try:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        candidates = []
        for worksheet in workbook.worksheets[:15]:
            rows = _expand_worksheet_merged_rows(worksheet)
            if not rows:
                continue
            sheet_names.append(worksheet.title)
            candidates.append((_boq_sheet_score(rows), worksheet.title, rows))
        candidates.sort(key=lambda item: item[0], reverse=True)
        if not candidates:
            return {"rows": [], "warnings": ["No readable rows found in this workbook."], "sheet_names": sheet_names}

        _score, sheet_name, rows = candidates[0]
        header_idx, mapping, map_score = _detect_smart_boq_header(rows)
        headers = _headers_from_row(rows[header_idx] if header_idx < len(rows) else [])
        mapping_source = "fuzzy"
        if map_score < 28 or "description" not in mapping:
            sample_rows = [
                [_stringify_boq_cell(cell) for cell in row]
                for row in rows[header_idx + 1 : header_idx + 6]
            ]
            ai_mapping = _ai_map_boq_columns(headers, sample_rows)
            if ai_mapping:
                mapping = ai_mapping
                mapping_source = "ai"

        preview_rows, warnings = _extract_smart_rows_with_mapping(sheet_name, rows, header_idx, headers, mapping, mapping_source)
        if preview_rows:
            return {
                "rows": preview_rows,
                "warnings": warnings,
                "sheet_names": sheet_names,
                "detected_sheet": sheet_name,
                "mapping": {field: headers[idx] for field, idx in mapping.items() if idx < len(headers)},
                "mapping_source": mapping_source,
                "method": "openpyxl_smart",
            }
    except Exception as error:
        app.logger.warning("openpyxl smart BOQ parse failed for %s: %s", filename, error)

    try:
        sheet_names, extracted_items, _flat_rows = parse_boq_excel_with_pandas(file_bytes, filename)
        preview_rows, warnings = _fallback_items_to_preview(extracted_items, "Excel Upload")
        return {
            "rows": preview_rows,
            "warnings": warnings,
            "sheet_names": sheet_names,
            "detected_sheet": sheet_names[0] if sheet_names else "",
            "mapping": {},
            "mapping_source": "fallback",
            "method": "existing_universal_parser",
        }
    except Exception as error:
        app.logger.exception("Smart BOQ Excel fallback failed")
        return {
            "rows": [],
            "warnings": ["Could not parse this Excel file. Check that it is not password-protected and contains BOQ rows."],
            "error": _friendly_ai_error(error) if "timeout" in str(error).lower() else "Could not parse this Excel file.",
            "sheet_names": sheet_names,
        }


def _parse_pdf_boq_text_with_ai(text_content, project_name="Construction Project"):
    if not text_content.strip():
        return [], ["No extractable PDF text found."]
    system_prompt = "You extract Bill of Quantities line items from construction PDFs. Respond in JSON only."
    user_prompt = f"""Extract BOQ line items from this PDF text for project: {project_name}.

Return JSON only:
{{
  "items": [
    {{
      "description": "string",
      "unit": "string",
      "quantity": number,
      "unit_rate": number,
      "total": number,
      "notes": "string"
    }}
  ],
  "warnings": ["string"]
}}

PDF TEXT:
{text_content[:9000]}"""
    success, parsed, _raw = call_openai_json(system_prompt, user_prompt, max_tokens=3500, temperature=0.0)
    if not success:
        return [], [parsed]
    items = parsed.get("items", []) if isinstance(parsed, dict) else []
    warnings = parsed.get("warnings", []) if isinstance(parsed, dict) else []
    rows, row_warnings = _fallback_items_to_preview(items, "PDF Upload")
    return rows, (warnings or []) + row_warnings


def _extract_pdf_text(file_bytes):
    import io

    parts = []
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages[:20]:
                text = page.extract_text() or ""
                if text:
                    parts.append(text)
                for table in page.extract_tables() or []:
                    for row in table or []:
                        if row:
                            parts.append(" | ".join(str(cell) if cell is not None else "" for cell in row))
    except Exception:
        app.logger.exception("PDF text extraction failed")
    return "\n".join(parts)


def _parse_scanned_pdf_boq_with_vision(file_bytes, project_name="Construction Project"):
    if client is None:
        return [], ["AI vision extraction is not configured. Please set OPENAI_API_KEY."]
    import base64
    import io

    try:
        import pypdfium2 as pdfium
    except Exception:
        return [], ["Scanned PDF vision fallback is unavailable because pypdfium2 is not installed."]

    try:
        pdf = pdfium.PdfDocument(io.BytesIO(file_bytes))
        image_parts = []
        page_count = min(len(pdf), 3)
        for page_index in range(page_count):
            page = pdf[page_index]
            bitmap = page.render(scale=1.6)
            pil_image = bitmap.to_pil()
            image_buffer = io.BytesIO()
            pil_image.save(image_buffer, format="JPEG", quality=82)
            image_parts.append(
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/jpeg;base64,{base64.b64encode(image_buffer.getvalue()).decode('ascii')}"
                    },
                }
            )
    except Exception:
        app.logger.exception("Scanned PDF render failed")
        return [], ["Could not render the scanned PDF for vision extraction."]

    if not image_parts:
        return [], ["No scanned PDF pages were available for vision extraction."]

    prompt = f"""Extract BOQ rows from these scanned PDF pages for project: {project_name}.
Return JSON only:
{{
  "items": [
    {{"description": "string", "unit": "string", "quantity": number, "unit_rate": number, "total": number, "notes": "string"}}
  ],
  "warnings": ["string"]
}}"""
    try:
        kwargs = {
            "model": app.config.get("OPENAI_MODEL", "gpt-4o-mini"),
            "messages": [
                {
                    "role": "user",
                    "content": [{"type": "text", "text": prompt}] + image_parts,
                }
            ],
            "max_tokens": 3500,
            "temperature": 0.0,
        }
        try:
            response = client.chat.completions.create(**kwargs, response_format={"type": "json_object"})
        except TypeError:
            response = client.chat.completions.create(**kwargs)
        session["ai_calls"] = session.get("ai_calls", 0) + 1
        session.modified = True
        parsed = _parse_json_like_response(response.choices[0].message.content or "")
        items = parsed.get("items", []) if isinstance(parsed, dict) else []
        warnings = parsed.get("warnings", []) if isinstance(parsed, dict) else []
        rows, row_warnings = _fallback_items_to_preview(items, "Scanned PDF Upload")
        return rows, (warnings or []) + row_warnings
    except Exception as error:
        app.logger.exception("Scanned PDF vision extraction failed")
        return [], [_friendly_ai_error(error)]


def parse_boq_pdf_smart(file_bytes, filename, project_name="Construction Project"):
    text_content = _extract_pdf_text(file_bytes)
    rows, warnings = _parse_pdf_boq_text_with_ai(text_content, project_name)
    if len(rows) < 10:
        vision_rows, vision_warnings = _parse_scanned_pdf_boq_with_vision(file_bytes, project_name)
        if len(vision_rows) > len(rows):
            rows = vision_rows
            warnings = (warnings or []) + ["Regular PDF extraction returned fewer than 10 items; vision extraction was used."] + (vision_warnings or [])
    return {
        "rows": rows,
        "warnings": warnings[:40],
        "sheet_names": ["PDF Upload"],
        "detected_sheet": "PDF Upload",
        "mapping": {},
        "mapping_source": "pdf_text_or_vision",
        "method": "pdf_smart",
        "raw_text_length": len(text_content or ""),
    }


def _compute_financial_summary(quotation_items):
    supply_subtotal = float(sum(float(i.get("total_sar", 0) or 0) for i in quotation_items))
    freight = round(supply_subtotal * 0.03, 2)
    installation = round(supply_subtotal * 0.15, 2)
    subtotal_before_vat = round(supply_subtotal + freight + installation, 2)
    vat = round(subtotal_before_vat * 0.15, 2)
    grand_total = round(subtotal_before_vat + vat, 2)
    return {
        "supply_subtotal": round(supply_subtotal, 2),
        "freight_handling_3pct": freight,
        "installation_15pct": installation,
        "subtotal_before_vat": subtotal_before_vat,
        "vat_15pct": vat,
        "grand_total": grand_total,
        "currency": "SAR",
    }


def _build_boq_ai_source_rows(items):
    rows = []
    for idx, item in enumerate(items, start=1):
        quantity = float(item.get("quantity", 0) or 0)
        input_rate = float(item.get("rate", 0) or 0)
        rows.append(
            {
                "source_index": idx,
                "no": str(item.get("no", idx) or idx),
                "model": str(item.get("model", "") or ""),
                "description": str(item.get("description", "") or ""),
                "unit": str(item.get("unit", "No.") or "No."),
                "quantity": quantity,
                "section": str(item.get("section", "General") or "General"),
                "source_sheet": str(item.get("source_sheet", "") or ""),
                "trade_category": str(item.get("trade_category", "") or ""),
                "input_rate": round(input_rate, 2),
                "input_total": round(quantity * input_rate, 2) if input_rate > 0 else 0,
            }
        )
    return rows


def _build_boq_section_summaries(quotation_items):
    grouped = {}
    for item in quotation_items:
        section = item.get("section", "General") or "General"
        bucket = grouped.setdefault(
            section,
            {
                "name": section,
                "item_count": 0,
                "estimated_value_sar": 0.0,
                "samples": [],
            },
        )
        bucket["item_count"] += 1
        bucket["estimated_value_sar"] += float(item.get("total_sar", 0) or 0)
        sample = item.get("description") or item.get("model") or ""
        if sample and len(bucket["samples"]) < 3:
            bucket["samples"].append(sample[:80])

    rows = []
    for bucket in grouped.values():
        rows.append(
            {
                "name": bucket["name"],
                "item_count": bucket["item_count"],
                "estimated_value_sar": round(bucket["estimated_value_sar"], 2),
                "key_items": ", ".join(bucket["samples"]),
            }
        )
    rows.sort(key=lambda row: row["estimated_value_sar"], reverse=True)
    return rows


def _infer_boq_measurement_method(item):
    description = _normalize_boq_text_token(item.get("description"))
    unit = _normalize_boq_text_token(item.get("unit"))
    if any(keyword in description for keyword in ["concrete", "excav", "backfill", "blinding", "earthwork"]) or unit in {"m3", "cum"}:
        return "Measure by executed volume in m3 against approved dimensions and pour/excavation records."
    if any(keyword in description for keyword in ["paint", "plaster", "tile", "ceiling", "partition", "waterproof", "screed"]) or unit in {"m2", "sqm", "sq m"}:
        return "Measure by completed surface area in m2 from approved drawings and site measurements."
    if any(keyword in description for keyword in ["cable", "wire", "pipe", "piping", "duct", "tray", "conduit"]) or unit in {"m", "mtr", "meter", "metre"}:
        return "Measure by installed running length in m after routing is approved on site."
    if unit in {"no", "nos", "nr", "pcs", "pc", "set", "sets", "item"} or any(keyword in description for keyword in ["panel", "pump", "ahu", "fcu", "db", "controller", "valve", "equipment"]):
        return "Measure by counted installed units against approved submittals and inspection records."
    if unit in {"ls", "lot"}:
        return "Measure as lump sum against defined scope boundaries, deliverables, and milestone completion."
    return f"Measure using the BOQ unit `{item.get('unit', 'No.')}` and maintain signed site measurement records."


def _build_boq_data_issues(quotation_items):
    issues = []
    seen = {}

    for item in quotation_items:
        desc = str(item.get("description", "") or "").strip()
        section = str(item.get("section", "General") or "General").strip()
        unit = str(item.get("unit", "") or "").strip()
        qty = float(item.get("quantity", 0) or 0)

        if qty <= 0:
            issues.append(
                {
                    "severity": "high",
                    "issue": f"Missing or zero quantity for `{desc or item.get('no', 'item')}`.",
                    "recommendation": "Verify the BOQ row and confirm the measured quantity before billing or procurement.",
                }
            )
        if not unit:
            issues.append(
                {
                    "severity": "high",
                    "issue": f"Missing unit of measurement for `{desc or item.get('no', 'item')}`.",
                    "recommendation": "Assign the correct billing unit such as m, m2, m3, No., or LS before using this item.",
                }
            )
        if not section or section.lower() == "general":
            issues.append(
                {
                    "severity": "medium",
                    "issue": f"Section is missing or generic for `{desc or item.get('no', 'item')}`.",
                    "recommendation": "Tag this BOQ row to the right trade package so planning, billing, and procurement stay aligned.",
                }
            )

        desc_key = _normalize_boq_text_token(desc)
        if desc_key:
            dup_key = (section.lower(), desc_key, unit.lower())
            seen[dup_key] = seen.get(dup_key, 0) + 1

        if qty > 0:
            desc_norm = _normalize_boq_text_token(desc)
            unit_norm = _normalize_boq_text_token(unit)
            if any(keyword in desc_norm for keyword in ["concrete", "excav", "backfill"]) and unit_norm not in {"m3", "cum"}:
                issues.append(
                    {
                        "severity": "medium",
                        "issue": f"Possible unit mismatch for `{desc}`.",
                        "recommendation": "Concrete and excavation items are usually billed in m3. Confirm that the uploaded unit is correct.",
                    }
                )
            if any(keyword in desc_norm for keyword in ["paint", "plaster", "tile", "ceiling", "partition"]) and unit_norm not in {"m2", "sqm", "sq m"}:
                issues.append(
                    {
                        "severity": "medium",
                        "issue": f"Possible area unit mismatch for `{desc}`.",
                        "recommendation": "Finishing items are commonly measured in m2. Recheck the BOQ unit and quantity basis.",
                    }
                )
            if qty > 100000:
                issues.append(
                    {
                        "severity": "medium",
                        "issue": f"Suspiciously high quantity detected for `{desc}`: {qty:,.0f}.",
                        "recommendation": "Review decimal placement, unit conversion, and whether this quantity should be split across multiple items.",
                    }
                )

    for (section, desc_key, unit), count in seen.items():
        if count > 1:
            issues.append(
                {
                    "severity": "medium",
                    "issue": f"Duplicate BOQ item detected in section `{section.title()}`.",
                    "recommendation": f"The same description/unit appears {count} times. Confirm whether these are genuine repeated items or duplicate uploads.",
                }
            )

    unique = []
    dedupe = set()
    for issue in issues:
        key = (issue["severity"], issue["issue"])
        if key in dedupe:
            continue
        dedupe.add(key)
        unique.append(issue)
    return unique[:12]


def _build_boq_billing_guidance(quotation_items):
    billable_items = [item for item in quotation_items if float(item.get("quantity", 0) or 0) > 0]
    high_value_items = sorted(
        billable_items,
        key=lambda item: float(item.get("total_sar", 0) or 0),
        reverse=True,
    )[:6]
    measurement_methods = []
    for item in high_value_items[:5]:
        measurement_methods.append(
            {
                "description": item.get("description", "") or item.get("model", ""),
                "section": item.get("section", "General"),
                "method": _infer_boq_measurement_method(item),
                "quantity": float(item.get("quantity", 0) or 0),
                "unit": item.get("unit", "No."),
            }
        )

    payment_notes = [
        "Prepare IPC / payment applications from signed site measurements, approved drawings, and inspection records.",
        "Separate confirmed billable quantities from provisional or missing-scope items before issuing the next valuation.",
        "Flag any row with zero quantity, blank unit, or missing section before using it for billing or subcontract payment.",
    ]
    if any(float(item.get("unit_rate_sar", 0) or 0) <= 0 for item in billable_items):
        payment_notes.append("Some rows still have no confirmed rate. Validate market rate support or contract rate references before final payment certification.")

    summary = (
        f"{len(billable_items)} billable BOQ rows were identified. "
        f"Use the uploaded quantity and unit as the billing basis, then certify payment using approved measurement sheets and executed progress."
    )
    return {
        "summary": summary,
        "billable_items_count": len(billable_items),
        "measurement_methods": measurement_methods,
        "payment_notes": payment_notes,
    }


def _build_boq_extraction_summary(sheet_names, quotation_items, data_issues):
    sections = sorted({str(item.get("section", "General") or "General") for item in quotation_items})
    return {
        "sheet_count": len(sheet_names or []),
        "sheet_names": list(sheet_names or [])[:8],
        "item_count": len(quotation_items),
        "section_count": len(sections),
        "section_names": sections[:8],
        "issue_count": len(data_issues or []),
    }


def enrich_boq_analysis_payload(analyzed, sheet_names):
    quotation_items = analyzed.get("quotation_items", []) or []
    data_issues = _build_boq_data_issues(quotation_items)
    billing_guidance = _build_boq_billing_guidance(quotation_items)
    extraction_summary = _build_boq_extraction_summary(sheet_names, quotation_items, data_issues)
    analyzed["data_issues"] = data_issues
    analyzed["billing_guidance"] = billing_guidance
    analyzed["extraction_summary"] = extraction_summary
    return analyzed


def _fallback_boq_analysis(items, summary_text, market_note):
    quotation_items = []
    for item in items:
        qty = float(item.get("quantity", 0) or 0)
        rate = float(item.get("rate", 0) or 0)
        total = round(qty * rate, 2)
        quotation_items.append(
            {
                "no": item.get("no", ""),
                "model": item.get("model", ""),
                "description": item.get("description", ""),
                "unit": item.get("unit", "No."),
                "quantity": qty,
                "unit_rate_sar": round(rate, 2),
                "total_sar": total,
                "section": item.get("section", "General"),
                "rate_basis": "Input" if rate > 0 else "TBC",
            }
        )
    return {
        "project_type": "Construction BOQ",
        "scope_summary": summary_text,
        "total_items_found": len(quotation_items),
        "sections": _build_boq_section_summaries(quotation_items),
        "quotation_items": quotation_items,
        "financial_summary": _compute_financial_summary(quotation_items),
        "executive_summary": summary_text,
        "risk_flags": [],
        "recommendations": [],
        "overall_rating": "Fair",
        "confidence_score": 55,
        "market_notes": market_note,
        "original_items": items,
    }


def ai_analyze_boq(items, project_name, client_name, sheet_count, sheets, ai_estimate_rates):
    """
    Sends extracted BOQ items to AI and gets quotation-ready analysis.
    """
    import json

    if not items:
        return {"error": "No items extracted", "items": [], "totals": {}}

    total_items = len(items)
    source_rows = _build_boq_ai_source_rows(items)
    source_rows_json = json.dumps(source_rows, ensure_ascii=False)
    rate_policy = (
        "AI must provide unit_rate_sar and total_sar for every source row using current Saudi market pricing."
        if ai_estimate_rates
        else "Prefer any provided input_rate as context, but still return unit_rate_sar and total_sar for every source row."
    )

    prompt = f"""You are BanaaIQ, an advanced BOQ (Bill of Quantities) Analysis AI for Saudi Arabian construction projects.
Act like a combination of quantity surveyor, site engineer, and cost engineer.
You must keep the uploaded BOQ data exact. Do NOT rewrite, merge, delete, reorder, or invent line items.

PROJECT: {project_name}
CLIENT: {client_name or 'Not specified'}
FILE SHEETS: {sheet_count} sheets ({', '.join(sheets[:8])})
TOTAL LINE ITEMS: {total_items}

SOURCE_ROWS_JSON:
{source_rows_json}

RULES:
1. SOURCE_ROWS_JSON is the master source of truth.
2. Keep the same number of rows and return one pricing row for every source_index.
3. Do not change description, model, quantity, unit, or section. Backend will preserve them exactly.
4. Only provide pricing and practical construction analysis.
5. {rate_policy}
6. total_sar should align with quantity * unit_rate_sar.
7. Highlight practical project risks, execution gaps, and commercial actions.
8. Use concise professional English.

Respond ONLY as valid JSON:
{{
  "project_type": "MEP Works",
  "scope_summary": "2-3 sentence description based on the uploaded BOQ",
  "pricing_rows": [
    {{
      "source_index": 1,
      "unit_rate_sar": 11500,
      "total_sar": 7613000,
      "rate_basis": "Saudi market estimate Q2 2026"
    }}
  ],
  "executive_summary": "Professional summary suitable for tender review",
  "risk_flags": [
    {{
      "issue": "Example issue",
      "severity": "medium",
      "recommendation": "Example recommendation"
    }}
  ],
  "recommendations": [
    "Recommendation 1",
    "Recommendation 2",
    "Recommendation 3"
  ],
  "overall_rating": "Good",
  "confidence_score": 82,
  "market_notes": "Saudi market-based estimate note"
}}"""

    success, result = call_openai(prompt=prompt, max_tokens=3500, temperature=0.1)
    if not success:
        return {"error": result, "quotation_items": items}

    try:
        parsed = json.loads(_normalize_json_response(result))
    except Exception:
        return _fallback_boq_analysis(
            items,
            "AI returned non-JSON; showing the exact parsed BOQ rows with available pricing.",
            "Fallback result generated from the uploaded BOQ rows.",
        )

    ai_rows = parsed.get("pricing_rows") or parsed.get("quotation_items") or []
    pricing_lookup = {}
    for idx, row in enumerate(ai_rows, start=1):
        source_index = int(row.get("source_index", idx) or idx)
        pricing_lookup[source_index] = row

    normalized_rows = []
    for idx, src in enumerate(items, start=1):
        pricing = pricing_lookup.get(idx, {})
        qty = float(src.get("quantity", 0) or 0)
        in_rate = float(src.get("rate", 0) or 0)
        ai_rate = float(pricing.get("unit_rate_sar", pricing.get("rate", 0)) or 0)
        rate = ai_rate if ai_rate > 0 else in_rate
        total = float(pricing.get("total_sar", 0) or 0)
        calc_total = round(qty * rate, 2)
        if total <= 0 or (calc_total > 0 and abs(total - calc_total) > max(1.0, calc_total * 0.02)):
            total = calc_total
        normalized_rows.append(
            {
                "no": str(src.get("no", idx) or idx),
                "model": str(src.get("model", "") or ""),
                "description": str(src.get("description", "") or ""),
                "unit": str(src.get("unit", "No.") or "No."),
                "quantity": qty,
                "unit_rate_sar": round(rate, 2),
                "total_sar": round(total, 2),
                "section": str(src.get("section", "General") or "General"),
                "rate_basis": pricing.get("rate_basis", "AI market estimate" if ai_rate > 0 else ("Input" if in_rate > 0 else "TBC")),
            }
        )

    parsed["quotation_items"] = normalized_rows
    parsed["sections"] = _build_boq_section_summaries(normalized_rows)
    parsed["total_items_found"] = len(normalized_rows)
    parsed["financial_summary"] = _compute_financial_summary(normalized_rows)
    parsed["original_items"] = items
    if not parsed.get("scope_summary"):
        parsed["scope_summary"] = "Uploaded BOQ parsed successfully. AI pricing has been applied to the exact extracted rows."
    if not parsed.get("executive_summary"):
        parsed["executive_summary"] = parsed["scope_summary"]
    parsed["pricing_mode"] = "ai_estimated"
    return parsed


def ai_analyze_pdf_boq(text_content, project_name, client_name):
    """Analyzes PDF BOQ text content and returns quotation-ready structure."""
    import json

    prompt = f"""You are BanaaIQ, an advanced BOQ Analysis AI for Saudi Arabian construction projects.
Extract every possible BOQ row from this PDF text, reconstruct broken table rows logically, and return a practical quotation-ready result.
Keep descriptions as close as possible to the source PDF wording and do not silently drop relevant rows.

PROJECT: {project_name}
CLIENT: {client_name or 'Not specified'}

PDF CONTENT:
{text_content[:4000]}

Respond ONLY as valid JSON:
{{
  "project_type": "type of works",
  "scope_summary": "what this covers",
  "total_items_found": 25,
  "quotation_items": [
    {{
      "no": "1",
      "model": "",
      "description": "item description",
      "unit": "m2",
      "quantity": 250,
      "unit_rate_sar": 85,
      "total_sar": 21250,
      "section": "Civil Works"
    }}
  ],
  "financial_summary": {{
    "supply_subtotal": 500000,
    "freight_handling_3pct": 15000,
    "installation_15pct": 75000,
    "subtotal_before_vat": 590000,
    "vat_15pct": 88500,
    "grand_total": 678500,
    "currency": "SAR"
  }},
  "executive_summary": "professional summary",
  "risk_flags": [],
  "recommendations": [],
  "overall_rating": "Good",
  "confidence_score": 75
}}"""

    success, result = call_openai(prompt=prompt, max_tokens=2000, temperature=0.2)
    if not success:
        return {"error": result}

    try:
        parsed = json.loads(_normalize_json_response(result))
        rows = parsed.get("quotation_items") or []
        normalized_rows = []
        for idx, row in enumerate(rows, start=1):
            qty = float(row.get("quantity", 0) or 0)
            rate = float(row.get("unit_rate_sar", row.get("rate", 0)) or 0)
            total = float(row.get("total_sar", 0) or 0)
            if total <= 0:
                total = round(qty * rate, 2)
            normalized_rows.append(
                {
                    "no": str(row.get("no", idx)),
                    "model": str(row.get("model", "") or ""),
                    "description": str(row.get("description", "") or ""),
                    "unit": str(row.get("unit", "No.") or "No."),
                    "quantity": qty,
                    "unit_rate_sar": round(rate, 2),
                    "total_sar": round(total, 2),
                    "section": str(row.get("section", "General") or "General"),
                }
            )
        parsed["quotation_items"] = normalized_rows
        parsed["total_items_found"] = parsed.get("total_items_found") or len(normalized_rows)
        if not parsed.get("financial_summary"):
            parsed["financial_summary"] = _compute_financial_summary(normalized_rows)
        return parsed
    except Exception:
        return {
            "project_type": "Construction BOQ",
            "scope_summary": "PDF analyzed",
            "executive_summary": result[:500],
            "quotation_items": [],
            "financial_summary": _compute_financial_summary([]),
            "risk_flags": [],
            "recommendations": [],
            "overall_rating": "Fair",
            "confidence_score": 50,
        }


def ai_analyze_boq_by_role(extracted_data, reviewer_name, engineer_role, project_name="", job_title=None):
    """
    Sends extracted BOQ data to GPT-4o-mini with role-specific instructions
    and returns a complete analysis tailored to the selected engineer role.
    """
    import json

    role_config = ENGINEER_ROLES.get(engineer_role, ENGINEER_ROLES["quantity_surveyor"])
    role_label = role_config["label"]
    # If job_title is provided, use it to enrich the role context
    effective_role = job_title or role_label
    focus_areas = ", ".join(role_config["focus"])
    key_questions = "\n".join([f"- {q}" for q in role_config["questions"]])
    bill_items = "\n".join([f"- {b}" for b in role_config["bill_items"]])

    sections_summary = ""
    for sec in extracted_data.get("sections", []):
        title = sec.get("title", "Section")
        items = sec.get("items", [])
        sections_summary += f"\n[{title}] - {len(items)} items\n"
        for item in items[:8]:
            sections_summary += f"  - {item.get('description', '')} (Qty: {item.get('quantity', 0)})\n"

    raw_text = extracted_data.get("raw_text", "")[:4000]
    prompt = f"""You are a senior construction consultant reviewing a BOQ (Bill of Quantities) for a Saudi Arabian project.

You are reviewing this BOQ specifically as a:
{effective_role.upper()} ({role_label})

Reviewer: {reviewer_name}
Project: {extracted_data.get('project_title') or project_name or 'Construction Project'}
File: {extracted_data.get('file_name', '')}

EXTRACTED BOQ CONTENT:
{sections_summary if sections_summary else raw_text}

FULL RAW DATA:
{raw_text}

As a {role_label}, your focus areas are:
{focus_areas}

YOUR TASK - Provide a complete structured analysis with ALL of the following sections:

1. BOQ OVERVIEW
   - What this BOQ covers (project type, scope)
   - Total number of line items found
   - Sections identified
   - Key equipment/materials listed

2. ROLE-SPECIFIC ANALYSIS for {role_label}
   - What parts of this BOQ are directly relevant to your role
   - Item-by-item review of relevant sections
   - Technical observations for your discipline

3. KEY QUESTIONS TO VERIFY
   Answer or flag each of these:
{key_questions}

4. BILL OF WORKS REQUIRED
   Based on this BOQ, as {role_label} you need to include or verify these cost items in your bill:
{bill_items}
   Add any additional items specific to this BOQ content.

5. COST ESTIMATE GUIDANCE (SAR)
   Provide Saudi market 2025/2026 unit rates and estimated costs for each relevant item.
   Format as a table:
   | Item | Unit | Qty | Unit Rate SAR | Total SAR |
   Include installation labour at 15-20%.
   Include VAT 15% at the end.

6. RISKS AND GAPS
   - What is missing from this BOQ
   - Items that need clarification
   - Quantities that seem incorrect
   - Saudi compliance requirements (SASO, etc.)

7. RECOMMENDED ACTIONS
   List 5-10 specific actions this {role_label} should take before signing off this BOQ.

8. EXECUTIVE SUMMARY
   2-3 paragraphs suitable for a formal report. Professional language.

Respond with clear section headers.
Be specific - reference actual model numbers, quantities and items from the BOQ content.
All costs in SAR."""

    success, result = call_openai(prompt=prompt, max_tokens=2500, temperature=0.2)
    return success, result


def build_boq_analysis_pdf_buffer(title, project_name, status, analysis_text):
    import io

    navy = colors.HexColor("#0a0a0a")
    gold = colors.HexColor("#e8c547")
    light = colors.HexColor("#f6f3ec")

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    story = []

    hdr = [[
        Paragraph(
            '<b><font color="white" size="14">BanaaIQ | BOQ Analysis Report</font></b>',
            ps("H", fontName="Helvetica-Bold", textColor=colors.white, fontSize=14),
        ),
        Paragraph(
            f'<font color="#e8c547" size="9"><b>ROLE-BASED REVIEW</b></font><br/>'
            f'<font color="#aaa" size="8">{datetime.now().strftime("%d %B %Y")}</font>',
            ps("HR", fontName="Helvetica", fontSize=9, textColor=colors.white, alignment=TA_RIGHT),
        ),
    ]]
    ht = Table(hdr, colWidths=[11 * cm, 6 * cm])
    ht.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), navy), ("PADDING", (0, 0), (-1, -1), 14), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(ht)
    story.append(Spacer(1, 0.4 * cm))

    info_data = [
        ["Project:", title or "—"],
        ["File / Project:", project_name or "—"],
        ["Status:", status or "Analyzed"],
        ["Generated:", datetime.now().strftime("%d %B %Y %H:%M")],
    ]
    info_table = Table(info_data, colWidths=[3 * cm, 14 * cm])
    info_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, -1), light),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e4ddd0")),
            ]
        )
    )
    story.append(info_table)
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=gold))
    story.append(Spacer(1, 0.3 * cm))

    for line in (analysis_text or "").split("\n"):
        line = line.strip()
        if not line:
            story.append(Spacer(1, 0.08 * cm))
            continue
        if line.startswith(tuple(f"{i}." for i in range(1, 9))) or (line.isupper() and len(line) > 5):
            story.append(Spacer(1, 0.2 * cm))
            story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e4ddd0")))
            story.append(Paragraph(line, ps("SH", fontName="Helvetica-Bold", fontSize=11, textColor=navy, spaceBefore=6, spaceAfter=4)))
        elif line.startswith(("- ", "• ", "* ")):
            story.append(Paragraph("  " + line[2:], ps("BL", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#1a1a1a"), leftIndent=12, spaceBefore=2, leading=13)))
        elif "|" in line:
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if len(cells) >= 3:
                row_data = [Paragraph(c, ps(f"TC{i}", fontName="Helvetica", fontSize=8)) for i, c in enumerate(cells)]
                col_w = [17 * cm / len(cells)] * len(cells)
                table = Table([row_data], colWidths=col_w)
                table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 5), ("FONTSIZE", (0, 0), (-1, -1), 8)]))
                story.append(table)
            else:
                story.append(Paragraph(line, ps("NL", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#1a1a1a"), leading=13)))
        else:
            story.append(Paragraph(line, ps("NM", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#1a1a1a"), leading=14, spaceAfter=2)))

    story.append(Spacer(1, 0.5 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e4ddd0")))
    story.append(Paragraph("Generated by BanaaIQ AI | For guidance only - verify all quantities and rates independently", ps("FT", fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#999"), alignment=TA_CENTER)))
    doc.build(story)
    buffer.seek(0)
    return buffer


class BoqScoreGauge(Flowable):
    def __init__(self, score, size=82):
        super().__init__()
        self.score = max(0, min(100, int(score or 0)))
        self.size = size
        self.width = size
        self.height = size

    def draw(self):
        color = colors.HexColor("#198754") if self.score >= 80 else colors.HexColor("#e8c547") if self.score >= 50 else colors.HexColor("#dc3545")
        c = self.canv
        radius = self.size / 2
        c.setStrokeColor(colors.HexColor("#e5e7eb"))
        c.setLineWidth(9)
        c.circle(radius, radius, radius - 6, stroke=1, fill=0)
        c.setStrokeColor(color)
        c.setLineWidth(9)
        extent = 360 * (self.score / 100)
        c.arc(6, 6, self.size - 6, self.size - 6, startAng=90, extent=-extent)
        c.setFillColor(colors.HexColor("#0a0a0a"))
        c.setFont("Helvetica-Bold", 18)
        c.drawCentredString(radius, radius - 3, str(self.score))
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#6b7280"))
        c.drawCentredString(radius, radius - 17, "score")


def build_boq_audit_pdf_buffer(boq_obj, audit):
    buffer = BytesIO()
    setup_arabic_font()
    styles = getSampleStyleSheet()
    navy = colors.HexColor("#0a0a0a")
    gold = colors.HexColor("#e8c547")
    light = colors.HexColor("#f6f3ec")
    green = colors.HexColor("#198754")
    amber = colors.HexColor("#e8c547")
    red = colors.HexColor("#dc3545")

    def esc(value):
        return str(value or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def fmt_sar(value):
        try:
            return f"SAR {float(value or 0):,.0f}"
        except Exception:
            return "SAR 0"

    normal = ParagraphStyle("AuditNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor("#1a1a1a"))
    small = ParagraphStyle("AuditSmall", parent=styles["BodyText"], fontName="Helvetica", fontSize=7.5, leading=9, textColor=colors.HexColor("#555555"))
    heading = ParagraphStyle("AuditHeading", parent=styles["Heading3"], fontName="Helvetica-Bold", fontSize=10.5, textColor=navy, spaceBefore=8, spaceAfter=4)
    header_style = ParagraphStyle("AuditHeader", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=14, textColor=colors.white)
    right_header = ParagraphStyle("AuditRightHeader", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, textColor=colors.white, alignment=TA_RIGHT)
    arabic = ParagraphStyle("AuditArabic", parent=styles["BodyText"], fontName="Cairo" if setup_arabic_font() else "Helvetica", fontSize=9, leading=13, alignment=TA_RIGHT)

    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    story = []
    project_name = getattr(boq_obj, "project", "") or getattr(boq_obj, "title", "") or "BOQ Audit"

    header = Table(
        [[
            Paragraph("<b>BanaaIQ | BOQ Saudi Market Audit</b>", header_style),
            Paragraph(f"<font color='#e8c547'><b>{esc(project_name)}</b></font><br/><font color='#dddddd'>{datetime.utcnow().strftime('%d %B %Y')}</font>", right_header),
        ]],
        colWidths=[9.5 * cm, 7.5 * cm],
    )
    header.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), navy), ("PADDING", (0, 0), (-1, -1), 13), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(header)
    story.append(Spacer(1, 0.35 * cm))

    score = int(audit.get("overall_score", 0) or 0)
    breakdown = audit.get("score_breakdown", {}) or {}
    score_table = Table(
        [[
            BoqScoreGauge(score),
            Paragraph(
                f"<b>Overall Assessment</b><br/>{esc(audit.get('summary', 'No summary available.'))}<br/><br/>"
                f"<b>Potential Saving:</b> {fmt_sar(audit.get('total_potential_saving_sar', 0))}<br/>"
                f"<b>Errors Found:</b> {int(audit.get('total_errors_found', 0) or 0)}",
                normal,
            ),
        ]],
        colWidths=[3 * cm, 14 * cm],
    )
    score_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), light), ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 10), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(score_table)
    story.append(Spacer(1, 0.25 * cm))

    story.append(Paragraph("Executive Summary", heading))
    ar_summary = audit.get("summary_arabic", "")
    summary_rows = [[
        Paragraph(esc(audit.get("summary", "-")), normal),
        Paragraph(esc(process_arabic_text(ar_summary) if is_arabic_text(ar_summary) else ar_summary or "-"), arabic),
    ]]
    summary_table = Table(summary_rows, colWidths=[8.4 * cm, 8.4 * cm])
    summary_table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 8), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.append(summary_table)

    story.append(Paragraph("Score Breakdown", heading))
    breakdown_rows = [[Paragraph("<b>Category</b>", small), Paragraph("<b>Score</b>", small)]]
    labels = {
        "rate_accuracy": "Rate Accuracy",
        "completeness": "Completeness",
        "description_clarity": "Description Clarity",
        "calculation_accuracy": "Calculation Accuracy",
    }
    for key, label in labels.items():
        breakdown_rows.append([Paragraph(label, small), Paragraph(f"{int(breakdown.get(key, 0) or 0)} / 25", small)])
    breakdown_table = Table(breakdown_rows, colWidths=[12 * cm, 5 * cm])
    breakdown_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), navy), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 5)]))
    story.append(breakdown_table)

    story.append(Paragraph("Itemized Audit", heading))
    item_rows = [[Paragraph("<b></b>", small), Paragraph("<b>#</b>", small), Paragraph("<b>Description</b>", small), Paragraph("<b>Rating</b>", small), Paragraph("<b>Market Rate</b>", small), Paragraph("<b>Recommendation</b>", small)]]
    color_lookup = {"green": green, "amber": amber, "red": red}
    row_styles = [("BACKGROUND", (0, 0), (-1, 0), navy), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 4), ("VALIGN", (0, 0), (-1, -1), "TOP")]
    for idx, item in enumerate((audit.get("item_audits") or [])[:80], start=1):
        rating = str(item.get("rating", "amber")).lower()
        item_rows.append([
            "",
            Paragraph(esc(item.get("item_no", idx)), small),
            Paragraph(esc(item.get("description", ""))[:180], small),
            Paragraph(rating.title(), small),
            Paragraph(fmt_sar(item.get("market_rate_sar", 0)), small),
            Paragraph(esc(item.get("recommendation", ""))[:220], small),
        ])
        row_styles.append(("BACKGROUND", (0, idx), (0, idx), color_lookup.get(rating, amber)))
    item_table = Table(item_rows, colWidths=[0.18 * cm, 1.0 * cm, 6.2 * cm, 1.7 * cm, 2.5 * cm, 5.4 * cm], repeatRows=1)
    item_table.setStyle(TableStyle(row_styles))
    story.append(item_table)

    missing_items = audit.get("missing_items") or []
    if missing_items:
        story.append(Paragraph("Missing Items", heading))
        missing_rows = [[Paragraph("<b>Description</b>", small), Paragraph("<b>Reason</b>", small), Paragraph("<b>Suggested Total</b>", small)]]
        for item in missing_items[:25]:
            missing_rows.append([
                Paragraph(esc(item.get("description", ""))[:160], small),
                Paragraph(esc(item.get("reason", ""))[:180], small),
                Paragraph(fmt_sar(item.get("suggested_total_sar", 0)), small),
            ])
        missing_table = Table(missing_rows, colWidths=[6 * cm, 7 * cm, 4 * cm])
        missing_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), amber), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 5), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.append(missing_table)

    critical_issues = audit.get("critical_issues") or []
    if critical_issues:
        story.append(Paragraph("Critical Issues", heading))
        for issue in critical_issues[:10]:
            story.append(Paragraph(f"- {esc(issue)}", normal))

    saving = audit.get("total_potential_saving_sar", 0) or 0
    story.append(Spacer(1, 0.25 * cm))
    saving_table = Table([[Paragraph(f"<b>If flagged items are corrected: {fmt_sar(saving)} saved</b>", ParagraphStyle("Saving", parent=normal, fontName="Helvetica-Bold", textColor=navy, alignment=TA_CENTER))]], colWidths=[17 * cm])
    saving_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff8e7")), ("BOX", (0, 0), (-1, -1), 1, gold), ("PADDING", (0, 0), (-1, -1), 9)]))
    story.append(saving_table)

    story.append(Spacer(1, 0.35 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e4ddd0")))
    story.append(Paragraph(f"Audited by BanaaIQ AI - {datetime.utcnow().strftime('%d %B %Y')}", ParagraphStyle("AuditFooter", parent=styles["BodyText"], fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#888888"), alignment=TA_CENTER)))
    doc.build(story)
    buffer.seek(0)
    return buffer


def call_openai(prompt, max_tokens=500, temperature=0.4):
    if client is None:
        return False, "AI service is not configured. Please set OPENAI_API_KEY."
    try:
        response = client.chat.completions.create(
            model=app.config.get("OPENAI_MODEL", "gpt-4o-mini"),
            messages=[{"role": "user", "content": sanitize_input(prompt)}],
            max_tokens=max_tokens,
            temperature=temperature,
        )
        session["ai_calls"] = session.get("ai_calls", 0) + 1
        session.modified = True
        return True, (response.choices[0].message.content or "").strip()
    except Exception as e:
        app.logger.error(f"OpenAI error: {str(e)}")
        return False, "AI service temporarily unavailable. Please try again."


def process_translation(original_text, input_type, primary_lang, output_lang):
    import json

    if client is None:
        return {"success": False, "error": "AI service is not configured. Please set OPENAI_API_KEY."}
    if not original_text or not original_text.strip():
        return {"success": False, "error": "No text to translate."}

    prompt = f"""You are BanaaIQ's expert construction
translator for Saudi Arabian projects.
You specialize in Arabic and English translation for
construction site communications, contracts,
site instructions, and team messages.

INPUT TYPE: {input_type}
USER PRIMARY LANGUAGE: {primary_lang}
TRANSLATE INTO: {output_lang}

ORIGINAL TEXT:
\"\"\"
{original_text}
\"\"\"

Complete ALL 4 tasks. Respond ONLY as valid JSON.

TASK 1 - DETECT: What language is the text in?

TASK 2 - TRANSLATE to {output_lang}:
- Keep all construction technical terms accurate
- Use formal register for Saudi Arabian projects
- If already in {output_lang} still provide it

TASK 3 - SUMMARIZE in {primary_lang}:
- Maximum 2 sentences
- Focus on KEY MESSAGE and ACTION REQUIRED
- Plain language for a busy site engineer

TASK 4 - SUGGEST 3 REPLIES:
- Construction communication context
- Each reply maximum 40 words
- Reply 1: Professional formal response
- Reply 2: Short quick acknowledgment
- Reply 3: Ask for more info/clarification
- For each suggested reply, provide the reply text in both English (text_en) and Arabic (text_ar).
- The Arabic version must be a natural professional translation of the English version.

Respond ONLY as this exact JSON:
{{
  "detected_language": "Arabic",
  "translation": "Full translated text here",
  "summary": "This message is about X. Action: Y.",
  "suggested_replies": [
    {{
      "type": "formal",
      "label": "Formal Response",
      "text_en": "Reply text here",
      "text_ar": "النص العربي هنا"
    }},
    {{
      "type": "brief",
      "label": "Quick Acknowledgment",
      "text_en": "Reply text here",
      "text_ar": "النص العربي هنا"
    }},
    {{
      "type": "clarification",
      "label": "Ask for Clarification",
      "text_en": "Reply text here",
      "text_ar": "النص العربي هنا"
    }}
  ],
  "key_action": "What needs to happen next or None"
}}"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=900,
            temperature=0.3,
        )

        session["ai_calls"] = session.get("ai_calls", 0) + 1
        session.modified = True

        result_text = (response.choices[0].message.content or "").strip()

        if "```" in result_text:
            parts = result_text.split("```")
            for part in parts:
                if "{" in part:
                    result_text = part
                    if result_text.startswith("json"):
                        result_text = result_text[4:]
                    break

        parsed = json.loads(result_text)
        replies = []
        for r in parsed.get("suggested_replies", []):
            text_en = (r.get("text_en") or "").strip()
            text_ar = (r.get("text_ar") or "").strip()
            if not text_en and r.get("text"):
                if primary_lang == "Arabic":
                    text_ar = r.get("text", "").strip()
                else:
                    text_en = r.get("text", "").strip()
            replies.append(
                {
                    "type": r.get("type") or r.get("tone") or "general",
                    "label": r.get("label") or "Reply",
                    "text_en": text_en,
                    "text_ar": text_ar,
                }
            )
        return {
            "success": True,
            "input_type": input_type,
            "original_text": original_text,
            "detected_language": parsed.get("detected_language", "Unknown"),
            "translation": parsed.get("translation", ""),
            "summary": parsed.get("summary", ""),
            "suggested_replies": replies,
            "key_action": parsed.get("key_action", ""),
        }
    except json.JSONDecodeError:
        return {"success": False, "error": "AI response error. Please try again."}
    except Exception as e:
        app.logger.error(f"Translation error: {str(e)}")
        return {"success": False, "error": "Translation service unavailable."}


@login_manager.user_loader
def load_user(user_id):
    try:
        return db.session.get(User, int(user_id))
    except (TypeError, ValueError) as error:
        app.logger.error(str(error))
        return None
    except SQLAlchemyError as error:
        rollback_db_session()
        traceback.print_exc()
        app.logger.error(str(error))
        return None


@app.before_request
def ensure_language():
    if current_user.is_authenticated or session.get("is_guest"):
        session.permanent = True
    if session.get("lang") != "en":
        session["lang"] = "en"
        session.modified = True


def get_dashboard_shell_context():
    user_name, user_company = dashboard_identity()
    backlog_tasks = []
    backlog_count = 0
    if session.get("is_guest"):
        backlog_source = [task for task in TASKS_DATA if task.get("status") == "backlog"]
        backlog_tasks = backlog_source[:5]
        backlog_count = len(backlog_source)
    elif current_user.is_authenticated:
        try:
            from sqlalchemy import func as _sf
            backlog_count = Task.query.filter(
                Task.assigned_to_id == current_user.id,
                Task.status == "not_started",
            ).count()
            backlog_tasks = (
                Task.query.filter(
                    Task.assigned_to_id == current_user.id,
                    Task.status == "not_started",
                )
                .order_by(Task.due_date.asc().nullslast())
                .limit(5)
                .all()
            )
        except SQLAlchemyError as error:
            rollback_db_session()
            traceback.print_exc()
            app.logger.error(str(error))

    return {
        "current_user": current_user,
        "user_name": user_name,
        "user_company": user_company,
        "username": user_name,
        "company": user_company,
        "dashboard_first_name": (user_name or "User").split()[0],
        "projects": [],
        "selected_project": "all",
        "backlog_count": backlog_count,
        "backlog_tasks": backlog_tasks,
    }


@app.context_processor
def inject_helpers():
    # Force English globally for all templates.
    lang = "en"
    today = datetime.now()
    hy, hm, hd = gregorian_to_hijri(today)
    is_demo = is_demo_mode()
    notifications = get_localized_notifications(lang)
    unread_count = sum(1 for n in notifications if not n["read"])

    def tr(en_text, ar_text):
        return ar_text if lang == "ar" else en_text

    return {
        "lang": lang,
        "is_rtl": lang == "ar",
        "tr": tr,
        "current_date": today.strftime("%B %d, %Y"),
        "today_gregorian": today.strftime("%B %d, %Y"),
        "hijri_date": f"{hd:02d}-{hm:02d}-{hy} AH",
        "today_hijri": f"{hd:02d}-{hm:02d}-{hy} AH",
        "ai_chat_calls": int(session.get("ai_chat_calls", 0)),
        "is_demo": is_demo,
        "notifications": notifications,
        "unread_count": unread_count,
        **get_dashboard_shell_context(),
    }


@app.context_processor
def inject_language():
    """
    Makes current language available in ALL templates automatically.
    No need to pass it manually in each route.
    """
    # Force English globally for all templates.
    lang = "en"
    return {
        "current_lang": lang,
        "is_arabic": False,
        "is_english": True,
    }


@app.context_processor
def inject_now():
    subscription_status = None
    if current_user.is_authenticated and getattr(current_user, "subscription", None):
        subscription_status = normalize_subscription_status(get_subscription_status(current_user))
    return {
        "now": datetime.utcnow(),
        "current_subscription_status": subscription_status,
    }


def dashboard_access_required(view_func):
    protected_view = login_required(view_func)

    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if session.get("is_guest"):
            return view_func(*args, **kwargs)
        return protected_view(*args, **kwargs)

    return wrapped


def dashboard_identity():
    if session.get("is_guest"):
        return session.get("guest_name", "Guest User"), session.get("guest_company", "Demo Company")
    if current_user.is_authenticated:
        return current_user.full_name, current_user.company
    return "User", ""


def _dashboard_int(value, default=0):
    try:
        return int(value or default)
    except (TypeError, ValueError):
        return default


def _dashboard_worker_count(record):
    if isinstance(record, dict):
        return _dashboard_int(record.get("workers"))
    workers = getattr(record, "workers", []) or []
    if isinstance(workers, int):
        return workers
    if isinstance(workers, list):
        return len([worker for worker in workers if worker.get("present", True)])
    return 0


def _dashboard_activity_item(activity):
    if isinstance(activity, dict):
        return {
            "title": activity.get("details") or activity.get("title") or "Activity logged",
            "time": activity.get("time") or "Live",
        }
    created_at = getattr(activity, "created_at", None)
    return {
        "title": getattr(activity, "details", None) or getattr(activity, "action", None) or "Activity logged",
        "time": created_at.strftime("%d %b %Y | %H:%M") if created_at else "Live",
    }


def build_dashboard_intelligence(stats, recent_dprs, recent_activity, backlog_tasks, health_projects, ai_used=0):
    stats = stats or {}
    total_tasks = _dashboard_int(stats.get("total_tasks"))
    done_tasks = _dashboard_int(stats.get("tasks_done"))
    backlog_count = _dashboard_int(stats.get("backlog_tasks"))
    active_tasks = _dashboard_int(stats.get("tasks_in_progress"))
    critical_items = _dashboard_int(stats.get("critical_items"))
    low_stock_items = _dashboard_int(stats.get("low_stock_items"))
    total_inventory = _dashboard_int(stats.get("total_inventory_items"))
    total_projects = _dashboard_int(stats.get("total_projects"))
    active_projects = _dashboard_int(stats.get("active_projects"))
    pending_approvals = _dashboard_int(stats.get("pending_stock_requests"))
    active_packages = _dashboard_int(stats.get("active_packages"))
    total_boqs = _dashboard_int(stats.get("total_boqs"))

    progress_pct = round((done_tasks / total_tasks) * 100) if total_tasks else 0
    manpower_total = sum(_dashboard_worker_count(record) for record in (recent_dprs or [])[:5])
    equipment_ready = max(total_inventory - critical_items - low_stock_items, 0)

    health_scores = []
    for project in health_projects or []:
        score = project.get("health_score") if isinstance(project, dict) else getattr(project, "health_score", None)
        if score is not None:
            health_scores.append(_dashboard_int(score, 100))
    health_avg = round(sum(health_scores) / len(health_scores)) if health_scores else (92 if active_projects else 0)

    risk_pressure = min(100, (critical_items * 18) + (low_stock_items * 8) + (backlog_count * 4) + (active_tasks * 2))
    safety_alerts = critical_items + min(backlog_count, 5)

    if critical_items:
        ai_primary = {
            "title": "Material risk detected",
            "body": f"{critical_items} critical stock item{'s' if critical_items != 1 else ''} need procurement attention before site productivity is affected.",
            "icon": "fa-triangle-exclamation",
        }
    elif backlog_count:
        ai_primary = {
            "title": "Backlog pressure rising",
            "body": f"{backlog_count} backlog task{'s' if backlog_count != 1 else ''} should be sequenced against the next DPR cycle.",
            "icon": "fa-list-check",
        }
    else:
        ai_primary = {
            "title": "Execution rhythm is stable",
            "body": "No critical dashboard signals are active. Keep monitoring DPR and inventory deltas.",
            "icon": "fa-circle-check",
        }

    ai_used_count = _dashboard_int(ai_used)
    ai_insights = [
        ai_primary,
        {
            "title": "Forecast confidence",
            "body": f"{total_boqs} BOQ record{'s' if total_boqs != 1 else ''} and {total_inventory} inventory item{'s' if total_inventory != 1 else ''} are feeding the command view.",
            "icon": "fa-chart-line",
        },
        {
            "title": "AI usage",
            "body": f"{ai_used_count} AI-assisted action{'s' if ai_used_count != 1 else ''} logged this month.",
            "icon": "fa-robot",
        },
    ]

    recommendations = [
        "Review critical materials before releasing tomorrow's work front.",
        "Use DPR notes to validate manpower variance against active tasks.",
        "Prioritize approval ageing for BOQ, inventory, and site requests.",
    ]
    if not critical_items and not backlog_count:
        recommendations[0] = "Maintain the current cadence and schedule the next health-score recalculation."

    timeline = [
        {"label": "Planning", "value": 100 if total_projects else 20},
        {"label": "Procurement", "value": min(100, (total_boqs * 18) + (total_inventory * 2))},
        {"label": "Execution", "value": progress_pct},
        {"label": "Closeout", "value": min(100, done_tasks * 10)},
    ]

    return {
        "progress_pct": progress_pct,
        "health_avg": health_avg,
        "risk_pressure": risk_pressure,
        "manpower_total": manpower_total,
        "equipment_ready": equipment_ready,
        "safety_alerts": safety_alerts,
        "pending_approvals": pending_approvals,
        "active_packages": active_packages,
        "ai_insights": ai_insights,
        "recommendations": recommendations,
        "timeline": timeline,
        "activity_preview": [_dashboard_activity_item(item) for item in (recent_activity or [])[:5]],
    }


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/features")
def features():
    return redirect(url_for("index") + "#features")


@app.route("/tutorials")
@dashboard_access_required
def tutorials():
    username, company = dashboard_identity()
    return render_template("tutorials/index.html", username=username, company=company, active_dashboard="tutorials", active_tutorial="getting-started")


@app.route("/about")
def about():
    return render_template("about.html")


@app.route("/privacy-policy")
def privacy_policy():
    return render_template("legal/privacy.html")


@app.route("/terms-of-service")
def terms_of_service():
    return render_template("legal/terms.html")


@app.route("/data-ai-disclosure")
def data_ai_disclosure():
    return render_template("legal/ai_disclosure.html")


@app.route("/contact")
def contact():
    return render_template("contact.html")


@app.route("/contact/submit", methods=["POST"])
def contact_submit():
    payload = (request.get_json(silent=True) or {}) if request.is_json else request.form
    wants_json = bool(
        request.is_json
        or request.headers.get("X-Requested-With")
        or "application/json" in request.headers.get("Accept", "")
    )

    name = sanitize_input(payload.get("name", ""), 100).strip()
    email = sanitize_input(payload.get("email", ""), 100).strip()
    company = sanitize_input(payload.get("company", ""), 100).strip()
    message = sanitize_input(payload.get("message", ""), 2000).strip()
    subject = sanitize_input(payload.get("subject", "Contact Form Message"), 200).strip() or "Contact Form Message"

    email_pattern = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
    if not name or not email or not message or not re.match(email_pattern, email):
        if wants_json:
            return jsonify({"success": False, "error": "Name, email and message are required"}), 400
        return redirect(url_for("contact", error="missing"))

    email_sent = False
    safe_name = html.escape(name)
    safe_email = html.escape(email)
    safe_company = html.escape(company or "Not provided")
    safe_subject = html.escape(subject)
    safe_message = html.escape(message)

    try:
        msg = Message(
            subject=f"BanaaIQ Contact: {subject}",
            recipients=["iqbaana@gmail.com"],
            reply_to=email,
            body=f"""
New contact form submission from BanaaIQ website.

Name:    {name}
Email:   {email}
Company: {company or 'Not provided'}
Subject: {subject}

Message:
{message}

---
Sent from BanaaIQ Contact Form
banaaiq.com
            """.strip(),
            html=f"""
<div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;">
  <div style="background:#0a0a0a;padding:20px;border-radius:8px 8px 0 0;">
    <h2 style="color:#e8c547;margin:0;">BanaaIQ - New Contact Message</h2>
  </div>
  <div style="background:#f8f9fa;padding:24px;border:1px solid #dee2e6;border-radius:0 0 8px 8px;">
    <table style="width:100%;border-collapse:collapse;">
      <tr>
        <td style="padding:8px 0;font-weight:bold;color:#0a0a0a;width:100px;">Name:</td>
        <td style="padding:8px 0;color:#333;">{safe_name}</td>
      </tr>
      <tr>
        <td style="padding:8px 0;font-weight:bold;color:#0a0a0a;">Email:</td>
        <td style="padding:8px 0;color:#333;"><a href="mailto:{safe_email}">{safe_email}</a></td>
      </tr>
      <tr>
        <td style="padding:8px 0;font-weight:bold;color:#0a0a0a;">Company:</td>
        <td style="padding:8px 0;color:#333;">{safe_company}</td>
      </tr>
      <tr>
        <td style="padding:8px 0;font-weight:bold;color:#0a0a0a;">Subject:</td>
        <td style="padding:8px 0;color:#333;">{safe_subject}</td>
      </tr>
    </table>
    <hr style="border:1px solid #dee2e6;margin:16px 0;">
    <h4 style="color:#0a0a0a;margin-bottom:8px;">Message:</h4>
    <p style="color:#333;line-height:1.7;white-space:pre-wrap;background:white;padding:12px;border-radius:6px;border:1px solid #dee2e6;">{safe_message}</p>
    <hr style="border:1px solid #dee2e6;margin:16px 0;">
    <p style="color:#999;font-size:12px;margin:0;">Sent from BanaaIQ Contact Form</p>
  </div>
</div>
            """.strip(),
        )
        mail.send(msg)
        email_sent = True

        auto_reply = Message(
            subject="We received your message - BanaaIQ",
            recipients=[email],
            body=f"""
Hi {name},

Thank you for reaching out to BanaaIQ!

We have received your message and will get back to you within 24 hours.

Your message:
"{message[:200]}{'...' if len(message) > 200 else ''}"

Best regards,
The BanaaIQ Team
iqbaana@gmail.com
            """.strip(),
        )
        mail.send(auto_reply)
    except Exception as e:
        app.logger.error(f"Email send error: {str(e)}")
        email_sent = False

    if wants_json:
        return jsonify(
            {
                "success": True,
                "message": "Message sent successfully",
                "email_sent": email_sent,
            }
        )

    return redirect(url_for("contact", sent="true"))


@app.route("/pricing")
def pricing():
    currency = get_active_currency()
    plan_lookup = {plan.name: plan for plan in Plan.query.order_by(Plan.id).all()}
    ordered_plans = [
        build_plan_view_model(plan_lookup[name], currency["code"])
        for name in ["starter", "professional", "enterprise"]
        if name in plan_lookup
    ]

    _region_both = (
        "🇮🇳 INR pricing · GST invoice · AWS Mumbai · Razorpay"
        "  |  "
        "🇸🇦 SAR pricing · ZATCA VAT · Stripe"
    )
    plan_copy = {
        "starter": {
            "tag": "Best for small teams",
            "arabic": "المبتدئ",
            "summary": (
                "Everything you need to digitize site reporting, task tracking, "
                "and inventory for a single project team."
            ),
            "region": _region_both,
            "quote": (
                "Replace WhatsApp and Excel with structured daily reports, "
                "live inventory, and AI-assisted BOQs."
            ),
        },
        "professional": {
            "tag": "Best for growing firms",
            "arabic": "المحترف",
            "summary": (
                "Full bilingual control across multiple projects — AI translation, "
                "priority support, and unlimited AI queries."
            ),
            "region": _region_both,
            "quote": (
                "One platform for your PM, QS, engineers, and site supervisors "
                "— in English and Arabic."
            ),
        },
        "enterprise": {
            "tag": "For large organisations",
            "arabic": "المؤسسة",
            "summary": (
                "Custom deployment, dedicated onboarding, SLA guarantee, and SSO "
                "integrations for programme-scale rollouts."
            ),
            "region": (
                "🇮🇳 Custom INR agreement · GST compliant"
                "  |  "
                "🇸🇦 Custom SAR agreement · ZATCA compliant"
            ),
            "quote": (
                "Built for regional developers, main contractors, "
                "and programme management offices."
            ),
        },
    }

    comparison_rows = [
        {"feature": "DPR",                  "label": "Daily Progress Reports",       "arabic": "تقارير التقدم اليومي"},
        {"feature": "BOQ",                  "label": "Bill of Quantities (BOQ)",      "arabic": "جدول الكميات"},
        {"feature": "Inventory",            "label": "Inventory Management",          "arabic": "إدارة المخزون"},
        {"feature": "Tasks",                "label": "Task Management (Kanban)",      "arabic": "إدارة المهام"},
        {"feature": "Translator",           "label": "AI Translator (EN ↔ AR)",       "arabic": "ترجمة الذكاء الاصطناعي"},
        {"feature": "Priority Support",     "label": "Priority Support",              "arabic": "دعم ذو أولوية"},
        {"feature": "SLA",                  "label": "Service Level Agreement",       "arabic": "اتفاقية مستوى الخدمة"},
        {"feature": "Dedicated Onboarding", "label": "Dedicated Onboarding",          "arabic": "تأهيل مخصص"},
        {"feature": "Custom Integrations",  "label": "Custom Integrations",           "arabic": "تكاملات مخصصة"},
    ]

    faq_items = [
        {
            "question": "Do I need a credit card to start a trial?",
            "answer": (
                "No. BanaaIQ is free during beta. Start a trial and explore every "
                "feature without entering payment details."
            ),
        },
        {
            "question": "Which payment gateway is used for India?",
            "answer": (
                "Razorpay — RBI-authorised, supports UPI, cards, and net banking in INR. "
                "GST invoices are generated automatically at 18% IGST."
            ),
        },
        {
            "question": "Which payment gateway is used for Saudi Arabia and GCC?",
            "answer": (
                "Stripe is configured for KSA and GCC markets with ZATCA-compliant "
                "VAT invoicing at 15%."
            ),
        },
        {
            "question": "Can I switch between monthly and annual billing?",
            "answer": (
                "Yes. You can upgrade to annual billing at any time from the Billing page. "
                "Annual plans save up to 17% compared to monthly."
            ),
        },
        {
            "question": "What happens when my trial ends?",
            "answer": (
                "You will receive reminder emails 3 days and 1 day before your trial ends. "
                "After expiry your data is retained and you can activate a paid plan at any time."
            ),
        },
        {
            "question": "Is my project data secure?",
            "answer": (
                "Yes. BanaaIQ runs on AWS Mumbai (ap-south-1), uses HTTPS everywhere, "
                "and complies with India's DPDP Act 2023 and Saudi Arabia's PDPL 2021."
            ),
        },
    ]

    return render_template(
        "pricing.html",
        plans=ordered_plans,
        currency=currency,
        plan_copy=plan_copy,
        comparison_rows=comparison_rows,
        faq_items=faq_items,
    )


@app.route("/set-currency", methods=["POST"])
def set_currency():
    selected_currency = str(request.form.get("currency") or "").upper()
    next_url = get_safe_next_url(request.form.get("next")) or url_for("pricing")

    if selected_currency in CURRENCY_METADATA:
        session["selected_currency"] = selected_currency
    else:
        session.pop("selected_currency", None)
    session.modified = True
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return ("", 204)
    return redirect(next_url)


@app.route("/start-trial/<plan_name>", methods=["GET", "POST"])
def start_trial(plan_name):
    plan = Plan.query.filter_by(name=plan_name).first_or_404()
    if plan.name == "enterprise":
        return redirect(url_for("contact"))

    if not current_user.is_authenticated:
        return redirect(url_for("register", next=url_for("start_trial", plan_name=plan_name)))

    subscription_obj = current_user.subscription
    status_value = get_subscription_status(current_user)
    normalized_status = normalize_subscription_status(status_value)
    current_plan_name = subscription_obj.plan.name if subscription_obj and subscription_obj.plan else None
    is_plan_change = bool(
        subscription_obj
        and normalized_status in ("trialing", "active")
        and current_plan_name
        and current_plan_name != plan.name
    )

    if subscription_obj and normalized_status in ("trialing", "active") and current_plan_name == plan.name:
        flash("You already have an active subscription.", "info")
        return redirect(url_for("billing"))

    billing_cycle = request.form.get("billing_cycle") or getattr(subscription_obj, "billing_cycle", None) or "monthly"
    if billing_cycle not in ("monthly", "annual"):
        billing_cycle = "monthly"

    preview_trial_end = datetime.utcnow() + timedelta(days=14)
    if subscription_obj and normalized_status == "trialing" and subscription_obj.trial_end:
        preview_trial_end = subscription_obj.trial_end

    if request.method == "POST":
        now = datetime.utcnow()
        # Capture optional GSTIN for Indian B2B customers
        buyer_gstin = (request.form.get("buyer_gstin") or "").strip().upper() or None
        if buyer_gstin and len(buyer_gstin) == 15:
            current_user.gstin = buyer_gstin

        customer_id = (
            subscription_obj.gateway_customer_id
            if subscription_obj and subscription_obj.gateway_customer_id
            else create_customer(current_user)
        )

        if subscription_obj and is_plan_change:
            subscription_obj.plan_id = plan.id
            subscription_obj.billing_cycle = billing_cycle
            subscription_obj.gateway_customer_id = customer_id
            subscription_obj.cancelled_at = None
            subscription_obj.cancellation_reason = None
            if subscription_obj.status == "active" and not subscription_obj.current_period_start:
                subscription_obj.current_period_start = now
            if subscription_obj.status == "active" and not subscription_obj.current_period_end:
                subscription_obj.current_period_end = get_billing_period_end(now, billing_cycle)
            db.session.commit()
            flash(f"Your plan has been updated to {plan.display_name}.", "success")
            return redirect(url_for("billing"))

        if subscription_obj:
            sub = subscription_obj
            sub.plan_id = plan.id
            sub.billing_cycle = billing_cycle
            sub.status = "trialing"
            sub.trial_start = now
            sub.trial_end = now + timedelta(days=14)
            sub.current_period_start = now
            sub.current_period_end = now + timedelta(days=14)
            sub.cancelled_at = None
            sub.cancellation_reason = None
            sub.gateway_customer_id = customer_id
            sub.gateway_subscription_id = None
        else:
            sub = Subscription(
                user_id=current_user.id,
                plan_id=plan.id,
                billing_cycle=billing_cycle,
                status="trialing",
                trial_start=now,
                trial_end=now + timedelta(days=14),
                current_period_start=now,
                current_period_end=now + timedelta(days=14),
                gateway_customer_id=customer_id,
            )
            db.session.add(sub)

        db.session.commit()
        send_trial_welcome_email(current_user, plan.display_name, sub.trial_end)
        flash(f"Your 14-day {plan.display_name} trial is active!", "success")
        return redirect(url_for("dashboard_overview"))

    currency = get_active_currency()
    return render_template(
        "start_trial.html",
        plan=plan,
        plan_data=build_plan_view_model(plan, currency["code"]),
        trial_end=preview_trial_end,
        billing_cycle=billing_cycle,
        is_plan_change=is_plan_change,
        current_subscription=subscription_obj,
        current_status=normalized_status,
        currency=currency,
    )


@app.route("/billing")
@login_required
def billing():
    currency = get_active_currency()
    username, company = dashboard_identity()
    subscription_obj = current_user.subscription
    status_value = get_subscription_status(current_user)
    normalized_status = normalize_subscription_status(status_value)
    subscription_days_left = get_subscription_days_left(status_value)
    payment_method = get_default_payment_method(current_user)
    invoices = (
        Invoice.query.filter_by(user_id=current_user.id)
        .order_by(Invoice.created_at.desc())
        .all()
    )
    ai_quota = get_ai_quota_snapshot(current_user)

    plan_price = None
    next_billing_date = None
    trial_progress = 0
    trial_days_used = 0
    trial_total_days = 14

    if subscription_obj and subscription_obj.plan:
        plan_price = get_plan_amount(subscription_obj.plan, subscription_obj.billing_cycle)
        next_billing_date = subscription_obj.current_period_end or subscription_obj.trial_end
        if normalized_status == "trialing" and subscription_obj.trial_start and subscription_obj.trial_end:
            trial_total_days = max((subscription_obj.trial_end - subscription_obj.trial_start).days, 14)
            trial_days_used = min(max((datetime.utcnow() - subscription_obj.trial_start).days, 0), trial_total_days)
            trial_progress = min(int((trial_days_used / max(trial_total_days, 1)) * 100), 100)

    return render_template(
        "billing.html",
        username=username,
        company=company,
        active_dashboard="billing",
        sub=subscription_obj,
        subscription_status=normalized_status,
        subscription_days_left=subscription_days_left,
        payment_method=payment_method,
        invoices=invoices,
        plan_price=plan_price,
        next_billing_date=next_billing_date,
        trial_progress=trial_progress,
        trial_days_used=trial_days_used,
        trial_total_days=trial_total_days,
        publishable_key=get_publishable_key(),
        ai_quota=ai_quota,
        currency=currency,
    )


@app.route("/billing/payment-method", methods=["POST"])
@login_required
def billing_payment_method():
    flash(
        "Payment coming soon. You will be notified when payments are activated for your region.",
        "info",
    )
    return redirect(url_for("billing"))


@app.route("/billing/invoice/<int:invoice_id>/pdf")
@login_required
def billing_invoice_pdf(invoice_id):
    from reportlab.pdfgen import canvas as pdf_canvas

    invoice = Invoice.query.filter_by(id=invoice_id, user_id=current_user.id).first_or_404()
    plan_name = (
        invoice.subscription.plan.display_name
        if invoice.subscription and invoice.subscription.plan
        else "Subscription"
    )
    period_start_text = invoice.period_start.strftime("%d %B %Y") if invoice.period_start else "-"
    period_end_text = invoice.period_end.strftime("%d %B %Y") if invoice.period_end else "-"
    invoice_date_text = invoice.created_at.strftime("%d %B %Y") if invoice.created_at else "-"
    due_date_text = invoice.due_date.strftime("%d %B %Y") if invoice.due_date else "-"

    # Determine invoice format: India (INR/GST) vs KSA (SAR/ZATCA)
    invoice_currency = (invoice.currency or "INR").upper()
    is_india_invoice = invoice_currency == "INR"

    if is_india_invoice:
        tax_ctx = get_payment_provider().create_invoice_tax_context(invoice)
        tax_label = tax_ctx.get("tax_label", "IGST 18%")
        tax_rate = tax_ctx.get("tax_rate", 18)
        company_name = tax_ctx.get("company_legal_name", "BanaaIQ Technologies")
        company_gstin = tax_ctx.get("gstin", "GSTIN-PENDING")
        invoice_title = "TAX INVOICE"
        currency_symbol = "₹"
        hsn_sac = tax_ctx.get("hsn_sac", "998313")
    else:
        tax_label = "VAT 15% (ZATCA)"
        tax_rate = 15
        company_name = "BanaaIQ"
        company_gstin = ""
        invoice_title = "TAX INVOICE"
        currency_symbol = "SAR "
        hsn_sac = ""

    def money(value):
        return f"{currency_symbol}{float(value or 0):,.2f}"

    buffer = BytesIO()
    pdf = pdf_canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    navy = colors.HexColor("#0a0a0a")
    gold = colors.HexColor("#D4AF37")
    green = colors.HexColor("#198754")
    orange = colors.HexColor("#FD7E14")
    border = colors.HexColor("#D9DEE7")
    text_color = colors.HexColor("#2F3542")

    watermark_text = "PAID" if invoice.status == "paid" else "PENDING"
    watermark_color = green if invoice.status == "paid" else orange
    pdf.saveState()
    if hasattr(pdf, "setFillAlpha"):
        pdf.setFillAlpha(0.14)
    pdf.setFillColor(watermark_color)
    pdf.setFont("Helvetica-Bold", 78)
    pdf.translate(width / 2, height / 2)
    pdf.rotate(32)
    pdf.drawCentredString(0, 0, watermark_text)
    pdf.restoreState()

    pdf.setFillColor(navy)
    pdf.rect(0, height - 92, width, 92, fill=1, stroke=0)
    pdf.setFillColor(gold)
    pdf.setFont("Helvetica-Bold", 26)
    pdf.drawString(42, height - 56, "BanaaIQ")
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 24)
    pdf.drawRightString(width - 42, height - 56, invoice_title)

    y = height - 132
    pdf.setFillColor(text_color)
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(42, y, "Invoice Number")
    pdf.drawString(230, y, "Invoice Date")
    pdf.drawString(390, y, "Due Date")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(42, y - 18, invoice.invoice_number or "-")
    pdf.drawString(230, y - 18, invoice_date_text)
    pdf.drawString(390, y - 18, due_date_text)

    y -= 60
    pdf.setStrokeColor(border)
    pdf.setFillColor(colors.white)
    pdf.roundRect(42, y - 72, width - 84, 72, 10, fill=1, stroke=1)
    pdf.setFillColor(text_color)
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(56, y - 18, "Bill To")
    pdf.setFont("Helvetica", 11)
    pdf.drawString(56, y - 38, current_user.full_name or current_user.email or "User")
    pdf.drawString(56, y - 56, current_user.company or "-")

    y -= 108
    table_x = 42
    table_width = width - 84
    table_header_y = y
    col_widths = [240, 50, 90, 90]
    headers = ["Description", "Qty", "Unit Price", "Total"]

    pdf.setFillColor(navy)
    pdf.rect(table_x, table_header_y - 24, table_width, 24, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 10)
    header_x = table_x + 8
    for index, header in enumerate(headers):
        pdf.drawString(header_x, table_header_y - 16, header)
        header_x += col_widths[index]

    row_y = table_header_y - 24
    pdf.setFillColor(colors.white)
    pdf.setStrokeColor(border)
    pdf.rect(table_x, row_y - 42, table_width, 42, fill=1, stroke=1)
    pdf.setFillColor(text_color)
    pdf.setFont("Helvetica", 10)
    description = f"{plan_name} Subscription ({period_start_text} to {period_end_text})"
    pdf.drawString(table_x + 8, row_y - 18, description[:56])
    pdf.drawString(table_x + col_widths[0] + 8, row_y - 18, "1")
    pdf.drawString(table_x + col_widths[0] + col_widths[1] + 8, row_y - 18, money(invoice.amount))
    pdf.drawString(table_x + col_widths[0] + col_widths[1] + col_widths[2] + 8, row_y - 18, money(invoice.amount))

    totals_y = row_y - 78
    totals_x = width - 250
    label_x = totals_x
    value_x = width - 52
    pdf.setFont("Helvetica", 10)
    pdf.drawString(label_x, totals_y, "Subtotal")
    pdf.drawRightString(value_x, totals_y, money(invoice.amount))
    pdf.drawString(label_x, totals_y - 20, tax_label)
    pdf.drawRightString(value_x, totals_y - 20, money(invoice.vat_amount))
    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(label_x, totals_y - 44, "Grand Total")
    pdf.drawRightString(value_x, totals_y - 44, money(invoice.total_amount))
    pdf.line(label_x, totals_y - 28, value_x, totals_y - 28)

    # India-specific: GST details block
    if is_india_invoice:
        gst_y = totals_y - 68
        pdf.setFont("Helvetica", 8)
        pdf.setFillColor(colors.HexColor("#6C757D"))
        pdf.drawString(42, gst_y, f"SAC: {hsn_sac}  |  Reverse Charge: No  |  Place of Supply: {tax_ctx.get('company_state', 'India')}")
        if company_gstin:
            pdf.drawString(42, gst_y - 12, f"Seller GSTIN: {company_gstin}")
        buyer_gstin = getattr(invoice, "buyer_gstin", None)
        if buyer_gstin:
            pdf.drawString(42, gst_y - 24, f"Buyer GSTIN: {buyer_gstin}")
        pdf.drawString(42, gst_y - 36, "This is a computer-generated invoice. GSTIN to be updated once registered.")

    footer_y = 52
    pdf.setStrokeColor(border)
    pdf.line(42, footer_y + 18, width - 42, footer_y + 18)
    pdf.setFont("Helvetica", 9)
    pdf.setFillColor(colors.HexColor("#6C757D"))
    seller_line = f"{company_name} | {company_gstin}" if company_gstin else company_name
    pdf.drawString(42, footer_y, seller_line)
    pdf.drawRightString(width - 42, footer_y, f"Status: {invoice.status.upper()}")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{invoice.invoice_number or 'invoice'}.pdf",
        mimetype="application/pdf",
    )


@app.route("/billing/cancel", methods=["POST"])
@login_required
def billing_cancel():
    subscription_obj = current_user.subscription
    normalized_status = normalize_subscription_status(get_subscription_status(current_user))
    if not subscription_obj or normalized_status not in ("trialing", "active"):
        flash("There is no active subscription to cancel.", "warning")
        return redirect(url_for("billing"))

    reason = sanitize_input(request.form.get("reason", ""), 120)
    feedback = sanitize_input(request.form.get("feedback", ""), 1200)
    if not reason or reason.lower().startswith("select"):
        flash("Please choose a cancellation reason before confirming.", "danger")
        return redirect(url_for("billing"))

    if subscription_obj.gateway_subscription_id:
        cancel_subscription(subscription_obj.gateway_subscription_id)

    subscription_obj.status = "cancelled"
    subscription_obj.cancelled_at = datetime.utcnow()
    subscription_obj.cancellation_reason = reason if not feedback else f"{reason}\n\n{feedback}"
    db.session.commit()

    access_until = subscription_obj.current_period_end or subscription_obj.trial_end or datetime.utcnow()
    send_cancellation_email(current_user, access_until)
    flash(
        "Subscription cancelled. You have access until " + access_until.strftime("%d %B %Y"),
        "info",
    )
    return redirect(url_for("billing"))


@app.route("/webhooks/razorpay", methods=["POST"])
@csrf.exempt
def razorpay_webhook():
    """Razorpay payment event webhook.
    Signature verified via RAZORPAY_WEBHOOK_SECRET before any DB changes.
    """
    signature = request.headers.get("X-Razorpay-Signature", "")
    payload = request.get_data(as_text=True)

    provider = get_payment_provider()
    if not provider.verify_webhook(payload, signature):
        app.logger.warning("Razorpay webhook: invalid signature")
        return jsonify({"error": "invalid signature"}), 400

    try:
        event = request.get_json(force=True) or {}
    except Exception:
        return jsonify({"error": "bad json"}), 400

    event_type = event.get("event", "")
    entity = event.get("payload", {}).get("subscription", {}).get("entity", {})
    gateway_sub_id = entity.get("id")

    app.logger.info("Razorpay webhook: %s sub=%s", event_type, gateway_sub_id)

    if gateway_sub_id:
        sub_obj = Subscription.query.filter_by(gateway_subscription_id=gateway_sub_id).first()
    else:
        sub_obj = None

    if event_type == "subscription.activated" and sub_obj:
        sub_obj.status = "active"
        db.session.commit()

    elif event_type == "subscription.charged" and sub_obj:
        sub_obj.status = "active"
        # Mark most-recent pending invoice as paid
        pending_inv = Invoice.query.filter_by(
            subscription_id=sub_obj.id, status="pending"
        ).order_by(Invoice.created_at.desc()).first()
        if pending_inv:
            pending_inv.status = "paid"
            pending_inv.paid_at = datetime.utcnow()
        db.session.commit()

    elif event_type == "subscription.cancelled" and sub_obj:
        sub_obj.status = "cancelled"
        sub_obj.cancelled_at = datetime.utcnow()
        db.session.commit()
        if sub_obj.user:
            access_until = sub_obj.current_period_end or sub_obj.trial_end or datetime.utcnow()
            try:
                send_cancellation_email(sub_obj.user, access_until)
            except Exception as e:
                app.logger.error("Razorpay webhook: cancellation email failed: %s", e)

    elif event_type == "payment.failed" and sub_obj:
        pending_inv = Invoice.query.filter_by(
            subscription_id=sub_obj.id, status="pending"
        ).order_by(Invoice.created_at.desc()).first()
        if pending_inv:
            pending_inv.status = "failed"
            pending_inv.retry_count = (pending_inv.retry_count or 0) + 1
        sub_obj.status = "past_due"
        db.session.commit()

    else:
        app.logger.info("Razorpay webhook: unhandled event %s", event_type)

    return jsonify({"status": "ok"}), 200


@app.route("/login", methods=["GET", "POST"], endpoint="login")
@app.route("/auth/login", methods=["GET", "POST"])
@limiter.limit("20 per hour")
@limiter.limit("5 per minute")
def auth_login():
    login_form = LoginForm(prefix="login")
    register_form = RegisterForm(prefix="register")
    next_url = get_safe_next_url(request.args.get("next") or request.form.get("next"))
    try:
        if current_user.is_authenticated and not session.get("is_guest"):
            return redirect(next_url or url_for("dashboard_overview"))
        if request.method == "POST" and request.form.get("form_type") == "login" and login_form.validate_on_submit():
            email = login_form.email.data.strip().lower()
            user = User.query.filter_by(email=email).first()

            # ── Account lockout check ───────────────────────────────────────────
            _MAX_FAILED = 5
            _LOCKOUT_MINUTES = 15
            if user:
                locked_until = getattr(user, "locked_until", None)
                if locked_until and datetime.utcnow() < locked_until:
                    remaining = int((locked_until - datetime.utcnow()).total_seconds() // 60) + 1
                    flash(
                        f"Account locked due to too many failed attempts. "
                        f"Try again in {remaining} minute(s) or reset your password.",
                        "danger",
                    )
                    return render_template(
                        "auth/login.html",
                        login_form=login_form,
                        register_form=register_form,
                        active_tab="signin",
                        next_url=next_url,
                    )
            # ────────────────────────────────────────────────────────────────────

            if not user or not check_password_hash(user.password_hash, login_form.password.data):
                # Increment failed attempt counter
                if user:
                    try:
                        attempts = (getattr(user, "failed_login_attempts", 0) or 0) + 1
                        user.failed_login_attempts = attempts
                        if attempts >= _MAX_FAILED:
                            user.locked_until = datetime.utcnow() + timedelta(minutes=_LOCKOUT_MINUTES)
                            app.logger.warning(
                                "Account locked after %d failed attempts: user_id=%s email=%s",
                                attempts, user.id, user.email,
                            )
                        db.session.commit()
                    except Exception as _le:
                        rollback_db_session()
                        app.logger.error("Failed to update lockout counter: %s", _le)
                flash("Invalid email or password.", "danger")
                return render_template(
                    "auth/login.html",
                    login_form=login_form,
                    register_form=register_form,
                    active_tab="signin",
                    next_url=next_url,
                )

            # Successful login — reset counter and lockout
            try:
                user.failed_login_attempts = 0
                user.locked_until = None
                db.session.commit()
            except Exception as _re:
                rollback_db_session()
                app.logger.error("Failed to reset lockout counter: %s", _re)

            remember_me = bool(request.form.get("remember_me"))
            # Ensure user has a session token (backfill for accounts created before S5)
            if not getattr(user, "session_token", None):
                user.session_token = secrets.token_hex(32)
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
            session.clear()
            session.permanent = True
            login_user(user, remember=remember_me, duration=timedelta(days=30))
            session["session_started_at"] = datetime.utcnow().isoformat()
            session["lang"] = user.preferred_lang or "en"
            session["session_token"] = user.session_token
            flash("Signed in successfully.", "success")
            return redirect(next_url or url_for("dashboard_overview"))
        return render_template(
            "auth/login.html",
            login_form=login_form,
            register_form=register_form,
            active_tab="signin",
            next_url=next_url,
        )
    except Exception as e:
        rollback_db_session()
        traceback.print_exc()
        app.logger.error(str(e))
        flash(auth_error_message("sign you in", e), "danger")
        return render_template(
            "auth/login.html",
            login_form=login_form,
            register_form=register_form,
            active_tab="signin",
            next_url=next_url,
        ), 500


@app.route("/auth/social/<provider>")
def auth_social(provider):
    # Social OAuth is not yet configured. The previous stub silently created a shared
    # account (all Google users → same email/account) which is a security vulnerability.
    # This route now returns a clear error until real OAuth is implemented.
    app.logger.warning("Social auth attempted for provider=%s — not configured", provider)
    flash(
        "Social sign-in is not yet available. Please use email and password to sign in.",
        "warning",
    )
    return redirect(url_for("auth_login"))


@app.route("/auth/google/callback")
def auth_google_callback():
    return redirect(url_for("auth_social", provider="google"))


@app.route("/auth/linkedin/callback")
def auth_linkedin_callback():
    return redirect(url_for("auth_social", provider="linkedin"))


@app.route("/register", methods=["GET", "POST"], endpoint="register")
@app.route("/auth/register", methods=["GET", "POST"])
@limiter.limit("10 per hour")
@limiter.limit("3 per minute")
def auth_register():
    login_form = LoginForm(prefix="login")
    register_form = RegisterForm(prefix="register")
    next_url = get_safe_next_url(request.args.get("next") or request.form.get("next"))
    try:
        if current_user.is_authenticated and not session.get("is_guest"):
            return redirect(next_url or url_for("dashboard_overview"))
        if request.method == "POST":
            legal_consent = request.form.get("legal_consent")
            if not legal_consent:
                return render_template(
                    "auth/login.html",
                    login_form=login_form,
                    register_form=register_form,
                    register_error=(
                        "You must agree to the Terms of Service "
                        "and Privacy Policy to continue."
                    ),
                    active_tab="register",
                    next_url=next_url,
                )
        if request.method == "POST":
            name = sanitize(request.form.get(register_form.full_name.name, ""), 100).strip()
            company = sanitize(request.form.get(register_form.company.name, ""), 100).strip()
            # account_type → User.role (access-control)
            account_type = sanitize(request.form.get(register_form.account_type.name, ""), 20).strip()
            # job_title → User.job_title (professional designation, required for site_engineer)
            job_title_val = sanitize(request.form.get(register_form.job_title.name, ""), 50).strip()
            country_val = sanitize(request.form.get(register_form.country.name, "IN"), 10).strip().upper()
            if country_val not in {c[0] for c in REGISTER_COUNTRY_CHOICES}:
                country_val = "IN"
            email = sanitize(request.form.get(register_form.email.name, ""), 120).strip().lower()
            phone = sanitize(request.form.get("phone", ""), 30).strip()
            password = request.form.get(register_form.password.name, "")
            confirm_password = request.form.get(register_form.confirm_password.name, "")

            if not all([name, company, account_type, email, phone, password, confirm_password]):
                flash("Please complete all required fields.", "danger")
                return redirect(url_for("register"))

            if account_type not in (ROLE_PROJECT_MANAGER, ROLE_SITE_ENGINEER):
                flash("Please select a valid account type.", "danger")
                return redirect(url_for("register"))

            if account_type == ROLE_SITE_ENGINEER and not job_title_val:
                flash("Please select your job title.", "danger")
                return redirect(url_for("register"))

            if job_title_val and job_title_val not in ROLE_CHOICES:
                flash("Please select a valid job title.", "danger")
                return redirect(url_for("register"))

            if not re.fullmatch(r"[0-9+\s\-]{7,20}", phone):
                flash("Please enter a valid phone number with country code.", "danger")
                return redirect(url_for("register"))

            if len(password) < 8:
                flash("Password must be at least 8 characters.", "danger")
                return redirect(url_for("register"))

            if password != confirm_password:
                flash("Passwords do not match.", "danger")
                return redirect(url_for("register"))

            if User.query.filter_by(email=email).first():
                flash("Email already registered.", "danger")
                return redirect(url_for("register"))

            hashed = generate_password_hash(
                password,
                method="pbkdf2:sha256",
                salt_length=16,
            )
            preferred_currency_val = map_country_code_to_currency(country_val)
            user = User(
                full_name=name or "User",
                company=company,
                job_title=job_title_val,
                role=account_type,
                email=email,
                phone=phone,
                password_hash=hashed,
                preferred_lang=session.get("lang", "en"),
                country=country_val,
                preferred_currency=preferred_currency_val,
            )
            db.session.add(user)
            try:
                db.session.commit()
            except SQLAlchemyError as e:
                rollback_db_session()
                traceback.print_exc()
                app.logger.error(str(e))
                raise
            # Generate session token for new user
            user.session_token = secrets.token_hex(32)
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
            session.clear()
            session.permanent = True
            login_user(user)
            session["session_started_at"] = datetime.utcnow().isoformat()
            session["lang"] = user.preferred_lang or "en"
            session["session_token"] = user.session_token
            # Send welcome email — non-blocking; failure does not stop registration
            try:
                from mailer import send_email as _send_email, MailError as _MailError
                _send_email(
                    to=user.email,
                    subject="Welcome to BanaaIQ",
                    html_body=render_template("emails/welcome.html", user=user),
                )
            except Exception as _wel_err:
                app.logger.warning("Welcome email failed for user_id=%s: %s", user.id, _wel_err)
            flash("Welcome to BanaaIQ! / مرحباً بك في BanaaIQ", "success")
            return redirect(url_for("pricing"))
        return render_template(
            "auth/register.html",
            login_form=login_form,
            register_form=register_form,
            active_tab="register",
            next_url=next_url,
        )
    except Exception as e:
        rollback_db_session()
        traceback.print_exc()
        app.logger.error(str(e))
        flash(auth_error_message("create your account", e), "danger")
        return render_template(
            "auth/register.html",
            login_form=login_form,
            register_form=register_form,
            active_tab="register",
            next_url=next_url,
        ), 500


@app.route("/auth/logout", methods=["POST"])
@login_required
def auth_logout():
    logout_user()
    session.clear()
    flash("You've been logged out. / تم تسجيل خروجك.", "info")
    response = redirect(url_for("index"))
    response.delete_cookie("remember_token")
    return apply_no_store_headers(response)


@app.route("/auth/heartbeat", methods=["POST"])
@login_required
@limiter.limit("60 per hour")
def auth_heartbeat():
    """Touch the session to reset the inactivity timer."""
    session.modified = True
    return jsonify({"ok": True})


@app.route("/auth/change-password", methods=["GET", "POST"])
@login_required
@limiter.limit("10 per hour")
def auth_change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not check_password_hash(current_user.password_hash, current_password):
            flash("Current password is incorrect. / كلمة المرور الحالية غير صحيحة.", "danger")
            return redirect(url_for("auth_change_password"))

        if new_password != confirm_password:
            flash("New passwords don't match. / كلمتا المرور الجديدتان غير متطابقتين.", "danger")
            return redirect(url_for("auth_change_password"))

        ok, error = validate_password_strength(new_password)
        if not ok:
            flash(error, "danger")
            return redirect(url_for("auth_change_password"))

        current_user.password_hash = generate_password_hash(new_password, method="pbkdf2:sha256", salt_length=16)
        current_user.password_updated_at = datetime.utcnow()
        # Rotate session token so other devices are signed out
        current_user.session_token = secrets.token_hex(32)
        session["session_token"] = current_user.session_token
        try:
            db.session.commit()
        except Exception as _cpe:
            db.session.rollback()
            app.logger.error("Change password DB error: %s", _cpe)
            flash("Could not update password. Please try again.", "danger")
            return redirect(url_for("auth_change_password"))

        # Security notification — non-blocking
        try:
            from mailer import send_email as _send_email, MailError as _MailError
            _send_email(
                to=current_user.email,
                subject="Your BanaaIQ password was changed",
                html_body=render_template(
                    "emails/password_changed.html",
                    user=current_user,
                    changed_at=datetime.utcnow(),
                ),
            )
        except Exception as _cpe_mail:
            app.logger.warning("Password-changed email failed: %s", _cpe_mail)

        flash("Password updated. / تم تحديث كلمة المرور.", "success")
        return redirect(url_for("settings_security"))

    return render_template("auth/change_password.html")


_PASSWORD_RESET_SALT = "banaaiq-pw-reset-v1"
_PASSWORD_RESET_MAX_AGE = 3600  # 1 hour


def _make_reset_token(user):
    """Generate a signed, time-limited, single-use reset token.

    The token stores a short password-hash fingerprint, so changing the
    password invalidates outstanding reset links without a separate DB column.
    """
    s = _URLSafeTimedSerializer(app.config["SECRET_KEY"])
    password_fingerprint = hashlib.sha256((user.password_hash or "").encode("utf-8")).hexdigest()
    return s.dumps(
        {
            "uid": user.id,
            "email": user.email,
            "hash": password_fingerprint,
        },
        salt=_PASSWORD_RESET_SALT,
    )


def _verify_reset_token(token):
    """Return the User if the token is valid and unexpired, else None."""
    s = _URLSafeTimedSerializer(app.config["SECRET_KEY"])
    try:
        payload = s.loads(token, salt=_PASSWORD_RESET_SALT, max_age=_PASSWORD_RESET_MAX_AGE)
    except (SignatureExpired, BadSignature):
        return None
    if not isinstance(payload, dict):
        return None
    user = db.session.get(User, payload.get("uid"))
    if not user:
        return None
    if user.email != payload.get("email"):
        return None
    password_fingerprint = hashlib.sha256((user.password_hash or "").encode("utf-8")).hexdigest()
    if password_fingerprint != payload.get("hash"):
        return None
    return user


@app.route("/auth/forgot", methods=["GET", "POST"])
@limiter.limit("5 per hour", methods=["POST"])
def auth_forgot():
    form = ForgotForm()
    if form.validate_on_submit():
        email = (form.email.data or "").strip().lower()
        user = User.query.filter_by(email=email).first()
        # Always show the same success message — prevents email enumeration
        if user:
            try:
                token = _make_reset_token(user)
                reset_url = url_for("auth_reset_password", token=token, _external=True)
                from mailer import send_email as _send_email, MailError as _MailError
                _send_email(
                    to=user.email,
                    subject="Reset your BanaaIQ password",
                    html_body=render_template(
                        "emails/password_reset.html",
                        user=user,
                        reset_url=reset_url,
                    ),
                )
                app.logger.info("Password reset email sent to user_id=%s", user.id)
            except Exception as e:
                app.logger.error("Password reset email failed for user_id=%s: %s", user.id if user else "?", e)
                if _sentry_sdk:
                    _sentry_sdk.capture_exception(e)
        flash(
            "If an account with that email exists, a reset link has been sent. "
            "Please check your inbox (and spam folder).",
            "success",
        )
        return redirect(url_for("auth_forgot"))
    return render_template("auth/forgot.html", form=form)


@app.route("/auth/reset/<token>", methods=["GET", "POST"])
@limiter.limit("10 per hour")
def auth_reset_password(token):
    user = _verify_reset_token(token)
    if not user:
        flash("This password reset link is invalid or has expired. Please request a new one.", "danger")
        return redirect(url_for("auth_forgot"))

    if request.method == "POST":
        new_password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")
        if len(new_password) < 8:
            flash("Password must be at least 8 characters.", "danger")
            return render_template("auth/reset.html", token=token)
        if new_password != confirm:
            flash("Passwords do not match.", "danger")
            return render_template("auth/reset.html", token=token)
        user.set_password(new_password)
        db.session.commit()
        app.logger.info("Password reset completed for user_id=%s", user.id)
        # Security notification — inform user their password was changed
        try:
            from mailer import send_email as _send_email, MailError as _MailError
            _send_email(
                to=user.email,
                subject="Your BanaaIQ password was changed",
                html_body=render_template(
                    "emails/password_changed.html",
                    user=user,
                    changed_at=datetime.utcnow(),
                ),
            )
        except Exception as _notif_err:
            app.logger.warning("Password-changed notification failed for user_id=%s: %s", user.id, _notif_err)
        flash("Your password has been reset. Please sign in with your new password.", "success")
        return redirect(url_for("auth_login"))

    return render_template("auth/reset.html", token=token)


@app.route("/auth/guest")
def auth_guest():
    if current_user.is_authenticated:
        logout_user()
    session.clear()
    session.permanent = True
    session["is_guest"] = True
    session["guest_name"] = "Mohammed Al-Rashidi"
    session["guest_company"] = "Al Noor Development Co."
    session["session_started_at"] = datetime.utcnow().isoformat()
    return redirect(url_for("dashboard_overview"))


@app.route("/dashboard", endpoint="dashboard")
@app.route("/dashboard", endpoint="dashboard_overview")
@dashboard_access_required
def dashboard():
    lang = session.get("lang", "en")
    notifications = get_localized_notifications(lang)
    if current_user.is_authenticated and not is_demo_mode():
        user_name = current_user.full_name
        user_company = current_user.company
        is_demo = False
    else:
        user_name = session.get("guest_name", "Guest")
        user_company = session.get("guest_company", "Al Noor Development Co.")
        is_demo = True
    unread_count = sum(1 for n in notifications if not n["read"])
    if is_demo:
        stats = {
            "total_dprs": 3,
            "dprs_this_month": 2,
            "total_boqs": 4,
            "boqs_total_value": 2456780,
            "total_inventory_items": 12,
            "low_stock_items": 4,
            "critical_items": 2,
            "total_tasks": 10,
            "backlog_tasks": 3,
            "tasks_in_progress": 3,
            "tasks_done": 2,
            "total_projects": 2,
            "active_projects": 2,
            "pending_stock_requests": 4,
            "active_packages": 6,
        }
        backlog_tasks = []
        my_tasks_widget = []
        backlog_count = 0
        recent_dprs = DPR_RECORDS[:3]
        recent_activity = []
        recent_projects = []
        health_projects = []
    else:
        try:
            uid = current_user.id
            month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            total_dprs = DPR.query.filter_by(user_id=uid).count()
            dprs_this_month = DPR.query.filter(DPR.user_id == uid, DPR.created_at >= month_start).count()
            if current_user.role == ROLE_SITE_ENGINEER:
                total_boqs = BOQ.query.filter(
                    BOQ.assigned_to_user_id == uid,
                    BOQ.status.in_(["distributed", "revised"]),
                ).count()
                boqs_value = db.session.query(db.func.sum(BOQ.grand_total)).filter(
                    BOQ.assigned_to_user_id == uid,
                    BOQ.status.in_(["distributed", "revised"]),
                ).scalar() or 0
            else:
                total_boqs = BOQ.query.filter_by(user_id=uid).count()
                boqs_value = db.session.query(db.func.sum(BOQ.grand_total)).filter_by(user_id=uid).scalar() or 0
            inventory_items = InventoryItem.query.filter_by(user_id=uid).all()
            total_inventory_items = len(inventory_items)
            low_stock_items = sum(1 for item in inventory_items if item.status == "low")
            critical_items = sum(1 for item in inventory_items if item.status == "critical")
            # Task counts — new schema uses project_id + assigned_to_id (no user_id)
            from sqlalchemy import func as _sa_func
            _task_counts = db.session.query(
                Task.status, _sa_func.count(Task.id)
            ).join(Project, Task.project_id == Project.id).filter(
                Project.user_id == uid
            ).group_by(Task.status).all()
            _tc = {s: c for s, c in _task_counts}
            all_task_count = sum(_tc.values())
            backlog_count = _tc.get("not_started", 0)
            in_progress_count = _tc.get("in_progress", 0)
            done_count = _tc.get("done", 0)
            total_projects = Project.query.filter_by(user_id=uid).count()
            active_projects_count = Project.query.filter_by(user_id=uid, status="active").count()
            pending_stock_requests = (
                StockRequest.query.join(Project, StockRequest.project_id == Project.id)
                .filter(Project.user_id == uid, StockRequest.status == "pending")
                .count()
            )
            active_packages = (
                EngineerPackage.query.join(Project, EngineerPackage.project_id == Project.id)
                .filter(Project.user_id == uid, EngineerPackage.status == "active")
                .count()
            )
            stats = {
                "total_dprs": total_dprs,
                "dprs_this_month": dprs_this_month,
                "total_boqs": total_boqs,
                "boqs_total_value": round(boqs_value, 0),
                "total_inventory_items": total_inventory_items,
                "low_stock_items": low_stock_items,
                "critical_items": critical_items,
                "total_tasks": all_task_count,
                "backlog_tasks": backlog_count,
                "tasks_in_progress": in_progress_count,
                "tasks_done": done_count,
                "total_projects": total_projects,
                "active_projects": active_projects_count,
                "pending_stock_requests": pending_stock_requests,
                "active_packages": active_packages,
            }
            # My Tasks widget: tasks assigned to me (any role), not done, sorted overdue first
            from sqlalchemy.orm import selectinload as _sel
            my_tasks_widget = (
                Task.query.filter(
                    Task.assigned_to_id == uid,
                    Task.status != "done",
                ).options(
                    _sel(Task.project_rel),
                ).order_by(Task.due_date.asc().nullslast())
                .limit(5)
                .all()
            )
            backlog_tasks = my_tasks_widget  # kept for base_dashboard.html compat
            recent_dprs = (
                DPR.query.filter_by(user_id=uid)
                .order_by(DPR.created_at.desc())
                .limit(5)
                .all()
            )
            recent_activity = []  # TaskActivity is per-task; no global feed needed here
            recent_projects = (
                Project.query.filter_by(user_id=uid)
                .order_by(Project.created_at.desc())
                .limit(3)
                .all()
            )
            health_projects = (
                Project.query.filter_by(user_id=uid, status="active")
                .order_by(Project.created_at.desc())
                .all()
            )
        except SQLAlchemyError as error:
            rollback_db_session()
            traceback.print_exc()
            app.logger.error(str(error))
            flash("Signed in, but some dashboard data is still loading. Please check the server logs.", "warning")
            stats = {
                "total_dprs": 0,
                "dprs_this_month": 0,
                "total_boqs": 0,
                "boqs_total_value": 0,
                "total_inventory_items": 0,
                "low_stock_items": 0,
                "critical_items": 0,
                "total_tasks": 0,
                "backlog_tasks": 0,
                "tasks_in_progress": 0,
                "tasks_done": 0,
                "total_projects": 0,
                "active_projects": 0,
                "pending_stock_requests": 0,
                "active_packages": 0,
            }
            backlog_tasks = []
            my_tasks_widget = []
            backlog_count = 0
            recent_dprs = []
            recent_activity = []
            recent_projects = []
            health_projects = []
    subscription_obj = current_user.subscription if current_user.is_authenticated and not is_demo else None
    subscription_status = (
        normalize_subscription_status(get_subscription_status(current_user))
        if subscription_obj
        else "none"
    )
    show_trial_btn = not subscription_obj or subscription_status in ("expired", "cancelled", "none")
    ai_used = 0
    if subscription_obj and subscription_obj.plan and subscription_obj.plan.max_ai_queries:
        ai_used = get_monthly_ai_usage(current_user)
    dashboard_intelligence = build_dashboard_intelligence(
        stats,
        recent_dprs,
        recent_activity,
        backlog_tasks,
        health_projects,
        ai_used,
    )
    return render_template(
        "dashboard/index.html",
        user_name=user_name,
        user_company=user_company,
        username=user_name,
        company=user_company,
        is_demo=is_demo,
        show_demo_data=is_demo_mode(),
        notifications=notifications,
        unread_count=unread_count,
        stats=stats,
        backlog_tasks=backlog_tasks,
        backlog_count=backlog_count,
        recent_dprs=recent_dprs,
        recent_activity=[],
        my_tasks_widget=my_tasks_widget if not is_demo else [],
        recent_projects=recent_projects,
        health_projects=health_projects,
        dashboard_intelligence=dashboard_intelligence,
        active_dashboard="overview",
        sub=subscription_obj,
        show_trial_btn=show_trial_btn,
        ai_used=ai_used,
    )



@app.route("/projects")
@login_required
@role_required(ROLE_PROJECT_MANAGER, ROLE_SITE_ENGINEER)
def projects_list():
    selected_status = sanitize_input(request.args.get("status", "all"), 20)

    if current_user.role == ROLE_PROJECT_MANAGER:
        # PM sees only projects they created.
        query = Project.query.filter_by(user_id=current_user.id)
        if selected_status in {"active", "on_hold", "completed"}:
            query = query.filter_by(status=selected_status)
        projects = query.order_by(Project.created_at.desc()).all()
    else:
        # SE sees projects where assigned via EngineerPackage OR explicit ProjectAssignment.
        pkg_ids = {
            r[0] for r in db.session.query(EngineerPackage.project_id)
            .filter(EngineerPackage.assigned_user_id == current_user.id)
            .distinct()
            .all()
        }
        assign_ids = {
            r[0] for r in db.session.query(ProjectAssignment.project_id)
            .filter(ProjectAssignment.user_id == current_user.id)
            .distinct()
            .all()
        }
        all_project_ids = list(pkg_ids | assign_ids)
        query = Project.query.filter(Project.id.in_(all_project_ids))
        if selected_status in {"active", "on_hold", "completed"}:
            query = query.filter_by(status=selected_status)
        projects = query.order_by(Project.created_at.desc()).all()
    # N+1 fix: preload aggregate counts in bulk queries
    from sqlalchemy import func as _func
    project_ids = [p.id for p in projects]
    if project_ids:
        boq_counts = dict(
            db.session.query(BOQ.project_id, _func.count(BOQ.id))
            .filter(BOQ.project_id.in_(project_ids))
            .filter(BOQ.assigned_to_user_id == None)
            .group_by(BOQ.project_id).all()
        )
        task_counts = dict(
            db.session.query(Task.project_id, _func.count(Task.id))
            .filter(Task.project_id.in_(project_ids))
            .group_by(Task.project_id).all()
        )
        inv_counts = dict(
            db.session.query(InventoryItem.project_id, _func.count(InventoryItem.id))
            .filter(InventoryItem.project_id.in_(project_ids))
            .group_by(InventoryItem.project_id).all()
        )
        dpr_counts = dict(
            db.session.query(DPR.project_id, _func.count(DPR.id))
            .filter(DPR.project_id.in_(project_ids))
            .group_by(DPR.project_id).all()
        )
    else:
        boq_counts = task_counts = inv_counts = dpr_counts = {}

    project_cards = [
        {
            "project": project,
            "progress": project_timeline_progress(project),
            "status_label": project_status_label(project.status),
            "type_label": project_type_label(project.project_type),
        }
        for project in projects
    ]
    return render_template(
        "projects/list.html",
        active_dashboard="projects",
        project_cards=project_cards,
        selected_status=selected_status,
        boq_counts=boq_counts,
        task_counts=task_counts,
        inv_counts=inv_counts,
        dpr_counts=dpr_counts,
        status_filters=[
            ("all", "All"),
            ("active", "Active"),
            ("on_hold", "On Hold"),
            ("completed", "Completed"),
        ],
    )


@app.route("/projects/create", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def projects_create():
    if request.method == "POST":
        form_data = project_form_data()
        if not form_data["name"]:
            flash("Project name is required.", "danger")
            return redirect(url_for("projects_create"))
        project = Project(user_id=current_user.id, project_code=generate_project_code(), status="active")
        apply_project_form(project, form_data)
        db.session.add(project)
        db.session.flush()
        create_default_project_milestones(project)
        add_custom_project_milestones(project)
        db.session.commit()
        flash("Project workspace created.", "success")
        return redirect(url_for("project_workspace", id=project.id))

    return render_template(
        "projects/create.html",
        active_dashboard="projects",
        is_edit=False,
        project=None,
        project_type_options=PROJECT_TYPE_OPTIONS,
        project_status_options=PROJECT_STATUS_OPTIONS,
        form_action=url_for("projects_create"),
    )


@app.route("/projects/<int:id>")
@login_required
def project_workspace(id):
    project = get_project_or_403(id)
    workspace = project_workspace_data(project)
    users = User.query.order_by(User.full_name).limit(100).all()

    # Build unified activity feed
    events = []
    try:
        # DPRs
        for d in DPR.query.filter_by(project_id=project.id).order_by(DPR.created_at.desc()).limit(20).all():
            actor_user = getattr(d, "author", None)
            actor = actor_user.full_name if actor_user else 'unknown'
            events.append({
                'ts': d.created_at or datetime.utcnow(),
                'icon': 'fa-clipboard-list',
                'color': '#3b82f6',
                'msg_en': f"DPR submitted by {actor}",
                'msg_ar': f"تقرير يومي مقدم من {actor}",
            })
        # BOQs
        for b in BOQ.query.filter_by(project_id=project.id).order_by(BOQ.created_at.desc()).limit(20).all():
            if b.status == 'master':
                msg_en = f"Master BOQ '{b.title}' created"
                msg_ar = f"تم إنشاء جدول الكميات الرئيسي '{b.title}'"
            elif b.status == 'distributed':
                msg_en = f"BOQ section '{b.trade_section or '—'}' distributed"
                msg_ar = f"تم توزيع قسم '{b.trade_section or '—'}'"
            elif b.status == 'revised':
                eng = b.assigned_engineer.full_name if b.assigned_engineer else 'engineer'
                msg_en = f"BOQ revision v{b.version} by {eng}"
                msg_ar = f"تعديل جدول الكميات v{b.version}"
            else:
                continue
            events.append({'ts': b.created_at or datetime.utcnow(), 'icon': 'fa-calculator', 'color': '#c9a44d', 'msg_en': msg_en, 'msg_ar': msg_ar})
        # Inventory consumption
        for u in UsageLog.query.filter_by(project_id=project.id).order_by(UsageLog.created_at.desc()).limit(20).all():
            events.append({
                'ts': u.created_at or datetime.utcnow(),
                'icon': 'fa-boxes-stacked', 'color': '#8b5cf6',
                'msg_en': f"{u.used_by} used {u.quantity_used} {u.unit or ''} of {u.item_name}",
                'msg_ar': f"استُخدم {u.quantity_used} {u.unit or ''} من {u.item_name}",
            })
        # Stock requests
        for r in StockRequest.query.filter_by(project_id=project.id).order_by(StockRequest.created_at.desc()).limit(10).all():
            events.append({
                'ts': r.created_at or datetime.utcnow(),
                'icon': 'fa-clipboard-question', 'color': '#f59e0b',
                'msg_en': f"Stock request ({r.status}): {r.requested_qty} {r.unit or ''}",
                'msg_ar': f"طلب مخزون ({r.status}): {r.requested_qty}",
            })
        # Tasks
        for t in Task.query.filter_by(project_id=project.id).order_by(Task.created_at.desc()).limit(20).all():
            events.append({
                'ts': t.created_at or datetime.utcnow(),
                'icon': 'fa-list-check', 'color': '#10b981',
                'msg_en': f"Task '{t.title}' created",
                'msg_ar': f"مهمة '{t.title}' أنشئت",
            })
        # Engineer assignments
        for a in ProjectAssignment.query.filter_by(project_id=project.id).order_by(ProjectAssignment.assigned_at.desc()).limit(10).all():
            if a.user:
                events.append({
                    'ts': a.assigned_at or datetime.utcnow(),
                    'icon': 'fa-user-plus', 'color': '#0ea5e9',
                    'msg_en': f"{a.user.full_name} added to project",
                    'msg_ar': f"تمت إضافة {a.user.full_name} للمشروع",
                })
        events.sort(key=lambda e: e['ts'], reverse=True)
        events = events[:15]
    except Exception as _e:
        app.logger.warning("Activity feed error: %s", _e)
        events = []

    return render_template(
        "projects/workspace.html",
        active_dashboard="projects",
        project=project,
        workspace=workspace,
        events=events,
        health=project_health_payload(project),
        project_type_label=project_type_label(project.project_type),
        project_status_label=project_status_label(project.status),
        task_status_labels=TASK_STATUS_LABELS,
        trade_options=TRADE_OPTIONS,
        users=users,
    )


@app.route("/projects/<int:id>/recalculate-health", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_recalculate_health(id):
    project = get_project_or_403(id)
    invalidate_health_cache(project.id)
    data = get_cached_health_score(project.id)
    return jsonify(data)


@app.route("/api/projects/<int:id>/health")
@login_required
def api_project_health(id):
    import hashlib
    project = get_project_or_403(id)
    data = get_cached_health_score(project.id)
    if not data:
        abort(404)
    # ETag based on last_calculated
    last_calculated = data.get("last_calculated") or ""
    etag = hashlib.md5(last_calculated.encode()).hexdigest()
    if_none_match = request.headers.get("If-None-Match", "")
    if if_none_match == etag:
        return "", 304
    response = jsonify(data)
    response.headers["ETag"] = etag
    return response


@app.route("/projects/<int:id>/edit", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_edit(id):
    project = get_project_or_403(id)
    if request.method == "POST":
        form_data = project_form_data(project)
        if not form_data["name"]:
            flash("Project name is required.", "danger")
            return redirect(url_for("project_edit", id=project.id))
        apply_project_form(project, form_data)
        db.session.commit()
        flash("Project details updated.", "success")
        return redirect(url_for("project_workspace", id=project.id))

    return render_template(
        "projects/create.html",
        active_dashboard="projects",
        is_edit=True,
        project=project,
        project_type_options=PROJECT_TYPE_OPTIONS,
        project_status_options=PROJECT_STATUS_OPTIONS,
        form_action=url_for("project_edit", id=project.id),
    )


@app.route("/projects/<int:id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_delete(id):
    project = get_project_or_403(id)
    DPR.query.filter_by(user_id=current_user.id, project_id=project.id).update({"project_id": None})
    BOQ.query.filter_by(user_id=current_user.id, project_id=project.id).update({"project_id": None})
    Task.query.filter_by(project_id=project.id).delete(synchronize_session=False)
    InventoryItem.query.filter_by(user_id=current_user.id, project_id=project.id).update({"project_id": None})
    db.session.delete(project)
    db.session.commit()
    flash("Project deleted. Linked DPR, BOQ, task, and inventory records were kept unassigned.", "info")
    return redirect(url_for("projects_list"))


@app.route("/projects/<int:id>/complete", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_complete(id):
    project = get_project_or_403(id)
    project.status = "completed"
    project.actual_completion = date.today()
    for milestone in project.milestones:
        if milestone.status != "completed" and milestone.planned_date and milestone.planned_date <= date.today():
            milestone.actual_date = date.today()
            milestone.status = "completed"
    db.session.commit()
    calculate_health_score(project)
    flash("Project marked complete. Generate the completion report from the workspace banner.", "success")
    return redirect(url_for("project_workspace", id=project.id))


@app.route("/projects/<int:project_id>/oversight")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_oversight(project_id):
    project = get_project_or_403(project_id)
    health = get_cached_health_score(project.id)
    today = date.today()

    # ── Per-engineer card data ──────────────────────────────────────────────
    assignments = ProjectAssignment.query.filter_by(project_id=project.id).all()
    assigned_user_ids = {a.user_id for a in assignments}
    pkg_user_ids = {
        p.assigned_user_id for p in EngineerPackage.query.filter_by(project_id=project.id).all()
        if p.assigned_user_id
    }
    all_engineer_ids = assigned_user_ids | pkg_user_ids
    engineers_list = User.query.filter(User.id.in_(all_engineer_ids)).all() if all_engineer_ids else []

    engineer_cards = []
    for eng in engineers_list:
        role_row = ProjectAssignment.query.filter_by(project_id=project.id, user_id=eng.id).first()
        role_on_project = (role_row.role_on_project or eng.job_title or "Engineer") if role_row else (eng.job_title or "Engineer")
        # EngineerPackage items progress
        eng_packages = EngineerPackage.query.filter_by(project_id=project.id, assigned_user_id=eng.id).all()
        total_items = 0
        complete_items = 0
        for pkg in eng_packages:
            statuses = pkg.item_statuses
            for v in statuses.values():
                total_items += 1
                if v == "complete":
                    complete_items += 1
        progress_pct = round(complete_items / total_items * 100, 1) if total_items else 0.0
        # Tasks
        eng_tasks = Task.query.filter(
            Task.project_id == project.id,
            Task.assigned_to_id == eng.id,
        ).all()
        tasks_done = sum(1 for t in eng_tasks if t.status == "done")
        tasks_open = len(eng_tasks) - tasks_done
        # Last DPR
        last_dpr = DPR.query.filter_by(project_id=project.id, user_id=eng.id).order_by(DPR.date.desc()).first()
        last_dpr_date = last_dpr.date.strftime("%d %b %Y") if last_dpr and last_dpr.date else "—"
        # Materials
        inv_assignments = InventoryAssignment.query.filter_by(project_id=project.id, assigned_user_id=eng.id).all()
        mat_ok = mat_low = mat_critical = 0
        for ia in inv_assignments:
            item = db.session.get(InventoryItem, ia.inventory_item_id)
            if item:
                s = item.status
                if s == "critical":
                    mat_critical += 1
                elif s == "low":
                    mat_low += 1
                else:
                    mat_ok += 1
        engineer_cards.append({
            "user": eng,
            "name": eng.full_name,
            "email": eng.email,
            "job_title": eng.job_title or "",
            "role_on_project": role_on_project,
            "progress_pct": progress_pct,
            "tasks_open": tasks_open,
            "tasks_done": tasks_done,
            "last_dpr_date": last_dpr_date,
            "materials_status": {"ok": mat_ok, "low": mat_low, "critical": mat_critical},
        })

    # ── Project summary ─────────────────────────────────────────────────────
    all_packages = EngineerPackage.query.filter_by(project_id=project.id).all()
    total_pkg_items = 0
    complete_pkg_items = 0
    for pkg in all_packages:
        for v in pkg.item_statuses.values():
            total_pkg_items += 1
            if v == "complete":
                complete_pkg_items += 1
    boq_completion_pct = round(complete_pkg_items / total_pkg_items * 100, 1) if total_pkg_items else 0.0

    boqs = BOQ.query.filter_by(project_id=project.id).all()
    master_boqs = [b for b in boqs if b.is_master]
    budget_boqs = master_boqs if master_boqs else boqs
    budget_sar = sum(float(b.grand_total or 0) for b in budget_boqs)
    boq_ids = [b.id for b in boqs]
    actuals_rows = BOQActual.query.filter(BOQActual.boq_id.in_(boq_ids)).all() if boq_ids else []
    spend_sar = sum(float(r.actual_total_sar or 0) for r in actuals_rows)

    all_tasks_proj = Task.query.filter_by(project_id=project.id).all()
    active_tasks_count = sum(1 for t in all_tasks_proj if t.status not in ("complete", "done"))
    inv_items_proj = InventoryItem.query.filter_by(project_id=project.id).all()
    inv_low_critical = sum(1 for i in inv_items_proj if i.status in ("low", "critical"))
    overdue_tasks = sum(1 for t in all_tasks_proj if t.due and t.due < today and t.status not in ("complete", "done"))
    alerts_count = inv_low_critical + overdue_tasks

    project_summary = {
        "boq_completion_pct": boq_completion_pct,
        "spend_sar": spend_sar,
        "budget_sar": budget_sar,
        "active_tasks_count": active_tasks_count,
        "alerts_count": alerts_count,
    }

    # ── Activity feed (no new schema) ────────────────────────────────────────
    feed_items = []
    recent_dprs = DPR.query.filter_by(project_id=project.id).order_by(DPR.created_at.desc()).limit(5).all()
    for dpr in recent_dprs:
        actor = User.query.get(dpr.user_id)
        feed_items.append({
            "timestamp": dpr.created_at or datetime.utcnow(),
            "type": "dpr",
            "actor_name": actor.full_name if actor else "Engineer",
            "message": f"Submitted DPR for {dpr.date.strftime('%d %b %Y') if dpr.date else '—'}: {(dpr.progress_notes or '')[:80]}",
            "message_ar": f"تقرير يومي بتاريخ {dpr.date.strftime('%d %b %Y') if dpr.date else '—'}",
        })
    recent_tasks = Task.query.filter_by(project_id=project.id).order_by(Task.updated_at.desc()).limit(5).all()
    for task in recent_tasks:
        feed_items.append({
            "timestamp": task.updated_at or task.created_at or datetime.utcnow(),
            "type": "task",
            "actor_name": (task.assigned_to.full_name if task.assigned_to else None) or "Engineer",
            "message": f"Task '{task.name}' → {task.status}",
            "message_ar": f"مهمة '{task.name}' → {task.status}",
        })
    recent_packages = EngineerPackage.query.filter_by(project_id=project.id).order_by(EngineerPackage.created_at.desc()).limit(5).all()
    for pkg in recent_packages:
        feed_items.append({
            "timestamp": pkg.created_at or datetime.utcnow(),
            "type": "package",
            "actor_name": pkg.assigned_engineer_name or "Engineer",
            "message": f"Package '{pkg.trade or 'Package'}' {int(pkg.completion_percentage or 0)}% complete",
            "message_ar": f"حزمة '{pkg.trade or 'حزمة'}' {int(pkg.completion_percentage or 0)}% مكتملة",
        })
    inv_notifications = Notification.query.filter(
        Notification.user_id == project.user_id,
        Notification.message.ilike("%stock%"),
    ).order_by(Notification.created_at.desc()).limit(5).all()
    for notif in inv_notifications:
        feed_items.append({
            "timestamp": notif.created_at or datetime.utcnow(),
            "type": "inventory",
            "actor_name": "System",
            "message": notif.message or "Inventory alert",
            "message_ar": notif.message or "تنبيه المخزون",
        })
    feed_items.sort(key=lambda x: x["timestamp"], reverse=True)
    activity_feed = feed_items[:20]

    return render_template(
        "projects/oversight.html",
        active_dashboard="projects",
        project=project,
        health=health,
        engineer_cards=engineer_cards,
        project_summary=project_summary,
        activity_feed=activity_feed,
    )


@app.route("/projects/<int:id>/completion-report")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_completion_report(id):
    project = get_project_or_403(id)
    if project.status != "completed":
        flash("Mark the project complete before downloading the completion report.", "warning")
        return redirect(url_for("project_workspace", id=project.id))
    if not project.actual_completion:
        project.actual_completion = date.today()
        db.session.commit()
    calculate_health_score(project)
    buffer = build_completion_report_pdf(project)
    filename = f"BanaaIQ_Completion_{project.project_code or project.id}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@app.route("/projects/<int:id>/packages")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_packages(id):
    project = get_project_or_403(id)
    packages = EngineerPackage.query.filter_by(project_id=project.id).order_by(EngineerPackage.created_at.desc()).all()
    boqs = BOQ.query.filter_by(user_id=current_user.id, project_id=project.id).order_by(BOQ.created_at.desc()).all()
    users = User.query.order_by(User.full_name).limit(100).all()
    return render_template(
        "projects/packages.html",
        active_dashboard="projects",
        project=project,
        packages=packages,
        boqs=boqs,
        users=users,
        trade_options=TRADE_OPTIONS,
    )


@app.route("/projects/<int:id>/packages/assign", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_package_assign(id):
    project = get_project_or_403(id)
    engineer_name = sanitize_input(request.form.get("engineer_name", ""), 200).strip()
    engineer_email = sanitize_input(request.form.get("engineer_email", ""), 200).strip().lower()
    trade = sanitize_input(request.form.get("trade", "Other"), 50).strip()
    if trade not in TRADE_OPTIONS:
        trade = "Other"
    package_value = parse_decimal_field("package_value")
    scope_description = sanitize_input(request.form.get("scope_description", ""), 3000).strip()
    boq_id_raw = sanitize_input(request.form.get("boq_id", ""), 20).strip()
    boq = None
    if boq_id_raw:
        try:
            boq_id = int(boq_id_raw)
        except ValueError:
            boq_id = None
        if boq_id:
            boq = BOQ.query.filter_by(id=boq_id, user_id=current_user.id, project_id=project.id).first()

    assigned_user = None
    if engineer_email:
        assigned_user = User.query.filter(db.func.lower(User.email) == engineer_email).first()
    if assigned_user and not engineer_name:
        engineer_name = assigned_user.full_name
    if not engineer_name:
        flash("Engineer name is required to assign a package.", "danger")
        return redirect(url_for("project_workspace", id=project.id))

    package = EngineerPackage(
        project_id=project.id,
        boq_id=boq.id if boq else None,
        assigned_user_id=assigned_user.id if assigned_user else None,
        assigned_engineer_name=engineer_name,
        trade=trade,
        package_value=package_value,
        scope_description=scope_description,
        completion_percentage=0.0,
        status="active",
    )
    db.session.add(package)
    db.session.flush()

    if assigned_user:
        message = f"Your {trade} package has been assigned - SAR {float(package_value or 0):,.2f} scope on {project.name}"
        db.session.add(
            Notification(
                user_id=assigned_user.id,
                message=message,
                link=url_for("engineer_package_workspace", package_id=package.id),
                is_read=False,
            )
        )
        package.notified_at = datetime.utcnow()

    db.session.commit()
    flash("Engineer package assigned.", "success")
    return redirect(url_for("project_workspace", id=project.id))


@app.route("/projects/<int:project_id>/engineers/add", methods=["POST"])
@login_required
@limiter.limit("30 per hour")
@role_required(ROLE_PROJECT_MANAGER)
def project_engineer_add(project_id):
    project = get_project_or_403(project_id)
    try:
        user_id = int(sanitize_input(request.form.get("user_id", ""), 20).strip())
    except (ValueError, TypeError):
        flash("Invalid engineer selected.", "danger")
        return redirect(url_for("project_workspace", id=project.id))
    se = User.query.filter_by(id=user_id, role=ROLE_SITE_ENGINEER).first()
    if not se:
        flash("Selected user is not a site engineer.", "danger")
        return redirect(url_for("project_workspace", id=project.id))
    role_on_project = sanitize_input(request.form.get("role_on_project", ""), 50).strip() or None
    try:
        assignment = ProjectAssignment(
            project_id=project.id,
            user_id=se.id,
            role_on_project=role_on_project,
        )
        db.session.add(assignment)
        db.session.commit()
        flash(f"{se.full_name} has been added to the project.", "success")
    except SQLAlchemyError as exc:
        db.session.rollback()
        if "uq_project_user" in str(exc).lower() or "unique" in str(exc).lower():
            flash("Engineer already assigned to this project.", "warning")
        else:
            app.logger.error("project_engineer_add error: %s", exc)
            flash("Could not add engineer. Please try again.", "danger")
    return redirect(url_for("project_workspace", id=project.id))


@app.route("/projects/<int:project_id>/engineers/<int:user_id>/remove", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_engineer_remove(project_id, user_id):
    project = get_project_or_403(project_id)
    assignment = ProjectAssignment.query.filter_by(
        project_id=project.id,
        user_id=user_id,
    ).first_or_404()
    engineer_name = assignment.user.full_name if assignment.user else f"User #{user_id}"
    removed_packages = EngineerPackage.query.filter_by(
        project_id=project.id,
        assigned_user_id=user_id,
    ).all()
    pkg_count = len(removed_packages)
    for pkg in removed_packages:
        db.session.delete(pkg)
    db.session.delete(assignment)
    db.session.commit()
    if pkg_count:
        flash(f"Removed {engineer_name}. {pkg_count} package(s) also removed.", "success")
    else:
        flash(f"Removed {engineer_name} from the project.", "success")
    return redirect(url_for("project_workspace", id=project.id))


@app.route("/api/users/site-engineers")
@login_required
@limiter.limit("60 per hour")
@role_required(ROLE_PROJECT_MANAGER)
def api_site_engineers():
    engineers = (
        User.query.filter_by(role=ROLE_SITE_ENGINEER)
        .order_by(User.full_name)
        .all()
    )
    return jsonify([
        {
            "id": u.id,
            "full_name": u.full_name,
            "email": u.email,
            "job_title": u.job_title or "",
        }
        for u in engineers
    ])


# ─── STEP 2: PM Distribute Work ───────────────────────────────────────────────

@app.route("/projects/<int:project_id>/distribute")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def project_distribute(project_id):
    project = get_project_or_403(project_id)
    engineers = [a.user for a in project.assignments if a.user and a.user.role == ROLE_SITE_ENGINEER]
    boqs = BOQ.query.filter(
        BOQ.project_id == project.id,
        BOQ.user_id == current_user.id,
        BOQ.status == 'master',
        BOQ.assigned_to_user_id.is_(None),
    ).order_by(BOQ.created_at.desc()).all()
    inventory_items = InventoryItem.query.filter_by(project_id=project.id).order_by(InventoryItem.name).all()
    tasks = Task.query.filter_by(project_id=project.id).order_by(Task.created_at.desc()).all()
    return render_template(
        "projects/distribute.html",
        active_dashboard="projects",
        project=project,
        engineers=engineers,
        boqs=boqs,
        inventory_items=inventory_items,
        tasks=tasks,
        trade_options=TRADE_OPTIONS,
    )


@app.route("/projects/<int:project_id>/distribute/boq", methods=["POST"])
@login_required
@limiter.limit("30 per hour")
@role_required(ROLE_PROJECT_MANAGER)
def project_distribute_boq(project_id):
    project = get_project_or_403(project_id)
    assigned_user_id_raw = sanitize_input(request.form.get("assigned_user_id", ""), 20).strip()
    boq_id_raw = sanitize_input(request.form.get("boq_id", ""), 20).strip()
    trade = sanitize_input(request.form.get("trade", "Other"), 50).strip()
    if trade not in TRADE_OPTIONS:
        trade = "Other"
    item_indexes = request.form.getlist("item_indexes")

    try:
        assigned_user_id = int(assigned_user_id_raw)
    except (ValueError, TypeError):
        flash("Invalid engineer selected.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    # Validate engineer is assigned to this project
    assignment = ProjectAssignment.query.filter_by(project_id=project.id, user_id=assigned_user_id).first()
    if not assignment:
        flash("Selected engineer is not assigned to this project.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    engineer = db.session.get(User, assigned_user_id)
    if not engineer:
        flash("Engineer not found.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    try:
        boq_id = int(boq_id_raw)
    except (ValueError, TypeError):
        flash("Invalid BOQ selected.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    boq = BOQ.query.filter_by(id=boq_id, project_id=project.id, user_id=current_user.id).first()
    if not boq:
        flash("BOQ not found.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    all_items = boq.items
    try:
        selected_indexes = [int(i) for i in item_indexes]
    except (ValueError, TypeError):
        selected_indexes = []

    selected_items = [all_items[i] for i in selected_indexes if 0 <= i < len(all_items)]
    if not selected_items:
        flash("No BOQ items selected.", "warning")
        return redirect(url_for("project_distribute", project_id=project.id))

    package_value = sum(float(item.get("total", item.get("amount", 0)) or 0) for item in selected_items)
    initial_statuses = {str(i): "not_started" for i in range(len(selected_items))}

    package = EngineerPackage(
        project_id=project.id,
        boq_id=boq.id,
        assigned_user_id=engineer.id,
        assigned_engineer_name=engineer.full_name,
        trade=trade,
        package_value=package_value,
        scope_description=json.dumps(selected_items),
        completion_percentage=0.0,
        status="active",
        notified_at=datetime.utcnow(),
    )
    package.item_statuses = initial_statuses
    db.session.add(package)
    db.session.flush()

    db.session.add(Notification(
        user_id=engineer.id,
        message=f"You have been assigned a {trade} BOQ package on {project.name}",
        link=url_for("engineer_workspace"),
        is_read=False,
    ))
    db.session.commit()
    flash(f"BOQ package assigned to {engineer.full_name}.", "success")
    return redirect(url_for("project_distribute", project_id=project.id))


@app.route("/projects/<int:project_id>/distribute/inventory", methods=["POST"])
@login_required
@limiter.limit("30 per hour")
@role_required(ROLE_PROJECT_MANAGER)
def project_distribute_inventory(project_id):
    project = get_project_or_403(project_id)
    # Form sends parallel arrays: inventory_item_id[], assigned_user_id[], allocated_qty[]
    item_ids = request.form.getlist("inventory_item_id")
    user_ids = request.form.getlist("assigned_user_id")
    qtys = request.form.getlist("allocated_qty")

    # Get valid project engineer IDs
    valid_user_ids = {a.user_id for a in project.assignments}

    created_count = 0
    for i, item_id_raw in enumerate(item_ids):
        try:
            item_id = int(sanitize_input(item_id_raw, 20).strip())
            user_id = int(sanitize_input(user_ids[i] if i < len(user_ids) else "", 20).strip())
            qty_raw = sanitize_input(qtys[i] if i < len(qtys) else "0", 20).strip()
            qty = Decimal(qty_raw or "0")
        except (ValueError, TypeError, InvalidOperation):
            continue

        if user_id not in valid_user_ids:
            continue

        inv_item = InventoryItem.query.filter_by(id=item_id, project_id=project.id).first()
        if not inv_item:
            continue

        engineer = db.session.get(User, user_id)
        if not engineer:
            continue

        existing = InventoryAssignment.query.filter_by(
            project_id=project.id,
            inventory_item_id=item_id,
            assigned_user_id=user_id,
        ).first()

        if existing:
            existing.allocated_qty = qty
        else:
            assignment_row = InventoryAssignment(
                project_id=project.id,
                inventory_item_id=item_id,
                assigned_user_id=user_id,
                allocated_qty=qty,
                status="allocated",
            )
            db.session.add(assignment_row)
            created_count += 1

        db.session.flush()
        db.session.add(Notification(
            user_id=user_id,
            message=f"{float(qty):,.2f} {inv_item.unit or 'units'} of {inv_item.name} allocated to you on {project.name}",
            link=url_for("my_workspace_project", project_id=project.id),
            is_read=False,
        ))

    db.session.commit()
    flash(f"Inventory assignments saved.", "success")
    return redirect(url_for("project_distribute", project_id=project.id))


@app.route("/projects/<int:project_id>/distribute/task", methods=["POST"])
@login_required
@limiter.limit("30 per hour")
@role_required(ROLE_PROJECT_MANAGER)
def project_distribute_task(project_id):
    project = get_project_or_403(project_id)
    title = sanitize_input(request.form.get("title", ""), 200).strip()
    description = sanitize_input(request.form.get("description", ""), 2000).strip()
    assignee_id_raw = sanitize_input(request.form.get("assignee", ""), 20).strip()
    due_raw = sanitize_input(request.form.get("due", ""), 20).strip()
    priority = sanitize_input(request.form.get("priority", "medium"), 20).strip()
    category = sanitize_input(request.form.get("category", "Admin"), 50).strip()
    is_private = request.form.get("is_private_to_engineer") == "1"

    if not title:
        flash("Task title is required.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    try:
        assignee_id = int(assignee_id_raw)
    except (ValueError, TypeError):
        flash("Invalid assignee.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    # Validate assignee is in project
    assignment = ProjectAssignment.query.filter_by(project_id=project.id, user_id=assignee_id).first()
    if not assignment:
        flash("Selected assignee is not assigned to this project.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    engineer = db.session.get(User, assignee_id)
    if not engineer:
        flash("Engineer not found.", "danger")
        return redirect(url_for("project_distribute", project_id=project.id))

    due_date = None
    if due_raw:
        try:
            due_date = datetime.strptime(due_raw, "%Y-%m-%d").date()
        except ValueError:
            pass

    task = Task(
        user_id=current_user.id,
        project_id=project.id,
        title=title,
        description=description,
        assignee=engineer.full_name,
        due=due_date,
        priority=priority if priority in ["low", "medium", "high", "critical"] else "medium",
        category=category or "Admin",
        status="backlog",
        created_by=current_user.full_name,
        created_by_id=current_user.id,
        is_private_to_engineer=is_private,
    )
    db.session.add(task)
    db.session.flush()

    db.session.add(Notification(
        user_id=engineer.id,
        message=f"New task assigned: {title} on {project.name}",
        link=url_for("my_workspace_project", project_id=project.id),
        is_read=False,
    ))
    db.session.commit()
    flash(f"Task assigned to {engineer.full_name}.", "success")
    return redirect(url_for("project_distribute", project_id=project.id))


# ─── STEP 3: Site Engineer My Workspace ──────────────────────────────────────

@app.route("/my-workspace")
@login_required
@role_required(ROLE_SITE_ENGINEER)
def my_workspace():
    # Projects via ProjectAssignment
    pa_project_ids = {a.project_id for a in current_user.project_assignments}
    # Projects via EngineerPackage
    ep_project_ids = {p.project_id for p in current_user.engineer_packages if p.project_id}
    all_project_ids = pa_project_ids | ep_project_ids

    projects_data = []
    for pid in all_project_ids:
        proj = db.session.get(Project, pid)
        if not proj:
            continue
        # Count child BOQs (new distribution system); fall back to EngineerPackage for legacy
        child_boq_count = BOQ.query.filter(
            BOQ.assigned_to_user_id == current_user.id,
            BOQ.project_id == pid,
            BOQ.status.in_(["distributed", "revised"]),
        ).count()
        package_count = child_boq_count or EngineerPackage.query.filter_by(project_id=pid, assigned_user_id=current_user.id).count()
        materials_count = InventoryAssignment.query.filter_by(project_id=pid, assigned_user_id=current_user.id).count()
        open_task_count = Task.query.filter(
            Task.project_id == pid,
            Task.assigned_to_id == current_user.id,
            Task.status != "done",
        ).count()
        last_dpr = DPR.query.filter_by(project_id=pid, user_id=current_user.id).order_by(DPR.date.desc()).first()
        projects_data.append({
            "project": proj,
            "package_count": package_count,
            "materials_count": materials_count,
            "open_task_count": open_task_count,
            "last_dpr_date": last_dpr.date.strftime("%d %b %Y") if last_dpr and last_dpr.date else None,
        })

    return render_template(
        "projects/my_workspace.html",
        active_dashboard="workspace",
        projects_data=projects_data,
    )


@app.route("/my-workspace/<int:project_id>")
@login_required
@role_required(ROLE_SITE_ENGINEER)
def my_workspace_project(project_id):
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    # 403 if engineer not in project
    has_assignment = ProjectAssignment.query.filter_by(project_id=project_id, user_id=current_user.id).first()
    has_package = EngineerPackage.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).first()
    if not has_assignment and not has_package:
        abort(403)

    packages = EngineerPackage.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).order_by(EngineerPackage.created_at.desc()).all()
    inventory_assignments = InventoryAssignment.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).all()
    tasks = Task.query.filter(
        Task.project_id == project_id,
        Task.assigned_to_id == current_user.id,
    ).order_by(Task.created_at.desc()).all()
    dprs = DPR.query.filter_by(project_id=project_id, user_id=current_user.id).order_by(DPR.date.desc(), DPR.created_at.desc()).all()

    # Child BOQs assigned via the new distribution system (single source of truth)
    _raw_child_boqs = BOQ.query.filter(
        BOQ.assigned_to_user_id == current_user.id,
        BOQ.project_id == project_id,
        BOQ.status.in_(["distributed", "revised"]),
    ).order_by(BOQ.parent_master_boq_id, BOQ.version.desc()).all()
    _seen_boq_keys = set()
    child_boqs = []
    for _b in _raw_child_boqs:
        _key = (_b.parent_master_boq_id, _b.trade_section)
        if _key not in _seen_boq_keys:
            _seen_boq_keys.add(_key)
            child_boqs.append(_b)

    return render_template(
        "projects/my_workspace_project.html",
        active_dashboard="workspace",
        project=project,
        packages=packages,
        child_boqs=child_boqs,
        inventory_assignments=inventory_assignments,
        tasks=tasks,
        dprs=dprs,
        task_status_labels=TASK_STATUS_LABELS,
        item_status_labels=ENGINEER_ITEM_STATUS_LABELS,
    )


@app.route("/my-workspace/<int:project_id>/boq/<int:package_id>/items/<int:item_index>/status", methods=["POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
def my_workspace_boq_item_status(project_id, package_id, item_index):
    package = db.session.get(EngineerPackage, package_id)
    if not package:
        abort(404)
    if package.assigned_user_id != current_user.id:
        abort(403)
    if package.project_id != project_id:
        abort(403)
    new_status = sanitize_input(request.form.get("status", ""), 30)
    if new_status not in ENGINEER_ITEM_STATUS_LABELS:
        abort(400)
    statuses = package.item_statuses
    statuses[str(item_index)] = new_status
    package.item_statuses = statuses
    refresh_package_completion(package)
    db.session.commit()
    return jsonify({
        "success": True,
        "package_completion_percentage": package.completion_percentage,
        "package_status": package.status,
    })


@app.route("/my-workspace/<int:project_id>/inventory/<int:assignment_id>/notify", methods=["POST"])
@login_required
@limiter.limit("10 per hour")
@role_required(ROLE_SITE_ENGINEER)
def my_workspace_inventory_notify(project_id, assignment_id):
    assignment = db.session.get(InventoryAssignment, assignment_id)
    if not assignment:
        abort(404)
    if assignment.assigned_user_id != current_user.id:
        abort(403)
    if assignment.project_id != project_id:
        abort(403)
    project = db.session.get(Project, project_id)
    if not project:
        abort(404)
    db.session.add(Notification(
        user_id=project.user_id,
        message=f"{current_user.full_name} requests inventory update for assignment #{assignment_id} on {project.name}",
        link=url_for("project_distribute", project_id=project_id),
        is_read=False,
    ))
    assignment.notified_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"success": True, "notified_at": assignment.notified_at.strftime("%d %b %Y %H:%M")})


@app.route("/my-workspace/<int:project_id>/tasks/<int:task_id>/notes", methods=["POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
def my_workspace_task_notes(project_id, task_id):
    # Verify engineer is in project
    has_assignment = ProjectAssignment.query.filter_by(project_id=project_id, user_id=current_user.id).first()
    has_package = EngineerPackage.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).first()
    if not has_assignment and not has_package:
        abort(403)
    task = db.session.get(Task, task_id)
    if not task:
        abort(404)
    if task.project_id != project_id:
        abort(403)
    if task.assigned_to_id != current_user.id:
        abort(403)
    notes = sanitize_input(request.form.get("engineer_notes", ""), 5000)
    task.remarks = notes
    task.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"success": True})


@app.route("/my-workspace/<int:project_id>/tasks/<int:task_id>/status", methods=["POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
def my_workspace_task_status(project_id, task_id):
    has_assignment = ProjectAssignment.query.filter_by(project_id=project_id, user_id=current_user.id).first()
    has_package = EngineerPackage.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).first()
    if not has_assignment and not has_package:
        abort(403)
    task = db.session.get(Task, task_id)
    if not task:
        abort(404)
    if task.project_id != project_id:
        abort(403)
    if task.assigned_to_id != current_user.id:
        abort(403)
    new_status = sanitize_input(request.form.get("status", ""), 30)
    if new_status not in TASK_STATUS_LABELS:
        abort(400)

    old_status = task.status or "not_started"
    if old_status != new_status:
        task.status = new_status
        task.updated_at = datetime.utcnow()
        create_task_log(
            task,
            "status_changed",
            f'{current_user.full_name} changed Status from "{TASK_STATUS_LABELS.get(old_status, old_status)}" to "{TASK_STATUS_LABELS[new_status]}"',
            "status",
            old_status,
            new_status,
        )
        db.session.commit()
        invalidate_health_cache(project_id)
        project = db.session.get(Project, project_id)
        if project:
            calculate_health_score(project)

    return jsonify({
        "success": True,
        "status": new_status,
        "label": TASK_STATUS_LABELS[new_status],
    })


@app.route("/my-workspace/<int:project_id>/dprs")
@login_required
@role_required(ROLE_SITE_ENGINEER)
def my_workspace_dprs(project_id):
    # Verify access
    has_assignment = ProjectAssignment.query.filter_by(project_id=project_id, user_id=current_user.id).first()
    has_package = EngineerPackage.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).first()
    if not has_assignment and not has_package:
        abort(403)
    dprs = DPR.query.filter_by(project_id=project_id, user_id=current_user.id).order_by(DPR.date.desc()).all()
    return jsonify([d.to_dict() for d in dprs])



@app.route("/workspace")
@login_required
@role_required(ROLE_SITE_ENGINEER)
def engineer_workspace():
    packages = (
        EngineerPackage.query.filter_by(assigned_user_id=current_user.id)
        .order_by(EngineerPackage.created_at.desc())
        .all()
    )
    package = packages[0] if packages else None
    return render_template(
        "projects/workspace_engineer.html",
        active_dashboard="workspace",
        packages=packages,
        package=package,
        workspace=engineer_workspace_data(package) if package else {},
        task_status_labels=TASK_STATUS_LABELS,
        item_status_labels=ENGINEER_ITEM_STATUS_LABELS,
    )


@app.route("/workspace/<int:package_id>")
@login_required
@role_required(ROLE_SITE_ENGINEER)
def engineer_package_workspace(package_id):
    package = get_package_or_403(package_id)
    packages = (
        EngineerPackage.query.filter_by(assigned_user_id=current_user.id)
        .order_by(EngineerPackage.created_at.desc())
        .all()
    )
    return render_template(
        "projects/workspace_engineer.html",
        active_dashboard="workspace",
        packages=packages,
        package=package,
        workspace=engineer_workspace_data(package),
        task_status_labels=TASK_STATUS_LABELS,
        item_status_labels=ENGINEER_ITEM_STATUS_LABELS,
    )


@app.route("/workspace/<int:package_id>/items/<int:item_index>/status", methods=["POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
def engineer_package_item_status(package_id, item_index):
    package = get_package_or_403(package_id, allow_owner=False)
    new_status = sanitize_input(request.form.get("status", ""), 30)
    if new_status not in ENGINEER_ITEM_STATUS_LABELS:
        abort(400)
    statuses = package.item_statuses
    statuses[str(item_index)] = new_status
    package.item_statuses = statuses
    refresh_package_completion(package)
    db.session.commit()
    if package.project_id:
        invalidate_health_cache(package.project_id)
    return redirect(url_for("engineer_package_workspace", package_id=package.id) + "#my-boq")


@app.route("/workspace/<int:package_id>/tasks/<int:task_id>/status", methods=["POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
def engineer_workspace_task_status(package_id, task_id):
    package = get_package_or_403(package_id, allow_owner=False)
    task = Task.query.filter_by(id=task_id, project_id=package.project_id).first_or_404()
    if task.assigned_to_id != current_user.id:
        abort(403)
    new_status = sanitize_input(request.form.get("status", ""), 30)
    if new_status not in TASK_STATUS_LABELS:
        abort(400)
    old_status = task.status
    task.status = new_status
    task.updated_at = datetime.utcnow()
    create_task_log(task, "status_changed", f"{current_user.full_name} changed task status from {old_status} to {new_status}", "status", old_status, new_status)
    db.session.commit()
    if task.project_id:
        invalidate_health_cache(task.project_id)
        project = db.session.get(Project, task.project_id)
        if project:
            calculate_health_score(project)
    return redirect(url_for("engineer_package_workspace", package_id=package.id) + "#my-tasks")


@app.route("/settings/security", methods=["GET"])
@login_required
def settings_security():
    return render_template("settings/security.html")


@app.route("/settings/security/sign-out-everywhere", methods=["POST"])
@login_required
def sign_out_everywhere():
    current_user.session_token = secrets.token_hex(32)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("Could not complete action. Please try again.", "danger")
        return redirect(url_for("settings_security"))
    # Keep current session alive with the new token
    session["session_token"] = current_user.session_token
    # Security notification — non-blocking
    try:
        from mailer import send_email as _send_email, MailError as _MailError
        _send_email(
            to=current_user.email,
            subject="Signed out everywhere — BanaaIQ",
            html_body=render_template("emails/sign_out_everywhere.html", user=current_user),
        )
    except Exception as _soe_err:
        app.logger.warning("sign-out-everywhere email failed: %s", _soe_err)
    flash("Signed out on all other devices. / تم تسجيل الخروج من جميع الأجهزة الأخرى.", "success")
    return redirect(url_for("settings_security"))


@app.route("/dashboard/settings")
@dashboard_access_required
def dashboard_settings():
    return redirect(url_for("dashboard"))


@app.route("/dashboard/tutorials")
@dashboard_access_required
@role_required(ROLE_SITE_ENGINEER)
def dashboard_tutorials():
    return redirect(url_for("tutorials"))


@app.route("/dashboard/boq")
@dashboard_access_required
def dashboard_boq_index():
    username, company = dashboard_identity()
    if current_user.is_authenticated and current_user.role == ROLE_PROJECT_MANAGER:
        # Master BOQs (not distributed children — those have assigned_to_user_id set)
        master_boqs = BOQ.query.filter_by(
            user_id=current_user.id
        ).filter(BOQ.assigned_to_user_id == None).order_by(BOQ.created_at.desc()).all()
        boq_data = []
        for b in master_boqs:
            try:
                sections = json.loads(b.items_json or '[]')
                sections_count = len(sections) if isinstance(sections, list) and sections and isinstance(sections[0], dict) and 'trade' in sections[0] else 0
            except Exception:
                sections_count = 0
            boq_data.append({'boq': b, 'sections_count': sections_count})
        return render_template('dashboard/boq/index.html',
            username=username, company=company,
            active_dashboard='boq', boq_data=boq_data, role='pm')
    elif current_user.is_authenticated and current_user.role == ROLE_SITE_ENGINEER:
        # SE: BOQs assigned to them (latest version per parent/trade)
        # Exclude engineer_analysis BOQs — those are the SE's own uploads, shown separately
        assigned = BOQ.query.filter_by(
            assigned_to_user_id=current_user.id
        ).filter(
            BOQ.status.in_(['distributed', 'revised']),
            BOQ.source != 'engineer_analysis',
        ).order_by(
            BOQ.parent_master_boq_id, BOQ.version.desc()
        ).all()
        # Keep latest version per (parent_master_boq_id, trade_section) key
        seen_parents = set()
        my_boqs = []
        for b in assigned:
            key = (b.parent_master_boq_id, b.trade_section)
            if key not in seen_parents:
                seen_parents.add(key)
                my_boqs.append(b)
        # SE's own upload-analyze BOQs
        my_analyses = BOQ.query.filter_by(
            user_id=current_user.id,
            source="engineer_analysis",
        ).order_by(BOQ.created_at.desc()).all()
        # SE's own manually-created BOQs
        my_manual_boqs = BOQ.query.filter_by(
            user_id=current_user.id,
            source="engineer_manual",
        ).order_by(BOQ.created_at.desc()).all()
        return render_template('dashboard/boq/index.html',
            username=username, company=company,
            active_dashboard='boq', my_boqs=my_boqs, my_analyses=my_analyses,
            my_manual_boqs=my_manual_boqs, role='se')
    else:
        # Guest or unknown role — show empty PM view
        return render_template('dashboard/boq/index.html',
            username=username, company=company,
            active_dashboard='boq', boq_data=[], role='pm')


@app.route("/dashboard/tasks")
@dashboard_access_required
def dashboard_tasks():
    from sqlalchemy.orm import selectinload as _sl
    uid = current_user.id
    role = current_user.role

    # ── Query filters from QS ──────────────────────────────────────────────
    filter_project_id = request.args.get("project_id", type=int)
    filter_status     = request.args.get("status", "")
    filter_priority   = request.args.get("priority", "")
    filter_search     = sanitize_input(request.args.get("q", ""), 200).strip()

    # ── Base queryset ─────────────────────────────────────────────────────
    if role == ROLE_PROJECT_MANAGER:
        q = (Task.query
             .join(Project, Task.project_id == Project.id)
             .filter(Project.user_id == uid)
             .options(_sl(Task.assigned_to), _sl(Task.project_rel)))
        # PM-accessible projects for the filter dropdown
        projects = Project.query.filter_by(user_id=uid).order_by(Project.name).all()
    else:
        # SE: tasks assigned to them across all their projects
        q = (Task.query
             .filter(Task.assigned_to_id == uid)
             .options(_sl(Task.assigned_to), _sl(Task.project_rel)))
        # SE's projects (via assignments)
        from sqlalchemy.orm import selectinload as _sl2
        assigned_project_ids = [
            a.project_id for a in
            ProjectAssignment.query.filter_by(user_id=uid).all()
        ]
        projects = Project.query.filter(
            Project.id.in_(assigned_project_ids)
        ).order_by(Project.name).all()

    if filter_project_id:
        q = q.filter(Task.project_id == filter_project_id)
    if filter_status and filter_status in TASK_STATUS_LABELS:
        q = q.filter(Task.status == filter_status)
    if filter_priority:
        q = q.filter(Task.priority == filter_priority)
    if filter_search:
        q = q.filter(Task.name.ilike(f"%{filter_search}%"))

    tasks = q.order_by(Task.due_date.asc().nullslast(), Task.created_at.desc()).all()

    # ── Kanban columns ────────────────────────────────────────────────────
    kanban = {k: [] for k in TASK_STATUS_LABELS}
    for t in tasks:
        col = t.status if t.status in kanban else "not_started"
        kanban[col].append(t)

    username, company = dashboard_identity()
    return render_template(
        "dashboard/tasks/index.html",
        active_dashboard="tasks",
        username=username,
        company=company,
        tasks=tasks,
        kanban=kanban,
        projects=projects,
        task_status_labels=TASK_STATUS_LABELS,
        filter_project_id=filter_project_id,
        filter_status=filter_status,
        filter_priority=filter_priority,
        filter_search=filter_search,
        role=role,
    )


@app.route("/dashboard/tasks/new", methods=["GET", "POST"])
@dashboard_access_required
@role_required(ROLE_PROJECT_MANAGER)
def task_new():
    uid = current_user.id
    projects = Project.query.filter_by(user_id=uid).order_by(Project.name).all()
    preselect_project_id = request.args.get("project_id", type=int)

    if request.method == "POST":
        name = sanitize_input(request.form.get("name", ""), 200).strip()
        if not name:
            flash("Task name is required. / اسم المهمة مطلوب.", "error")
            return redirect(url_for("task_new"))

        project_id  = request.form.get("project_id", type=int)
        assignee_id = request.form.get("assigned_to_id", type=int)
        status      = sanitize_input(request.form.get("status", "not_started"), 20)
        priority    = sanitize_input(request.form.get("priority", "normal"), 10)
        due_str     = request.form.get("due_date", "").strip()
        description = sanitize_input(request.form.get("description", ""), 5000)
        remarks     = sanitize_input(request.form.get("remarks", ""), 2000)
        dep_task_id = request.form.get("depends_on_task_id", type=int)

        if not project_id:
            flash("Please select a project. / يرجى اختيار مشروع.", "error")
            return redirect(url_for("task_new"))

        project = Project.query.filter_by(id=project_id, user_id=uid).first()
        if not project:
            abort(403)

        if not assignee_id:
            flash("Please select an assignee. / يرجى اختيار مسند إليه.", "error")
            return redirect(url_for("task_new"))

        # Validate assignee is a member of the project or the PM themselves
        valid_assignees = {uid}  # PM can assign to themselves
        for a in ProjectAssignment.query.filter_by(project_id=project_id).all():
            valid_assignees.add(a.user_id)
        if assignee_id not in valid_assignees:
            abort(403)

        due_date = None
        if due_str:
            try:
                from datetime import date as _date
                due_date = _date.fromisoformat(due_str)
            except ValueError:
                pass

        if status not in TASK_STATUS_LABELS:
            status = "not_started"
        if priority not in ("low", "normal", "high", "urgent"):
            priority = "normal"

        task = Task(
            project_id=project_id,
            name=name,
            description=description or None,
            created_by_id=uid,
            assigned_to_id=assignee_id,
            status=status,
            priority=priority,
            due_date=due_date,
            remarks=remarks or None,
            depends_on_task_id=dep_task_id or None,
        )
        db.session.add(task)
        db.session.flush()  # get task.id before activity

        create_task_log(task, "created", f"Task created by {current_user.full_name}")
        db.session.commit()

        # Notify assignee — wrapped so it cannot kill the redirect
        try:
            if assignee_id != uid:
                assignee = db.session.get(User, assignee_id)
                if assignee:
                    db.session.add(Notification(
                        user_id=assignee_id,
                        title="New Task Assigned / مهمة جديدة",
                        message=f"{current_user.full_name} assigned you: {name}",
                        link=url_for("task_detail", task_id=task.id),
                        is_read=False,
                    ))
                    db.session.commit()
        except Exception as _ne:
            db.session.rollback()
            app.logger.warning("task_new: notification failed for task %d: %s", task.id, _ne)

        flash(f"Task '{name}' created. / تم إنشاء المهمة.", "success")
        return redirect(url_for("task_detail", task_id=task.id))

    username, company = dashboard_identity()
    return render_template(
        "dashboard/tasks/new.html",
        active_dashboard="tasks",
        username=username,
        company=company,
        projects=projects,
        preselect_project_id=preselect_project_id,
        task_status_labels=TASK_STATUS_LABELS,
    )


@app.route("/api/projects/<int:project_id>/members")
@login_required
def api_project_members(project_id):
    """Return assignable members for a project (PM + all SEs assigned)."""
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id and current_user.role != ROLE_PROJECT_MANAGER:
        abort(403)
    members = []
    # Include the PM (project owner)
    pm = db.session.get(User, project.user_id)
    if pm:
        members.append({"id": pm.id, "name": pm.full_name, "role": "PM"})
    # Assigned SEs
    for a in ProjectAssignment.query.filter_by(project_id=project_id).all():
        if a.user and a.user.id != project.user_id:
            members.append({"id": a.user.id, "name": a.user.full_name, "role": "SE"})
    return jsonify({"members": members})


@app.route("/api/projects/<int:project_id>/tasks")
@login_required
def api_project_tasks(project_id):
    """Return task stubs for a project (id + name) for dependency picker."""
    project = Project.query.get_or_404(project_id)
    if project.user_id != current_user.id and current_user.role != ROLE_PROJECT_MANAGER:
        abort(403)
    tasks = Task.query.filter_by(project_id=project_id).order_by(Task.name).all()
    return jsonify({"tasks": [{"id": t.id, "name": t.name} for t in tasks]})


@app.route("/api/tasks/transcribe", methods=["POST"])
@login_required
def task_transcribe():
    forbidden = ensure_ai_access("task")
    if forbidden:
        return forbidden
    if client is None:
        return jsonify({"success": False, "error": "AI not configured."}), 503
    audio_file = request.files.get("audio")
    if not audio_file:
        return jsonify({"success": False, "error": "No audio file"}), 400
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as tmp:
            audio_file.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            result = client.audio.transcriptions.create(
                model="whisper-1", file=f, response_format="verbose_json"
            )
        record_ai_usage("task_voice")
        return jsonify({"success": True, "text": result.text or "",
                        "language_detected": getattr(result, "language", "auto")})
    except Exception as e:
        app.logger.error(f"Task Whisper error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try: os.unlink(tmp_path)
            except Exception: pass


@app.route("/dashboard/tasks/<int:task_id>")
@dashboard_access_required
def task_detail(task_id):
    from sqlalchemy.orm import selectinload as _sl
    task = (Task.query
            .options(_sl(Task.assigned_to), _sl(Task.created_by),
                     _sl(Task.project_rel), _sl(Task.depends_on))
            .filter_by(id=task_id).first_or_404())
    uid = current_user.id
    role = current_user.role

    # PM can see any task in their projects; SE can see tasks assigned to them
    if role == ROLE_PROJECT_MANAGER:
        project = Project.query.filter_by(id=task.project_id, user_id=uid).first()
        if not project:
            abort(403)
    else:
        has_access = (
            task.assigned_to_id == uid or
            ProjectAssignment.query.filter_by(project_id=task.project_id, user_id=uid).first()
        )
        if not has_access:
            abort(403)

    # Use direct query instead of dynamic relationship to avoid lazy="dynamic" edge cases
    activities = (TaskActivity.query
                  .filter_by(task_id=task_id)
                  .order_by(TaskActivity.created_at.asc())
                  .all())

    # Dependency tasks in same project
    dep_candidates = Task.query.filter(
        Task.project_id == task.project_id,
        Task.id != task_id,
    ).order_by(Task.name).all() if role == ROLE_PROJECT_MANAGER else []

    # Project members for reassign
    members = []
    if role == ROLE_PROJECT_MANAGER:
        pm = db.session.get(User, task.project_rel.user_id if task.project_rel else uid)
        if pm:
            members.append({"id": pm.id, "name": pm.full_name, "role": "PM"})
        for a in ProjectAssignment.query.filter_by(project_id=task.project_id).all():
            if a.user and a.user.id != (pm.id if pm else None):
                members.append({"id": a.user.id, "name": a.user.full_name, "role": "SE"})

    username, company = dashboard_identity()
    return render_template(
        "dashboard/tasks/detail.html",
        active_dashboard="tasks",
        username=username,
        company=company,
        task=task,
        activities=activities,
        dep_candidates=dep_candidates,
        members=members,
        task_status_labels=TASK_STATUS_LABELS,
        role=role,
    )


@app.route("/dashboard/tasks/<int:task_id>/status", methods=["POST"])
@dashboard_access_required
def task_update_status(task_id):
    task = Task.query.filter_by(id=task_id).first_or_404()
    uid = current_user.id
    role = current_user.role

    if role == ROLE_PROJECT_MANAGER:
        if not Project.query.filter_by(id=task.project_id, user_id=uid).first():
            abort(403)
    else:
        if task.assigned_to_id != uid:
            abort(403)

    new_status = sanitize_input(request.form.get("status", ""), 20)
    if new_status not in TASK_STATUS_LABELS:
        return jsonify({"success": False, "error": "Invalid status"}), 400

    old_status = task.status
    if old_status == new_status:
        return jsonify({"success": True, "status": new_status, "label": TASK_STATUS_LABELS[new_status]})

    task.status = new_status
    task.updated_at = datetime.utcnow()
    if new_status == "done" and not task.completed_at:
        task.completed_at = datetime.utcnow()
    elif new_status != "done":
        task.completed_at = None

    create_task_log(task, "status_changed",
                    f"Status changed from {old_status} to {new_status}",
                    "status", old_status, new_status)
    db.session.commit()
    invalidate_health_cache(task.project_id)
    return jsonify({"success": True, "status": new_status,
                    "label": TASK_STATUS_LABELS[new_status]})


@app.route("/dashboard/tasks/<int:task_id>/edit", methods=["GET", "POST"])
@dashboard_access_required
@role_required(ROLE_PROJECT_MANAGER)
def task_edit(task_id):
    task = Task.query.filter_by(id=task_id).first_or_404()
    if not Project.query.filter_by(id=task.project_id, user_id=current_user.id).first():
        abort(403)

    if request.method == "POST":
        name = sanitize_input(request.form.get("name", ""), 200).strip()
        if not name:
            flash("Name is required.", "error")
            return redirect(url_for("task_edit", task_id=task_id))

        assignee_id = request.form.get("assigned_to_id", type=int)
        status      = sanitize_input(request.form.get("status", task.status), 20)
        priority    = sanitize_input(request.form.get("priority", task.priority), 10)
        due_str     = request.form.get("due_date", "").strip()
        description = sanitize_input(request.form.get("description", ""), 5000)
        remarks     = sanitize_input(request.form.get("remarks", ""), 2000)
        dep_id      = request.form.get("depends_on_task_id", type=int)

        # Validate assignee
        valid_ids = {current_user.id}
        for a in ProjectAssignment.query.filter_by(project_id=task.project_id).all():
            valid_ids.add(a.user_id)
        if assignee_id and assignee_id not in valid_ids:
            abort(403)

        due_date = task.due_date
        if due_str:
            try:
                from datetime import date as _d
                due_date = _d.fromisoformat(due_str)
            except ValueError:
                pass
        elif due_str == "":
            due_date = None

        changes = []
        if task.name != name: changes.append(f"name: {task.name!r} → {name!r}")
        if task.status != status: changes.append(f"status: {task.status} → {status}")
        if task.priority != priority: changes.append(f"priority: {task.priority} → {priority}")

        task.name        = name
        task.description = description or None
        task.assigned_to_id = assignee_id or task.assigned_to_id
        task.status      = status if status in TASK_STATUS_LABELS else task.status
        task.priority    = priority if priority in ("low","normal","high","urgent") else task.priority
        task.due_date    = due_date
        task.remarks     = remarks or None
        task.depends_on_task_id = dep_id if dep_id and dep_id != task_id else None
        task.updated_at  = datetime.utcnow()
        if task.status == "done" and not task.completed_at:
            task.completed_at = datetime.utcnow()

        if changes:
            create_task_log(task, "edited", "; ".join(changes))
        db.session.commit()
        flash("Task updated. / تم تحديث المهمة.", "success")
        return redirect(url_for("task_detail", task_id=task_id))

    # GET: render edit form
    dep_candidates = Task.query.filter(
        Task.project_id == task.project_id, Task.id != task_id
    ).order_by(Task.name).all()
    members = []
    proj = task.project_rel
    if proj:
        pm = db.session.get(User, proj.user_id)
        if pm: members.append({"id": pm.id, "name": pm.full_name, "role": "PM"})
    for a in ProjectAssignment.query.filter_by(project_id=task.project_id).all():
        if a.user:
            members.append({"id": a.user.id, "name": a.user.full_name, "role": "SE"})

    username, company = dashboard_identity()
    return render_template(
        "dashboard/tasks/edit.html",
        active_dashboard="tasks",
        username=username,
        company=company,
        task=task,
        members=members,
        dep_candidates=dep_candidates,
        task_status_labels=TASK_STATUS_LABELS,
    )


@app.route("/dashboard/tasks/<int:task_id>/delete", methods=["POST"])
@dashboard_access_required
@role_required(ROLE_PROJECT_MANAGER)
def task_delete(task_id):
    task = Task.query.filter_by(id=task_id).first_or_404()
    if not Project.query.filter_by(id=task.project_id, user_id=current_user.id).first():
        abort(403)
    project_id = task.project_id
    db.session.delete(task)
    db.session.commit()
    invalidate_health_cache(project_id)
    flash("Task deleted. / تم حذف المهمة.", "info")
    return redirect(url_for("dashboard_tasks"))


@app.route("/dashboard/boq/new")
@dashboard_access_required
@role_required(ROLE_PROJECT_MANAGER)
def dashboard_boq_new_legacy():
    return redirect(url_for("boq_create"))


@app.route("/dashboard/boq/upload")
@dashboard_access_required
@role_required(ROLE_SITE_ENGINEER)
def dashboard_boq_upload_legacy():
    return redirect(url_for("boq_upload_revise"))


@app.route("/boq/generate-from-description", methods=["POST"])
@dashboard_access_required
@role_required(ROLE_PROJECT_MANAGER)
def boq_generate_from_description_legacy():
    description = sanitize(request.form.get("description", "").strip(), max_length=4000)
    if description:
        session["boq_create_prefill_description"] = description
    return redirect(url_for("boq_create"))


# ── BOQ REBUILD — HELPERS ────────────────────────────────────────────────────

# Rich item detail block appended to every BOQ AI system prompt
_BOQ_ITEM_DETAIL_PROMPT = (
    "\n\nCRITICAL DETAIL REQUIREMENT — Engineers use this BOQ as their working site document.\n"
    "Every item MUST include ALL of these fields (generic items are unacceptable):\n"
    '- "specification": Material grade/standard/strength '
    '(e.g. "Grade C40 (40 MPa), sulfate-resistant cement Type V, SBC 304 compliant, 28-day cure")\n'
    '- "sub_items": Component breakdown array — '
    '[{"name":"Portland Cement Type V","qty_per_unit":350,"unit":"kg/m3"}, ...]\n'
    '- "application_notes": On-site installation and quality-control notes for engineers\n'
    '- "standard_reference": Applicable Saudi/international codes (e.g. "SBC 304-2018, ACI 318")\n'
    "For MEP sections (HVAC, Electrical, Plumbing, Fire Fighting), ALSO add per item:\n"
    '- "equipment_type": Equipment category (e.g. "VRF ODU/IDU", "LV Distribution Panel", "PPR Pipe")\n'
    '- "suggested_model": Real Saudi-available brand + model number '
    '(Gree, LG, Daikin, Carrier, ABB, Schneider, Saudi Cable, SABIC) + "or equivalent"\n'
    '  Example: "GMV-400WM/G1-X(S) or equivalent"\n'
    '- "accessories": Array of accessories '
    '["Copper pipe 3/8\'\' x 15m", "PE insulation sleeve", "Wall bracket", ...]\n\n'
    "Be specific like a real Saudi MEP/Civil/Structural engineer writing a tender document.\n"
    "Updated required item schema:\n"
    '{"item_no":"1.1","description":"Reinforced Concrete Foundation Footings",'
    '"description_ar":"\u062e\u0631\u0633\u0627\u0646\u0629 \u0645\u0633\u0644\u062d\u0629 \u0644\u0644\u0623\u0633\u0627\u0633\u0627\u062a",'
    '"specification":"Grade C40 (40 MPa), sulfate-resistant cement Type V, SBC 304 compliant, w/c 0.45 max",'
    '"sub_items":[{"name":"Portland Cement Type V","qty_per_unit":350,"unit":"kg/m3"},'
    '{"name":"Coarse Aggregate 20mm","qty_per_unit":1050,"unit":"kg/m3"},'
    '{"name":"Reinforcement B500B","qty_per_unit":120,"unit":"kg/m3"}],'
    '"application_notes":"Apply waterproof membrane below slab. Cure min 7 days. Structural inspection before backfill.",'
    '"standard_reference":"SBC 304-2018, ACI 318",'
    '"unit":"m3","qty":100,"rate_sar":450,"total_sar":45000}'
)

# ── Inventory item detail prompt (appended to inventory system prompts) ───────
_INVENTORY_ITEM_DETAIL_PROMPT = (
    "\n\nEVERY item must include these procurement fields:\n"
    '- "specification": Technical grade/standard/attributes (e.g. "ASTM C150 Type V, 28-day strength ≥40 MPa, low alkali"). '
    'NOT a generic description. This is what the procurement officer orders.\n'
    '- "brand_suggestions": Array of 3-5 real Saudi-market brands. Use actual companies operating in KSA:\n'
    '  Cement: ["Saudi Cement Company","Yamama Cement","Najran Cement","Qassim Cement","or equivalent"]\n'
    '  Steel/Rebar: ["SABIC HADEED","Al Rajhi Steel","Al Tuwairqi Steel","or equivalent"]\n'
    '  Electrical: ["ABB Saudi","Schneider Electric KSA","Saudi Cable Co.","or equivalent"]\n'
    '  HVAC: ["Gree Saudi","LG Saudi","Daikin","Carrier","or equivalent"]\n'
    '  Plumbing: ["SAFAT","Hepworth","Geberit","Saudi Modern Plastic","or equivalent"]\n'
    '  Paint: ["Jotun Saudi","National Paints","SIPES","or equivalent"]\n'
    '  Tiles: ["RAK Ceramics","Saudi Ceramics","Future Ceramic","or equivalent"]\n'
    '  Safety: ["3M Saudi","MSA Safety","Honeywell","or equivalent"]\n'
    '  Always end with "or equivalent".\n'
    '- "supplier_hint": String — specific Saudi distributor or region (e.g. "Saudi Cement Company — Riyadh distributor")\n'
    '- "storage_requirements": String — KSA-climate-appropriate storage (heat, dust, humidity). Specific to material type.\n'
    '- "reorder_lead_time_days": Integer — realistic KSA procurement lead time:\n'
    '  Cement/aggregate 3-7 days; standard rebar 7-14 days; imported electrical/MEP 21-45 days; specialty 30-60 days\n'
    '- "min_order_qty": Number — typical supplier minimum quantity\n'
    '- "min_order_unit": String — unit description (e.g. "bags (1 pallet)", "ton", "reel")\n'
    '- "safety_notes": String — PPE and handling hazards. Required under Saudi OSHA/HSE.\n'
    '- "alternative_items": Array of 1-3 objects: [{"name":"...","note":"trade-off or condition for substitution"}]\n'
    '- "shelf_life_days": Integer — for perishable materials (cement 90, paint 365, adhesive 180). Null for non-perishable.\n'
    '- "shelf_life_note": String — rotation and testing guidance. Null for non-perishable.\n'
    '- "notes": String — project-context notes (null is acceptable).\n\n'
    "Updated required item schema example:\n"
    '{"name":"Portland Cement Type V","name_ar":"إسمنت بورتلاندي نوع 5",'
    '"unit":"bag (50kg)","recommended_stock":500,"threshold":100,"value_sar":26,'
    '"specification":"Sulfate-resistant ASTM C150 Type V, 28-day compressive strength ≥40 MPa, low alkali content",'
    '"brand_suggestions":["Saudi Cement Company","Yamama Cement","Najran Cement","or equivalent"],'
    '"supplier_hint":"Saudi Cement Company — Riyadh region distributor",'
    '"storage_requirements":"Indoor dry storage. Stack max 10 bags. Pallets required. Avoid moisture and direct sun. FIFO rotation.",'
    '"reorder_lead_time_days":7,'
    '"min_order_qty":50,"min_order_unit":"bags (1 pallet)",'
    '"safety_notes":"Alkaline irritant. PPE: N95 dust mask, safety goggles, gloves, long sleeves.",'
    '"alternative_items":[{"name":"Portland Cement Type I","note":"Only if Type V unavailable; lower sulfate resistance — not for high-sulfate-soil foundations"}],'
    '"shelf_life_days":90,"shelf_life_note":"From manufacture date on bag. Test before use if over 90 days.",'
    '"notes":null}'
)

_BOQ_CHUNK_DETAIL_PROMPT = (
    "\n\nEach item should include the rich detail fields used by BanaaIQ: "
    "specification, sub_items, application_notes, standard_reference, unit, qty, rate_sar, total_sar. "
    "For MEP/service items also include equipment_type, suggested_model, and accessories where relevant. "
    "Keep JSON compact: specification and notes must be short phrases, not paragraphs. "
    "If a bonus field is not applicable, use [] or null instead of expanding prose."
)

_INVENTORY_CHUNK_DETAIL_PROMPT = (
    "\n\nEach item should include the rich procurement fields used by BanaaIQ: "
    "specification, brand_suggestions, supplier_hint, storage_requirements, reorder_lead_time_days, "
    "min_order_qty, min_order_unit, safety_notes, alternative_items, shelf_life_days, shelf_life_note, notes. "
    "Use concise Saudi-market brands and put null or [] for non-applicable bonus fields."
)

# India-market prompt extensions — appended to base prompts when user.country == "IN"
_BOQ_INDIA_EXTENSION = (
    "\n\nThis is an INDIA project. Use realistic 2025 Indian market rates in INR (₹). "
    "Add a 'description_hi' field to every item with the Hindi (Devanagari script) translation of the description — "
    "e.g. 'description_hi': 'आर.सी.सी. स्तंभ निर्माण'. "
    "Also add 'rate_inr' and 'total_inr' fields alongside rate_sar and total_sar (set SAR fields to 0). "
    "Use IS codes (Bureau of Indian Standards) for standard_reference where applicable. "
    "Use Indian construction terminology and metric units common on Indian sites."
)

_INVENTORY_INDIA_EXTENSION = (
    "\n\nThis is an INDIA project. Use Indian market suppliers and brand names (UltraTech, ACC, JSW, TATA Steel, "
    "Astral Pipes, Havells, Finolex etc.). Include a 'name_hi' field with the Hindi/Devanagari name for each item. "
    "Use INR values and Indian material grades (Fe415/Fe500 for rebar, IS:383 for aggregates etc.)."
)


def _extract_json_from_ai_response(raw_text):
    """Strip markdown fences and prose, return parsed dict or None."""
    if not raw_text:
        return None
    cleaned = re.sub(r'^```(?:json)?\s*', '', raw_text.strip(), flags=re.IGNORECASE)
    cleaned = re.sub(r'\s*```\s*$', '', cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    match = re.search(r'\{.*\}', cleaned, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            return None
    return None


AI_GENERATION_TIMEOUT_SECONDS = 60
AI_GENERATION_MAX_WORKERS = 5


def _ai_payload_text_length(payload):
    if isinstance(payload, str):
        return len(payload)
    return sum(len(p.get("text", "")) for p in payload if isinstance(p, dict))


def _ai_usage_tuple(response):
    usage = getattr(response, "usage", None)
    return (
        getattr(usage, "prompt_tokens", None),
        getattr(usage, "completion_tokens", None),
        getattr(usage, "total_tokens", None),
    )


def _openai_json_completion(call_id, system_prompt, user_payload, max_tokens, timeout_seconds=AI_GENERATION_TIMEOUT_SECONDS):
    api_client = client.with_options(timeout=timeout_seconds) if hasattr(client, "with_options") else client
    app.logger.info(
        "[%s] OpenAI JSON call system_len=%d user_len=%d max_tokens=%d timeout=%ss",
        call_id, len(system_prompt), _ai_payload_text_length(user_payload), max_tokens, timeout_seconds,
    )
    response = api_client.chat.completions.create(
        model="gpt-4o-mini",
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_payload},
        ],
        max_tokens=max_tokens,
        temperature=0.2,
    )
    raw = response.choices[0].message.content or ""
    finish_reason = response.choices[0].finish_reason
    prompt_tokens, completion_tokens, total_tokens = _ai_usage_tuple(response)
    app.logger.info(
        "[%s] finish_reason=%s prompt_tokens=%s completion_tokens=%s total_tokens=%s raw_len=%d",
        call_id, finish_reason, prompt_tokens, completion_tokens, total_tokens, len(raw),
    )
    return response, raw, finish_reason


def _clean_ai_label_list(values, fallback_key=None):
    if isinstance(values, dict) and fallback_key:
        values = values.get(fallback_key)
    if isinstance(values, str):
        values = [v.strip() for v in re.split(r"[,;\n]", values)]
    if not isinstance(values, list):
        return []
    labels = []
    seen = set()
    for value in values:
        if not isinstance(value, str):
            continue
        cleaned = re.sub(r"\s+", " ", value).strip(" -:;,.")
        if not cleaned:
            continue
        key = cleaned.casefold()
        if key in seen:
            continue
        seen.add(key)
        labels.append(cleaned[:80])
    return labels


def _add_unique_label(labels, label):
    if label.casefold() not in {item.casefold() for item in labels}:
        labels.append(label)


def _consolidate_boq_trades(trades, description=""):
    grouped = []
    text = " ".join(trades + [description or ""]).lower()
    complex_project = any(term in text for term in (
        "tower", "kafd", "basement", "hospital", "hotel", "industrial", "airport",
        "mall", "high-rise", "high rise", "mixed-use", "mixed use"
    ))
    for trade in trades:
        lowered = trade.lower()
        if any(term in lowered for term in ("civil", "earth", "site prep", "excavation")):
            _add_unique_label(grouped, "Civil")
        elif any(term in lowered for term in ("structural", "concrete", "steel", "rebar")):
            _add_unique_label(grouped, "Structural")
        elif any(term in lowered for term in (
            "mep", "electrical", "plumbing", "hvac", "fire", "bms", "elv", "mechanical", "low current"
        )):
            _add_unique_label(grouped, "MEP")
        elif any(term in lowered for term in ("finish", "fit-out", "fitout", "tiling", "paint", "ceiling", "floor")):
            _add_unique_label(grouped, "Finishing")
        elif any(term in lowered for term in ("external", "landscape", "parking", "utility", "road", "boundary")):
            _add_unique_label(grouped, "External Works")
        elif any(term in lowered for term in ("facade", "curtain wall", "glazing")):
            _add_unique_label(grouped, "Facade")
        elif any(term in lowered for term in ("elevator", "lift", "vertical")):
            _add_unique_label(grouped, "Vertical Transportation")
        else:
            _add_unique_label(grouped, trade)

    if "villa" in text or "building" in text or "tower" in text or "office" in text:
        for fallback in ("Civil", "Structural", "MEP", "Finishing"):
            _add_unique_label(grouped, fallback)
    if any(term in text for term in ("external", "landscape", "parking", "utility", "tower", "kafd")):
        _add_unique_label(grouped, "External Works")
    if "facade" in text or "curtain wall" in text:
        _add_unique_label(grouped, "Facade")
    if not complex_project:
        compact = []
        if "Civil" in grouped or "Structural" in grouped:
            _add_unique_label(compact, "Civil & Structural")
        if "MEP" in grouped:
            _add_unique_label(compact, "MEP")
        if "Finishing" in grouped:
            _add_unique_label(compact, "Finishing")
        if "External Works" in grouped:
            _add_unique_label(compact, "External Works")
        if len(compact) >= 2:
            return compact[:4]
    if len(grouped) > 5 and "Facade" in grouped and "External Works" in grouped:
        grouped.remove("Facade")
    return grouped[:5]


def _consolidate_inventory_categories(categories, description=""):
    grouped = []
    text = " ".join(categories + [description or ""]).lower()
    for category in categories:
        lowered = category.lower()
        if any(term in lowered for term in ("cement", "concrete", "aggregate", "block")):
            _add_unique_label(grouped, "Cement & Concrete")
        elif any(term in lowered for term in ("steel", "rebar", "structural")):
            _add_unique_label(grouped, "Steel & Rebar")
        elif any(term in lowered for term in ("electrical", "plumbing", "hvac", "fire", "mep", "bms", "elv")):
            _add_unique_label(grouped, "MEP Materials")
        elif any(term in lowered for term in ("finish", "tiling", "paint", "coating", "ceiling", "floor")):
            _add_unique_label(grouped, "Finishing Materials")
        elif any(term in lowered for term in ("door", "window", "glass", "facade", "curtain")):
            _add_unique_label(grouped, "Doors, Windows & Facade")
        elif any(term in lowered for term in ("safety", "hse", "ppe")):
            _add_unique_label(grouped, "Safety Equipment")
        elif any(term in lowered for term in ("external", "landscape", "parking", "utility")):
            _add_unique_label(grouped, "External Works Materials")
        else:
            _add_unique_label(grouped, category)

    if "villa" in text or "building" in text or "tower" in text or "office" in text:
        for fallback in ("Cement & Concrete", "Steel & Rebar", "MEP Materials", "Finishing Materials"):
            _add_unique_label(grouped, fallback)
    if any(term in text for term in ("facade", "curtain wall", "tower", "kafd")):
        _add_unique_label(grouped, "Doors, Windows & Facade")
    if any(term in text for term in ("safety", "tower", "kafd")):
        _add_unique_label(grouped, "Safety Equipment")
    return grouped[:5]


def _description_is_too_vague(description):
    words = re.findall(r"[A-Za-z0-9\u0600-\u06FF]+", description or "")
    if len(words) <= 2:
        return True
    construction_terms = (
        "villa", "tower", "building", "warehouse", "office", "hotel", "hospital",
        "school", "sqm", "m2", "floor", "floors", "structural", "civil", "mep",
        "electrical", "plumbing", "finishing", "riyadh", "jeddah", "dammam",
        "construction", "renovation", "fitout", "fit-out",
    )
    lowered = (description or "").lower()
    return len(description or "") < 25 and not any(term in lowered for term in construction_terms)


def _coerce_float(value, default=0.0):
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _validate_boq_response(parsed, min_sections=1, require_detail=False, min_total_items=None):
    """Return (is_valid, reason)."""
    if not isinstance(parsed, dict):
        return False, "Response is not a JSON object"
    sections = parsed.get('sections')
    if not isinstance(sections, list) or len(sections) == 0:
        return False, "No trade sections returned"
    if len(sections) < min_sections:
        return False, f"Only {len(sections)} trade section(s) returned; {min_sections} required for this project"
    total_items = sum(len(s.get('items', [])) for s in sections if isinstance(s.get('items'), list))
    required_total = min_total_items if min_total_items is not None else (5 if min_sections >= 2 else 1)
    if total_items < required_total:
        return False, f"Only {total_items} BOQ item(s) returned; {required_total} required for this project"

    all_items = [
        item for s in sections if isinstance(s, dict)
        for item in s.get('items', []) if isinstance(item, dict)
    ]
    if all_items:
        missing_spec = sum(1 for item in all_items if not str(item.get('specification') or '').strip())
        missing_spec_ratio = missing_spec / len(all_items)
        if missing_spec_ratio > 0.40:
            app.logger.warning(
                "BOQ validator warning: specification missing on %.0f%% of items (%d/%d).",
                missing_spec_ratio * 100, missing_spec, len(all_items),
            )

        bonus_fields = ("sub_items", "application_notes", "standard_reference")
        for field in bonus_fields:
            missing_count = sum(1 for item in all_items if not item.get(field))
            if missing_count:
                app.logger.info(
                    "BOQ validator bonus-field warning: %s missing on %d/%d items.",
                    field, missing_count, len(all_items),
                )
    return True, "ok"


def _call_ai_for_master_boq(system_prompt, user_payload, attempt=1, min_sections=1, require_detail=True):
    """Call GPT with json_object mode; retry once on parse/validation failure."""
    import time as _time
    call_id = f"boq-{int(_time.time()*1000)}"

    if client is None:
        return None, "AI service is not configured. Please set OPENAI_API_KEY.", ""

    system_len = len(system_prompt)
    user_len = len(user_payload) if isinstance(user_payload, str) else sum(
        len(p.get("text", "")) for p in user_payload if isinstance(p, dict)
    )
    app.logger.info(
        "[%s] BOQ AI attempt=%d system_len=%d user_len=%d",
        call_id, attempt, system_len, user_len,
    )

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_payload},
            ],
            max_tokens=16000,
            temperature=0.2,
        )
    except Exception as e:
        app.logger.error("[%s] OpenAI exception: %s: %s", call_id, type(e).__name__, e)
        return None, f"openai_error: {e}", ""

    finish_reason = response.choices[0].finish_reason
    raw = response.choices[0].message.content or ""
    usage = response.usage

    app.logger.info(
        "[%s] finish_reason=%s prompt_tokens=%s completion_tokens=%s total_tokens=%s raw_len=%d",
        call_id, finish_reason,
        usage.prompt_tokens, usage.completion_tokens, usage.total_tokens,
        len(raw),
    )
    app.logger.info("[%s] raw_start: %s", call_id, raw[:400])
    app.logger.info("[%s] raw_end: %s", call_id, raw[-400:] if len(raw) > 400 else raw)

    if finish_reason == "length":
        app.logger.error(
            "[%s] BOQ AI hit token limit (finish_reason=length) on attempt %d — "
            "response truncated at %d chars. JSON will likely be invalid.",
            call_id, attempt, len(raw),
        )

    parsed = _extract_json_from_ai_response(raw)

    if parsed is None:
        app.logger.error("[%s] JSON parse FAILED on attempt %d", call_id, attempt)
    else:
        sec_count = len(parsed.get('sections', []))
        item_count = sum(len(s.get('items', [])) for s in parsed.get('sections', []) if isinstance(s, dict))
        app.logger.info("[%s] PARSED OK: sections=%d items=%d", call_id, sec_count, item_count)

    # Detect AI explicitly saying the description is too vague
    if isinstance(parsed, dict) and parsed.get("error") == "description_too_vague":
        app.logger.warning("[%s] BOQ AI returned description_too_vague on attempt %d", call_id, attempt)
        return None, "description_too_vague", raw

    if finish_reason == "length" and parsed is None:
        # Truncated response — no point retrying with same or stricter prompt;
        # return a clear reason so the caller can show the right message.
        return None, "token_limit_truncation", raw

    is_valid, reason = (
        _validate_boq_response(parsed, min_sections=min_sections, require_detail=require_detail)
        if parsed is not None else (False, "parse failed")
    )
    app.logger.info("[%s] validation: is_valid=%s reason='%s'", call_id, is_valid, reason)

    if is_valid:
        app.logger.info(
            "[%s] BOQ AI success: trades_detected=%s sections=%d items=%d",
            call_id,
            parsed.get('trades_detected'),
            len(parsed.get('sections', [])),
            sum(len(s.get('items', [])) for s in parsed.get('sections', []) if isinstance(s, dict)),
        )
        return parsed, None, raw

    app.logger.warning("[%s] BOQ AI attempt=%d FAILED validation: reason=%s", call_id, attempt, reason)

    if attempt < 2:
        # Retry: ask for correct structure only — do NOT demand richer detail
        # (that would make the response larger and risk truncation again).
        app.logger.warning("[%s] Retrying with stricter prompt", call_id)
        retry_system = system_prompt + (
            "\n\nCRITICAL: Your previous response was rejected. Return ONLY a valid JSON object matching the schema. "
            "No markdown, no prose, no code fences. "
            f"You MUST include at least {min_sections} separate trade sections, each with at least 2 items. "
            "Do NOT return only Civil — a real construction project always needs Civil, Structural, "
            "MEP, and Finishing at minimum. "
            "It is acceptable for some items to omit optional fields (sub_items, application_notes) "
            "if needed to keep the response within limits — structure and coverage matter most."
        )
        return _call_ai_for_master_boq(
            retry_system, user_payload, attempt=2,
            min_sections=min_sections, require_detail=require_detail
        )

    app.logger.error("[%s] FINAL FAILURE after %d attempts", call_id, attempt)
    return None, reason, raw


def _normalize_inventory_category(category, fallback_category):
    if not isinstance(category, dict):
        return None
    items = category.get("items")
    if not isinstance(items, list):
        return None
    normalized_items = []
    for item in items:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        row.setdefault("name", row.get("item") or fallback_category)
        row.setdefault("name_ar", "")
        row.setdefault("unit", row.get("uom") or "pcs")
        row["recommended_stock"] = _coerce_float(row.get("recommended_stock", row.get("stock", row.get("quantity", 0))))
        row["threshold"] = _coerce_float(row.get("threshold", 0))
        row["value_sar"] = _coerce_float(row.get("value_sar", row.get("unit_rate_sar", row.get("rate_sar", 0))))
        normalized_items.append(row)
    if not normalized_items:
        return None
    return {
        "category": sanitize(str(category.get("category") or fallback_category), 80),
        "category_ar": sanitize(str(category.get("category_ar") or ""), 80),
        "items": normalized_items,
    }


def _detect_inventory_categories(project, description, user_payload):
    if client is None:
        return None, "AI service is not configured. Please set OPENAI_API_KEY.", ""
    if _description_is_too_vague(description):
        app.logger.warning("Inventory category detection skipped: description_too_vague")
        return [], "description_too_vague", ""
    call_id = f"inventory-detect-{uuid.uuid4().hex[:8]}"
    system_prompt = (
        "Read the Saudi construction project description. Return strict JSON only with the list of "
        "master inventory categories this project needs. Be comprehensive but realistic. Use categories like "
        "Cement & Concrete, Steel & Rebar, Electrical, Plumbing, HVAC, Fire Fighting, Finishing & Tiling, "
        "Paints & Coatings, Doors & Windows, Safety Equipment. If the description is too vague to identify "
        "a construction project, return {\"categories\":[],\"error\":\"description_too_vague\"}."
    )
    detect_payload = _prepend_text_to_user_payload(
        user_payload,
        f"Project: {project.name}.\nTask: Detect inventory categories only. Return JSON like "
        "{\"categories\":[\"Cement & Concrete\",\"Steel & Rebar\"]}.",
    )
    try:
        _, raw, finish_reason = _openai_json_completion(call_id, system_prompt, detect_payload, max_tokens=500)
    except Exception as error:
        app.logger.error("[%s] Inventory category detection failed: %s: %s", call_id, type(error).__name__, error)
        return None, f"openai_error: {error}", ""
    parsed = _extract_inventory_json(raw)
    if not isinstance(parsed, dict):
        app.logger.error("[%s] Inventory category detection JSON parse failed raw_len=%d", call_id, len(raw))
        return None, "parse failed", raw
    if parsed.get("error") == "description_too_vague":
        return [], "description_too_vague", raw
    categories = _clean_ai_label_list(parsed.get("categories") or parsed.get("categories_detected"))
    if finish_reason == "length":
        categories = categories[:10]
    categories = _consolidate_inventory_categories(categories, description)
    if len(categories) < 3:
        for fallback in ("Cement & Concrete", "Steel & Rebar", "MEP Materials", "Finishing Materials"):
            if fallback.casefold() not in {c.casefold() for c in categories}:
                categories.append(fallback)
            if len(categories) >= 5:
                break
    app.logger.info("[%s] Inventory categories detected: %s", call_id, categories)
    return categories, None, raw


def _generate_inventory_category(category, project, description, user_payload):
    call_id = f"inventory-category-{uuid.uuid4().hex[:8]}"
    system_prompt = (
        "You are a senior Saudi construction site materials manager. Generate exactly one master inventory "
        f"category for {category}. Include 5-6 materials typical for {category} on this project. Use realistic "
        "2025 Saudi market unit rates in SAR. Return strict JSON only in this shape: "
        "{\"category\":\"Cement & Concrete\",\"category_ar\":\"...\",\"items\":[{\"name\":\"...\","
        "\"name_ar\":\"...\",\"unit\":\"bag\",\"recommended_stock\":100,\"threshold\":20,\"value_sar\":25,"
        "\"specification\":\"...\",\"brand_suggestions\":[],\"supplier_hint\":\"...\","
        "\"storage_requirements\":\"...\",\"reorder_lead_time_days\":7,\"min_order_qty\":50,"
        "\"min_order_unit\":\"bags\",\"safety_notes\":\"...\",\"alternative_items\":[],"
        "\"shelf_life_days\":90,\"shelf_life_note\":\"...\",\"notes\":null}]}. "
        "Do not include other categories in this response."
    ) + _INVENTORY_CHUNK_DETAIL_PROMPT
    category_payload = _prepend_text_to_user_payload(
        user_payload,
        f"Project: {project.name}.\nDescription: {description}\nGenerate only the {category} inventory category.",
    )
    try:
        _, raw, finish_reason = _openai_json_completion(call_id, system_prompt, category_payload, max_tokens=6000)
    except Exception as error:
        app.logger.error("Category %s failed: %s: %s", category, type(error).__name__, error)
        return None, f"{category}: {type(error).__name__}", ""
    parsed = _extract_inventory_json(raw)
    if not isinstance(parsed, dict):
        app.logger.error("[%s] Category %s JSON parse failed raw_len=%d", call_id, category, len(raw))
        return None, f"{category}: parse failed", raw
    if finish_reason == "length":
        app.logger.warning("[%s] Category %s reached token limit; attempting parsed partial.", call_id, category)
    if isinstance(parsed.get("categories"), list) and parsed["categories"]:
        parsed = parsed["categories"][0]
    normalized = _normalize_inventory_category(parsed, category)
    if not normalized:
        app.logger.error("[%s] Category %s validation failed: no usable items", call_id, category)
        return None, f"{category}: no usable items", raw
    app.logger.info("[%s] Category %s generated items=%d", call_id, category, len(normalized.get("items", [])))
    return normalized, None, raw


def _generate_chunked_master_inventory(project, description, user_payload):
    categories, reason, detect_raw = _detect_inventory_categories(project, description, user_payload)
    raw_parts = [detect_raw or ""]
    if reason:
        return None, reason, detect_raw or ""
    if not categories:
        return None, "description_too_vague", detect_raw or ""

    category_payloads = []
    failed_categories = []
    pool = ThreadPoolExecutor(max_workers=AI_GENERATION_MAX_WORKERS)
    futures = {
        pool.submit(_generate_inventory_category, category, project, description, user_payload): category
        for category in categories
    }
    try:
        for future in as_completed(futures, timeout=AI_GENERATION_TIMEOUT_SECONDS + 5):
            category = futures[future]
            try:
                result, category_reason, category_raw = future.result()
            except Exception as error:
                result, category_reason, category_raw = None, f"{category}: {type(error).__name__}", ""
                app.logger.error("Category %s failed outside handler: %s", category, error)
            raw_parts.append(category_raw or "")
            if result:
                category_payloads.append(result)
            else:
                failed_categories.append(category_reason or f"{category}: failed")
    except FuturesTimeoutError:
        for future, category in futures.items():
            if not future.done():
                future.cancel()
                failed_categories.append(f"{category}: timeout")
                app.logger.error("Category %s failed: timeout after %ss", category, AI_GENERATION_TIMEOUT_SECONDS)
    finally:
        pool.shutdown(wait=False, cancel_futures=True)

    if not category_payloads:
        return None, "all_categories_failed", "\n\n".join(raw_parts)

    category_payloads.sort(
        key=lambda c: categories.index(c.get("category")) if c.get("category") in categories else len(categories)
    )
    total_items = sum(len(category.get("items", [])) for category in category_payloads)
    total_value = round(sum(
        _coerce_float(item.get("recommended_stock")) * _coerce_float(item.get("value_sar"))
        for category in category_payloads
        for item in category.get("items", [])
        if isinstance(item, dict)
    ), 2)
    result = {
        "categories_detected": [category.get("category") for category in category_payloads],
        "categories": category_payloads,
        "total_items": total_items,
        "total_value_sar": total_value,
    }
    if failed_categories:
        result["_partial_generation_warnings"] = failed_categories
        app.logger.warning("Inventory partial generation warnings: %s", failed_categories)

    is_valid, validation_reason = _validate_inventory_response(
        result, min_categories=3, require_detail=True, min_total_items=10
    )
    app.logger.info(
        "Inventory chunked validation: is_valid=%s reason=%s categories=%d items=%d",
        is_valid, validation_reason, len(category_payloads), total_items,
    )
    if not is_valid:
        return None, validation_reason, "\n\n".join(raw_parts)
    return result, None, "\n\n".join(raw_parts)


def _prepend_text_to_user_payload(user_payload, prefix):
    if isinstance(user_payload, str):
        return f"{prefix}\n\n{user_payload}"
    if not isinstance(user_payload, list):
        return prefix
    payload = []
    inserted = False
    for part in user_payload:
        if isinstance(part, dict):
            copied = dict(part)
            if not inserted and copied.get("type") == "text":
                copied["text"] = f"{prefix}\n\n{copied.get('text', '')}"
                inserted = True
            payload.append(copied)
        else:
            payload.append(part)
    if not inserted:
        payload.insert(0, {"type": "text", "text": prefix})
    return payload


def _normalize_boq_section(section, fallback_trade, section_index=1):
    if not isinstance(section, dict):
        return None
    items = section.get("items")
    if not isinstance(items, list):
        return None
    normalized_items = []
    for item_index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            continue
        row = dict(item)
        row.setdefault("item_no", f"{section_index}.{item_index}")
        row.setdefault("description", row.get("desc") or row.get("item_description") or fallback_trade)
        row.setdefault("description_ar", "")
        row.setdefault("unit", row.get("uom") or "No.")
        qty = _coerce_float(row.get("qty", row.get("quantity", 0)))
        rate = _coerce_float(row.get("rate_sar", row.get("rate", row.get("unit_rate", 0))))
        total = _coerce_float(row.get("total_sar", row.get("total", row.get("amount", 0))))
        if total <= 0 and qty > 0 and rate > 0:
            total = round(qty * rate, 2)
        row["qty"] = qty
        row["rate_sar"] = rate
        row["total_sar"] = total
        normalized_items.append(row)
    if not normalized_items:
        return None
    return {
        "trade": sanitize(str(section.get("trade") or fallback_trade), 80),
        "trade_ar": sanitize(str(section.get("trade_ar") or ""), 80),
        "items": normalized_items,
    }


def _detect_boq_trades(project, description, user_payload):
    if client is None:
        return None, "AI service is not configured. Please set OPENAI_API_KEY.", ""
    if _description_is_too_vague(description):
        app.logger.warning("BOQ trade detection skipped: description_too_vague")
        return [], "description_too_vague", ""
    call_id = f"boq-detect-{uuid.uuid4().hex[:8]}"
    system_prompt = (
        "Read the Saudi construction project description. Return strict JSON only with the list of "
        "BOQ trades this project needs. Be comprehensive but realistic. Use common trade names like "
        "Civil, Structural, Electrical, Plumbing, HVAC, Fire Fighting, Finishing, External Works. "
        "If the description is too vague to identify a construction project, return "
        "{\"trades\":[],\"error\":\"description_too_vague\"}."
    )
    detect_payload = _prepend_text_to_user_payload(
        user_payload,
        f"Project: {project.name}.\nTask: Detect BOQ trades only. Return JSON like {{\"trades\":[\"Civil\",\"Structural\"]}}.",
    )
    try:
        _, raw, finish_reason = _openai_json_completion(call_id, system_prompt, detect_payload, max_tokens=500)
    except Exception as error:
        app.logger.error("[%s] BOQ trade detection failed: %s: %s", call_id, type(error).__name__, error)
        return None, f"openai_error: {error}", ""
    parsed = _extract_json_from_ai_response(raw)
    if not isinstance(parsed, dict):
        app.logger.error("[%s] BOQ trade detection JSON parse failed raw_len=%d", call_id, len(raw))
        return None, "parse failed", raw
    if parsed.get("error") == "description_too_vague":
        return [], "description_too_vague", raw
    trades = _clean_ai_label_list(parsed.get("trades") or parsed.get("trades_detected"))
    if finish_reason == "length":
        trades = trades[:8]
    trades = _consolidate_boq_trades(trades, description)
    if len(trades) < 2:
        for fallback in ("Civil", "Structural", "MEP", "Finishing"):
            if fallback.casefold() not in {t.casefold() for t in trades}:
                trades.append(fallback)
            if len(trades) >= 4:
                break
    app.logger.info("[%s] BOQ trades detected: %s", call_id, trades)
    return trades, None, raw


def _generate_boq_section_for_trade(trade, project, description, user_payload, section_index=1, item_count_hint="6-8", market="KSA"):
    call_id = f"boq-section-{uuid.uuid4().hex[:8]}"
    market_role = "Indian" if market == "IN" else "Saudi"
    system_prompt = (
        f"You are a senior {market_role} quantity surveyor. Generate exactly one BOQ section for the requested trade. "
        f"Trade: {trade}.\n"
        f"Include {item_count_hint} items typical for {trade} on this project. Use realistic 2025 {market_role} market rates in {'INR' if market == 'IN' else 'SAR'}. "
        "Return strict JSON only in this shape: "
        "{\"trade\":\"Civil\",\"trade_ar\":\"...\",\"items\":[{\"item_no\":\"1.1\",\"description\":\"...\","
        "\"description_ar\":\"...\",\"specification\":\"...\",\"sub_items\":[],\"application_notes\":\"...\","
        "\"standard_reference\":\"...\",\"unit\":\"m3\",\"qty\":100,\"rate_sar\":250,\"total_sar\":25000}]}. "
        "Do not include other trades in this response."
    ) + _BOQ_CHUNK_DETAIL_PROMPT
    if market == "IN":
        system_prompt += _BOQ_INDIA_EXTENSION
    section_payload = _prepend_text_to_user_payload(
        user_payload,
        f"Project: {project.name}.\nDescription: {description}\nGenerate only the {trade} BOQ section.",
    )
    try:
        _, raw, finish_reason = _openai_json_completion(call_id, system_prompt, section_payload, max_tokens=6000)
    except Exception as error:
        app.logger.error("Section %s failed: %s: %s", trade, type(error).__name__, error)
        return None, f"{trade}: {type(error).__name__}", ""
    parsed = _extract_json_from_ai_response(raw)
    if not isinstance(parsed, dict):
        app.logger.error("[%s] Section %s JSON parse failed raw_len=%d", call_id, trade, len(raw))
        return None, f"{trade}: parse failed", raw
    if finish_reason == "length":
        app.logger.warning("[%s] Section %s reached token limit; attempting parsed partial.", call_id, trade)
    if isinstance(parsed.get("sections"), list) and parsed["sections"]:
        parsed = parsed["sections"][0]
    section = _normalize_boq_section(parsed, trade, section_index=section_index)
    if not section:
        app.logger.error("[%s] Section %s validation failed: no usable items", call_id, trade)
        return None, f"{trade}: no usable items", raw
    app.logger.info("[%s] Section %s generated items=%d", call_id, trade, len(section.get("items", [])))
    return section, None, raw


def _generate_chunked_master_boq(project, description, user_payload, market="KSA"):
    trades, reason, detect_raw = _detect_boq_trades(project, description, user_payload)
    raw_parts = [detect_raw or ""]
    if reason:
        return None, reason, detect_raw or ""
    if not trades:
        return None, "description_too_vague", detect_raw or ""

    sections = []
    failed_sections = []
    complex_project = any(term in (description or "").lower() for term in (
        "tower", "kafd", "basement", "hospital", "hotel", "industrial", "airport",
        "mall", "high-rise", "high rise", "mixed-use", "mixed use"
    ))
    item_count_hint = "6-8" if complex_project else "5"
    pool = ThreadPoolExecutor(max_workers=AI_GENERATION_MAX_WORKERS)
    futures = {
        pool.submit(
            _generate_boq_section_for_trade,
            trade, project, description, user_payload, idx, item_count_hint, market
        ): trade
        for idx, trade in enumerate(trades, start=1)
    }
    try:
        for future in as_completed(futures, timeout=AI_GENERATION_TIMEOUT_SECONDS + 5):
            trade = futures[future]
            try:
                section, section_reason, section_raw = future.result()
            except Exception as error:
                section, section_reason, section_raw = None, f"{trade}: {type(error).__name__}", ""
                app.logger.error("Section %s failed outside handler: %s", trade, error)
            raw_parts.append(section_raw or "")
            if section:
                sections.append(section)
            else:
                failed_sections.append(section_reason or f"{trade}: failed")
    except FuturesTimeoutError:
        for future, trade in futures.items():
            if not future.done():
                future.cancel()
                failed_sections.append(f"{trade}: timeout")
                app.logger.error("Section %s failed: timeout after %ss", trade, AI_GENERATION_TIMEOUT_SECONDS)
    finally:
        pool.shutdown(wait=False, cancel_futures=True)

    if not sections:
        return None, "all_sections_failed", "\n\n".join(raw_parts)

    sections.sort(key=lambda s: trades.index(s.get("trade")) if s.get("trade") in trades else len(trades))
    subtotal = round(sum(
        _coerce_float(item.get("total_sar"))
        for section in sections
        for item in section.get("items", [])
        if isinstance(item, dict)
    ), 2)
    vat_amount = round(subtotal * 0.15, 2)
    grand_total = round(subtotal + vat_amount, 2)
    result = {
        "trades_detected": [section.get("trade") for section in sections],
        "sections": sections,
        "subtotal": subtotal,
        "vat_amount": vat_amount,
        "grand_total": grand_total,
    }
    if failed_sections:
        result["_partial_generation_warnings"] = failed_sections
        app.logger.warning("BOQ partial generation warnings: %s", failed_sections)

    is_valid, validation_reason = _validate_boq_response(
        result, min_sections=2, require_detail=True, min_total_items=5
    )
    app.logger.info(
        "BOQ chunked validation: is_valid=%s reason=%s sections=%d items=%d",
        is_valid, validation_reason, len(sections),
        sum(len(section.get("items", [])) for section in sections),
    )
    if not is_valid:
        return None, validation_reason, "\n\n".join(raw_parts)
    return result, None, "\n\n".join(raw_parts)


@app.template_filter('sar_format')
def _tpl_sar_format(value):
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        amount = 0.0
    if amount >= 1_000_000:
        return f"{amount / 1_000_000:.1f}M"
    if amount >= 1_000:
        return f"{amount / 1_000:.0f}K"
    return f"{amount:,.0f}"


@app.template_filter('from_json')
def _tpl_from_json(s):
    try:
        return json.loads(s) if s else []
    except Exception:
        return []


@app.template_filter('sum_section_items')
def _tpl_sum_section_items(parsed):
    """Count total items across all trade sections."""
    if isinstance(parsed, dict):
        sections = parsed.get('sections', [])
    elif isinstance(parsed, list):
        sections = parsed
    else:
        return 0
    total = 0
    for s in sections:
        if isinstance(s, dict):
            items = s.get('items', [])
            if isinstance(items, list):
                total += len(items)
    return total


@app.template_filter('flat_boq_items')
def _tpl_flat_boq_items(parsed):
    """Flatten sections → items into a single list for iteration."""
    if isinstance(parsed, dict):
        sections = parsed.get('sections', [])
    elif isinstance(parsed, list):
        sections = parsed
    else:
        return []
    result = []
    for s in sections:
        if isinstance(s, dict):
            items = s.get('items', [])
            if isinstance(items, list):
                result.extend(items)
    return result


# ── BOQ REBUILD — NEW ROUTES ─────────────────────────────────────────────────

@app.route("/boq/package/<int:package_id>")
@login_required
def boq_package_view(package_id):
    package = db.session.get(BOQPackage, package_id) or abort(404)
    if not package_access_allowed(package):
        abort(403)
    tracker = compute_boq_tracker_summary(package.boq, package.id) if package.boq else {}
    return render_template(
        "dashboard/boq/package.html",
        active_dashboard="boq",
        package=package,
        boq=package.boq,
        tracker=tracker,
    )


@app.route("/boq/manual-actual", methods=["POST"])
@login_required
def boq_manual_actual():
    data = request.get_json(silent=True) or {}
    try:
        boq_id = int(data.get("boq_id") or 0)
    except (TypeError, ValueError):
        boq_id = 0
    boq_obj = db.session.get(BOQ, boq_id) if boq_id else None
    if not boq_obj:
        return jsonify({"success": False, "error": "BOQ not found."}), 404

    allowed = boq_obj.user_id == current_user.id
    package_id = None
    raw_package_id = data.get("package_id")
    if raw_package_id not in (None, ""):
        try:
            package_id = int(raw_package_id)
        except (TypeError, ValueError):
            return jsonify({"success": False, "error": "Invalid package."}), 400
        package = db.session.get(BOQPackage, package_id)
        if not package or package.boq_id != boq_obj.id:
            return jsonify({"success": False, "error": "Package not found."}), 404
        allowed = allowed or package_access_allowed(package)
    else:
        package = BOQPackage.query.filter_by(
            boq_id=boq_obj.id,
            assigned_engineer_email=current_user.email,
        ).first()
        if package:
            package_id = package.id
            allowed = True

    if not allowed:
        abort(403)

    item_no = sanitize_input(data.get("item_no", ""), 80).strip()
    actual_qty = _boq_float(data.get("actual_qty", data.get("qty", 0)))
    if not item_no or actual_qty <= 0:
        return jsonify({"success": False, "error": "Select a BOQ item and enter a positive quantity."}), 400

    matched_item = None
    for item in _boq_prompt_items(boq_obj.items):
        if _boq_item_no(item).lower() == item_no.lower():
            matched_item = item
            break
    if not matched_item:
        return jsonify({"success": False, "error": "BOQ item not found."}), 404

    recorded_at = datetime.utcnow()
    raw_recorded_at = data.get("recorded_at")
    if raw_recorded_at:
        parsed = parse_demo_date(raw_recorded_at)
        if parsed:
            recorded_at = datetime.combine(parsed, datetime.min.time())

    try:
        actual = _create_boq_actual(
            boq_obj,
            matched_item,
            actual_qty,
            source=sanitize_input(data.get("source", "manual"), 30) or "manual",
            source_id=None,
            recorded_by=sanitize_input(data.get("engineer_name", ""), 120) or current_user.full_name,
            package_id=package_id,
            actual_rate=data.get("actual_rate"),
            recorded_at=recorded_at,
        )
        db.session.commit()
        if boq_obj.project_id:
            project = db.session.get(Project, boq_obj.project_id)
            if project:
                calculate_health_score(project)
        return jsonify(
            {
                "success": True,
                "actual": actual.to_dict(),
                "tracker": compute_boq_tracker_summary(boq_obj, package_id),
            }
        )
    except SQLAlchemyError as error:
        rollback_db_session()
        app.logger.error("Manual BOQ actual failed: %s", error)
        return jsonify({"success": False, "error": "Could not save BOQ progress."}), 500


@app.route("/dashboard/boq/create", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
@limiter.limit("10 per hour", methods=["POST"])
def boq_create():
    username, company = dashboard_identity()
    if request.method == "GET":
        prefill_description = session.pop("boq_create_prefill_description", "")
        prefill_project_id = request.args.get("project_id", type=int)
        projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()
        return render_template("dashboard/boq/create.html",
            username=username, company=company,
            active_dashboard="boq", projects=projects,
            prefill_description=prefill_description,
            prefill_project_id=prefill_project_id)

    # POST — generate master BOQ via AI
    try:
        ok, msg = can_use_ai(current_user)
        if not ok:
            flash(f"AI quota: {msg}", "warning")
            return redirect(url_for("boq_create"))

        project_id = request.form.get("project_id", type=int)
        title = sanitize(request.form.get("title", "").strip())
        description = sanitize(request.form.get("description", "").strip(), max_length=4000)
        if not project_id or not description:
            flash("Project and description are required. / المشروع والوصف مطلوبان.", "danger")
            return redirect(url_for("boq_create"))

        project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()

        # Save uploaded files and build AI context
        saved_files = []
        upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "boq_designs", str(current_user.id))
        os.makedirs(upload_dir, exist_ok=True)
        image_parts = []
        text_context = ""

        files = request.files.getlist("design_files")
        for f in files[:5]:
            if not f or not f.filename:
                continue
            filename = secure_filename(f.filename)
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ['.pdf', '.xlsx', '.xls', '.jpg', '.jpeg', '.png']:
                continue
            file_data = f.read()
            if len(file_data) > 10 * 1024 * 1024:
                flash(f"File {filename} exceeds 10MB limit / الملف {filename} يتجاوز حد 10 ميغابايت", "warning")
                continue
            unique_name = f"{uuid.uuid4().hex}_{filename}"
            save_path = os.path.join(upload_dir, unique_name)
            with open(save_path, 'wb') as out:
                out.write(file_data)
            rel_path = f"uploads/boq_designs/{current_user.id}/{unique_name}"
            saved_files.append(rel_path)

            if ext in ['.jpg', '.jpeg', '.png']:
                b64 = base64.b64encode(file_data).decode('utf-8')
                mime = 'image/jpeg' if ext in ['.jpg', '.jpeg'] else 'image/png'
                image_parts.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})
            elif ext == '.pdf':
                try:
                    import pdfplumber
                    with pdfplumber.open(BytesIO(file_data)) as pdf:
                        for page in pdf.pages[:5]:
                            text_context += page.extract_text() or ""
                except Exception:
                    pass
            elif ext in ['.xlsx', '.xls']:
                try:
                    wb_tmp = load_workbook(BytesIO(file_data), read_only=True)
                    for ws_tmp in list(wb_tmp.worksheets)[:3]:
                        for row in ws_tmp.iter_rows(max_row=50, values_only=True):
                            text_context += " ".join(str(c) for c in row if c) + "\n"
                except Exception:
                    pass

        # Build AI messages
        user_content = [{"type": "text", "text": f"Project: {project.name}. Description: {description}."}]
        if text_context:
            user_content[0]["text"] += f"\n\nFile context:\n{text_context[:3000]}"
        user_content.extend(image_parts)

        min_sections = 3 if len(description) > 100 else 1
        system_prompt = (
            "You are a senior Saudi quantity surveyor generating a complete construction Bill of Quantities.\n\n"
            "CRITICAL RULES — failure to follow these will cause project delays:\n\n"
            "1. Analyze the project description carefully and identify EVERY engineering trade required. "
            "Construction projects almost always need MULTIPLE trades, not just one.\n\n"
            "2. For a typical building project, you MUST include ALL of these trades when applicable "
            "(which is almost always):\n"
            "   - Civil (excavation, earthwork, foundations)\n"
            "   - Structural (concrete, reinforcement, columns, slabs)\n"
            "   - MEP (electrical, plumbing, HVAC, fire fighting)\n"
            "   - Finishing (flooring, paint, ceilings, doors, windows)\n"
            "   - External (landscaping, parking, boundary walls)\n\n"
            "3. For specialty projects, ADD relevant trades:\n"
            "   - Hotel/hospitality: spa fit-out, kitchen equipment, specialty lighting\n"
            "   - Industrial: cranes, specialized flooring\n"
            "   - Infrastructure: drainage, asphalt, signage\n"
            "   - Healthcare: medical gas, specialty HVAC\n\n"
            "4. Each trade section MUST have AT LEAST 3-5 line items with realistic 2025 Saudi market rates in SAR.\n\n"
            "5. Apply 15% VAT. Include bilingual EN+AR item descriptions.\n\n"
            "6. Only return {\"error\": \"description_too_vague\"} if the description is under 50 characters "
            "or contains no identifiable project type at all. "
            "For descriptions over 100 characters, ALWAYS generate a BOQ — infer missing details "
            "from the project type rather than refusing.\n"
        ) + _BOQ_ITEM_DETAIL_PROMPT + (
            "\n\nReturn strict JSON ONLY. No prose, no markdown fences.\n"
            "Required schema:\n"
            '{"trades_detected":["Civil","Structural","MEP","Finishing"],'
            '"sections":[{"trade":"Civil","trade_ar":"\u0623\u0639\u0645\u0627\u0644 \u0645\u062f\u0646\u064a\u0629",'
            '"items":[{"item_no":"1.1","description":"...","description_ar":"...",'
            '"specification":"...","sub_items":[{"name":"...","qty_per_unit":0,"unit":"..."}],'
            '"application_notes":"...","standard_reference":"...",'
            '"unit":"m3","qty":100,"rate_sar":250,"total_sar":25000}]}],'
            '"subtotal":25000,"vat_amount":3750,"grand_total":28750}'
        )

        # user_content may include image parts; for json_object mode the text part must carry the schema instruction
        user_market = getattr(current_user, "country", "SA") or "SA"
        boq_data, reason, raw = _generate_chunked_master_boq(project, description, user_content, market=user_market)

        if boq_data is None:
            app.logger.error(
                "Master BOQ AI failed after retry: reason=%s, description_len=%d, raw_preview=%s",
                reason, len(description), (raw[:500] if raw else "None")
            )
            session["boq_create_prefill_description"] = description
            # Distinguish user-input problems from technical failures
            if reason == "description_too_vague" or len(description) < 200:
                flash(
                    "AI couldn't generate a BOQ from that description. Please write at least "
                    "2-3 sentences describing the project type, size, floors, and any specific "
                    "trades needed. Example: 'Two-floor villa, 250 sqm, 4 bedrooms, with full "
                    "MEP, plumbing, and finishing works.'",
                    "warning"
                )
                flash(
                    "تعذر على الذكاء الاصطناعي إنشاء جدول الكميات من هذا الوصف. "
                    "يرجى كتابة 2-3 جمل على الأقل تصف نوع المشروع والمساحة "
                    "والطوابق والتخصصات المطلوبة.",
                    "warning"
                )
            else:
                flash(
                    "BOQ generation failed due to a technical issue — your description was good. "
                    "Please try again. If the problem persists, contact support.",
                    "warning"
                )
                flash(
                    "تعذر إنشاء جدول الكميات بسبب مشكلة تقنية. وصفك كان جيداً — حاول مرة أخرى. "
                    "إذا استمرت المشكلة، تواصل مع الدعم الفني.",
                    "warning"
                )
            return redirect(url_for("boq_create"))

        if boq_data.get("_partial_generation_warnings"):
            flash(
                "Some sections couldn't be generated - you can add them manually using Edit Items.",
                "warning"
            )
            flash(
                "تعذر إنشاء بعض الأقسام - يمكنك إضافتها يدويًا من تعديل البنود.",
                "warning"
            )

        sections = boq_data.get("sections", [])
        subtotal = float(boq_data.get("subtotal", 0))
        vat_amount = float(boq_data.get("vat_amount", 0))
        grand_total = float(boq_data.get("grand_total", 0))

        new_boq = BOQ(
            user_id=current_user.id,
            project_id=project_id,
            title=title or f"Master BOQ — {project.name}",
            items_json=json.dumps(sections),
            subtotal=subtotal,
            vat_amount=vat_amount,
            grand_total=grand_total,
            status="master",
            is_master=True,
            source="ai_master",
            generation_mode="ai_master",
            version=1,
            design_files_json=json.dumps(saved_files) if saved_files else None,
            project_description=description,
        )
        db.session.add(new_boq)
        db.session.commit()

        record_ai_usage("boq_master_generation")
        flash("Master BOQ generated successfully. / تم إنشاء جدول الكميات الرئيسي بنجاح.", "success")
        return redirect(url_for("boq_view", boq_id=new_boq.id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"BOQ create error: {e}")
        traceback.print_exc()
        flash("An error occurred. Please try again. / حدث خطأ. حاول مرة أخرى.", "danger")
        return redirect(url_for("boq_create"))


@app.route("/api/boq/transcribe", methods=["POST"])
@login_required
@limiter.limit("30 per hour")
def boq_transcribe():
    forbidden = ensure_ai_access("boq")
    if forbidden:
        return forbidden
    if client is None:
        return jsonify({"success": False, "error": "AI service is not configured. Please set OPENAI_API_KEY."}), 503
    audio_file = request.files.get("audio")
    if not audio_file:
        return jsonify({"success": False, "error": "No audio file"}), 400
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as tmp:
            audio_file.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            result = client.audio.transcriptions.create(
                model="whisper-1",
                file=f,
                response_format="verbose_json",
            )
        text = result.text or ""
        record_ai_usage("boq_voice")
        return jsonify({"success": True, "text": text, "language_detected": getattr(result, "language", "auto")})
    except Exception as e:
        app.logger.error(f"Whisper transcription error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@app.route("/dashboard/boq/<int:boq_id>")
@login_required
def boq_view(boq_id):
    boq = db.session.get(BOQ, boq_id) or abort(404)
    is_owner = (boq.user_id == current_user.id)
    is_assigned = (boq.assigned_to_user_id == current_user.id)
    if not is_owner and not is_assigned:
        # Also allow PM to view distributed children of their master BOQs
        if current_user.role == ROLE_PROJECT_MANAGER and boq.parent_master_boq_id:
            parent = db.session.get(BOQ, boq.parent_master_boq_id)
            if parent and parent.user_id == current_user.id:
                is_owner = True
        if not is_owner and not is_assigned:
            abort(403)

    try:
        sections = json.loads(boq.items_json or '[]')
    except Exception:
        sections = []

    # Version history for SE's assigned BOQ chain
    version_history = []
    if boq.parent_master_boq_id and boq.assigned_to_user_id:
        version_history = BOQ.query.filter_by(
            parent_master_boq_id=boq.parent_master_boq_id,
            assigned_to_user_id=boq.assigned_to_user_id,
            trade_section=boq.trade_section
        ).order_by(BOQ.version).all()

    design_files = []
    if boq.design_files_json:
        try:
            design_files = json.loads(boq.design_files_json)
        except Exception:
            pass

    # Distributed children for PM master-BOQ view — keyed on parent_master_boq_id
    distributed_children = []
    if boq.is_master and is_owner:
        distributed_children = BOQ.query.filter_by(
            parent_master_boq_id=boq.id
        ).order_by(BOQ.trade_section, BOQ.version.desc()).all()

    # PM viewing an engineer_manual BOQ: read-only
    pm_readonly = (
        current_user.role == ROLE_PROJECT_MANAGER and
        boq.source == "engineer_manual"
    )
    username, company = dashboard_identity()
    return render_template("dashboard/boq/view.html",
        boq=boq, sections=sections, version_history=version_history,
        design_files=design_files, is_owner=is_owner, is_assigned=is_assigned,
        distributed_children=distributed_children, pm_readonly=pm_readonly,
        username=username, company=company, active_dashboard="boq")


@app.route("/dashboard/boq/<int:boq_id>/delete", methods=["POST"])
@login_required
def boq_delete(boq_id):
    boq = db.session.get(BOQ, boq_id) or abort(404)
    if current_user.role == ROLE_PROJECT_MANAGER:
        # PM can delete master BOQs they own (and child BOQs they own as PM)
        if boq.user_id != current_user.id:
            abort(403)
        # Cascade: delete all child BOQs distributed from this master
        if boq.is_master:
            children = BOQ.query.filter_by(parent_master_boq_id=boq.id).all()
            for child in children:
                db.session.delete(child)
            app.logger.info(
                "boq_delete: PM user_id=%d deleted master boq_id=%d with %d children",
                current_user.id, boq.id, len(children),
            )
        else:
            app.logger.info(
                "boq_delete: PM user_id=%d deleted child boq_id=%d", current_user.id, boq.id
            )
    elif current_user.role == ROLE_SITE_ENGINEER:
        # SE can delete: (a) their own analysis/manual BOQs, (b) assigned child BOQs (revised/distributed)
        is_own_se_boq = (
            boq.user_id == current_user.id and
            boq.source in ("engineer_analysis", "engineer_manual")
        )
        is_assigned_child = (
            boq.assigned_to_user_id == current_user.id
            and boq.parent_master_boq_id is not None
            and boq.status in ("revised", "distributed")
        )
        is_own_analysis = is_own_se_boq  # compat alias
        if not is_own_se_boq and not is_assigned_child:
            abort(403)
        # Revision restore: if deleting a revised BOQ, unarchive the previous version
        if boq.status == "revised" and boq.parent_revision_id:
            prev = db.session.get(BOQ, boq.parent_revision_id)
            if prev and prev.status == "archived":
                prev.status = "distributed"
                app.logger.info(
                    "boq_delete: restored boq_id=%d to 'distributed' after SE deleted revision boq_id=%d",
                    prev.id, boq.id,
                )
        app.logger.info(
            "boq_delete: SE user_id=%d deleted %s boq_id=%d",
            current_user.id, boq.status, boq.id,
        )
    else:
        abort(403)
    db.session.delete(boq)
    db.session.commit()
    flash("BOQ deleted. / تم حذف جدول الكميات.", "success")
    next_url = request.form.get("next", "").strip()
    # Validate: must be a safe internal relative path (no open-redirect)
    if next_url and next_url.startswith("/") and not next_url.startswith("//"):
        return redirect(next_url)
    return redirect(url_for("dashboard_boq_index"))


@app.route("/dashboard/boq/<int:boq_id>/download/excel")
@login_required
def boq_download_excel(boq_id):
    boq = db.session.get(BOQ, boq_id) or abort(404)
    # Access check
    is_owner = (boq.user_id == current_user.id)
    is_assigned = (boq.assigned_to_user_id == current_user.id)
    if not is_owner and not is_assigned:
        if current_user.role == ROLE_PROJECT_MANAGER and boq.parent_master_boq_id:
            parent = db.session.get(BOQ, boq.parent_master_boq_id)
            if parent and parent.user_id == current_user.id:
                is_owner = True
        if not is_owner and not is_assigned:
            abort(403)

    try:
        sections = json.loads(boq.items_json or '[]')
    except Exception:
        sections = []

    from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover"
    ws['A1'] = "BanaaIQ — Bill of Quantities"
    ws['A1'].font = XFont(bold=True, size=16)
    ws['A2'] = boq.title
    ws['A3'] = f"Project: {boq.dna_project.name if boq.dna_project else (boq.project or '—')}"
    ws['A4'] = f"Version: v{boq.version or 1}"
    ws['A5'] = f"Grand Total SAR: {float(boq.grand_total or 0):,.2f}"
    ws['A6'] = f"Generated: {boq.created_at.strftime('%Y-%m-%d') if boq.created_at else '—'}"
    ws.column_dimensions['A'].width = 40

    NAVY = "0A1628"

    for sec in sections:
        if not isinstance(sec, dict):
            continue
        trade = (sec.get('trade', 'General') or 'General')[:31]
        ws2 = wb.create_sheet(title=trade)
        headers = [
            "#", "Description", "Arabic Description", "Unit", "Qty", "Rate SAR", "Total SAR",
            "Specification", "Sub-Items", "Application Notes", "Standard Reference",
        ]
        ws2.append(headers)
        for cell in ws2[1]:
            cell.font = XFont(bold=True, color="FFFFFF")
            cell.fill = XFill("solid", fgColor=NAVY)
        ws2.column_dimensions['B'].width = 40
        ws2.column_dimensions['H'].width = 45
        ws2.column_dimensions['I'].width = 40
        ws2.column_dimensions['J'].width = 40
        ws2.column_dimensions['K'].width = 25
        for item in sec.get('items', []):
            if not isinstance(item, dict):
                continue
            sub_items = item.get('sub_items', [])
            sub_items_str = ''
            if isinstance(sub_items, list):
                sub_items_str = '; '.join(
                    f"{s.get('name', '')} ({s.get('qty_per_unit', '')} {s.get('unit', '')})"
                    for s in sub_items if isinstance(s, dict) and s.get('name')
                )
            ws2.append([
                safe_excel_cell(str(item.get('item_no', ''))),
                safe_excel_cell(str(item.get('description', ''))),
                safe_excel_cell(str(item.get('description_ar', ''))),
                safe_excel_cell(str(item.get('unit', ''))),
                item.get('qty', 0),
                item.get('rate_sar', 0),
                item.get('total_sar', 0),
                safe_excel_cell(str(item.get('specification', ''))),
                safe_excel_cell(sub_items_str),
                safe_excel_cell(str(item.get('application_notes', ''))),
                safe_excel_cell(str(item.get('standard_reference', ''))),
            ])

    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["", "Amount SAR"])
    ws_sum.append(["Subtotal", float(boq.subtotal or 0)])
    ws_sum.append(["VAT (15%)", float(boq.vat_amount or 0)])
    ws_sum.append(["Grand Total", float(boq.grand_total or 0)])
    for row in ws_sum.iter_rows(min_row=4, max_row=4):
        for cell in row:
            cell.font = XFont(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"BanaaIQ_BOQ_{boq.id}_v{boq.version or 1}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/dashboard/boq/<int:boq_id>/download/pdf")
@login_required
def boq_download_pdf(boq_id):
    boq = db.session.get(BOQ, boq_id) or abort(404)
    # Access check
    is_owner = (boq.user_id == current_user.id)
    is_assigned = (boq.assigned_to_user_id == current_user.id)
    if not is_owner and not is_assigned:
        if current_user.role == ROLE_PROJECT_MANAGER and boq.parent_master_boq_id:
            parent = db.session.get(BOQ, boq.parent_master_boq_id)
            if parent and parent.user_id == current_user.id:
                is_owner = True
        if not is_owner and not is_assigned:
            abort(403)

    try:
        sections = json.loads(boq.items_json or '[]')
    except Exception:
        sections = []

    arabic_available = setup_arabic_font()
    hindi_available = setup_hindi_font()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    normal_style = ParagraphStyle("BoqNormal", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=12)
    arabic_style = ParagraphStyle(
        "BoqArabic", parent=styles["BodyText"],
        fontName="Cairo" if arabic_available else "Helvetica",
        fontSize=10, leading=14, alignment=TA_RIGHT,
    )
    hindi_style = ParagraphStyle(
        "BoqHindi", parent=styles["BodyText"],
        fontName="NotoDevanagari" if hindi_available else "Helvetica",
        fontSize=8, leading=12,
    )

    def _esc(text):
        text = str(text or "")
        return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    def _para(text):
        val = str(text or "")
        if is_arabic_text(val) and arabic_available:
            return Paragraph(_esc(process_arabic_text(val)), arabic_style)
        return Paragraph(_esc(val), normal_style)

    story = [
        Paragraph("<b>BanaaIQ</b> | BILL OF QUANTITIES", styles["Title"]),
        Spacer(1, 6),
        Paragraph(f"<b>Title:</b> {_esc(boq.title)}", styles["BodyText"]),
        Paragraph(f"<b>Project:</b> {_esc(boq.dna_project.name if boq.dna_project else (boq.project or '—'))}", styles["BodyText"]),
        Paragraph(f"<b>Version:</b> v{boq.version or 1} | <b>Status:</b> {_esc(boq.status or 'draft')}", styles["BodyText"]),
        Spacer(1, 10),
    ]

    for sec in sections:
        if not isinstance(sec, dict):
            continue
        trade = sec.get('trade', 'General')
        trade_ar = sec.get('trade_ar', '')
        if trade_ar and arabic_available:
            section_header = Paragraph(
                f'<b>{_esc(trade)}</b> / <font face="Cairo">{_esc(process_arabic_text(trade_ar))}</font>',
                normal_style,
            )
        else:
            section_header = Paragraph(f"<b>{_esc(trade)}</b>", styles["Heading2"])
        story.append(section_header)

        if arabic_available:
            desc_hdr = Paragraph(
                'Description / <font face="Cairo">\u0627\u0644\u0648\u0635\u0641</font>',
                ParagraphStyle("BoqHdrCell", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white),
            )
        else:
            desc_hdr = "Description / \u0627\u0644\u0648\u0635\u0641"
        table_rows = [["#", desc_hdr, "Unit", "Qty", "Rate SAR", "Total SAR"]]
        for item in sec.get('items', []):
            if not isinstance(item, dict):
                continue
            desc_en = item.get('description', '')
            desc_ar = item.get('description_ar', '')
            desc_hi = item.get('description_hi', '')
            spec = str(item.get('specification') or '')
            notes = str(item.get('application_notes') or '')
            std_ref = str(item.get('standard_reference') or '')
            sub_items = item.get('sub_items', [])

            # Build description cell with optional detail lines
            desc_parts = [_esc(desc_en)]
            if desc_ar and arabic_available:
                desc_parts.append(
                    f'<font face="Cairo" size="8">{_esc(process_arabic_text(desc_ar))}</font>'
                )
            elif desc_ar:
                desc_parts.append(_esc(desc_ar))
            if desc_hi and hindi_available:
                desc_parts.append(
                    f'<font face="NotoDevanagari" size="8">{_esc(desc_hi)}</font>'
                )
            elif desc_hi:
                desc_parts.append(_esc(desc_hi))
            if spec:
                desc_parts.append(
                    f'<i><font size="7" color="#444444">{_esc(spec)}</font></i>'
                )
            if isinstance(sub_items, list) and sub_items:
                sub_line = ', '.join(
                    f"{s.get('name', '')} ({s.get('qty_per_unit', '')} {s.get('unit', '')})"
                    for s in sub_items[:4] if isinstance(s, dict) and s.get('name')
                )
                if sub_line:
                    desc_parts.append(
                        f'<font size="6" color="#666666">\u2022 {_esc(sub_line)}</font>'
                    )
            if std_ref:
                desc_parts.append(
                    f'<font size="6" color="#888888">Ref: {_esc(std_ref)}</font>'
                )
            if notes:
                desc_parts.append(
                    f'<font size="6" color="#555555">Note: {_esc(notes)}</font>'
                )
            desc_cell = Paragraph('<br/>'.join(desc_parts), normal_style)

            table_rows.append([
                str(item.get('item_no', '')),
                desc_cell,
                str(item.get('unit', '')),
                f"{float(item.get('qty', 0)):,.2f}",
                f"{float(item.get('rate_sar', 0)):,.2f}",
                f"{float(item.get('total_sar', 0)):,.2f}",
            ])

        if len(table_rows) > 1:
            t = Table(table_rows, colWidths=[28, 200, 42, 52, 72, 72])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0a0a0a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e4ddd0")),
                ("PADDING", (0, 0), (-1, -1), 4),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            story.append(t)
        story.append(Spacer(1, 8))

    story.extend([
        HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e8c547")),
        Spacer(1, 6),
        Paragraph(f"<b>Subtotal:</b> SAR {float(boq.subtotal or 0):,.2f}", styles["BodyText"]),
        Paragraph(f"<b>VAT (15%):</b> SAR {float(boq.vat_amount or 0):,.2f}", styles["BodyText"]),
        Paragraph(f"<b>Grand Total:</b> SAR {float(boq.grand_total or 0):,.2f}", styles["Heading3"]),
    ])
    doc.build(story)
    buffer.seek(0)
    fname = f"BanaaIQ_BOQ_{boq.id}_v{boq.version or 1}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=fname)


@app.route("/dashboard/boq/<int:boq_id>/distribute", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def boq_distribute(boq_id):
    boq = db.session.get(BOQ, boq_id) or abort(404)
    if boq.user_id != current_user.id or not boq.is_master:
        abort(403)
    try:
        sections = json.loads(boq.items_json or '[]')
    except Exception:
        sections = []
    # Get project's assigned site engineers
    assignments = ProjectAssignment.query.filter_by(project_id=boq.project_id).all()
    engineers = []
    for a in assignments:
        u = db.session.get(User, a.user_id)
        if u and u.role == ROLE_SITE_ENGINEER:
            engineers.append(u)

    if request.method == "GET":
        username, company = dashboard_identity()
        return render_template("dashboard/boq/distribute.html",
            boq=boq, sections=sections, engineers=engineers,
            username=username, company=company, active_dashboard="boq")

    # POST — save distribution
    try:
        any_distributed = False
        for sec in sections:
            if not isinstance(sec, dict):
                continue
            trade = sec.get('trade', 'General')
            field_name = f"engineer_{trade.replace(' ', '_').lower()}"
            engineer_id = request.form.get(field_name, type=int)
            if not engineer_id:
                continue
            engineer = db.session.get(User, engineer_id)
            if not engineer or engineer.role != ROLE_SITE_ENGINEER:
                continue

            # Calculate section totals
            items = sec.get('items', [])
            sec_subtotal = sum(float(i.get('total_sar', 0)) for i in items if isinstance(i, dict))
            sec_vat = sec_subtotal * 0.15
            sec_grand = sec_subtotal + sec_vat

            # Create distributed child BOQ
            child_boq = BOQ(
                user_id=current_user.id,
                project_id=boq.project_id,
                title=f"{boq.title} — {trade}",
                items_json=json.dumps([sec]),  # single section
                subtotal=sec_subtotal,
                vat_amount=sec_vat,
                grand_total=sec_grand,
                status="distributed",
                is_master=False,
                source="distributed",
                generation_mode="distributed",
                parent_master_boq_id=boq.id,
                assigned_to_user_id=engineer.id,
                trade_section=trade,
                version=1,
                project_description=boq.project_description,
            )
            db.session.add(child_boq)
            db.session.flush()  # populate child_boq.id before using in notification link

            # Notify engineer
            notif = Notification(
                user_id=engineer.id,
                message=f"A BOQ section ({trade}) has been distributed to you for project: {boq.dna_project.name if boq.dna_project else 'N/A'}",
                link=f"/dashboard/boq/{child_boq.id}",
            )
            db.session.add(notif)
            any_distributed = True

        if any_distributed:
            boq.status = "distributed"

        db.session.commit()
        flash("BOQ distributed to engineers successfully. / تم توزيع جدول الكميات على المهندسين بنجاح.", "success")
        return redirect(url_for("boq_view", boq_id=boq.id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"BOQ distribute error: {e}")
        flash("Distribution failed. Please try again. / فشل التوزيع. حاول مجددًا.", "danger")
        return redirect(url_for("boq_distribute", boq_id=boq.id))


@app.route("/dashboard/boq/<int:boq_id>/revise", methods=["GET", "POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
def boq_revise(boq_id):
    boq = db.session.get(BOQ, boq_id) or abort(404)
    if boq.assigned_to_user_id != current_user.id or boq.is_master:
        abort(403)
    # SE's own upload-analyze BOQs cannot be "revised" — they have no PM to notify
    if boq.source == "engineer_analysis":
        abort(403)

    try:
        sections = json.loads(boq.items_json or '[]')
    except Exception:
        sections = []

    username, company = dashboard_identity()

    if request.method == "GET":
        return render_template("dashboard/boq/revise.html",
            boq=boq, sections=sections,
            username=username, company=company, active_dashboard="boq")

    # POST — create new revision using AI to apply the description
    try:
        revision_notes = sanitize(request.form.get("revision_notes", "").strip(), max_length=2000)
        if not revision_notes:
            flash("Please describe what changed in this revision. / يرجى وصف ما تغير.", "warning")
            return redirect(url_for("boq_revise", boq_id=boq_id))

        ok, quota_msg = can_use_ai(current_user)
        if not ok:
            flash(f"AI quota: {quota_msg}", "warning")
            return redirect(url_for("boq_revise", boq_id=boq_id))

        # Call AI to apply the revision description to the existing items
        revision_system = (
            "You are revising an existing BOQ section based on a site engineer's feedback.\n"
            "You will receive the current BOQ items (JSON) and the engineer's revision request.\n\n"
            "Apply the requested changes precisely:\n"
            "- If asked to add items: add them with realistic 2025 Saudi market rates in SAR\n"
            "- If asked to change quantities: update only those items\n"
            "- If asked to remove items: remove them\n"
            "- Keep all unrelated items unchanged\n"
            "- Recompute section subtotals, apply 15% VAT, update grand_total\n"
            "- Include bilingual EN+AR descriptions for all items"
            + _BOQ_ITEM_DETAIL_PROMPT
            + "\nReturn ONLY strict JSON, no markdown fences.\n"
            "Required schema:\n"
            '{"trades_detected":["Civil"],"sections":[{"trade":"Civil","trade_ar":"\u0623\u0639\u0645\u0627\u0644 \u0645\u062f\u0646\u064a\u0629",'
            '"items":[{"item_no":"1.1","description":"...","description_ar":"...",'
            '"specification":"...","sub_items":[{"name":"...","qty_per_unit":0,"unit":"..."}],'
            '"application_notes":"...","standard_reference":"...","unit":"m3","qty":100,'
            '"rate_sar":250,"total_sar":25000}]}],'
            '"subtotal":25000,"vat_amount":3750,"grand_total":28750}'
        )
        revision_user = (
            "EXISTING BOQ ITEMS:\n"
            + json.dumps(sections, ensure_ascii=False, indent=2)[:4000]
            + "\n\nENGINEER REVISION REQUEST:\n"
            + revision_notes
        )

        revised_data, revision_err, _ = _call_ai_for_master_boq(
            revision_system, revision_user, min_sections=1, require_detail=True
        )

        if revised_data is None:
            app.logger.error("BOQ revision AI failed: boq_id=%d reason=%s", boq_id, revision_err)
            flash(
                "Could not apply revision. Please describe changes more specifically "
                "(which items to add/modify/remove, with quantities). / "
                "\u062a\u0639\u0630\u0651\u0631 \u062a\u0637\u0628\u064a\u0642 \u0627\u0644\u062a\u0639\u062f\u064a\u0644. \u064a\u0631\u062c\u0649 \u0648\u0635\u0641 \u0627\u0644\u062a\u063a\u064a\u064a\u0631\u0627\u062a \u0628\u062f\u0642\u0629 (\u0623\u064a \u0627\u0644\u0628\u0646\u0648\u062f\u060c \u0648\u0627\u0644\u0643\u0645\u064a\u0627\u062a).",
                "danger"
            )
            return redirect(url_for("boq_revise", boq_id=boq_id))

        new_sections = revised_data.get("sections", [])

        # Detect whether AI actually changed the items
        def _content_hash(secs):
            simplified = []
            for sec in (secs or []):
                if not isinstance(sec, dict):
                    continue
                simplified.append({
                    "trade": sec.get("trade", ""),
                    "items": [
                        {
                            "item_no": it.get("item_no", ""),
                            "description": str(it.get("description", "")).strip().lower(),
                            "unit": it.get("unit", ""),
                            "qty": round(float(it.get("qty", 0) or 0), 3),
                            "rate_sar": round(float(it.get("rate_sar", 0) or 0), 2),
                        }
                        for it in sec.get("items", []) if isinstance(it, dict)
                    ],
                })
            return hashlib.md5(
                json.dumps(simplified, sort_keys=True, ensure_ascii=False).encode()
            ).hexdigest()

        if _content_hash(new_sections) == _content_hash(sections):
            nudge_user = (
                revision_user
                + "\n\nWARNING: Your previous response returned IDENTICAL items — no changes were applied. "
                "You MUST apply the engineer's revision request. The output MUST differ from the input."
            )
            revised_data2, _, _ = _call_ai_for_master_boq(
                revision_system, nudge_user, min_sections=1, require_detail=True
            )
            if revised_data2 and _content_hash(revised_data2.get("sections", [])) != _content_hash(sections):
                revised_data = revised_data2
                new_sections = revised_data.get("sections", [])
            else:
                app.logger.warning(
                    "BOQ revision unchanged after retry: boq_id=%d desc=%r",
                    boq_id, revision_notes[:100]
                )
                flash(
                    "Could not apply revision — AI returned unchanged items. "
                    "Please describe changes more specifically "
                    "(which items to add/modify/remove, with quantities). / "
                    "\u062a\u0639\u0630\u0651\u0631 \u0627\u0644\u062a\u0639\u062f\u064a\u0644. \u064a\u0631\u062c\u0649 \u0648\u0635\u0641 \u0627\u0644\u062a\u063a\u064a\u064a\u0631\u0627\u062a \u0628\u062f\u0642\u0629 \u0623\u0643\u0628\u0631.",
                    "danger"
                )
                return redirect(url_for("boq_revise", boq_id=boq_id))

        orig_count = sum(len(s.get("items", [])) for s in sections if isinstance(s, dict))
        new_count = sum(len(s.get("items", [])) for s in new_sections if isinstance(s, dict))
        app.logger.info(
            "BOQ revision applied: boq_id=%d orig_items=%d new_items=%d desc_len=%d",
            boq_id, orig_count, new_count, len(revision_notes)
        )

        new_version = (boq.version or 1) + 1
        revised_boq = BOQ(
            user_id=boq.user_id,
            project_id=boq.project_id,
            title=boq.title,
            items_json=json.dumps(new_sections),
            subtotal=float(revised_data.get("subtotal", 0)),
            vat_amount=float(revised_data.get("vat_amount", 0)),
            grand_total=float(revised_data.get("grand_total", 0)),
            status="revised",
            is_master=False,
            source="revised",
            generation_mode="revised",
            parent_master_boq_id=boq.parent_master_boq_id,
            assigned_to_user_id=current_user.id,
            trade_section=boq.trade_section,
            version=new_version,
            parent_revision_id=boq.id,
            project_description=revision_notes,
        )
        db.session.add(revised_boq)
        db.session.flush()

        boq.status = "archived"

        pm = db.session.get(User, boq.user_id)
        if pm:
            notif = Notification(
                user_id=pm.id,
                message=f"{current_user.full_name} submitted a revision (v{new_version}) for BOQ section: {boq.trade_section or boq.title}",
                link=f"/dashboard/boq/{revised_boq.id}",
            )
            db.session.add(notif)

        db.session.commit()
        record_ai_usage("boq_revision")
        flash(
            f"Revision v{new_version} submitted with AI-applied changes. / "
            f"\u062a\u0645 \u062a\u0642\u062f\u064a\u0645 \u0627\u0644\u0645\u0631\u0627\u062c\u0639\u0629 v{new_version}.",
            "success"
        )
        return redirect(url_for("boq_view", boq_id=revised_boq.id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"BOQ revise error: {e}")
        flash("Revision failed. Please try again. / \u0641\u0634\u0644\u062a \u0627\u0644\u0645\u0631\u0627\u062c\u0639\u0629.", "danger")
        return redirect(url_for("boq_revise", boq_id=boq_id))


# ── BOQ EDIT ITEMS ──────────────────────────────────────────────────────────

@app.route("/dashboard/boq/<int:boq_id>/edit-items", methods=["GET", "POST"])
@login_required
def boq_edit_items(boq_id):
    boq = db.session.get(BOQ, boq_id) or abort(404)

    # Permission check
    can_edit = False
    is_se_editing = False
    if current_user.role == ROLE_PROJECT_MANAGER:
        # PM cannot edit engineer_manual BOQs
        if boq.source == "engineer_manual":
            abort(403)
        if boq.user_id == current_user.id:
            can_edit = True
        elif boq.parent_master_boq_id:
            parent = db.session.get(BOQ, boq.parent_master_boq_id)
            if parent and parent.user_id == current_user.id:
                can_edit = True
    elif current_user.role == ROLE_SITE_ENGINEER:
        if boq.assigned_to_user_id == current_user.id and boq.status != "archived":
            can_edit = True
            is_se_editing = True
        # SE can also edit their own manual BOQs
        elif boq.user_id == current_user.id and boq.source == "engineer_manual":
            can_edit = True
            is_se_editing = True
    if not can_edit:
        abort(403)

    try:
        sections = json.loads(boq.items_json or "[]")
    except Exception:
        sections = []

    username, company = dashboard_identity()

    if request.method == "GET":
        return render_template("dashboard/boq/edit_items.html",
            boq=boq, sections=sections, is_se_editing=is_se_editing,
            username=username, company=company, active_dashboard="boq")

    # POST
    try:
        sections_json_str = request.form.get("sections_json", "")
        if not sections_json_str:
            flash("No data received. / لم يتم استقبال بيانات.", "danger")
            return redirect(url_for("boq_edit_items", boq_id=boq_id))

        new_sections = json.loads(sections_json_str)
        if not isinstance(new_sections, list) or not new_sections:
            flash("At least one section with one item is required.", "danger")
            return redirect(url_for("boq_edit_items", boq_id=boq_id))

        subtotal = 0.0
        for s_idx, sec in enumerate(new_sections):
            if not isinstance(sec, dict):
                continue
            sec["trade"] = sanitize(str(sec.get("trade", "")), 50)
            sec["trade_ar"] = sanitize(str(sec.get("trade_ar", "")), 80)
            items = sec.get("items", [])
            if not items:
                flash(f"Section '{sec.get('trade', 'Unknown')}' must have at least one item.", "warning")
                return redirect(url_for("boq_edit_items", boq_id=boq_id))
            for i_idx, item in enumerate(items, 1):
                if not isinstance(item, dict):
                    continue
                item["description"] = sanitize(str(item.get("description", "")), 300)
                item["description_ar"] = sanitize(str(item.get("description_ar", "")), 300)
                item["unit"] = sanitize(str(item.get("unit", "")), 30)
                item["item_no"] = item.get("item_no") or f"{s_idx + 1}.{i_idx}"
                item["specification"] = sanitize(str(item.get("specification") or ""), 500)
                item["application_notes"] = sanitize(str(item.get("application_notes") or ""), 500)
                item["standard_reference"] = sanitize(str(item.get("standard_reference") or ""), 200)
                # Preserve sub_items list (cap at 20); MEP-specific fields passed through as-is
                sub_items = item.get("sub_items")
                if isinstance(sub_items, list):
                    item["sub_items"] = sub_items[:20]
                elif "sub_items" in item:
                    item["sub_items"] = []
                for mep_field in ("equipment_type", "suggested_model", "accessories"):
                    if item.get(mep_field) is not None:
                        if mep_field == "accessories" and isinstance(item[mep_field], list):
                            item[mep_field] = item[mep_field][:20]
                        elif mep_field != "accessories":
                            item[mep_field] = sanitize(str(item[mep_field] or ""), 200)
                try:
                    qty = float(item.get("qty", 0) or 0)
                    rate = float(item.get("rate_sar", 0) or 0)
                    total = round(qty * rate, 2)
                    item["qty"] = qty
                    item["rate_sar"] = rate
                    item["total_sar"] = total
                    subtotal += total
                except (ValueError, TypeError):
                    flash("Qty and Rate must be numeric. / يجب أن تكون الكمية والسعر أرقاماً.", "danger")
                    return redirect(url_for("boq_edit_items", boq_id=boq_id))

        vat_amount = round(subtotal * 0.15, 2)
        grand_total = round(subtotal + vat_amount, 2)

        if is_se_editing:
            new_version = (boq.version or 1) + 1
            revised_boq = BOQ(
                user_id=boq.user_id,
                project_id=boq.project_id,
                title=boq.title,
                items_json=json.dumps(new_sections),
                subtotal=subtotal,
                vat_amount=vat_amount,
                grand_total=grand_total,
                status="revised",
                is_master=False,
                source="revised",
                generation_mode="revised",
                parent_master_boq_id=boq.parent_master_boq_id,
                assigned_to_user_id=current_user.id,
                trade_section=boq.trade_section,
                version=new_version,
                parent_revision_id=boq.id,
            )
            db.session.add(revised_boq)
            boq.status = "archived"
            db.session.commit()
            flash(f"Saved as revision v{new_version}. / تم الحفظ كمراجعة v{new_version}.", "success")
            return redirect(url_for("boq_view", boq_id=revised_boq.id))
        else:
            boq.items_json = json.dumps(new_sections)
            boq.subtotal = subtotal
            boq.vat_amount = vat_amount
            boq.grand_total = grand_total
            boq.updated_at = datetime.utcnow()
            db.session.commit()
            flash("BOQ items updated. / تم تحديث البنود.", "success")
            return redirect(url_for("boq_view", boq_id=boq_id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"BOQ edit items error: {e}")
        flash("Save failed. Please try again. / فشل الحفظ.", "danger")
        return redirect(url_for("boq_edit_items", boq_id=boq_id))


# ── BOQ MANUAL CREATE ────────────────────────────────────────────────────────

@app.route("/dashboard/boq/create-manual", methods=["GET", "POST"])
@login_required
def boq_create_manual():
    if current_user.role not in (ROLE_PROJECT_MANAGER, ROLE_SITE_ENGINEER):
        abort(403)
    username, company = dashboard_identity()

    # Project list differs by role
    if current_user.role == ROLE_PROJECT_MANAGER:
        projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()
    else:
        _assigned_ids = {a.project_id for a in ProjectAssignment.query.filter_by(user_id=current_user.id).all()}
        _pkg_ids = {p.project_id for p in EngineerPackage.query.filter_by(assigned_user_id=current_user.id).all()}
        _se_proj_ids = _assigned_ids | _pkg_ids
        projects = Project.query.filter(Project.id.in_(_se_proj_ids)).order_by(Project.name).all()

    if request.method == "GET":
        return render_template("dashboard/boq/create_manual.html",
            username=username, company=company,
            active_dashboard="boq", projects=projects,
            is_se=(current_user.role == ROLE_SITE_ENGINEER))

    try:
        project_id = request.form.get("project_id", type=int)
        title = sanitize(request.form.get("title", "").strip(), 200)
        sections_json_str = request.form.get("sections_json", "")

        if not project_id or not title:
            flash("Project and title are required. / المشروع والعنوان مطلوبان.", "danger")
            return redirect(url_for("boq_create_manual"))

        # Access check per role
        if current_user.role == ROLE_PROJECT_MANAGER:
            project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()
        else:
            # SE: must be assigned to the project
            _has_se_access = (
                ProjectAssignment.query.filter_by(project_id=project_id, user_id=current_user.id).first() or
                EngineerPackage.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).first()
            )
            if not _has_se_access:
                abort(403)
            project = Project.query.get_or_404(project_id)

        if not sections_json_str:
            flash("Please add at least one section with one item. / يرجى إضافة قسم واحد على الأقل.", "danger")
            return redirect(url_for("boq_create_manual"))

        new_sections = json.loads(sections_json_str)
        if not new_sections:
            flash("Please add at least one section with one item.", "danger")
            return redirect(url_for("boq_create_manual"))

        subtotal = 0.0
        for s_idx, sec in enumerate(new_sections):
            if not isinstance(sec, dict):
                continue
            sec["trade"] = sanitize(str(sec.get("trade", "")), 50)
            sec["trade_ar"] = sanitize(str(sec.get("trade_ar", "")), 80)
            items = sec.get("items", [])
            if not items:
                flash(f"Section '{sec.get('trade', '')}' must have at least one item.", "warning")
                return redirect(url_for("boq_create_manual"))
            for i_idx, item in enumerate(items, 1):
                if not isinstance(item, dict):
                    continue
                item["description"] = sanitize(str(item.get("description", "")), 300)
                item["description_ar"] = sanitize(str(item.get("description_ar", "")), 300)
                item["unit"] = sanitize(str(item.get("unit", "")), 30)
                item["item_no"] = item.get("item_no") or f"{s_idx + 1}.{i_idx}"
                item["specification"] = sanitize(str(item.get("specification") or ""), 500)
                item["application_notes"] = sanitize(str(item.get("application_notes") or ""), 500)
                item["standard_reference"] = sanitize(str(item.get("standard_reference") or ""), 200)
                sub_items = item.get("sub_items")
                if isinstance(sub_items, list):
                    item["sub_items"] = sub_items[:20]
                elif "sub_items" in item:
                    item["sub_items"] = []
                try:
                    qty = float(item.get("qty", 0) or 0)
                    rate = float(item.get("rate_sar", 0) or 0)
                    total = round(qty * rate, 2)
                    item["qty"] = qty
                    item["rate_sar"] = rate
                    item["total_sar"] = total
                    subtotal += total
                except (ValueError, TypeError):
                    pass

        vat_amount = round(subtotal * 0.15, 2)
        grand_total = round(subtotal + vat_amount, 2)

        _is_se = (current_user.role == ROLE_SITE_ENGINEER)
        new_boq = BOQ(
            user_id=current_user.id,
            project_id=project_id,
            title=title,
            items_json=json.dumps(new_sections),
            subtotal=subtotal,
            vat_amount=vat_amount,
            grand_total=grand_total,
            # SE creates their own working BOQ — not a PM master
            status="engineer_manual" if _is_se else "master",
            is_master=not _is_se,
            assigned_to_user_id=current_user.id if _is_se else None,
            source="engineer_manual" if _is_se else "manual",
            generation_mode="manual",
            version=1,
        )
        db.session.add(new_boq)
        db.session.commit()
        flash("Manual BOQ created. / تم إنشاء جدول الكميات يدوياً.", "success")
        return redirect(url_for("boq_view", boq_id=new_boq.id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"BOQ create manual error: {e}")
        flash("Save failed. Please try again. / فشل الحفظ.", "danger")
        return redirect(url_for("boq_create_manual"))


# ── BOQ UPLOAD + AI REVISE ───────────────────────────────────────────────────

@app.route("/dashboard/boq/upload-revise", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
@limiter.limit("10 per hour", methods=["POST"])
def boq_upload_revise():
    username, company = dashboard_identity()
    projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()

    if request.method == "GET":
        return render_template("dashboard/boq/upload_revise.html",
            username=username, company=company,
            active_dashboard="boq", projects=projects)

    try:
        ok, msg = can_use_ai(current_user)
        if not ok:
            flash(f"AI quota: {msg}", "warning")
            return redirect(url_for("boq_upload_revise"))

        project_id = request.form.get("project_id", type=int)
        title = sanitize(request.form.get("title", "").strip(), 200)
        revision_description = sanitize(request.form.get("revision_description", "").strip(), max_length=3000)

        if not project_id or not title:
            flash("Project and title are required. / المشروع والعنوان مطلوبان.", "danger")
            return redirect(url_for("boq_upload_revise"))

        project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()

        f = request.files.get("boq_file")
        if not f or not f.filename:
            flash("Please upload a BOQ file (.xlsx, .xls, or .pdf). / يرجى رفع ملف جدول الكميات.", "danger")
            return redirect(url_for("boq_upload_revise"))

        filename = secure_filename(f.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".xlsx", ".xls", ".pdf"]:
            flash("Only .xlsx, .xls, and .pdf files are supported.", "danger")
            return redirect(url_for("boq_upload_revise"))

        file_data = f.read()
        if len(file_data) > 10 * 1024 * 1024:
            flash("File exceeds 10MB limit. / الملف يتجاوز حد 10 ميغابايت.", "danger")
            return redirect(url_for("boq_upload_revise"))

        upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "boq_designs", str(current_user.id))
        os.makedirs(upload_dir, exist_ok=True)
        unique_name = f"{uuid.uuid4().hex}_{filename}"
        with open(os.path.join(upload_dir, unique_name), "wb") as out:
            out.write(file_data)
        rel_path = f"uploads/boq_designs/{current_user.id}/{unique_name}"

        # Parse uploaded file
        extracted_text = ""
        try:
            if ext in [".xlsx", ".xls"]:
                wb_tmp = load_workbook(BytesIO(file_data), read_only=True)
                for ws_tmp in list(wb_tmp.worksheets)[:5]:
                    for row in ws_tmp.iter_rows(max_row=200, values_only=True):
                        extracted_text += " ".join(str(c) for c in row if c is not None) + "\n"
            elif ext == ".pdf":
                import pdfplumber
                with pdfplumber.open(BytesIO(file_data)) as pdf:
                    for page in pdf.pages[:10]:
                        extracted_text += page.extract_text() or ""
        except Exception as parse_err:
            app.logger.warning(f"BOQ file parse error: {parse_err}")
            flash("Could not read the uploaded file. Ensure it is a valid Excel or PDF with text content.", "danger")
            return redirect(url_for("boq_upload_revise"))

        if not extracted_text.strip():
            flash("The uploaded file appears empty or has no extractable text.", "danger")
            return redirect(url_for("boq_upload_revise"))

        # STEP 1 — structure the file as v1
        structure_sys = (
            "You are a senior Saudi construction quantity surveyor. "
            "Parse the following raw extracted text from a BOQ file into a structured JSON. "
            "Group items into trade sections (detect naturally). "
            "Include bilingual EN+AR descriptions — translate if only one language. "
            "Preserve existing quantities and rates exactly as found in the file. "
            "For each item, add specification, sub_items, application_notes, and standard_reference "
            "based on your knowledge of Saudi construction standards for that item type. "
            "Return ONLY valid JSON, no markdown fences.\n"
            "Required JSON schema:\n"
            '{"trades_detected":["Civil"],"sections":[{"trade":"Civil","trade_ar":"\u0623\u0639\u0645\u0627\u0644 \u0645\u062f\u0646\u064a\u0629",'
            '"items":[{"item_no":"1.1","description":"...","description_ar":"...",'
            '"specification":"...","sub_items":[{"name":"...","qty_per_unit":0,"unit":"..."}],'
            '"application_notes":"...","standard_reference":"...",'
            '"unit":"m3","qty":100,"rate_sar":250,"total_sar":25000}]}],'
            '"subtotal":25000,"vat_amount":3750,"grand_total":28750}'
        )
        structure_user = f"Project: {project.name}.\n\nRaw BOQ text:\n{extracted_text[:4000]}"

        v1_data, reason_v1, _ = _call_ai_for_master_boq(structure_sys, structure_user, require_detail=False)

        if v1_data is None:
            app.logger.error(f"BOQ structure AI failed: {reason_v1}")
            flash("Could not parse the uploaded BOQ. Ensure the file contains readable BOQ data with items, quantities, and rates.", "danger")
            return redirect(url_for("boq_upload_revise"))

        v1_sections = v1_data.get("sections", [])
        boq_v1 = BOQ(
            user_id=current_user.id,
            project_id=project_id,
            title=f"{title} (v1 — Uploaded)",
            items_json=json.dumps(v1_sections),
            subtotal=float(v1_data.get("subtotal", 0)),
            vat_amount=float(v1_data.get("vat_amount", 0)),
            grand_total=float(v1_data.get("grand_total", 0)),
            status="master",
            is_master=True,
            source="uploaded_v1",
            generation_mode="uploaded_v1",
            version=1,
            design_files_json=json.dumps([rel_path]),
        )
        db.session.add(boq_v1)
        db.session.flush()
        record_ai_usage("boq_master_generation")

        if not revision_description:
            db.session.commit()
            flash("BOQ uploaded and structured as v1. Add a revision description to apply AI changes. / تم رفع الجدول كـ v1.", "info")
            return redirect(url_for("boq_view", boq_id=boq_v1.id))

        # STEP 2 — apply revision
        revision_sys = (
            "You are a senior Saudi construction quantity surveyor revising an existing BOQ. "
            "Apply the user's requested changes. Keep all unrelated items intact. "
            "Use 2025 Saudi market rates for any new items. "
            "Return ONLY valid JSON, no markdown fences."
        ) + _BOQ_ITEM_DETAIL_PROMPT + (
            "\nRequired JSON schema:\n"
            '{"trades_detected":["Civil"],"sections":[{"trade":"Civil","trade_ar":"\u0623\u0639\u0645\u0627\u0644 \u0645\u062f\u0646\u064a\u0629",'
            '"items":[{"item_no":"1.1","description":"...","description_ar":"...",'
            '"specification":"...","sub_items":[{"name":"...","qty_per_unit":0,"unit":"..."}],'
            '"application_notes":"...","standard_reference":"...",'
            '"unit":"m3","qty":100,"rate_sar":250,"total_sar":25000}]}],'
            '"subtotal":25000,"vat_amount":3750,"grand_total":28750}'
        )
        revision_user = (
            f"Current BOQ:\n{json.dumps(v1_sections, ensure_ascii=False)[:3000]}\n\n"
            f"Requested changes: {revision_description}"
        )

        v2_data, reason_v2, _ = _call_ai_for_master_boq(revision_sys, revision_user)

        if v2_data is None:
            db.session.commit()
            record_ai_usage("boq_master_generation")
            app.logger.error(f"BOQ revision AI failed: {reason_v2}")
            flash(
                "v1 saved (original upload). AI revision failed — changes not applied. "
                "You can edit items manually from the BOQ view. / "
                "تم حفظ v1. فشل تطبيق التعديل بالذكاء الاصطناعي.",
                "warning"
            )
            return redirect(url_for("boq_view", boq_id=boq_v1.id))

        v2_sections = v2_data.get("sections", [])
        boq_v2 = BOQ(
            user_id=current_user.id,
            project_id=project_id,
            title=f"{title} (v2 — Revised)",
            items_json=json.dumps(v2_sections),
            subtotal=float(v2_data.get("subtotal", 0)),
            vat_amount=float(v2_data.get("vat_amount", 0)),
            grand_total=float(v2_data.get("grand_total", 0)),
            status="master",
            is_master=True,
            source="uploaded_revised",
            generation_mode="uploaded_revised",
            version=2,
            parent_revision_id=boq_v1.id,
            design_files_json=json.dumps([rel_path]),
            project_description=revision_description,
        )
        db.session.add(boq_v2)
        db.session.commit()
        record_ai_usage("boq_master_generation")
        flash("BOQ uploaded as v1 and revised as v2. / تم رفع الجدول كـ v1 وتعديله كـ v2.", "success")
        return redirect(url_for("boq_view", boq_id=boq_v2.id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"BOQ upload revise error: {e}")
        traceback.print_exc()
        flash("An error occurred. Please try again. / حدث خطأ.", "danger")
        return redirect(url_for("boq_upload_revise"))


# ── SE UPLOAD & ANALYZE ──────────────────────────────────────────────────────

_SE_ANALYSIS_ROLES = [
    "MEP Engineer",
    "Civil Engineer",
    "Structural Engineer",
    "Quantity Surveyor",
    "Procurement Officer",
    "HSE Officer",
    "Site Engineer",
]


@app.route("/dashboard/boq/analyze-upload", methods=["GET", "POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
@limiter.limit("10 per hour", methods=["POST"])
def boq_analyze_upload():
    username, company = dashboard_identity()
    # Projects SE is assigned to
    pa_ids = {a.project_id for a in ProjectAssignment.query.filter_by(user_id=current_user.id).all()}
    ep_ids = {b.project_id for b in BOQ.query.filter(
        BOQ.assigned_to_user_id == current_user.id
    ).with_entities(BOQ.project_id).distinct().all() if b.project_id}
    all_pids = sorted(pa_ids | ep_ids)
    projects = Project.query.filter(Project.id.in_(all_pids)).order_by(Project.name).all() if all_pids else []

    if request.method == "GET":
        return render_template(
            "dashboard/boq/analyze_upload.html",
            username=username, company=company,
            active_dashboard="boq",
            projects=projects,
            se_roles=_SE_ANALYSIS_ROLES,
            default_role=current_user.job_title or "Site Engineer",
        )

    # POST
    try:
        ok, msg = can_use_ai(current_user)
        if not ok:
            flash(f"AI quota: {msg} / حصة الذكاء الاصطناعي: {msg}", "warning")
            return redirect(url_for("boq_analyze_upload"))

        project_id = request.form.get("project_id", type=int)
        selected_role = sanitize(request.form.get("role", "Site Engineer").strip(), 80)
        if selected_role not in _SE_ANALYSIS_ROLES:
            selected_role = "Site Engineer"
        adjustment = sanitize(request.form.get("adjustment", "").strip(), max_length=2000)

        if not project_id:
            flash("Please select a project. / يرجى اختيار المشروع.", "warning")
            return redirect(url_for("boq_analyze_upload"))

        # Verify project access
        has_pa = ProjectAssignment.query.filter_by(project_id=project_id, user_id=current_user.id).first()
        has_boq = BOQ.query.filter(
            BOQ.assigned_to_user_id == current_user.id, BOQ.project_id == project_id
        ).first()
        if not has_pa and not has_boq:
            abort(403)
        project = db.session.get(Project, project_id) or abort(404)

        f = request.files.get("boq_file")
        if not f or not f.filename:
            flash("Please upload a BOQ file (.xlsx, .xls, or .pdf). / يرجى رفع ملف جدول الكميات.", "danger")
            return redirect(url_for("boq_analyze_upload"))

        filename = secure_filename(f.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".xlsx", ".xls", ".pdf"]:
            flash("Only .xlsx, .xls, and .pdf files are supported. / يُدعم فقط .xlsx و .xls و .pdf", "danger")
            return redirect(url_for("boq_analyze_upload"))

        file_data = f.read()
        if len(file_data) > 10 * 1024 * 1024:
            flash("File exceeds 10 MB limit. / الملف يتجاوز حد 10 ميغابايت.", "danger")
            return redirect(url_for("boq_analyze_upload"))

        # Magic-byte check
        _MAGIC = {b"\x50\x4b": ".xlsx/.xls", b"\xd0\xcf": ".xls", b"\x25\x50": ".pdf"}
        sig = file_data[:2]
        if not any(file_data.startswith(k) for k in _MAGIC):
            flash("File signature does not match a valid Excel or PDF. / توقيع الملف غير صالح.", "danger")
            return redirect(url_for("boq_analyze_upload"))

        # Save file
        upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "boq_designs", str(current_user.id))
        os.makedirs(upload_dir, exist_ok=True)
        unique_name = f"{uuid.uuid4().hex}_{filename}"
        with open(os.path.join(upload_dir, unique_name), "wb") as out:
            out.write(file_data)
        rel_path = f"uploads/boq_designs/{current_user.id}/{unique_name}"

        # Parse file text
        extracted_text = ""
        try:
            if ext in [".xlsx", ".xls"]:
                wb_tmp = load_workbook(BytesIO(file_data), read_only=True)
                for ws_tmp in list(wb_tmp.worksheets)[:5]:
                    for row in ws_tmp.iter_rows(max_row=200, values_only=True):
                        extracted_text += " ".join(str(c) for c in row if c is not None) + "\n"
            elif ext == ".pdf":
                import pdfplumber
                with pdfplumber.open(BytesIO(file_data)) as pdf:
                    for page in pdf.pages[:10]:
                        extracted_text += page.extract_text() or ""
        except Exception as parse_err:
            app.logger.warning("SE BOQ analyze parse error: %s", parse_err)
            flash("Could not read the uploaded file. Ensure it is a valid Excel or PDF. / تعذّر قراءة الملف.", "danger")
            return redirect(url_for("boq_analyze_upload"))

        if not extracted_text.strip():
            flash("The uploaded file appears empty or has no extractable text. / الملف يبدو فارغاً.", "danger")
            return redirect(url_for("boq_analyze_upload"))

        # AI analysis
        adjustment_clause = f"\nUser adjustment request: {adjustment}" if adjustment else ""
        system_prompt = (
            f"You are a senior Saudi quantity surveyor analyzing a BOQ from the perspective of a {selected_role}. "
            f"Focus your analysis on items, quantities, and rates relevant to that role. "
            f"Identify gaps, errors, or missing items typical for a {selected_role} reviewing this BOQ. "
            f"Apply the user's adjustment description.{adjustment_clause}"
        ) + _BOQ_ITEM_DETAIL_PROMPT + (
            "\nReturn ONLY valid JSON, no markdown fences.\n"
            "Required JSON schema:\n"
            '{"trades_detected":["Civil"],"sections":[{"trade":"Civil","trade_ar":"\u0623\u0639\u0645\u0627\u0644 \u0645\u062f\u0646\u064a\u0629",'
            '"items":[{"item_no":"1.1","description":"...","description_ar":"...",'
            '"specification":"...","sub_items":[{"name":"...","qty_per_unit":0,"unit":"..."}],'
            '"application_notes":"...","standard_reference":"...",'
            '"unit":"m3","qty":100,"rate_sar":250,"total_sar":25000}]}],'
            '"subtotal":25000,"vat_amount":3750,"grand_total":28750}'
        )
        user_payload = (
            f"Project: {project.name}. Reviewing as: {selected_role}.\n\n"
            f"Raw BOQ text:\n{extracted_text[:4000]}"
        )

        analyzed, reason, _ = _call_ai_for_master_boq(system_prompt, user_payload)

        if analyzed is None:
            app.logger.error("SE analyze AI failed: %s", reason)
            flash(
                f"AI analysis failed: {reason}. Please try again. / فشل التحليل بالذكاء الاصطناعي.",
                "danger",
            )
            return redirect(url_for("boq_analyze_upload"))

        sections = analyzed.get("sections", [])
        new_boq = BOQ(
            user_id=current_user.id,
            project_id=project_id,
            title=f"Analysis: {filename} ({selected_role})",
            items_json=json.dumps(sections),
            subtotal=float(analyzed.get("subtotal", 0)),
            vat_amount=float(analyzed.get("vat_amount", 0)),
            grand_total=float(analyzed.get("grand_total", 0)),
            status="distributed",
            is_master=False,
            assigned_to_user_id=current_user.id,
            source="engineer_analysis",
            generation_mode="upload_analyze_by_role",
            design_files_json=json.dumps([rel_path]),
            project_description=adjustment or selected_role,
            version=1,
        )
        db.session.add(new_boq)
        db.session.commit()
        record_ai_usage("boq_engineer_analysis")
        flash(
            f"Analysis complete ({selected_role}). Review your results below. / "
            f"اكتمل التحليل ({selected_role}).",
            "success",
        )
        return redirect(url_for("boq_view", boq_id=new_boq.id))

    except Exception as exc:
        db.session.rollback()
        app.logger.error("SE analyze upload error: %s", exc)
        traceback.print_exc()
        flash("An error occurred. Please try again. / حدث خطأ، يرجى المحاولة.", "danger")
        return redirect(url_for("boq_analyze_upload"))


# ── END BOQ REBUILD ROUTES ────────────────────────────────────────────────────


# [BOQ rebuild] removed: api_boq_save, api_boq_transcribe_voice, boq_generate_from_description, boq_upload_smart, boq_modify_from_previous


# [BOQ rebuild] removed: normalize_boq_audit_result, boq_audit, boq_audit_export_pdf, old boq_delete


# [BOQ rebuild] removed: _package_id_for_boq_item, record_actuals_from_dpr, boq routes (tracker/actuals/packages/distribution)


# ── INVENTORY REBUILD — HELPERS ──────────────────────────────────────────────

def _extract_inventory_json(raw_text):
    """Strip markdown fences and extract JSON from AI inventory response."""
    if not raw_text:
        return None
    import re as _re
    cleaned = _re.sub(r'^```(?:json)?\s*', '', raw_text.strip(), flags=_re.IGNORECASE)
    cleaned = _re.sub(r'\s*```\s*$', '', cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass
    match = _re.search(r'\{.*\}', cleaned, _re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except json.JSONDecodeError:
            return None
    return None


def _validate_inventory_response(parsed, min_categories=1, require_detail=False, min_total_items=None):
    """Return (is_valid, reason) for an AI inventory response."""
    if not isinstance(parsed, dict):
        return False, "Response is not a JSON object"
    categories = parsed.get('categories')
    if not isinstance(categories, list) or len(categories) == 0:
        return False, "No categories returned"
    if len(categories) < min_categories:
        return False, f"Only {len(categories)} category/categories returned; {min_categories} required for this project"
    all_items = [item for c in categories if isinstance(c, dict)
                 for item in c.get('items', []) if isinstance(item, dict)]
    total_items = len(all_items)
    required_total = min_total_items if min_total_items is not None else (10 if min_categories >= 3 else 1)
    if total_items < required_total:
        return False, f"Only {total_items} inventory item(s) returned; {required_total} required for this project"
    if all_items:
        missing_spec = sum(1 for it in all_items if not str(it.get('specification') or '').strip())
        missing_spec_ratio = missing_spec / len(all_items)
        if missing_spec_ratio > 0.40:
            app.logger.warning(
                "Inventory validator warning: specification missing on %.0f%% of items (%d/%d).",
                missing_spec_ratio * 100, missing_spec, len(all_items),
            )
        missing_brands = sum(
            1 for it in all_items
            if not (isinstance(it.get('brand_suggestions'), list) and len(it.get('brand_suggestions') or []) > 0)
        )
        missing_brands_ratio = missing_brands / len(all_items)
        if missing_brands_ratio > 0.50:
            app.logger.warning(
                "Inventory validator warning: brand_suggestions missing on %.0f%% of items (%d/%d).",
                missing_brands_ratio * 100, missing_brands, len(all_items),
            )
        for field in ("alternative_items", "shelf_life_days"):
            missing_count = sum(1 for it in all_items if it.get(field) in (None, "", []))
            if missing_count:
                app.logger.info(
                    "Inventory validator bonus-field warning: %s missing on %d/%d items.",
                    field, missing_count, len(all_items),
                )
    return True, "ok"


def _call_ai_for_master_inventory(system_prompt, user_payload, attempt=1, min_categories=1, require_detail=False):
    """Call GPT with json_object mode; retry once on failure."""
    if client is None:
        return None, "AI service is not configured. Please set OPENAI_API_KEY.", ""

    payload_len = len(user_payload) if isinstance(user_payload, str) else sum(
        len(p.get("text", "")) for p in user_payload if isinstance(p, dict)
    )

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_payload},
        ],
        max_tokens=16000,
        temperature=0.2,
    )
    raw = response.choices[0].message.content or ""
    finish_reason = response.choices[0].finish_reason

    app.logger.info(
        "Inventory AI attempt=%d: input_len=%d, response_len=%d, finish_reason=%s, "
        "first500=%s, last500=%s",
        attempt, payload_len, len(raw), finish_reason,
        raw[:500], raw[-500:] if len(raw) > 500 else raw,
    )

    if finish_reason == "length":
        app.logger.error(
            "Inventory AI hit token limit (finish_reason=length) on attempt %d — "
            "response truncated at %d chars.",
            attempt, len(raw),
        )

    parsed = _extract_inventory_json(raw)

    if finish_reason == "length" and parsed is None:
        return None, "token_limit_truncation", raw

    is_valid, reason = (
        _validate_inventory_response(parsed, min_categories=min_categories, require_detail=require_detail)
        if parsed is not None else (False, "parse failed")
    )

    if is_valid:
        app.logger.info(
            "Inventory AI success: categories_detected=%s, categories_count=%d, items_total=%d",
            parsed.get('categories_detected'),
            len(parsed.get('categories', [])),
            sum(len(c.get('items', [])) for c in parsed.get('categories', [])),
        )
        return parsed, None, raw

    app.logger.warning("Inventory AI attempt=%d failed validation: reason=%s", attempt, reason)

    if attempt < 2:
        retry_sys = system_prompt + (
            "\n\nCRITICAL: Your previous response was rejected. Return ONLY a valid JSON object matching the schema. "
            "No markdown, no prose, no code fences. "
            f"You MUST include at least {min_categories} separate categories, each with at least 2 items. "
            "A construction project always needs multiple material categories — do not return only one. "
            "It is acceptable for some items to omit optional fields if needed to keep response length reasonable."
        )
        return _call_ai_for_master_inventory(
            retry_sys, user_payload, attempt=2,
            min_categories=min_categories, require_detail=require_detail
        )
    return None, reason, raw


# ── INVENTORY REBUILD — NEW ROUTES ───────────────────────────────────────────

@app.route("/dashboard/inventory")
@dashboard_access_required
def inventory_index():
    username, company = dashboard_identity()
    filter_project_id = request.args.get("project_id", type=int)
    if current_user.role == ROLE_PROJECT_MANAGER:
        # Group items by master_inventory_batch_id
        q = InventoryItem.query.filter_by(user_id=current_user.id).filter(
            InventoryItem.master_inventory_batch_id != None
        )
        if filter_project_id:
            q = q.filter_by(project_id=filter_project_id)
        items = q.order_by(InventoryItem.created_at.desc()).all()
        # Build batch summary
        batch_map = {}
        for item in items:
            bid = item.master_inventory_batch_id
            if bid not in batch_map:
                batch_map[bid] = {
                    'batch_id': bid,
                    'project_id': item.project_id,
                    'project': item.project,
                    'item_count': 0,
                    'total_sar': 0.0,
                    'created_at': item.created_at,
                    'distributed_count': 0,
                }
            batch_map[bid]['item_count'] += 1
            batch_map[bid]['total_sar'] += float(item.value_sar or 0) * float(item.stock or 0)
        # Count distributed (have at least one InventoryAssignment)
        for bid, data in batch_map.items():
            batch_item_ids = [i.id for i in items if i.master_inventory_batch_id == bid]
            data['distributed_count'] = InventoryAssignment.query.filter(
                InventoryAssignment.inventory_item_id.in_(batch_item_ids)
            ).count() if batch_item_ids else 0
        batches = sorted(batch_map.values(), key=lambda x: x['created_at'] or '', reverse=True)
        # Pending stock requests count
        pm_project_ids = [p.id for p in Project.query.filter_by(user_id=current_user.id).all()]
        pending_requests = StockRequest.query.filter(
            StockRequest.project_id.in_(pm_project_ids),
            StockRequest.status == 'pending'
        ).count() if pm_project_ids else 0
        return render_template('dashboard/inventory/index.html',
            username=username, company=company,
            active_dashboard='inventory',
            role='pm', batches=batches,
            pending_requests=pending_requests,
            filter_project_id=filter_project_id)
    else:
        # SE: items assigned to them
        assignments = InventoryAssignment.query.filter_by(
            assigned_user_id=current_user.id
        ).all()
        item_ids = [a.inventory_item_id for a in assignments]
        q2 = InventoryItem.query.filter(InventoryItem.id.in_(item_ids))
        if filter_project_id:
            q2 = q2.filter_by(project_id=filter_project_id)
        assigned_items = q2.order_by(InventoryItem.project_id, InventoryItem.category).all() if item_ids else []
        # Build per-item data with assignment info
        assign_map = {a.inventory_item_id: a for a in assignments}
        items_data = []
        for item in assigned_items:
            a = assign_map.get(item.id)
            stock = float(item.stock or 0)
            threshold = float(item.threshold or 0)
            if threshold > 0 and stock < threshold * 0.25:
                alert = 'critical'
            elif threshold > 0 and stock < threshold:
                alert = 'low'
            else:
                alert = 'ok'
            items_data.append({'item': item, 'assignment': a, 'alert': alert})
        # SE's own pending stock requests (can be deleted)
        my_pending_requests = StockRequest.query.filter_by(
            requested_by_user_id=current_user.id,
            status='pending'
        ).order_by(StockRequest.created_at.desc()).all()
        return render_template('dashboard/inventory/index.html',
            username=username, company=company,
            active_dashboard='inventory',
        role='se', items_data=items_data,
            my_pending_requests=my_pending_requests)


def _legacy_inventory_fallback_items(project, description):
    return [
        {
            "name": "Portland Cement",
            "name_ar": "\u0625\u0633\u0645\u0646\u062a \u0628\u0648\u0631\u062a\u0644\u0627\u0646\u062f",
            "category": "Concrete",
            "category_ar": "\u0627\u0644\u062e\u0631\u0633\u0627\u0646\u0629",
            "unit": "bag",
            "recommended_stock": 500,
            "threshold": 100,
            "value_sar": 18.50,
            "supplier_hint": "Saudi Cement Co.",
            "notes": "Baseline master inventory item generated from project description.",
        },
        {
            "name": "Steel Rebar",
            "name_ar": "\u062d\u062f\u064a\u062f \u062a\u0633\u0644\u064a\u062d",
            "category": "Steel",
            "category_ar": "\u0627\u0644\u062d\u062f\u064a\u062f",
            "unit": "ton",
            "recommended_stock": 50,
            "threshold": 10,
            "value_sar": 2800.0,
            "supplier_hint": "Hadeed Steel",
            "notes": "Baseline master inventory item generated from project description.",
        },
    ]


def _flatten_inventory_payload(inv_data):
    if isinstance(inv_data, list):
        return inv_data
    if not isinstance(inv_data, dict):
        return []
    flattened = []
    for category in inv_data.get("categories", []):
        if not isinstance(category, dict):
            continue
        category_name = category.get("category") or "General"
        category_name_ar = category.get("category_ar") or ""
        for item in category.get("items", []):
            if isinstance(item, dict):
                row = dict(item)
                row.setdefault("category", category_name)
                row.setdefault("category_ar", category_name_ar)
                flattened.append(row)
    return flattened


def _legacy_inventory_items(project, description):
    api_key = (app.config.get("OPENAI_API_KEY") or "").strip()
    if client is not None and api_key and not api_key.startswith("sk-test"):
        ok, msg = can_use_ai(current_user)
        if not ok:
            app.logger.info("Inventory generate legacy fallback due to quota: %s", msg)
        else:
            system_prompt = (
                "Generate a concise master inventory for a Saudi construction project. "
                "Return JSON with categories[].items[] and realistic 2025 SAR values."
            )
            user_payload = [{"type": "text", "text": f"Project: {project.name}. Description: {description}."}]
            inv_data, reason, raw = _call_ai_for_master_inventory(system_prompt, user_payload, min_categories=1)
            items = _flatten_inventory_payload(inv_data)
            if items:
                record_ai_usage("inventory_master_generation")
                return items
            app.logger.warning("Inventory generate legacy fallback: %s", reason or "empty AI response")
    return _legacy_inventory_fallback_items(project, description)


def _render_inventory_generate_preview(project, description, items):
    csrf_value = generate_csrf()
    rows = []
    for item in items:
        name = html.escape(str(item.get("name") or "Unnamed"))
        category = html.escape(str(item.get("category") or "General"))
        unit = html.escape(str(item.get("unit") or "pcs"))
        qty = html.escape(str(item.get("recommended_stock") or item.get("stock") or 0))
        value = html.escape(str(item.get("value_sar") or 0))
        rows.append(
            f"<tr><td>{name}</td><td>{category}</td><td>{qty}</td><td>{unit}</td><td>{value}</td></tr>"
        )
    description_attr = html.escape(description, quote=True)
    project_name = html.escape(project.name)
    return f"""<!doctype html>
<html lang="en">
<head><meta charset="utf-8"><title>Inventory Preview</title></head>
<body>
  <main>
    <h1>Master Inventory Preview</h1>
    <p>{project_name}</p>
    <table>
      <thead><tr><th>Name</th><th>Category</th><th>Stock</th><th>Unit</th><th>Value SAR</th></tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
    <form method="post" action="{url_for('inventory_generate_legacy')}">
      <input type="hidden" name="csrf_token" value="{csrf_value}">
      <input type="hidden" name="project_id" value="{project.id}">
      <input type="hidden" name="project_description" value="{description_attr}">
      <input type="hidden" name="confirm" value="1">
      <button type="submit">Save All</button>
    </form>
  </main>
</body>
</html>"""


@app.route("/dashboard/inventory/generate", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
@limiter.limit("10 per hour", methods=["POST"])
def inventory_generate_legacy():
    username, company = dashboard_identity()
    projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()
    if request.method == "GET":
        return render_template(
            "dashboard/inventory/create.html",
            username=username,
            company=company,
            active_dashboard="inventory",
            projects=projects,
            prefill_description="",
            prefill_project_id=request.args.get("project_id", type=int),
        )

    project_id = request.form.get("project_id", type=int)
    description = sanitize(
        (request.form.get("project_description") or request.form.get("description") or "").strip(),
        max_length=3000,
    )
    if not project_id or not description:
        flash("Project and description are required.", "danger")
        return redirect(url_for("inventory_generate_legacy"))
    project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()
    items = _legacy_inventory_items(project, description)

    if request.form.get("confirm") != "1":
        return _render_inventory_generate_preview(project, description, items)

    batch_id = uuid.uuid4().hex[:32]
    saved_count = 0
    try:
        for item in items:
            inv = InventoryItem(
                user_id=current_user.id,
                project_id=project.id,
                name=sanitize(str(item.get("name") or "Unnamed"), 200),
                name_ar=sanitize(str(item.get("name_ar") or ""), 200),
                category=sanitize(str(item.get("category") or "General"), 50),
                category_ar=sanitize(str(item.get("category_ar") or ""), 80),
                unit=sanitize(str(item.get("unit") or "pcs"), 30),
                stock=float(item.get("recommended_stock") or item.get("stock") or 0),
                threshold=float(item.get("threshold") or 0),
                value_sar=float(item.get("value_sar") or 0),
                supplier=sanitize(str(item.get("supplier_hint") or item.get("supplier") or ""), 100),
                notes=sanitize(str(item.get("notes") or ""), 500),
                source="ai_generated_master",
                master_inventory_batch_id=batch_id,
            )
            db.session.add(inv)
            saved_count += 1
        db.session.commit()
        flash(f"Master inventory created: {saved_count} items.", "success")
        return redirect(url_for("inventory_batch_view", batch_id=batch_id))
    except Exception as error:
        db.session.rollback()
        app.logger.error("Legacy inventory generate save failed: %s", error)
        flash("Could not save generated inventory.", "danger")
        return redirect(url_for("inventory_generate_legacy"))


@app.route("/dashboard/inventory/create", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
@limiter.limit("10 per hour", methods=["POST"])
def inventory_create():
    username, company = dashboard_identity()
    projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()

    if request.method == "GET":
        prefill_description = session.pop("inv_create_prefill_description", "")
        prefill_project_id = request.args.get("project_id", type=int)
        return render_template("dashboard/inventory/create.html",
            username=username, company=company,
            active_dashboard="inventory", projects=projects,
            prefill_description=prefill_description,
            prefill_project_id=prefill_project_id)

    # POST
    try:
        ok, msg = can_use_ai(current_user)
        if not ok:
            flash(f"AI quota: {msg}", "warning")
            return redirect(url_for("inventory_create"))

        project_id = request.form.get("project_id", type=int)
        description = sanitize(request.form.get("description", "").strip(), max_length=3000)
        if not project_id or not description:
            flash("Project and description are required. / المشروع والوصف مطلوبان.", "danger")
            return redirect(url_for("inventory_create"))

        project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()

        # Handle file uploads
        saved_files = []
        upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "inventory_designs", str(current_user.id))
        os.makedirs(upload_dir, exist_ok=True)
        image_parts = []
        text_context = ""

        files = request.files.getlist("design_files")
        for f in files[:5]:
            if not f or not f.filename:
                continue
            filename = secure_filename(f.filename)
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ['.pdf', '.xlsx', '.xls', '.jpg', '.jpeg', '.png']:
                continue
            file_data = f.read()
            if len(file_data) > 10 * 1024 * 1024:
                flash(f"File {filename} exceeds 10MB limit.", "warning")
                continue
            unique_name = f"{uuid.uuid4().hex}_{filename}"
            save_path = os.path.join(upload_dir, unique_name)
            with open(save_path, 'wb') as out:
                out.write(file_data)
            saved_files.append(f"uploads/inventory_designs/{current_user.id}/{unique_name}")
            if ext in ['.jpg', '.jpeg', '.png']:
                b64 = base64.b64encode(file_data).decode('utf-8')
                mime = 'image/jpeg' if ext in ['.jpg', '.jpeg'] else 'image/png'
                image_parts.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}})
            elif ext == '.pdf':
                try:
                    import pdfplumber
                    with pdfplumber.open(BytesIO(file_data)) as pdf:
                        for page in pdf.pages[:5]:
                            text_context += page.extract_text() or ""
                except Exception:
                    pass
            elif ext in ['.xlsx', '.xls']:
                try:
                    wb_tmp = load_workbook(BytesIO(file_data), read_only=True)
                    for ws_tmp in list(wb_tmp.worksheets)[:3]:
                        for row in ws_tmp.iter_rows(max_row=50, values_only=True):
                            text_context += " ".join(str(c) for c in row if c) + "\n"
                except Exception:
                    pass

        min_categories = 3 if len(description) > 100 else 1
        system_prompt = (
            "You are a senior Saudi construction site materials manager generating a complete, "
            "procurement-ready Master Inventory list.\n\n"
            "CRITICAL RULES — failure to follow these will delay procurement:\n\n"
            "1. Analyze the project description carefully and identify EVERY material category required. "
            "Construction projects almost always need MULTIPLE categories, not just one.\n\n"
            "2. For a typical building project, you MUST include ALL of these categories when applicable "
            "(which is almost always):\n"
            "   - Cement & Concrete (cement, ready-mix, additives)\n"
            "   - Steel & Rebar (reinforcement bars, structural steel)\n"
            "   - Electrical (cables, breakers, conduits, fixtures)\n"
            "   - Plumbing (pipes, fittings, valves, fixtures)\n"
            "   - Finishing & Tiling (tiles, marble, adhesive, grout)\n"
            "   - Paints & Coatings (primer, emulsion, weathershield)\n"
            "   - Doors & Windows (frames, glass, hardware)\n"
            "   - Safety Equipment (helmets, vests, harnesses)\n\n"
            "3. For specialty projects, ADD relevant categories:\n"
            "   - HVAC: ductwork, units, insulation\n"
            "   - Landscaping: irrigation, plants, soil\n"
            "   - Fire Fighting: pipes, sprinklers, extinguishers\n\n"
            "4. Each category MUST have AT LEAST 3-5 items with realistic 2025 Saudi market rates in SAR.\n\n"
            "5. Include bilingual EN+AR item names.\n\n"
            "6. Only return {\"error\": \"description_too_vague\"} if the description is under 50 characters "
            "or contains no identifiable project type at all. "
            "For descriptions over 100 characters, ALWAYS generate an inventory.\n\n"
            "7. Return strict JSON. NO prose, NO markdown fences.\n"
        ) + _INVENTORY_ITEM_DETAIL_PROMPT + (
            "\n\nRequired top-level schema:\n"
            '{"categories_detected":["Cement & Concrete","Steel & Rebar","Electrical"],'
            '"categories":[{"category":"Cement & Concrete","category_ar":"الإسمنت والخرسانة",'
            '"items":[<item objects matching schema above>]}],'
            '"total_items":15,"total_value_sar":50000}'
        )

        user_content = [{"type": "text", "text": f"Project: {project.name}. Description: {description}."}]
        if text_context:
            user_content[0]["text"] += f"\n\nFile context:\n{text_context[:3000]}"
        user_content.extend(image_parts)

        inv_data, reason, raw = _generate_chunked_master_inventory(project, description, user_content)

        if inv_data is None:
            app.logger.error(
                "Master Inventory AI failed: reason=%s, description_len=%d, raw_preview=%s",
                reason, len(description), (raw[:500] if raw else "None")
            )
            session["inv_create_prefill_description"] = description
            if reason == "description_too_vague" or len(description) < 200:
                flash(
                    "AI couldn't generate an inventory from that description. Please write at least "
                    "2-3 sentences describing the project type, size, and materials needed.",
                    "warning"
                )
                flash(
                    "تعذر إنشاء المخزون. يرجى كتابة وصف أكثر تفصيلاً يتضمن نوع المشروع والمواد المطلوبة.",
                    "warning"
                )
            else:
                flash(
                    "Inventory generation failed due to a technical issue — your description was good. "
                    "Please try again.",
                    "warning"
                )
                flash(
                    "تعذر إنشاء المخزون بسبب مشكلة تقنية. وصفك كان جيداً — حاول مرة أخرى.",
                    "warning"
                )
            return redirect(url_for("inventory_create"))

        if inv_data.get("_partial_generation_warnings"):
            flash(
                "Some inventory categories couldn't be generated - you can add them manually using Edit Batch.",
                "warning"
            )
            flash(
                "تعذر إنشاء بعض فئات المخزون - يمكنك إضافتها يدويًا من تعديل الدفعة.",
                "warning"
            )

        batch_id = uuid.uuid4().hex[:32]
        saved_count = 0
        for cat in inv_data.get("categories", []):
            cat_name = sanitize(str(cat.get("category", "General")), 50)
            cat_name_ar = sanitize(str(cat.get("category_ar", "")), 80)
            for item_data in cat.get("items", []):
                try:
                    _brands = item_data.get("brand_suggestions")
                    _alts = item_data.get("alternative_items")
                    inv = InventoryItem(
                        user_id=current_user.id,
                        project_id=project_id,
                        name=sanitize(str(item_data.get("name", "Unnamed")), 200),
                        name_ar=sanitize(str(item_data.get("name_ar", "")), 200),
                        category=cat_name,
                        category_ar=cat_name_ar,
                        unit=sanitize(str(item_data.get("unit", "pcs")), 30),
                        stock=float(item_data.get("recommended_stock", 0) or 0),
                        threshold=float(item_data.get("threshold", 0) or 0),
                        value_sar=float(item_data.get("value_sar", 0) or 0),
                        supplier=sanitize(str(item_data.get("supplier_hint", "")), 100),
                        notes=sanitize(str(item_data.get("notes", "") or ""), 500),
                        source="ai_generated_master",
                        master_inventory_batch_id=batch_id,
                        design_files_json=json.dumps(saved_files) if saved_files else None,
                        # Rich procurement fields
                        specification=sanitize(str(item_data.get("specification", "") or ""), 2000) or None,
                        brand_suggestions_json=json.dumps(_brands) if isinstance(_brands, list) else None,
                        storage_requirements=sanitize(str(item_data.get("storage_requirements", "") or ""), 1000) or None,
                        reorder_lead_time_days=int(item_data["reorder_lead_time_days"]) if item_data.get("reorder_lead_time_days") else None,
                        min_order_qty=float(item_data["min_order_qty"]) if item_data.get("min_order_qty") else None,
                        min_order_unit=sanitize(str(item_data.get("min_order_unit", "") or ""), 50) or None,
                        safety_notes=sanitize(str(item_data.get("safety_notes", "") or ""), 1000) or None,
                        alternative_items_json=json.dumps(_alts) if isinstance(_alts, list) else None,
                        shelf_life_days=int(item_data["shelf_life_days"]) if item_data.get("shelf_life_days") else None,
                        shelf_life_note=sanitize(str(item_data.get("shelf_life_note", "") or ""), 500) or None,
                    )
                    db.session.add(inv)
                    saved_count += 1
                except Exception as e:
                    app.logger.warning(f"Inventory item save error: {e}")

        db.session.commit()
        record_ai_usage("inventory_master_generation")
        flash(f"Master inventory created: {saved_count} items across {len(inv_data.get('categories', []))} categories. / تم إنشاء المخزون الرئيسي.", "success")
        return redirect(url_for("inventory_batch_view", batch_id=batch_id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Inventory create error: {e}")
        traceback.print_exc()
        flash("An error occurred. Please try again. / حدث خطأ.", "danger")
        return redirect(url_for("inventory_create"))


@app.route("/api/inventory/transcribe", methods=["POST"])
@dashboard_access_required
@limiter.limit("30 per hour")
def inventory_transcribe():
    forbidden = ensure_ai_access("inventory")
    if forbidden:
        return forbidden
    if client is None:
        return jsonify({"success": False, "error": "AI service is not configured. Please set OPENAI_API_KEY."}), 503
    audio_file = request.files.get("audio")
    if not audio_file:
        return jsonify({"success": False, "error": "No audio file"}), 400
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as tmp:
            audio_file.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            result = client.audio.transcriptions.create(
                model="whisper-1",
                file=f,
                response_format="verbose_json",
            )
        text = result.text or ""
        record_ai_usage("inventory_voice")
        return jsonify({"success": True, "text": text, "language_detected": getattr(result, "language", "auto")})
    except Exception as e:
        app.logger.error(f"Inventory transcription error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


@app.route("/dashboard/inventory/batch/<batch_id>")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_batch_view(batch_id):
    username, company = dashboard_identity()
    items = InventoryItem.query.filter_by(user_id=current_user.id, master_inventory_batch_id=batch_id).order_by(InventoryItem.category, InventoryItem.name).all()
    if not items:
        abort(404)
    # Group by category
    categories = {}
    for item in items:
        cat = item.category or "General"
        if cat not in categories:
            categories[cat] = {'category': cat, 'category_ar': item.category_ar or '', 'items': [], 'total_sar': 0.0}
        categories[cat]['items'].append(item)
        categories[cat]['total_sar'] += float(item.value_sar or 0) * float(item.stock or 0)
    project = db.session.get(Project, items[0].project_id) if items and items[0].project_id else None
    total_items = len(items)
    total_sar = sum(float(i.value_sar or 0) * float(i.stock or 0) for i in items)
    # Check if this batch has a parent (came from upload-revise)
    parent_batch_id = items[0].parent_batch_id if items else None
    return render_template('dashboard/inventory/batch_view.html',
        username=username, company=company,
        active_dashboard='inventory',
        batch_id=batch_id,
        categories=list(categories.values()),
        project=project,
        total_items=total_items,
        total_sar=total_sar,
        items=items,
        parent_batch_id=parent_batch_id)


@app.route("/dashboard/inventory/batch/<batch_id>/edit", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_batch_edit(batch_id):
    username, company = dashboard_identity()
    items = InventoryItem.query.filter_by(user_id=current_user.id, master_inventory_batch_id=batch_id).order_by(InventoryItem.category, InventoryItem.name).all()
    if not items:
        abort(404)

    if request.method == "GET":
        categories = {}
        for item in items:
            cat = item.category or "General"
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item)
        return render_template('dashboard/inventory/batch_edit.html',
            username=username, company=company,
            active_dashboard='inventory',
            batch_id=batch_id,
            categories=categories)

    # POST: save edits + handle add/delete rows
    # 1. Process deletes
    delete_ids_raw = request.form.get("delete_ids", "")
    if delete_ids_raw:
        for del_id_str in delete_ids_raw.split(","):
            del_id_str = del_id_str.strip()
            if not del_id_str.isdigit():
                continue
            del_item = InventoryItem.query.filter_by(
                id=int(del_id_str), user_id=current_user.id, master_inventory_batch_id=batch_id
            ).first()
            if del_item:
                db.session.delete(del_item)

    # 2. Update existing items
    item_map = {str(i.id): i for i in items}
    for item_id_str, item in item_map.items():
        # Skip if marked for delete
        if item_id_str in delete_ids_raw.split(","):
            continue
        item.name = sanitize(request.form.get(f"name_{item_id_str}", item.name), 200) or item.name
        item.name_ar = sanitize(request.form.get(f"name_ar_{item_id_str}", item.name_ar or ""), 200)
        item.category = sanitize(request.form.get(f"category_{item_id_str}", item.category or ""), 50)
        item.unit = sanitize(request.form.get(f"unit_{item_id_str}", item.unit or ""), 30)
        try:
            item.stock = float(request.form.get(f"stock_{item_id_str}", item.stock) or 0)
        except (ValueError, TypeError):
            pass
        try:
            item.threshold = float(request.form.get(f"threshold_{item_id_str}", item.threshold) or 0)
        except (ValueError, TypeError):
            pass
        try:
            item.value_sar = float(request.form.get(f"value_sar_{item_id_str}", item.value_sar) or 0)
        except (ValueError, TypeError):
            pass
        item.supplier = sanitize(request.form.get(f"supplier_{item_id_str}", item.supplier or ""), 100)
        item.notes = sanitize(request.form.get(f"notes_{item_id_str}", item.notes or ""), 500)
        # Rich fields
        spec_val = sanitize(request.form.get(f"specification_{item_id_str}", ""), 2000)
        if spec_val:
            item.specification = spec_val
        brands_raw = request.form.get(f"brand_suggestions_{item_id_str}", "")
        if brands_raw.strip():
            item.brand_suggestions_json = json.dumps([s.strip() for s in brands_raw.split(",") if s.strip()])
        stor_val = sanitize(request.form.get(f"storage_requirements_{item_id_str}", ""), 1000)
        if stor_val:
            item.storage_requirements = stor_val
        rlt_val = request.form.get(f"reorder_lead_time_days_{item_id_str}", "")
        if rlt_val.strip().isdigit():
            item.reorder_lead_time_days = int(rlt_val)
        moq_val = request.form.get(f"min_order_qty_{item_id_str}", "")
        try:
            if moq_val.strip():
                item.min_order_qty = float(moq_val)
        except (ValueError, TypeError):
            pass
        mou_val = sanitize(request.form.get(f"min_order_unit_{item_id_str}", ""), 50)
        if mou_val:
            item.min_order_unit = mou_val
        sn_val = sanitize(request.form.get(f"safety_notes_{item_id_str}", ""), 1000)
        if sn_val:
            item.safety_notes = sn_val
        sld_val = request.form.get(f"shelf_life_days_{item_id_str}", "")
        if sld_val.strip().isdigit():
            item.shelf_life_days = int(sld_val)
        sln_val = sanitize(request.form.get(f"shelf_life_note_{item_id_str}", ""), 500)
        if sln_val:
            item.shelf_life_note = sln_val

    # 3. Add new items
    try:
        new_count = int(request.form.get("new_count", 0))
    except (ValueError, TypeError):
        new_count = 0
    project_id_for_new = items[0].project_id if items else None
    for n in range(1, new_count + 1):
        new_name = sanitize(request.form.get(f"new_{n}_name", "").strip(), 200)
        if not new_name:
            continue
        try:
            _new_brands_raw = request.form.get(f"new_{n}_brand_suggestions", "")
            _new_brands = [s.strip() for s in _new_brands_raw.split(",") if s.strip()] if _new_brands_raw.strip() else None
            _rlt = request.form.get(f"new_{n}_reorder_lead_time_days", "")
            _moq = request.form.get(f"new_{n}_min_order_qty", "")
            _sld = request.form.get(f"new_{n}_shelf_life_days", "")
            db.session.add(InventoryItem(
                user_id=current_user.id,
                project_id=project_id_for_new,
                name=new_name,
                name_ar=sanitize(request.form.get(f"new_{n}_name_ar", ""), 200),
                category=sanitize(request.form.get(f"new_{n}_category", "General"), 50),
                unit=sanitize(request.form.get(f"new_{n}_unit", ""), 30),
                stock=float(request.form.get(f"new_{n}_stock", 0) or 0),
                threshold=float(request.form.get(f"new_{n}_threshold", 0) or 0),
                value_sar=float(request.form.get(f"new_{n}_value_sar", 0) or 0),
                supplier=sanitize(request.form.get(f"new_{n}_supplier", ""), 100),
                notes=sanitize(request.form.get(f"new_{n}_notes", ""), 500),
                source="manual",
                master_inventory_batch_id=batch_id,
                specification=sanitize(request.form.get(f"new_{n}_specification", ""), 2000) or None,
                brand_suggestions_json=json.dumps(_new_brands) if _new_brands else None,
                storage_requirements=sanitize(request.form.get(f"new_{n}_storage_requirements", ""), 1000) or None,
                reorder_lead_time_days=int(_rlt) if _rlt.strip().isdigit() else None,
                min_order_qty=float(_moq) if _moq.strip() else None,
                min_order_unit=sanitize(request.form.get(f"new_{n}_min_order_unit", ""), 50) or None,
                safety_notes=sanitize(request.form.get(f"new_{n}_safety_notes", ""), 1000) or None,
                shelf_life_days=int(_sld) if _sld.strip().isdigit() else None,
                shelf_life_note=sanitize(request.form.get(f"new_{n}_shelf_life_note", ""), 500) or None,
            ))
        except Exception as new_item_err:
            app.logger.warning(f"New inventory item error: {new_item_err}")

    db.session.commit()
    flash("Inventory batch updated. / تم تحديث المخزون.", "success")
    return redirect(url_for("inventory_batch_view", batch_id=batch_id))


@app.route("/dashboard/inventory/batch/<batch_id>/download/excel")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_batch_download_excel(batch_id):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    items = InventoryItem.query.filter_by(user_id=current_user.id, master_inventory_batch_id=batch_id).order_by(InventoryItem.category, InventoryItem.name).all()
    if not items:
        abort(404)

    project = items[0].project

    navy_fill = PatternFill("solid", fgColor="0A1628")
    gold_fill = PatternFill("solid", fgColor="E8C547")
    light_fill = PatternFill("solid", fgColor="F8F9FA")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # Cover sheet
    ws_cover = wb.active
    ws_cover.title = "Summary"
    ws_cover.merge_cells("A1:H1")
    ws_cover["A1"] = f"BanaaIQ Master Inventory — {project.name if project else 'N/A'}"
    ws_cover["A1"].fill = navy_fill
    ws_cover["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws_cover["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_cover.row_dimensions[1].height = 35
    ws_cover.merge_cells("A2:H2")
    ws_cover["A2"] = f"Batch: {batch_id[:8].upper()} | Generated: {datetime.now().strftime('%d %B %Y %H:%M')} | BanaaIQ"
    ws_cover["A2"].fill = gold_fill
    ws_cover["A2"].font = white_font
    ws_cover["A2"].alignment = Alignment(horizontal="center")

    total_sar = sum(float(i.value_sar or 0) * float(i.stock or 0) for i in items)
    ws_cover["A4"] = f"Total Items: {len(items)}"
    ws_cover["C4"] = f"Total Value: SAR {total_sar:,.0f}"
    ws_cover["E4"] = f"Project: {project.name if project else 'N/A'}"

    # Per-category sheets
    categories_seen = {}
    for item in items:
        cat = item.category or "General"
        if cat not in categories_seen:
            categories_seen[cat] = []
        categories_seen[cat].append(item)

    for cat_name, cat_items in categories_seen.items():
        ws = wb.create_sheet(title=cat_name[:31])
        ws.merge_cells("A1:J1")
        ws["A1"] = cat_name
        ws["A1"].fill = navy_fill
        ws["A1"].font = white_font
        ws["A1"].alignment = Alignment(horizontal="center")
        headers = ["#", "Item Name (EN)", "Item Name (AR)", "Category", "Unit", "Stock", "Threshold", "Value SAR", "Supplier", "Notes"]
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=hdr)
            cell.fill = gold_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for idx, item in enumerate(cat_items, 1):
            row_data = [idx, safe_excel_cell(item.name or ""), safe_excel_cell(item.name_ar or ""), safe_excel_cell(item.category or ""), safe_excel_cell(item.unit or ""), float(item.stock or 0), float(item.threshold or 0), float(item.value_sar or 0), safe_excel_cell(item.supplier or ""), safe_excel_cell(item.notes or "")]
            fill = light_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for col, val in enumerate(row_data, 1):
                c = ws.cell(row=idx+2, column=col, value=val)
                c.border = border
                c.fill = fill
                if col == 2:
                    c.font = Font(bold=True)
        widths = [4, 35, 30, 18, 10, 12, 12, 14, 25, 30]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    # Full Detail sheet — all procurement fields
    ws_det = wb.create_sheet(title="Inventory Full Detail")
    ws_det.merge_cells("A1:R1")
    ws_det["A1"] = f"BanaaIQ Master Inventory — Full Procurement Detail — {project.name if project else 'N/A'}"
    ws_det["A1"].fill = navy_fill
    ws_det["A1"].font = Font(color="FFFFFF", bold=True, size=12)
    ws_det["A1"].alignment = Alignment(horizontal="center")
    ws_det.row_dimensions[1].height = 28
    det_headers = [
        "#", "Name (EN)", "Name (AR)", "Category", "Unit",
        "Stock", "Threshold", "Unit SAR",
        "Specification", "Brand Suggestions",
        "Supplier Hint", "Storage Requirements",
        "Lead Time (days)", "Min Order Qty", "Min Order Unit",
        "Safety Notes", "Shelf Life (days)", "Shelf Life Note",
    ]
    for col, hdr in enumerate(det_headers, 1):
        cell = ws_det.cell(row=2, column=col, value=hdr)
        cell.fill = gold_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    ws_det.row_dimensions[2].height = 30
    for idx, item in enumerate(items, 1):
        brands_str = ""
        if item.brand_suggestions_json:
            try:
                brands_str = ", ".join(json.loads(item.brand_suggestions_json))
            except Exception:
                brands_str = item.brand_suggestions_json
        det_row = [
            idx,
            safe_excel_cell(item.name or ""),
            safe_excel_cell(item.name_ar or ""),
            safe_excel_cell(item.category or ""),
            safe_excel_cell(item.unit or ""),
            float(item.stock or 0),
            float(item.threshold or 0),
            float(item.value_sar or 0),
            safe_excel_cell(item.specification or ""),
            safe_excel_cell(brands_str),
            safe_excel_cell(item.supplier or ""),
            safe_excel_cell(item.storage_requirements or ""),
            item.reorder_lead_time_days or "",
            float(item.min_order_qty) if item.min_order_qty else "",
            safe_excel_cell(item.min_order_unit or ""),
            safe_excel_cell(item.safety_notes or ""),
            item.shelf_life_days or "",
            safe_excel_cell(item.shelf_life_note or ""),
        ]
        fill = light_fill if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(det_row, 1):
            c = ws_det.cell(row=idx + 2, column=col, value=val)
            c.border = border
            c.fill = fill
            if col in (9, 12, 16):  # spec, storage, safety — wrap
                c.alignment = Alignment(wrap_text=True, vertical="top")
    det_widths = [4, 30, 24, 18, 10, 10, 10, 10, 45, 40, 30, 45, 12, 12, 18, 45, 12, 35]
    for idx, w in enumerate(det_widths, 1):
        ws_det.column_dimensions[get_column_letter(idx)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
        download_name=f"BanaaIQ_Inventory_{batch_id[:8]}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/dashboard/inventory/batch/<batch_id>/download/pdf")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_batch_download_pdf(batch_id):
    items = InventoryItem.query.filter_by(user_id=current_user.id, master_inventory_batch_id=batch_id).order_by(InventoryItem.category, InventoryItem.name).all()
    if not items:
        abort(404)

    project = items[0].project
    navy = colors.HexColor("#0a0a0a")
    gold = colors.HexColor("#e8c547")
    light = colors.HexColor("#f6f3ec")
    has_arabic = setup_arabic_font()
    arabic_font = "Cairo" if has_arabic else "Helvetica"
    has_hindi = setup_hindi_font()
    hindi_font = "NotoDevanagari" if has_hindi else "Helvetica"
    show_hindi_col = has_hindi and any(getattr(i, "name_hi", None) for i in items)

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = []

    # Cover header
    hdr_data = [[
        Paragraph(f'<b><font color="white" size="14">BanaaIQ | Master Inventory</font></b>', ps("H", fontName="Helvetica-Bold", fontSize=14, textColor=colors.white)),
        Paragraph(
            f'<font color="#e8c547" size="9"><b>BATCH #{batch_id[:8].upper()}</b></font><br/>'
            f'<font color="#aaaaaa" size="8">{project.name if project else "No Project"}<br/>{datetime.now().strftime("%d %B %Y")}</font>',
            ps("HR", fontName="Helvetica", fontSize=9, textColor=colors.white, alignment=TA_RIGHT),
        ),
    ]]
    ht = Table(hdr_data, colWidths=[11*cm, 6*cm])
    ht.setStyle(TableStyle([("BACKGROUND", (0,0),(-1,-1), navy), ("PADDING",(0,0),(-1,-1),14), ("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story.append(ht)
    story.append(Spacer(1, 0.3*cm))

    total_sar = sum(float(i.value_sar or 0) * float(i.stock or 0) for i in items)
    summary_data = [[
        Paragraph(f"<b>{len(items)}</b><br/><font size='8' color='#666666'>Total Items</font>", ps("S1", fontName="Helvetica-Bold", fontSize=16, alignment=TA_CENTER)),
        Paragraph(f"<b>SAR {total_sar:,.0f}</b><br/><font size='8' color='#666666'>Total Value</font>", ps("S2", fontName="Helvetica-Bold", fontSize=11, alignment=TA_CENTER)),
        Paragraph(f"<b>{datetime.now().strftime('%d %b %Y')}</b><br/><font size='8' color='#666666'>Generated</font>", ps("S3", fontName="Helvetica-Bold", fontSize=11, alignment=TA_CENTER)),
    ]]
    st = Table(summary_data, colWidths=[5.67*cm]*3)
    st.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),light), ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#e4ddd0")), ("PADDING",(0,0),(-1,-1),10), ("VALIGN",(0,0),(-1,-1),"MIDDLE"), ("ALIGN",(0,0),(-1,-1),"CENTER")]))
    story.append(st)
    story.append(Spacer(1, 0.3*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=gold))
    story.append(Spacer(1, 0.2*cm))

    # Per-category tables
    categories_seen = {}
    for item in items:
        cat = item.category or "General"
        if cat not in categories_seen:
            categories_seen[cat] = []
        categories_seen[cat].append(item)

    for cat_name, cat_items in categories_seen.items():
        story.append(Paragraph(cat_name, ps(f"CAT_{cat_name[:10]}", fontName="Helvetica-Bold", fontSize=11, textColor=navy, spaceAfter=4, spaceBefore=8)))
        cat_sar = sum(float(i.value_sar or 0) * float(i.stock or 0) for i in cat_items)
        story.append(Paragraph(f"<font size='8' color='#666666'>{len(cat_items)} items — SAR {cat_sar:,.0f}</font>", ps(f"CATSUB_{cat_name[:8]}", fontName="Helvetica", fontSize=8, spaceAfter=4)))

        def lbl(t):
            return Paragraph(f"<b>{t}</b>", ps("L", fontName="Helvetica-Bold", fontSize=7, textColor=colors.white))
        def cell(t, bold=False, clr=None):
            return Paragraph(str(t)[:60], ps("C", fontName="Helvetica-Bold" if bold else "Helvetica", fontSize=8, textColor=clr or colors.HexColor("#1a1a1a")))
        def cell_ar(t):
            text = process_arabic_text(str(t)) if t else ""
            return Paragraph(text[:80], ps("CAR", fontName=arabic_font, fontSize=8, alignment=TA_RIGHT, textColor=colors.HexColor("#1a1a1a")))
        def cell_hi(t):
            return Paragraph((str(t) or "")[:60], ps("CHI", fontName=hindi_font, fontSize=8, textColor=colors.HexColor("#1a1a1a")))

        hdr_row = [lbl("#"), lbl("Item Name"), lbl("Arabic"), lbl("Unit"), lbl("Stock"), lbl("Threshold"), lbl("Unit SAR")]
        col_widths = [1*cm, 4*cm, 3.5*cm, 2*cm, 2*cm, 2*cm, 2*cm]
        if show_hindi_col:
            hdr_row.insert(3, lbl("Hindi / हिंदी"))
            col_widths.insert(3, 3*cm)
            col_widths[1] = 3.5*cm  # shrink EN name col to fit

        tbl = [hdr_row]
        for idx, item in enumerate(cat_items, 1):
            row = [
                cell(str(idx)),
                cell(item.name, bold=True),
                cell_ar(item.name_ar or ""),
                cell(item.unit or "—"),
                cell(f"{float(item.stock or 0):,.1f}"),
                cell(f"{float(item.threshold or 0):,.1f}"),
                cell(f"{float(item.value_sar or 0):,.0f}"),
            ]
            if show_hindi_col:
                row.insert(3, cell_hi(getattr(item, "name_hi", "") or ""))
            tbl.append(row)

        cat_table = Table(tbl, colWidths=col_widths, repeatRows=1)
        cat_table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),navy), ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#e4ddd0")),
            ("PADDING",(0,0),(-1,-1),5),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, light]),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        story.append(cat_table)
        story.append(Spacer(1, 0.2*cm))

    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e4ddd0")))
    story.append(Paragraph("Generated by BanaaIQ | Confidential", ps("FT", fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#999"), alignment=TA_CENTER)))

    # Appendix — full procurement detail per item (only for items with rich data)
    rich_items = [i for i in items if i.specification or i.brand_suggestions_json or i.safety_notes]
    if rich_items:
        from reportlab.platypus import PageBreak
        story.append(PageBreak())
        story.append(Paragraph(
            "Appendix — Item Procurement Details / ملحق: تفاصيل مشتريات كل بند",
            ps("APX_HDR", fontName="Helvetica-Bold", fontSize=13, textColor=navy, spaceAfter=6)
        ))
        story.append(HRFlowable(width="100%", thickness=2, color=gold))
        story.append(Spacer(1, 0.3*cm))

        for item in rich_items:
            # Item header
            story.append(Paragraph(
                f'<b><font color="#0a0a0a">{item.name}</font></b>'
                + (f' &nbsp;<font size="9" color="#666666">{item.category or ""}</font>' if item.category else ""),
                ps(f"IH_{item.id}", fontName="Helvetica-Bold", fontSize=10, spaceBefore=6, spaceAfter=2)
            ))
            if item.name_ar:
                ar_text = process_arabic_text(item.name_ar)
                story.append(Paragraph(
                    ar_text,
                    ps(f"IHAR_{item.id}", fontName=arabic_font, fontSize=9, alignment=TA_RIGHT,
                       textColor=colors.HexColor("#555555"), spaceAfter=2)
                ))
            if has_hindi and getattr(item, "name_hi", None):
                story.append(Paragraph(
                    str(item.name_hi),
                    ps(f"IHHI_{item.id}", fontName=hindi_font, fontSize=9,
                       textColor=colors.HexColor("#444444"), spaceAfter=4)
                ))

            rows = []
            def _lbl(t):
                return Paragraph(f"<b><font size='8'>{t}</font></b>",
                                 ps("DL", fontName="Helvetica-Bold", fontSize=8, textColor=navy))
            def _val(t, wrap=True):
                return Paragraph(str(t)[:300],
                                 ps("DV", fontName="Helvetica", fontSize=8,
                                    textColor=colors.HexColor("#1a1a1a"), leading=11))

            if item.specification:
                rows.append([_lbl("Specification"), _val(item.specification)])
            if item.brand_suggestions_json:
                try:
                    brands = ", ".join(json.loads(item.brand_suggestions_json))
                except Exception:
                    brands = item.brand_suggestions_json
                rows.append([_lbl("Brand Suggestions"), _val(brands)])
            if item.supplier:
                rows.append([_lbl("Supplier Hint"), _val(item.supplier)])
            if item.storage_requirements:
                rows.append([_lbl("Storage Requirements"), _val(item.storage_requirements)])
            if item.reorder_lead_time_days or item.min_order_qty:
                proc_text = ""
                if item.reorder_lead_time_days:
                    proc_text += f"Lead time: {item.reorder_lead_time_days} days  "
                if item.min_order_qty:
                    proc_text += f"Min order: {float(item.min_order_qty):,.0f} {item.min_order_unit or ''}"
                rows.append([_lbl("Procurement"), _val(proc_text)])
            if item.safety_notes:
                rows.append([_lbl("Safety Notes"), _val(item.safety_notes)])
            if item.shelf_life_days:
                sl_text = f"{item.shelf_life_days} days"
                if item.shelf_life_note:
                    sl_text += f" — {item.shelf_life_note}"
                rows.append([_lbl("Shelf Life"), _val(sl_text)])
            if item.notes:
                rows.append([_lbl("Notes"), _val(item.notes)])

            if rows:
                det_table = Table(rows, colWidths=[4*cm, 13*cm])
                det_table.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e4ddd0")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f6f3ec")),
                    ("PADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(det_table)
            story.append(Spacer(1, 0.25*cm))

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
        download_name=f"BanaaIQ_Inventory_{batch_id[:8]}.pdf",
        mimetype="application/pdf")


@app.route("/dashboard/inventory/batch/<batch_id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_batch_delete(batch_id):
    items = InventoryItem.query.filter_by(user_id=current_user.id, master_inventory_batch_id=batch_id).all()
    if not items:
        abort(404)
    item_ids = [i.id for i in items]
    InventoryAssignment.query.filter(InventoryAssignment.inventory_item_id.in_(item_ids)).delete(synchronize_session=False)
    UsageLog.query.filter(UsageLog.item_id.in_(item_ids)).delete(synchronize_session=False)
    InventoryItem.query.filter(InventoryItem.id.in_(item_ids)).delete(synchronize_session=False)
    db.session.commit()
    flash("Inventory batch deleted. / تم حذف دفعة المخزون.", "success")
    return redirect(url_for("inventory_index"))


@app.route("/dashboard/inventory/batch/<batch_id>/distribute", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_batch_distribute(batch_id):
    username, company = dashboard_identity()
    items = InventoryItem.query.filter_by(user_id=current_user.id, master_inventory_batch_id=batch_id).order_by(InventoryItem.category, InventoryItem.name).all()
    if not items:
        abort(404)

    project = items[0].project
    project_id = items[0].project_id

    # Get engineers on this project
    engineers = []
    if project_id:
        assignments = ProjectAssignment.query.filter_by(project_id=project_id).all()
        engineer_user_ids = [a.user_id for a in assignments]
        engineers = User.query.filter(
            User.id.in_(engineer_user_ids),
            User.role == ROLE_SITE_ENGINEER
        ).all() if engineer_user_ids else []
        # Also include engineers from EngineerPackage
        pkg_user_ids = [ep.assigned_user_id for ep in EngineerPackage.query.filter_by(project_id=project_id).all() if ep.assigned_user_id]
        extra_engineers = User.query.filter(
            User.id.in_(pkg_user_ids),
            User.role == ROLE_SITE_ENGINEER,
            User.id.notin_([e.id for e in engineers])
        ).all() if pkg_user_ids else []
        engineers = engineers + extra_engineers

    if request.method == "GET":
        categories = {}
        for item in items:
            cat = item.category or "General"
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(item)
        # Existing assignments
        existing = {a.inventory_item_id: a for a in InventoryAssignment.query.filter(
            InventoryAssignment.inventory_item_id.in_([i.id for i in items])
        ).all()}
        return render_template('dashboard/inventory/distribute.html',
            username=username, company=company,
            active_dashboard='inventory',
            batch_id=batch_id,
            categories=categories,
            project=project,
            engineers=engineers,
            existing_assignments=existing)

    # POST: save distribution
    item_ids = request.form.getlist("item_id")
    engineer_ids = request.form.getlist("engineer_id")
    allocated_qtys = request.form.getlist("allocated_qty")

    allowed_engineer_ids = {engineer.id for engineer in engineers}
    notified_engineers = set()
    saved = 0
    for item_id_str, eng_id_str, qty_str in zip(item_ids, engineer_ids, allocated_qtys):
        if not eng_id_str or not item_id_str:
            continue
        try:
            item_id = int(item_id_str)
            eng_id = int(eng_id_str)
            qty = float(qty_str or 0)
        except (ValueError, TypeError):
            continue
        if qty <= 0:
            continue

        # Verify item belongs to this batch and user
        item = next((i for i in items if i.id == item_id), None)
        if not item:
            continue
        if eng_id not in allowed_engineer_ids:
            continue
        if qty > float(item.stock or 0):
            continue

        InventoryAssignment.query.filter(
            InventoryAssignment.project_id == (project_id or 0),
            InventoryAssignment.inventory_item_id == item_id,
            InventoryAssignment.assigned_user_id != eng_id,
        ).delete(synchronize_session=False)

        # Upsert InventoryAssignment
        existing_a = InventoryAssignment.query.filter_by(
            project_id=project_id or 0,
            inventory_item_id=item_id,
            assigned_user_id=eng_id
        ).first()
        if existing_a:
            existing_a.allocated_qty = qty
            existing_a.status = 'allocated'
        else:
            new_a = InventoryAssignment(
                project_id=project_id or 0,
                inventory_item_id=item_id,
                assigned_user_id=eng_id,
                allocated_qty=qty,
                status='allocated',
            )
            db.session.add(new_a)
        notified_engineers.add(eng_id)
        saved += 1

    db.session.flush()

    # Notify engineers
    for eng_id in notified_engineers:
        notif = Notification(
            user_id=eng_id,
            message=f"{current_user.full_name} allocated inventory items to you on project {project.name if project else 'N/A'}.",
            link=url_for('inventory_index'),
        )
        db.session.add(notif)

    db.session.commit()
    flash(f"Distribution saved: {saved} assignments. {len(notified_engineers)} engineers notified. / تم الحفظ.", "success")
    return redirect(url_for("inventory_batch_view", batch_id=batch_id))


@app.route("/dashboard/inventory/item/<int:item_id>")
@dashboard_access_required
def inventory_item_detail(item_id):
    username, company = dashboard_identity()
    item = db.session.get(InventoryItem, item_id) or abort(404)
    # Check access: item owner or assigned engineer
    is_owner = (item.user_id == current_user.id)
    assignment = InventoryAssignment.query.filter_by(
        inventory_item_id=item_id, assigned_user_id=current_user.id
    ).first()
    if not is_owner and not assignment:
        abort(403)
    usage_logs = UsageLog.query.filter_by(item_id=item_id).order_by(UsageLog.created_at.desc()).limit(20).all()
    return render_template('dashboard/inventory/item_detail.html',
        username=username, company=company,
        active_dashboard='inventory',
        item=item, assignment=assignment, usage_logs=usage_logs)


@app.route("/dashboard/inventory/item/<int:item_id>/consume", methods=["POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
@limiter.limit("60 per hour")
def inventory_item_consume(item_id):
    item = db.session.get(InventoryItem, item_id) or abort(404)
    assignment = InventoryAssignment.query.filter_by(
        inventory_item_id=item_id, assigned_user_id=current_user.id
    ).first()
    if not assignment:
        abort(403)

    try:
        qty = float(request.form.get("qty_used", 0) or 0)
    except (ValueError, TypeError):
        qty = 0.0

    if qty <= 0:
        flash("Quantity must be greater than zero. / الكمية يجب أن تكون أكبر من صفر.", "danger")
        return redirect(url_for("inventory_item_detail", item_id=item_id))

    current_stock = float(item.stock or 0)
    if qty > current_stock:
        flash(f"Cannot consume {qty} — only {current_stock:.2f} in stock. / لا يمكن استهلاك أكثر من المتوفر.", "danger")
        return redirect(url_for("inventory_item_detail", item_id=item_id))

    zone = sanitize(request.form.get("zone", ""), 100)
    notes = sanitize(request.form.get("notes", ""), 500)

    # Create usage log
    log = UsageLog(
        user_id=current_user.id,
        item_id=item_id,
        item_name=item.name,
        quantity_used=qty,
        unit=item.unit,
        used_by=current_user.full_name,
        zone=zone,
        project_id=item.project_id,
        notes=notes,
    )
    db.session.add(log)

    # Decrement stock
    item.stock = current_stock - qty
    new_stock = float(item.stock)
    threshold = float(item.threshold or 0)

    db.session.flush()

    # Send alert notifications to PM if stock is critical or low
    if item.user_id != current_user.id:
        if threshold > 0 and new_stock < threshold * 0.25:
            notif = Notification(
                user_id=item.user_id,
                message=f"CRITICAL: Stock of '{item.name}' is critically low ({new_stock:.1f} {item.unit or ''}). Engineer: {current_user.full_name}.",
                link=url_for('inventory_index'),
            )
            db.session.add(notif)
        elif threshold > 0 and new_stock < threshold:
            notif = Notification(
                user_id=item.user_id,
                message=f"LOW STOCK: '{item.name}' is below threshold ({new_stock:.1f} {item.unit or ''}). Engineer: {current_user.full_name}.",
                link=url_for('inventory_index'),
            )
            db.session.add(notif)

    db.session.commit()
    flash(f"Logged {qty} {item.unit or ''} consumed. Remaining: {new_stock:.2f}. / تم تسجيل الاستهلاك.", "success")
    return redirect(url_for("inventory_item_detail", item_id=item_id))


@app.route("/dashboard/inventory/item/<int:item_id>/request-more")
@dashboard_access_required
def inventory_item_request_more(item_id):
    return redirect(url_for("inventory_request", item_id=item_id))


@app.route("/dashboard/inventory/request", methods=["GET", "POST"])
@dashboard_access_required
def inventory_request():
    username, company = dashboard_identity()
    item_id = request.args.get("item_id", type=int)

    # Get projects accessible to this user
    if current_user.role == ROLE_SITE_ENGINEER:
        assigned_project_ids = {a.project_id for a in ProjectAssignment.query.filter_by(user_id=current_user.id).all()}
        package_project_ids = {
            p.project_id for p in EngineerPackage.query.filter_by(assigned_user_id=current_user.id).all()
            if p.project_id
        }
        project_ids = sorted(assigned_project_ids | package_project_ids)
        projects = Project.query.filter(Project.id.in_(project_ids)).all() if project_ids else []
    else:
        projects = Project.query.filter_by(user_id=current_user.id).all()

    prefill_item = None
    if item_id:
        prefill_item = db.session.get(InventoryItem, item_id)

    if request.method == "GET":
        return render_template('dashboard/inventory/request.html',
            username=username, company=company,
            active_dashboard='inventory',
            projects=projects,
            prefill_item=prefill_item)

    # POST: create stock request
    project_id = request.form.get("project_id", type=int)
    inv_item_id = request.form.get("inventory_item_id", type=int)
    proposed_name = sanitize(request.form.get("proposed_item_name", ""), 200)
    proposed_name_ar = sanitize(request.form.get("proposed_item_name_ar", ""), 200)
    description = sanitize(request.form.get("description", ""), 2000)
    unit = sanitize(request.form.get("unit", ""), 20)
    urgency = sanitize(request.form.get("urgency", "normal"), 20)
    if urgency not in ("normal", "urgent", "critical"):
        urgency = "normal"
    preferred_brand = sanitize(request.form.get("preferred_brand", ""), 200)
    specification_requested = sanitize(request.form.get("specification_requested", ""), 1000)
    try:
        requested_qty = float(request.form.get("requested_qty", 0) or 0)
    except (ValueError, TypeError):
        requested_qty = 0.0

    if not project_id or requested_qty <= 0:
        flash("Project and quantity are required. / المشروع والكمية مطلوبان.", "danger")
        return redirect(url_for("inventory_request", item_id=item_id or ""))

    # Verify project access
    if current_user.role == ROLE_SITE_ENGINEER:
        pa = ProjectAssignment.query.filter_by(project_id=project_id, user_id=current_user.id).first()
        pkg = EngineerPackage.query.filter_by(project_id=project_id, assigned_user_id=current_user.id).first()
        if not pa and not pkg:
            abort(403)
    else:
        project = Project.query.filter_by(id=project_id, user_id=current_user.id).first()
        if not project:
            abort(403)

    sr = StockRequest(
        project_id=project_id,
        inventory_item_id=inv_item_id,
        requested_by_user_id=current_user.id,
        requested_qty=requested_qty,
        unit=unit,
        proposed_item_name=proposed_name,
        proposed_item_name_ar=proposed_name_ar,
        description=description,
        status='pending',
        urgency=urgency,
        preferred_brand=preferred_brand or None,
        specification_requested=specification_requested or None,
    )
    db.session.add(sr)
    db.session.flush()

    # Notify PM
    project_obj = db.session.get(Project, project_id)
    if project_obj and project_obj.user_id != current_user.id:
        notif = Notification(
            user_id=project_obj.user_id,
            message=f"New stock request from {current_user.full_name} on project '{project_obj.name}': {proposed_name or (prefill_item.name if prefill_item else 'unknown item')} × {requested_qty}.",
            link=url_for('inventory_requests'),
        )
        db.session.add(notif)

    db.session.commit()
    flash("Stock request submitted. The PM will review it. / تم إرسال طلب المخزون.", "success")
    return redirect(url_for("inventory_index"))


@app.route("/dashboard/inventory/requests")
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_requests():
    username, company = dashboard_identity()
    pm_project_ids = [p.id for p in Project.query.filter_by(user_id=current_user.id).all()]
    if not pm_project_ids:
        all_requests = []
    else:
        all_requests = StockRequest.query.filter(
            StockRequest.project_id.in_(pm_project_ids)
        ).order_by(StockRequest.created_at.desc()).all()

    pending = [r for r in all_requests if r.status == 'pending']
    approved = [r for r in all_requests if r.status == 'approved']
    rejected = [r for r in all_requests if r.status == 'rejected']

    return render_template('dashboard/inventory/requests.html',
        username=username, company=company,
        active_dashboard='inventory',
        pending=pending, approved=approved, rejected=rejected)


@app.route("/dashboard/inventory/requests/<int:request_id>/approve", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_request_approve(request_id):
    sr = db.session.get(StockRequest, request_id) or abort(404)
    # Verify PM owns the project
    project = Project.query.filter_by(id=sr.project_id, user_id=current_user.id).first()
    if not project:
        abort(403)

    try:
        approved_qty = float(request.form.get("approved_qty") or sr.requested_qty)
    except (ValueError, TypeError):
        approved_qty = float(sr.requested_qty)

    review_notes = sanitize(request.form.get("review_notes", ""), 500)

    sr.status = 'approved'
    sr.approved_qty = approved_qty
    sr.reviewed_by_user_id = current_user.id
    sr.reviewed_at = datetime.utcnow()
    sr.review_notes = review_notes

    # Update or create inventory item stock
    if sr.inventory_item_id:
        item = db.session.get(InventoryItem, sr.inventory_item_id)
        if item:
            item.stock = float(item.stock or 0) + approved_qty
    else:
        # Create new inventory item for this project
        new_item = InventoryItem(
            user_id=current_user.id,
            project_id=sr.project_id,
            name=sr.proposed_item_name or "Stock Request Item",
            name_ar=sr.proposed_item_name_ar or "",
            unit=sr.unit or "pcs",
            stock=approved_qty,
            threshold=0,
            source='engineer_added',
        )
        db.session.add(new_item)
        db.session.flush()
        # Assign to requesting engineer
        new_assignment = InventoryAssignment(
            project_id=sr.project_id,
            inventory_item_id=new_item.id,
            assigned_user_id=sr.requested_by_user_id,
            allocated_qty=approved_qty,
            status='allocated',
        )
        db.session.add(new_assignment)

    # Notify engineer
    notif = Notification(
        user_id=sr.requested_by_user_id,
        message=f"Your stock request for '{sr.proposed_item_name or 'item'}' was APPROVED: {approved_qty} {sr.unit or ''}. {review_notes or ''}",
        link=url_for('inventory_index'),
    )
    db.session.add(notif)
    db.session.commit()
    flash("Stock request approved. / تمت الموافقة على الطلب.", "success")
    return redirect(url_for("inventory_requests"))


@app.route("/dashboard/inventory/requests/<int:request_id>/reject", methods=["POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_request_reject(request_id):
    sr = db.session.get(StockRequest, request_id) or abort(404)
    project = Project.query.filter_by(id=sr.project_id, user_id=current_user.id).first()
    if not project:
        abort(403)

    review_notes = sanitize(request.form.get("review_notes", ""), 500)
    sr.status = 'rejected'
    sr.reviewed_by_user_id = current_user.id
    sr.reviewed_at = datetime.utcnow()
    sr.review_notes = review_notes

    # Notify engineer
    notif = Notification(
        user_id=sr.requested_by_user_id,
        message=f"Your stock request for '{sr.proposed_item_name or 'item'}' was REJECTED. Notes: {review_notes or 'No notes provided.'}",
        link=url_for('inventory_index'),
    )
    db.session.add(notif)
    db.session.commit()
    flash("Stock request rejected. / تم رفض الطلب.", "info")
    return redirect(url_for("inventory_requests"))


@app.route("/dashboard/inventory/requests/<int:request_id>/delete", methods=["POST"])
@login_required
@role_required(ROLE_SITE_ENGINEER)
def inventory_request_delete(request_id):
    sr = db.session.get(StockRequest, request_id) or abort(404)
    if sr.requested_by_user_id != current_user.id:
        abort(403)
    if sr.status != 'pending':
        flash("Only pending requests can be deleted. / لا يمكن حذف الطلبات التي تمت مراجعتها.", "error")
        return redirect(url_for("inventory_index"))
    db.session.delete(sr)
    db.session.commit()
    flash("Stock request deleted. / تم حذف الطلب.", "success")
    return redirect(url_for("inventory_index"))


# ── INVENTORY MANUAL CREATE ─────────────────────────────────────────────────

@app.route("/dashboard/inventory/create-manual", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
def inventory_create_manual():
    username, company = dashboard_identity()
    projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()

    if request.method == "GET":
        return render_template("dashboard/inventory/create_manual.html",
            username=username, company=company,
            active_dashboard="inventory", projects=projects)

    try:
        project_id = request.form.get("project_id", type=int)
        if not project_id:
            flash("Project is required. / المشروع مطلوب.", "danger")
            return redirect(url_for("inventory_create_manual"))

        project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()

        items_json_str = request.form.get("items_json", "")
        if not items_json_str:
            flash("Please add at least one item. / يرجى إضافة عنصر واحد على الأقل.", "danger")
            return redirect(url_for("inventory_create_manual"))

        items_data = json.loads(items_json_str)
        if not items_data:
            flash("Please add at least one item.", "danger")
            return redirect(url_for("inventory_create_manual"))

        batch_id = uuid.uuid4().hex[:32]
        saved_count = 0
        for item_data in items_data:
            name = sanitize(str(item_data.get("name", "")).strip(), 200)
            if not name:
                continue
            try:
                _brands = item_data.get("brand_suggestions")
                _alts = item_data.get("alternative_items")
                inv = InventoryItem(
                    user_id=current_user.id,
                    project_id=project_id,
                    name=name,
                    name_ar=sanitize(str(item_data.get("name_ar", "")), 200),
                    category=sanitize(str(item_data.get("category", "General")), 50),
                    category_ar=sanitize(str(item_data.get("category_ar", "")), 80),
                    unit=sanitize(str(item_data.get("unit", "pcs")), 30),
                    stock=float(item_data.get("stock", 0) or 0),
                    threshold=float(item_data.get("threshold", 0) or 0),
                    value_sar=float(item_data.get("value_sar", 0) or 0),
                    supplier=sanitize(str(item_data.get("supplier", "")), 100),
                    notes=sanitize(str(item_data.get("notes", "")), 500),
                    source="manual",
                    master_inventory_batch_id=batch_id,
                    specification=sanitize(str(item_data.get("specification", "") or ""), 2000) or None,
                    brand_suggestions_json=json.dumps(_brands) if isinstance(_brands, list) else (
                        json.dumps([s.strip() for s in item_data["brand_suggestions"].split(",") if s.strip()])
                        if isinstance(item_data.get("brand_suggestions"), str) and item_data["brand_suggestions"].strip()
                        else None
                    ),
                    storage_requirements=sanitize(str(item_data.get("storage_requirements", "") or ""), 1000) or None,
                    reorder_lead_time_days=int(item_data["reorder_lead_time_days"]) if item_data.get("reorder_lead_time_days") else None,
                    min_order_qty=float(item_data["min_order_qty"]) if item_data.get("min_order_qty") else None,
                    min_order_unit=sanitize(str(item_data.get("min_order_unit", "") or ""), 50) or None,
                    safety_notes=sanitize(str(item_data.get("safety_notes", "") or ""), 1000) or None,
                    alternative_items_json=json.dumps(_alts) if isinstance(_alts, list) else None,
                    shelf_life_days=int(item_data["shelf_life_days"]) if item_data.get("shelf_life_days") else None,
                    shelf_life_note=sanitize(str(item_data.get("shelf_life_note", "") or ""), 500) or None,
                )
                db.session.add(inv)
                saved_count += 1
            except Exception as item_err:
                app.logger.warning(f"Manual inv item error: {item_err}")

        if saved_count == 0:
            flash("No valid items. Please add at least one item with a name.", "danger")
            return redirect(url_for("inventory_create_manual"))

        db.session.commit()
        flash(f"Manual inventory created: {saved_count} items. / تم إنشاء المخزون يدوياً.", "success")
        return redirect(url_for("inventory_batch_view", batch_id=batch_id))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Inventory create manual error: {e}")
        flash("Save failed. Please try again. / فشل الحفظ.", "danger")
        return redirect(url_for("inventory_create_manual"))


# ── INVENTORY UPLOAD + AI REVISE ─────────────────────────────────────────────

@app.route("/dashboard/inventory/upload-revise", methods=["GET", "POST"])
@login_required
@role_required(ROLE_PROJECT_MANAGER)
@limiter.limit("10 per hour", methods=["POST"])
def inventory_upload_revise():
    username, company = dashboard_identity()
    projects = Project.query.filter_by(user_id=current_user.id).order_by(Project.name).all()

    if request.method == "GET":
        return render_template("dashboard/inventory/upload_revise.html",
            username=username, company=company,
            active_dashboard="inventory", projects=projects)

    try:
        ok, msg = can_use_ai(current_user)
        if not ok:
            flash(f"AI quota: {msg}", "warning")
            return redirect(url_for("inventory_upload_revise"))

        project_id = request.form.get("project_id", type=int)
        title = sanitize(request.form.get("title", "").strip(), 200)
        revision_description = sanitize(request.form.get("revision_description", "").strip(), max_length=3000)

        if not project_id or not title:
            flash("Project and title are required. / المشروع والعنوان مطلوبان.", "danger")
            return redirect(url_for("inventory_upload_revise"))

        project = Project.query.filter_by(id=project_id, user_id=current_user.id).first_or_404()

        f = request.files.get("inventory_file")
        if not f or not f.filename:
            flash("Please upload an inventory file (.xlsx, .xls, or .pdf).", "danger")
            return redirect(url_for("inventory_upload_revise"))

        filename = secure_filename(f.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".xlsx", ".xls", ".pdf"]:
            flash("Only .xlsx, .xls, and .pdf files are supported.", "danger")
            return redirect(url_for("inventory_upload_revise"))

        file_data = f.read()
        if len(file_data) > 10 * 1024 * 1024:
            flash("File exceeds 10MB limit.", "danger")
            return redirect(url_for("inventory_upload_revise"))

        upload_dir = os.path.join(os.path.dirname(__file__), "uploads", "inventory_designs", str(current_user.id))
        os.makedirs(upload_dir, exist_ok=True)
        unique_name = f"{uuid.uuid4().hex}_{filename}"
        with open(os.path.join(upload_dir, unique_name), "wb") as out:
            out.write(file_data)

        extracted_text = ""
        try:
            if ext in [".xlsx", ".xls"]:
                wb_tmp = load_workbook(BytesIO(file_data), read_only=True)
                for ws_tmp in list(wb_tmp.worksheets)[:5]:
                    for row in ws_tmp.iter_rows(max_row=200, values_only=True):
                        extracted_text += " ".join(str(c) for c in row if c is not None) + "\n"
            elif ext == ".pdf":
                import pdfplumber
                with pdfplumber.open(BytesIO(file_data)) as pdf:
                    for page in pdf.pages[:10]:
                        extracted_text += page.extract_text() or ""
        except Exception as parse_err:
            app.logger.warning(f"Inventory file parse error: {parse_err}")
            flash("Could not read the uploaded file. Ensure it is a valid Excel or PDF.", "danger")
            return redirect(url_for("inventory_upload_revise"))

        if not extracted_text.strip():
            flash("The uploaded file appears empty or has no extractable text.", "danger")
            return redirect(url_for("inventory_upload_revise"))

        # STEP A — structure file as v1 (preserve data + add rich fields where inferable)
        structure_sys = (
            "You are a senior Saudi construction site materials manager. "
            "Parse the following raw extracted text from an inventory file into a structured JSON. "
            "Group items into categories (detect naturally). "
            "Include bilingual EN+AR names — translate if only one language present. "
            "Preserve existing quantities and prices exactly. "
            "Add specification, brand_suggestions, storage_requirements, reorder_lead_time_days, "
            "min_order_qty, min_order_unit, safety_notes, alternative_items, shelf_life_days, "
            "shelf_life_note for each item where you can infer them from the item name. "
            "Return ONLY valid JSON, no markdown fences.\n"
        ) + _INVENTORY_ITEM_DETAIL_PROMPT + (
            "\n\nRequired top-level schema:\n"
            '{"categories_detected":["Cement & Concrete"],"categories":[{"category":"Cement & Concrete",'
            '"category_ar":"الإسمنت والخرسانة","items":[<item objects>]}],'
            '"total_items":1,"total_value_sar":11000}'
        )
        structure_user = f"Project: {project.name}.\n\nRaw inventory text:\n{extracted_text[:4000]}"

        v1_data, reason_v1, _ = _call_ai_for_master_inventory(structure_sys, structure_user, require_detail=False)

        if v1_data is None:
            app.logger.error(f"Inventory structure AI failed: {reason_v1}")
            flash("Could not parse the uploaded inventory file. Ensure it contains readable item data.", "danger")
            return redirect(url_for("inventory_upload_revise"))

        batch_id_v1 = uuid.uuid4().hex[:32]
        saved_v1 = 0
        for cat in v1_data.get("categories", []):
            cat_name = sanitize(str(cat.get("category", "General")), 50)
            cat_name_ar = sanitize(str(cat.get("category_ar", "")), 80)
            for item_data in cat.get("items", []):
                try:
                    _brands = item_data.get("brand_suggestions")
                    _alts = item_data.get("alternative_items")
                    db.session.add(InventoryItem(
                        user_id=current_user.id,
                        project_id=project_id,
                        name=sanitize(str(item_data.get("name", "Unnamed")), 200),
                        name_ar=sanitize(str(item_data.get("name_ar", "")), 200),
                        category=cat_name,
                        category_ar=cat_name_ar,
                        unit=sanitize(str(item_data.get("unit", "pcs")), 30),
                        stock=float(item_data.get("recommended_stock", 0) or 0),
                        threshold=float(item_data.get("threshold", 0) or 0),
                        value_sar=float(item_data.get("value_sar", 0) or 0),
                        supplier=sanitize(str(item_data.get("supplier_hint", "")), 100),
                        notes=sanitize(str(item_data.get("notes", "") or ""), 500),
                        source="uploaded_v1",
                        master_inventory_batch_id=batch_id_v1,
                        specification=sanitize(str(item_data.get("specification", "") or ""), 2000) or None,
                        brand_suggestions_json=json.dumps(_brands) if isinstance(_brands, list) else None,
                        storage_requirements=sanitize(str(item_data.get("storage_requirements", "") or ""), 1000) or None,
                        reorder_lead_time_days=int(item_data["reorder_lead_time_days"]) if item_data.get("reorder_lead_time_days") else None,
                        min_order_qty=float(item_data["min_order_qty"]) if item_data.get("min_order_qty") else None,
                        min_order_unit=sanitize(str(item_data.get("min_order_unit", "") or ""), 50) or None,
                        safety_notes=sanitize(str(item_data.get("safety_notes", "") or ""), 1000) or None,
                        alternative_items_json=json.dumps(_alts) if isinstance(_alts, list) else None,
                        shelf_life_days=int(item_data["shelf_life_days"]) if item_data.get("shelf_life_days") else None,
                        shelf_life_note=sanitize(str(item_data.get("shelf_life_note", "") or ""), 500) or None,
                    ))
                    saved_v1 += 1
                except Exception:
                    pass

        db.session.flush()
        record_ai_usage("inventory_master_generation")

        if not revision_description:
            db.session.commit()
            flash(f"Inventory uploaded as v1: {saved_v1} items. No revision requested. / تم رفع المخزون كـ v1.", "info")
            return redirect(url_for("inventory_batch_view", batch_id=batch_id_v1))

        # STEP B — apply revision as v2 (add full rich fields)
        revision_sys = (
            "You are a senior Saudi construction site materials manager revising an inventory list. "
            "Apply the user's requested changes. Keep all unrelated items intact. "
            "Use 2025 Saudi market rates for any new items. "
            "Enrich ALL items with specification, brand_suggestions, storage_requirements, "
            "reorder_lead_time_days, min_order_qty, min_order_unit, safety_notes, "
            "alternative_items, shelf_life_days, shelf_life_note. "
            "Return ONLY valid JSON, no markdown fences.\n"
        ) + _INVENTORY_ITEM_DETAIL_PROMPT + (
            "\n\nRequired top-level schema:\n"
            '{"categories_detected":["Cement & Concrete"],"categories":[{"category":"Cement & Concrete",'
            '"category_ar":"الإسمنت والخرسانة","items":[<item objects>]}],'
            '"total_items":1,"total_value_sar":11000}'
        )
        v1_summary = json.dumps(v1_data.get("categories", []), ensure_ascii=False)[:3000]
        revision_user = f"Current inventory:\n{v1_summary}\n\nRequested changes: {revision_description}"

        v2_data, reason_v2, _ = _call_ai_for_master_inventory(revision_sys, revision_user)

        if v2_data is None:
            db.session.commit()
            record_ai_usage("inventory_master_generation")
            app.logger.error(f"Inventory revision AI failed: {reason_v2}")
            flash(
                f"v1 saved ({saved_v1} items). AI revision failed — changes not applied. "
                f"You can edit items manually from the batch view. / "
                f"تم حفظ v1. فشل تطبيق التعديل بالذكاء الاصطناعي.",
                "warning"
            )
            return redirect(url_for("inventory_batch_view", batch_id=batch_id_v1))

        batch_id_v2 = uuid.uuid4().hex[:32]
        saved_v2 = 0
        for cat in v2_data.get("categories", []):
            cat_name = sanitize(str(cat.get("category", "General")), 50)
            cat_name_ar = sanitize(str(cat.get("category_ar", "")), 80)
            for item_data in cat.get("items", []):
                try:
                    _brands = item_data.get("brand_suggestions")
                    _alts = item_data.get("alternative_items")
                    db.session.add(InventoryItem(
                        user_id=current_user.id,
                        project_id=project_id,
                        name=sanitize(str(item_data.get("name", "Unnamed")), 200),
                        name_ar=sanitize(str(item_data.get("name_ar", "")), 200),
                        category=cat_name,
                        category_ar=cat_name_ar,
                        unit=sanitize(str(item_data.get("unit", "pcs")), 30),
                        stock=float(item_data.get("recommended_stock", 0) or 0),
                        threshold=float(item_data.get("threshold", 0) or 0),
                        value_sar=float(item_data.get("value_sar", 0) or 0),
                        supplier=sanitize(str(item_data.get("supplier_hint", "")), 100),
                        notes=sanitize(str(item_data.get("notes", "") or ""), 500),
                        source="uploaded_revised",
                        master_inventory_batch_id=batch_id_v2,
                        parent_batch_id=batch_id_v1,
                        specification=sanitize(str(item_data.get("specification", "") or ""), 2000) or None,
                        brand_suggestions_json=json.dumps(_brands) if isinstance(_brands, list) else None,
                        storage_requirements=sanitize(str(item_data.get("storage_requirements", "") or ""), 1000) or None,
                        reorder_lead_time_days=int(item_data["reorder_lead_time_days"]) if item_data.get("reorder_lead_time_days") else None,
                        min_order_qty=float(item_data["min_order_qty"]) if item_data.get("min_order_qty") else None,
                        min_order_unit=sanitize(str(item_data.get("min_order_unit", "") or ""), 50) or None,
                        safety_notes=sanitize(str(item_data.get("safety_notes", "") or ""), 1000) or None,
                        alternative_items_json=json.dumps(_alts) if isinstance(_alts, list) else None,
                        shelf_life_days=int(item_data["shelf_life_days"]) if item_data.get("shelf_life_days") else None,
                        shelf_life_note=sanitize(str(item_data.get("shelf_life_note", "") or ""), 500) or None,
                    ))
                    saved_v2 += 1
                except Exception:
                    pass

        db.session.commit()
        record_ai_usage("inventory_master_generation")
        flash(
            f"Inventory uploaded (v1: {saved_v1} items) and revised (v2: {saved_v2} items). "
            f"/ تم رفع المخزون كـ v1 ({saved_v1} عناصر) وتعديله كـ v2 ({saved_v2} عناصر).",
            "success"
        )
        return redirect(url_for("inventory_batch_view", batch_id=batch_id_v2))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Inventory upload revise error: {e}")
        traceback.print_exc()
        flash("An error occurred. Please try again. / حدث خطأ.", "danger")
        return redirect(url_for("inventory_upload_revise"))


# ── END INVENTORY REBUILD ROUTES ─────────────────────────────────────────────

@app.route("/dashboard/dpr")
@dashboard_access_required
def dashboard_dpr_index():
    username, company = dashboard_identity()
    show_demo_data = is_demo_mode()
    all_records = get_dpr_records()
    project_options = [p.to_dict() for p in get_feature_projects("dpr")] if not show_demo_data else []
    selected_project = request.args.get("project", "all" if not show_demo_data else "All Projects")
    selected_status = request.args.get("status", "All")
    selected_from = request.args.get("from", "")
    selected_to = request.args.get("to", "")
    try:
        from_date = datetime.strptime(selected_from, "%Y-%m-%d").date() if selected_from else None
    except ValueError:
        from_date = None
    try:
        to_date = datetime.strptime(selected_to, "%Y-%m-%d").date() if selected_to else None
    except ValueError:
        to_date = None
    filtered = []
    if show_demo_data:
        for row in all_records:
            if selected_project != "All Projects" and row["project"] != selected_project:
                continue
            if selected_status != "All" and row["status"] != selected_status:
                continue
            date_obj = parse_demo_date(row["date"])
            if from_date and date_obj and date_obj < from_date:
                continue
            if to_date and date_obj and date_obj > to_date:
                continue
            filtered.append(row)
    else:
        selected_project, selected_project_obj = resolve_feature_project("dpr", selected_project)
        for row in all_records:
            if selected_project_obj and row.get("feature_project_id") != selected_project_obj.id:
                continue
            if selected_status != "All" and row.get("status") != selected_status:
                continue
            date_obj = parse_demo_date(row.get("date"))
            if from_date and date_obj and date_obj < from_date:
                continue
            if to_date and date_obj and date_obj > to_date:
                continue
            filtered.append(row)
    selected_project_name = (
        selected_project_obj.name if not show_demo_data and selected_project != "all" and selected_project_obj else "All Projects"
    )
    return render_template(
        "dashboard/dpr/index.html",
        username=username,
        company=company,
        active_dashboard="dpr",
        show_demo_data=show_demo_data,
        dpr_records=filtered,
        dpr_projects=DPR_PROJECTS,
        selected_project=selected_project,
        selected_project_name=selected_project_name,
        selected_status=selected_status,
        selected_from=selected_from,
        selected_to=selected_to,
        total_count=len(all_records),
        filtered_count=len(filtered),
        projects=project_options,
    )


@app.route("/dashboard/dpr/new")
@dashboard_access_required
def dashboard_dpr_new():
    username, company = dashboard_identity()
    selected_project = request.args.get("project", "all")
    selected_project, selected_project_obj = resolve_feature_project("dpr", selected_project)
    dna_project = get_selected_project(request.args.get("project_id"))
    project_options = [p.to_dict() for p in get_feature_projects("dpr")]
    dpr_projects = [p["name"] for p in project_options] if project_options else DPR_PROJECTS
    if dna_project and dna_project.name not in dpr_projects:
        dpr_projects = [dna_project.name] + dpr_projects
    return render_template(
        "dashboard/dpr/form.html",
        username=username,
        company=company,
        active_dashboard="dpr",
        dpr_projects=dpr_projects,
        projects=project_options,
        selected_project=selected_project,
        selected_project_name=dna_project.name if dna_project else (selected_project_obj.name if selected_project_obj else "All Projects"),
        dna_project_id=dna_project.id if dna_project else "",
    )


@app.route("/dashboard/dpr/<int:id>/edit", methods=["GET"])
@dashboard_access_required
def dpr_edit(id):
    if session.get("is_guest"):
        return redirect(url_for("dashboard_dpr_index"))
    username, company = dashboard_identity()
    dpr_obj = DPR.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    dpr_data = dpr_obj.to_dict()
    dpr_data["date_input"] = format_html_date(dpr_obj.date)
    project_options = [p.to_dict() for p in get_feature_projects("dpr")]
    dpr_projects = [p["name"] for p in project_options] if project_options else DPR_PROJECTS
    return render_template(
        "dashboard/dpr/form.html",
        username=username,
        company=company,
        active_dashboard="dpr",
        dpr_projects=dpr_projects,
        projects=project_options,
        selected_project=str(dpr_obj.feature_project_id or "all"),
        selected_project_name=dpr_obj.project,
        dna_project_id=dpr_obj.project_id or "",
        dpr=dpr_data,
        is_edit=True,
        edit_id=id,
    )


@app.route("/dashboard/dpr/submit", methods=["POST"])
@dashboard_access_required
def dashboard_dpr_submit():
    if session.get("is_guest"):
        return jsonify({"error": "Sign up to save reports", "success": False})

    names = request.form.getlist("worker_name[]")
    roles = request.form.getlist("worker_role[]")
    hours_list = request.form.getlist("worker_hours[]")
    present_states = request.form.getlist("worker_present_state[]")
    workers_data = []
    for idx, name in enumerate(names):
        safe_name = sanitize_input(name, 100).strip()
        if not safe_name:
            continue
        safe_role = sanitize_input(roles[idx] if idx < len(roles) else "", 100)
        safe_hours = sanitize_input(hours_list[idx] if idx < len(hours_list) else "0", 10)
        try:
            hours = float(safe_hours)
        except ValueError:
            hours = 0
        present_value = present_states[idx] if idx < len(present_states) else "1"
        is_present = str(present_value).strip().lower() not in {"0", "false", "no", "n", ""}
        workers_data.append({"name": safe_name, "role": safe_role, "hours": hours, "present": is_present})

    report_date = request.form.get("report_date", "")
    report_date_value = parse_demo_date(report_date) or datetime.utcnow().date()
    progress_notes = sanitize_input(request.form.get("progress_notes", request.form.get("notes", "")), 2500)
    issues = sanitize_input(request.form.get("issues", ""), 1500)
    weather = sanitize_input(request.form.get("weather", "Clear"), 20)
    selected_project = sanitize_input(request.form.get("project", DPR_PROJECTS[0]), 120)
    dna_project = None
    try:
        dna_project_id = int(request.form.get("project_id") or 0)
    except (TypeError, ValueError):
        dna_project_id = 0
    if dna_project_id:
        dna_project = get_selected_project(dna_project_id)
        if not dna_project:
            return jsonify({"error": "Select a valid Project DNA workspace.", "success": False}), 400
        selected_project = dna_project.name
    project_obj = None
    if not is_demo_mode():
        project_obj = FeatureProject.query.filter_by(user_id=current_user.id, feature="dpr", name=selected_project).first()
        if not project_obj and not dna_project:
            return jsonify({"error": "Select a valid project before saving this DPR.", "success": False}), 400

    new_dpr = DPR(
        user_id=current_user.id,
        project_id=dna_project.id if dna_project else None,
        feature_project_id=project_obj.id if project_obj else None,
        date=report_date_value,
        project=selected_project,
        zone=sanitize_input(request.form.get("zone", ""), 180),
        weather=weather,
        temperature=sanitize_input(request.form.get("temperature", ""), 30),
        progress_notes=progress_notes,
        issues=issues,
        status="Draft" if request.form.get("save_as_draft") == "1" else "Completed",
        ai_summary=request.form.get("ai_summary", ""),
        ai_key_insight=request.form.get("ai_key_insight", ""),
    )
    new_dpr.workers = workers_data
    db.session.add(new_dpr)
    db.session.commit()
    boq_actual_count = record_actuals_from_dpr(new_dpr)
    if new_dpr.project_id:
        project = db.session.get(Project, new_dpr.project_id)
        if project:
            calculate_health_score(project)

    if request.headers.get("X-Requested-With") == "XMLHttpRequest" or "application/json" in (request.headers.get("Accept") or ""):
        return jsonify({
            "success": True,
            "id": new_dpr.id,
            "message": "DPR saved successfully",
            "redirect": f"/dashboard/dpr/{new_dpr.id}",
            "boq_tracker_updated": boq_actual_count > 0,
            "boq_actual_count": boq_actual_count,
            "boq_tracker_message": "BOQ tracker updated based on today's progress report" if boq_actual_count else "",
        })

    flash(f"DPR #{new_dpr.id} submitted successfully.", "success")
    if boq_actual_count:
        flash("BOQ tracker updated based on today's progress report", "success")
    return redirect(url_for("dashboard_dpr_view", id=new_dpr.id))


@app.route("/dashboard/dpr/<int:id>/update", methods=["POST"])
@dashboard_access_required
def dpr_update(id):
    if session.get("is_guest"):
        return jsonify({"error": "Sign up to edit", "success": False})
    dpr_obj = DPR.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    try:
        dna_project_id = int(request.form.get("project_id") or 0)
    except (TypeError, ValueError):
        dna_project_id = 0
    if dna_project_id:
        dna_project = get_selected_project(dna_project_id)
        if not dna_project:
            return jsonify({"error": "Select a valid Project DNA workspace.", "success": False}), 400
        dpr_obj.project_id = dna_project.id
        dpr_obj.project = dna_project.name
    report_date = request.form.get("report_date", "")
    parsed_report_date = parse_demo_date(report_date)
    if parsed_report_date:
        dpr_obj.date = parsed_report_date
    if not dna_project_id:
        dpr_obj.project = sanitize_input(request.form.get("project", dpr_obj.project), 200)
    dpr_obj.zone = sanitize_input(request.form.get("zone", dpr_obj.zone), 100)
    dpr_obj.weather = sanitize_input(request.form.get("weather", dpr_obj.weather), 50)
    dpr_obj.temperature = sanitize_input(request.form.get("temperature", dpr_obj.temperature), 20)
    dpr_obj.progress_notes = sanitize_input(request.form.get("progress_notes", dpr_obj.progress_notes), 3000)
    dpr_obj.issues = sanitize_input(request.form.get("issues", dpr_obj.issues), 1000)

    project_obj = FeatureProject.query.filter_by(user_id=current_user.id, feature="dpr", name=dpr_obj.project).first()
    dpr_obj.feature_project_id = project_obj.id if project_obj else None

    worker_names = request.form.getlist("worker_name[]")
    worker_roles = request.form.getlist("worker_role[]")
    worker_hours = request.form.getlist("worker_hours[]")
    worker_present_states = request.form.getlist("worker_present_state[]")
    workers = []
    for i, name in enumerate(worker_names):
        if not name.strip():
            continue
        try:
            hours = float(worker_hours[i]) if i < len(worker_hours) and worker_hours[i] else 8
        except ValueError:
            hours = 8
        present = True
        if i < len(worker_present_states):
            present = str(worker_present_states[i]).strip().lower() not in {"0", "false", "no", "n", ""}
        workers.append(
            {
                "name": sanitize_input(name.strip(), 100),
                "role": sanitize_input(worker_roles[i] if i < len(worker_roles) else "", 100),
                "hours": hours,
                "present": present,
            }
        )
    dpr_obj.workers = workers

    ai_summary = request.form.get("ai_summary")
    if ai_summary:
        dpr_obj.ai_summary = ai_summary
    dpr_obj.updated_at = datetime.utcnow()
    db.session.commit()
    boq_actual_count = record_actuals_from_dpr(dpr_obj)
    if dpr_obj.project_id:
        project = db.session.get(Project, dpr_obj.project_id)
        if project:
            calculate_health_score(project)
    return jsonify({
        "success": True,
        "id": dpr_obj.id,
        "message": "DPR updated successfully",
        "redirect": f"/dashboard/dpr/{id}",
        "boq_tracker_updated": boq_actual_count > 0,
        "boq_actual_count": boq_actual_count,
        "boq_tracker_message": "BOQ tracker updated based on today's progress report" if boq_actual_count else "",
    })


@app.route("/dashboard/dpr/<int:id>")
@dashboard_access_required
def dashboard_dpr_view(id):
    username, company = dashboard_identity()
    report = get_dpr_record(id)
    if not report:
        abort(404)
    detail = get_dpr_detail(id)
    return render_template("dashboard/dpr/view.html", username=username, company=company, active_dashboard="dpr", report=report, detail=detail)


@app.route("/dashboard/dpr/<int:id>/delete", methods=["POST"])
@dashboard_access_required
def dpr_delete(id):
    if session.get("is_guest"):
        return jsonify({"error": "Cannot delete demo data", "success": False})

    dpr = DPR.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    db.session.delete(dpr)
    db.session.commit()
    return jsonify({"success": True, "message": "DPR deleted successfully"})


@app.route("/api/dpr/summarize", methods=["POST"])
@limiter.limit("10 per minute")
@dashboard_access_required
def api_dpr_summarize():
    forbidden = ensure_ai_access("dpr")
    if forbidden:
        return forbidden

    data = request.json or {}
    notes = sanitize_input(data.get("notes", ""))
    project = sanitize_input(data.get("project", ""))
    zone = sanitize_input(data.get("zone", ""))
    weather = sanitize_input(data.get("weather", ""))
    issues = sanitize_input(data.get("issues", ""))
    report_date = sanitize_input(data.get("date", ""))
    try:
        worker_count = int(data.get("worker_count", 0))
    except (TypeError, ValueError):
        worker_count = 0

    prompt = f"""You are a senior construction site engineer
writing a formal Daily Progress Report for a Saudi Arabian
construction project.

Project: {project}
Date: {report_date}
Zone/Location: {zone}
Weather Condition: {weather}
Workers Present: {worker_count}
Progress Notes: {notes}
Issues/Delays: {issues}

Generate a professional report summary with exactly
this structure:

ENGLISH SUMMARY:
[Write 3-4 sentences in formal engineering language
describing the day's progress, productivity, and
any notable events]

KEY INSIGHT:
[One sentence identifying a trend or recommendation
based on this data]

ARABIC SUMMARY:
[Translate the English Summary into formal Arabic
suitable for Saudi construction documentation]

Keep the total response under 300 words."""

    success, result = call_openai(prompt=prompt, max_tokens=500, temperature=0.3)
    if success:
        record_ai_usage("dpr")
        response_payload = {"summary": result, "success": True}
        response_payload.update(build_ai_response_meta(current_user))
        return jsonify(response_payload)
    return jsonify({"error": result, "success": False}), 200


@app.route("/api/dpr/transcribe-voice", methods=["POST"])
@limiter.limit("10 per minute")
@login_required
def dpr_transcribe_voice():
    import os
    import tempfile

    forbidden = ensure_ai_access("dpr")
    if forbidden:
        return forbidden
    if not app.config.get("OPENAI_API_KEY"):
        return jsonify({"error": "AI service is not configured. Please set OPENAI_API_KEY.", "success": False}), 503

    if "audio" not in request.files:
        return jsonify({"error": "No audio file", "success": False})

    audio_file = request.files["audio"]
    field = sanitize_input(request.form.get("field", "general"), 50)

    audio_file.seek(0, 2)
    size = audio_file.tell()
    audio_file.seek(0)
    if size > 25 * 1024 * 1024:
        return jsonify({"error": "Audio file too large. Max 25MB.", "success": False})

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".webm", delete=False) as tmp:
            audio_file.save(tmp.name)
            tmp_path = tmp.name

        whisper_client = OpenAI(api_key=app.config["OPENAI_API_KEY"])
        with open(tmp_path, "rb") as audio_stream:
            transcript = whisper_client.audio.transcriptions.create(
                model="whisper-1",
                file=audio_stream,
                response_format="verbose_json",
            )

        record_ai_usage("dpr")
        response_payload = {
            "success": True,
            "text": transcript.text,
            "language": getattr(transcript, "language", ""),
            "field": field,
        }
        response_payload.update(build_ai_response_meta(current_user))
        return jsonify(response_payload)
    except Exception as e:
        app.logger.error(f"Voice transcription error: {str(e)}")
        return jsonify({"error": f"Transcription failed: {str(e)}", "success": False})
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.route("/api/dpr/import-workers", methods=["POST"])
@login_required
def import_dpr_workers():
    import csv
    import io

    if "file" not in request.files:
        return jsonify({"error": "No file", "success": False})

    file = request.files["file"]
    filename = (file.filename or "").lower()
    workers = []

    try:
        if filename.endswith(".csv"):
            content = file.read().decode("utf-8")
            reader = csv.DictReader(content.splitlines())
            for row in reader:
                name = (row.get("Worker Name", "") or row.get("Name", "") or row.get("name", "")).strip()
                if not name:
                    continue
                raw_hours = row.get("Hours Worked", 8) or row.get("Hours", 8) or 8
                try:
                    hours = float(raw_hours)
                except (TypeError, ValueError):
                    hours = 8
                workers.append(
                    {
                        "name": name,
                        "role": (row.get("Role", "") or row.get("role", "")).strip(),
                        "hours": hours,
                        "present": str(row.get("Present (Yes/No)", "Yes")).lower() not in ["no", "false", "0", "n"],
                    }
                )
        elif filename.endswith((".xlsx", ".xls")):
            wb = load_workbook(io.BytesIO(file.read()), data_only=True)
            ws = wb.active
            header_row = None
            col_map = {}

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_lower = [str(cell).lower().strip() if cell else "" for cell in row]
                if any(key in row_lower for key in ["name", "worker"]):
                    header_row = row_idx
                    for col_idx, cell in enumerate(row_lower):
                        if "name" in cell or "worker" in cell:
                            col_map["name"] = col_idx
                        elif "role" in cell or "position" in cell or "trade" in cell:
                            col_map["role"] = col_idx
                        elif "hour" in cell:
                            col_map["hours"] = col_idx
                        elif "present" in cell or "attend" in cell:
                            col_map["present"] = col_idx
                    break

            if not header_row or "name" not in col_map:
                return jsonify(
                    {
                        "error": "Could not find Name/Worker column. Please use the template provided.",
                        "success": False,
                    }
                )

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                name_idx = col_map.get("name", 0)
                if name_idx >= len(row) or not row[name_idx]:
                    continue
                name = str(row[name_idx]).strip()
                if not name or name == "None":
                    continue

                role = ""
                if "role" in col_map and col_map["role"] < len(row):
                    role_val = row[col_map["role"]]
                    role = str(role_val).strip() if role_val else ""

                hours = 8
                if "hours" in col_map and col_map["hours"] < len(row):
                    try:
                        hours = float(row[col_map["hours"]] or 8)
                    except (TypeError, ValueError):
                        hours = 8

                present = True
                if "present" in col_map and col_map["present"] < len(row):
                    present = str(row[col_map["present"]]).lower() not in ["no", "false", "0", "n"]

                workers.append({"name": name, "role": role, "hours": hours, "present": present})
        else:
            return jsonify({"error": "Upload .xlsx, .xls or .csv", "success": False})

        if not workers:
            return jsonify({"error": "No workers found in file. Check column names match the template.", "success": False})

        return jsonify({"success": True, "workers": workers, "count": len(workers)})
    except Exception as e:
        return jsonify({"error": f"File error: {str(e)}", "success": False})


@app.route("/dashboard/dpr/<int:id>/pdf")
@dashboard_access_required
def dashboard_dpr_pdf(id):
    report = get_dpr_record(id)
    if not report:
        abort(404)
    detail = get_dpr_detail(id)

    def escape_html(text):
        text = str(text or "")
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DprTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.HexColor("#0a0a0a"),
        alignment=TA_CENTER,
        spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        "DprHeading",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        textColor=colors.HexColor("#0a0a0a"),
        spaceBefore=8,
        spaceAfter=4,
    )
    normal_style = ParagraphStyle(
        "DprNormal",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#1a1a1a"),
    )
    label_style = ParagraphStyle(
        "DprLabel",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=colors.HexColor("#666666"),
    )
    def build_text_para(text, default_style=normal_style):
        value = str(text or "-").strip()
        value = value if value else "-"
        return Paragraph(escape_html(value).replace("\n", "<br/>"), default_style)

    story = []
    header_data = [
        [
            Paragraph(
                '<font color="white"><b>BanaaIQ</b></font>',
                ParagraphStyle("HdrLeft", parent=styles["BodyText"], alignment=TA_LEFT, fontName="Helvetica-Bold", fontSize=14),
            ),
            Paragraph(
                '<font color="#e8c547"><b>DAILY PROGRESS REPORT</b></font>',
                ParagraphStyle("HdrRight", parent=styles["BodyText"], alignment=TA_RIGHT, fontName="Helvetica-Bold", fontSize=11),
            ),
        ]
    ]
    header = Table(header_data, colWidths=[8 * cm, 9 * cm])
    header.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0a0a0a")),
                ("PADDING", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(header)
    story.append(Spacer(1, 0.35 * cm))
    story.append(Paragraph("Daily Progress Report", title_style))
    story.append(Spacer(1, 0.25 * cm))

    def make_detail_row(label, value):
        return [Paragraph(f"<b>{escape_html(label)}</b>", label_style), build_text_para(value)]

    details_table = Table(
        [
            make_detail_row("Date", report.get("date", "")),
            make_detail_row("Project", report.get("project", "")),
            make_detail_row("Zone / Location", report.get("zone", "")),
            make_detail_row("Weather", report.get("weather", "")),
            make_detail_row("Workers", str(report.get("workers", ""))),
            make_detail_row("Report Status", report.get("status", "Completed")),
        ],
        colWidths=[5 * cm, 12 * cm],
    )
    details_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f6f3ec")),
                ("BACKGROUND", (1, 0), (1, -1), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e4ddd0")),
                ("PADDING", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(details_table)
    story.append(Spacer(1, 0.3 * cm))

    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#e8c547")))
    story.append(Spacer(1, 0.15 * cm))
    story.append(Paragraph("Progress Notes", heading_style))
    progress_text = detail.get("progress_notes") or detail.get("notes") or "N/A"
    story.append(build_text_para(progress_text))
    story.append(Spacer(1, 0.25 * cm))

    issues_text = detail.get("issues", "").strip()
    if issues_text:
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e4ddd0")))
        story.append(Spacer(1, 0.15 * cm))
        story.append(Paragraph("Issues / Delays", heading_style))
        story.append(build_text_para(issues_text))
        story.append(Spacer(1, 0.25 * cm))

    workers = detail.get("workers", [])
    if workers:
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e4ddd0")))
        story.append(Spacer(1, 0.15 * cm))
        story.append(Paragraph("Worker Attendance", heading_style))
        worker_rows = [[
            Paragraph("<b>Worker Name</b>", label_style),
            Paragraph("<b>Role</b>", label_style),
            Paragraph("<b>Hours</b>", label_style),
            Paragraph("<b>Present</b>", label_style),
        ]]
        for w in workers:
            worker_rows.append(
                [
                    build_text_para(w.get("name", "")),
                    build_text_para(w.get("role", "")),
                    build_text_para(str(w.get("hours", 0))),
                    build_text_para("Yes" if w.get("present") else "No"),
                ]
            )
        worker_table = Table(worker_rows, colWidths=[6 * cm, 5 * cm, 3 * cm, 3 * cm])
        worker_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0a0a0a")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e4ddd0")),
                    ("PADDING", (0, 0), (-1, -1), 6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f3ec")]),
                ]
            )
        )
        story.append(worker_table)
        story.append(Spacer(1, 0.25 * cm))

    ai_summary = (detail.get("ai_summary", "") or "").strip()
    if ai_summary:
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#e8c547")))
        story.append(Spacer(1, 0.15 * cm))
        story.append(Paragraph("AI-Generated Summary", heading_style))

        clean_summary = ai_summary.replace("**", "")
        marker = "ARABIC SUMMARY:"
        parts = clean_summary.split(marker, 1)

        english_part = parts[0].replace("ENGLISH SUMMARY:", "").replace("KEY INSIGHT:", "\nKey Insight: ").strip()
        summary_text = english_part or clean_summary
        story.append(build_text_para(summary_text))

        if False and len(parts) > 1:
            arabic_part = parts[1].strip()
            if arabic_part:
                story.append(Paragraph("الملخص بالعربية", arabic_heading_style))
                for line in [ln.strip() for ln in arabic_part.split("\n") if ln.strip()]:
                    if is_arabic_text(line) and arabic_available:
                        processed = process_arabic_text(line)
                        story.append(Paragraph(escape_html(processed), arabic_style))
                    else:
                        story.append(build_text_para(line))
        if False and len(parts) == 1:
            # Single language summary.
            story.append(build_text_para(clean_summary))
    else:
        story.append(Paragraph("AI summary has not been generated for this report.", normal_style))

    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e4ddd0")))
    footer = Table(
        [[
            Paragraph(f"Report ID: DPR-{id} | Confidential", ParagraphStyle("FooterL", parent=styles["BodyText"], fontSize=8, textColor=colors.HexColor("#999999"), alignment=TA_LEFT)),
            Paragraph("BanaaIQ", ParagraphStyle("FooterR", parent=styles["BodyText"], fontSize=8, textColor=colors.HexColor("#999999"), alignment=TA_RIGHT)),
        ]],
        colWidths=[8.5 * cm, 8.5 * cm],
    )
    footer.setStyle(TableStyle([("PADDING", (0, 0), (-1, -1), 3)]))
    story.append(footer)

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=f"DPR-{id}-BanaaIQ.pdf")


@app.route("/dashboard/dpr/<int:id>/excel")
@dashboard_access_required
def dashboard_dpr_excel(id):
    report = get_dpr_record(id)
    if not report:
        abort(404)
    detail = get_dpr_detail(id)
    workbook = Workbook()
    details_sheet = workbook.active
    details_sheet.title = "Report Details"
    details_sheet.merge_cells("A1:D1")
    details_sheet["A1"] = "BanaaIQ - Daily Progress Report"
    details_sheet["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    details_sheet["A1"].alignment = Alignment(horizontal="center")
    details_sheet["A1"].fill = PatternFill(start_color="0F2940", end_color="0F2940", fill_type="solid")
    kv_rows = [("Date", report["date"]), ("Project", report["project"]), ("Zone", report["zone"]), ("Weather", report["weather"]), ("Workers", report["workers"]), ("Status", report["status"])]
    row_idx = 3
    for key, value in kv_rows:
        details_sheet[f"A{row_idx}"] = key
        details_sheet[f"A{row_idx}"].font = Font(bold=True)
        details_sheet[f"B{row_idx}"] = value
        row_idx += 1
    details_sheet["A12"] = "Progress Notes"
    details_sheet["A12"].font = Font(bold=True)
    details_sheet["A13"] = detail.get("notes", "")
    details_sheet["A15"] = "Issues / Delays"
    details_sheet["A15"].font = Font(bold=True)
    details_sheet["A16"] = detail.get("issues", "")
    details_sheet["A18"] = "AI Summary"
    details_sheet["A18"].font = Font(bold=True)
    details_sheet["A19"] = detail.get("ai_summary", "") or "AI summary has not been generated for this report."
    details_sheet.column_dimensions["A"].width = 22
    details_sheet.column_dimensions["B"].width = 65
    worker_sheet = workbook.create_sheet("Worker Attendance")
    headers = ["Name", "Role", "Hours", "Present"]
    for col, header in enumerate(headers, start=1):
        cell = worker_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill(start_color="0F2940", end_color="0F2940", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for i, worker in enumerate(detail.get("workers", []), start=2):
        bg = "FFFFFF" if i % 2 == 0 else "F3F4F6"
        worker_sheet.cell(row=i, column=1, value=worker.get("name", ""))
        worker_sheet.cell(row=i, column=2, value=worker.get("role", ""))
        worker_sheet.cell(row=i, column=3, value=worker.get("hours", ""))
        worker_sheet.cell(row=i, column=4, value="Yes" if worker.get("present", False) else "No")
        for c in range(1, 5):
            worker_sheet.cell(row=i, column=c).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    worker_sheet.column_dimensions["A"].width = 24
    worker_sheet.column_dimensions["B"].width = 24
    worker_sheet.column_dimensions["C"].width = 12
    worker_sheet.column_dimensions["D"].width = 12
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"DPR-{id}-BanaaIQ.xlsx")


@app.route("/dashboard/dpr/export/pdf")
@dashboard_access_required
def dashboard_dpr_export_all_pdf():
    import io

    if session.get("is_guest"):
        return "Export unavailable in demo mode", 403

    dprs = DPR.query.filter_by(user_id=current_user.id).order_by(DPR.created_at.desc()).all()
    if not dprs:
        return "No DPR records found", 404

    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    navy = colors.HexColor("#0a0a0a")
    gold = colors.HexColor("#e8c547")
    light = colors.HexColor("#f6f3ec")

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    story = []

    hdr = [[
        Paragraph("<b><font color='white' size='14'>BanaaIQ | DPR Report</font></b>", ps("H", fontName="Helvetica-Bold", fontSize=14, textColor=colors.white)),
        Paragraph(f"<font color='#e8c547' size='9'><b>All Projects</b></font><br/><font color='#aaaaaa' size='8'>{datetime.now().strftime('%d %B %Y')}</font>", ps("HR", fontName="Helvetica", fontSize=9, alignment=TA_RIGHT, textColor=colors.white)),
    ]]
    header_table = Table(hdr, colWidths=[10 * cm, 7 * cm])
    header_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), navy), ("PADDING", (0, 0), (-1, -1), 14), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(header_table)
    story.append(Spacer(1, 0.4 * cm))

    story.append(
        Table(
            [[Paragraph(f"<b>{len(dprs)}</b><br/><font size='8' color='#666666'>Total Reports</font>", ps("S1", fontName="Helvetica-Bold", fontSize=16, alignment=TA_CENTER))]],
            colWidths=[17 * cm],
            style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), light), ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 10)]),
        )
    )
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=gold))
    story.append(Spacer(1, 0.2 * cm))

    def lbl(text):
        return Paragraph(f"<b>{text}</b>", ps("L", fontName="Helvetica-Bold", fontSize=8, textColor=colors.white))

    def cell(text):
        return Paragraph(str(text or "-"), ps("C", fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#1a1a1a")))

    rows = [[lbl("#"), lbl("Date"), lbl("Project"), lbl("Zone"), lbl("Weather"), lbl("Status")]]
    for idx, dpr in enumerate(dprs, start=1):
        rows.append([cell(idx), cell(format_display_date(dpr.date) or "-"), cell(dpr.project), cell(dpr.zone), cell(dpr.weather), cell(dpr.status)])

    table = Table(rows, colWidths=[1 * cm, 2.4 * cm, 4.3 * cm, 4.1 * cm, 2.2 * cm, 3 * cm], repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), navy), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e4ddd0")), ("PADDING", (0, 0), (-1, -1), 5), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light]), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="DPR-Report-BanaaIQ.pdf", mimetype="application/pdf")


@app.route("/dashboard/dpr/export/excel")
@dashboard_access_required
def dashboard_dpr_export_all_excel():
    import io
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if session.get("is_guest"):
        return "Export unavailable in demo mode", 403

    dprs = DPR.query.filter_by(user_id=current_user.id).order_by(DPR.created_at.desc()).all()
    if not dprs:
        return "No DPR records found", 404

    wb = Workbook()
    ws = wb.active
    ws.title = "DPR"

    navy = PatternFill("solid", fgColor="0A1628")
    gold = PatternFill("solid", fgColor="E8C547")
    light = PatternFill("solid", fgColor="F8F9FA")
    thin = Side(style="thin", color="DEE2E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:G1")
    ws["A1"] = "DAILY PROGRESS REPORTS - All Projects"
    ws["A1"].fill = navy
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y')} | BanaaIQ"
    ws["A2"].fill = gold
    ws["A2"].font = Font(color="FFFFFF", bold=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["#", "Date", "Project", "Zone", "Weather", "Temperature", "Status"]
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=header)
        c.fill = navy
        c.font = Font(color="FFFFFF", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for i, dpr in enumerate(dprs, start=5):
        row = [i - 4, format_display_date(dpr.date) or "-", safe_excel_cell(dpr.project or "-"), safe_excel_cell(dpr.zone or "-"), safe_excel_cell(dpr.weather or "-"), safe_excel_cell(str(dpr.temperature or "-")), safe_excel_cell(dpr.status or "-")]
        fill = light if (i - 5) % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.border = border
            c.alignment = Alignment(vertical="center")
            c.fill = fill
            if col == 3:
                c.font = Font(bold=True)

    widths = [5, 16, 36, 28, 14, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="DPR-Report-BanaaIQ.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/feature-projects/add", methods=["POST"])
@login_required
def add_feature_project():
    if session.get("is_guest"):
        return jsonify({"error": "Sign up to create projects", "success": False})

    data = request.get_json(silent=True) or {}
    feature = sanitize_input(data.get("feature", ""), 30).lower()
    name = sanitize_input(data.get("name", ""), 200).strip()
    color = sanitize_input(data.get("color", "#0a0a0a"), 10) or "#0a0a0a"
    if not feature or not name:
        return jsonify({"error": "Feature and name required", "success": False})
    if feature not in {"dpr", "boq", "inventory", "tasks"}:
        return jsonify({"error": "Invalid feature", "success": False})
    if not name:
        return jsonify({"error": "Project name is required", "success": False})

    existing = FeatureProject.query.filter_by(user_id=current_user.id, name=name, feature=feature).first()
    if existing:
        return jsonify({"error": "Project already exists", "success": False})

    new_project = FeatureProject(
        user_id=current_user.id,
        name=name,
        color=color,
        feature=feature,
    )
    db.session.add(new_project)
    db.session.commit()
    return jsonify({"success": True, "project": new_project.to_dict()})


@app.route("/api/feature-projects/list", methods=["GET"])
@login_required
def list_feature_projects():
    feature = sanitize_input(request.args.get("feature", ""), 30).lower()
    if not feature:
        return jsonify({"error": "feature required", "success": False})
    projects = FeatureProject.query.filter_by(user_id=current_user.id, feature=feature).order_by(FeatureProject.name).all()
    return jsonify({"success": True, "projects": [project.to_dict() for project in projects]})


@app.route("/api/feature-projects/delete/<int:id>", methods=["POST"])
@login_required
def delete_feature_project(id):
    proj = FeatureProject.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    feature = proj.feature
    if feature == "dpr":
        DPR.query.filter_by(user_id=current_user.id, feature_project_id=id).update({"feature_project_id": None, "project": ""})
    elif feature == "inventory":
        InventoryItem.query.filter_by(user_id=current_user.id, feature_project_id=id).update({"feature_project_id": None})
    elif feature == "tasks":
        pass  # tasks are now FK-cascaded from Project; feature_project_id no longer exists
    elif feature == "boq":
        BOQ.query.filter_by(user_id=current_user.id, feature_project_id=id).update({"feature_project_id": None, "project": ""})
    db.session.delete(proj)
    db.session.commit()
    return jsonify({"success": True, "message": f"Project deleted from {feature}"})


@app.route("/dashboard/translator")
@dashboard_access_required
@role_required(ROLE_SITE_ENGINEER)
def dashboard_translator():
    username, company = dashboard_identity()
    ai_quota = get_ai_quota_snapshot(current_user) if current_user.is_authenticated else {
        "used": 0,
        "remaining": 0,
        "limit": 0,
        "unlimited": False,
    }
    return render_template(
        "dashboard/translator/index.html",
        username=username,
        company=company,
        active_dashboard="translator",
        ai_quota=ai_quota,
    )


@app.route("/translate", methods=["POST"])
@limiter.limit("10 per minute")
def landing_translate():
    payload = request.get_json(silent=True) if request.is_json else request.form
    if not payload:
        return jsonify({"error": "No data received", "success": False}), 400

    text = sanitize_input(payload.get("text", ""), max_length=3000)
    direction = sanitize_input(payload.get("direction", "en-ar"), max_length=20).lower()

    if not text.strip():
        return jsonify({"error": "Please enter text to translate.", "success": False}), 400
    if len(text.strip()) < 3:
        return jsonify({"error": "Text too short. Enter more words.", "success": False}), 400

    if direction == "ar-en":
        primary_lang = "Arabic"
        output_lang = "English"
    else:
        primary_lang = "English"
        output_lang = "Arabic"

    if current_user.is_authenticated:
        forbidden = ensure_ai_access("translator")
        if forbidden:
            return forbidden

    result = process_translation(
        original_text=text,
        input_type="text",
        primary_lang=primary_lang,
        output_lang=output_lang,
    )

    if result.get("success"):
        record_ai_usage("translator")
        result.update(build_ai_response_meta(current_user))
        return jsonify(result)

    return jsonify(result), 400


@app.route("/api/translator/text", methods=["POST"])
@limiter.limit("15 per minute")
@dashboard_access_required
def api_translator_text():
    forbidden = ensure_ai_access("translator")
    if forbidden:
        return forbidden

    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "No data received", "success": False})

    text = sanitize_input(data.get("text", ""), max_length=3000)
    primary_lang = data.get("primary_lang", "English")
    output_lang = data.get("output_lang", "Arabic")

    if not text.strip():
        return jsonify({"error": "Please enter text to translate.", "success": False})
    if len(text.strip()) < 3:
        return jsonify({"error": "Text too short. Enter more words.", "success": False})

    result = process_translation(
        original_text=text,
        input_type="text",
        primary_lang=primary_lang,
        output_lang=output_lang,
    )
    if result.get("success"):
        record_ai_usage("translator")
        result.update(build_ai_response_meta(current_user))
    return jsonify(result)


@app.route("/api/translator/voice", methods=["POST"])
@limiter.limit("15 per minute")
@dashboard_access_required
def api_translator_voice():
    import os
    import tempfile

    forbidden = ensure_ai_access("translator")
    if forbidden:
        return forbidden
    if client is None:
        return jsonify({"error": "AI service is not configured. Please set OPENAI_API_KEY.", "success": False}), 503

    if "audio" not in request.files:
        return jsonify({"error": "No audio file uploaded.", "success": False})

    audio_file = request.files["audio"]
    primary_lang = request.form.get("primary_lang", "English")
    output_lang = request.form.get("output_lang", "Arabic")

    if audio_file.filename == "":
        return jsonify({"error": "No file selected.", "success": False})

    allowed = {"mp3", "mp4", "wav", "m4a", "ogg", "webm", "mpeg", "mpga"}
    filename = (audio_file.filename or "").lower()
    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""
    if ext not in allowed:
        return jsonify({"error": f"Unsupported format '.{ext}'. Use MP3, WAV, M4A or OGG.", "success": False})

    audio_file.seek(0, 2)
    size = audio_file.tell()
    audio_file.seek(0)
    if size > 25 * 1024 * 1024:
        return jsonify({"error": "File too large. Maximum 25MB.", "success": False})
    if size == 0:
        return jsonify({"error": "Audio file is empty.", "success": False})

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}") as tmp:
            audio_file.save(tmp.name)
            tmp_path = tmp.name

        with open(tmp_path, "rb") as audio:
            transcript = client.audio.transcriptions.create(
                model="whisper-1",
                file=audio,
                response_format="text",
            )

        extracted_text = str(transcript).strip()
        if not extracted_text:
            return jsonify(
                {
                    "error": "Could not transcribe audio. Please ensure clear speech with no background noise.",
                    "success": False,
                }
            )
        if len(extracted_text) < 5:
            return jsonify({"error": "Recording too short. Please record a longer message.", "success": False})

        result = process_translation(
            original_text=extracted_text,
            input_type="voice",
            primary_lang=primary_lang,
            output_lang=output_lang,
        )
        if result.get("success"):
            record_ai_usage("translator")
            result.update(build_ai_response_meta(current_user))
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Voice transcription error: {str(e)}")
        error_str = str(e).lower()
        if "api" in error_str:
            msg = "AI transcription unavailable. Try again."
        elif "format" in error_str:
            msg = "Audio format not supported. Use MP3."
        else:
            msg = "Voice processing failed. Try text input."
        return jsonify({"error": msg, "success": False})
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


# [BOQ rebuild] removed: api_boq_suggest, api_boq_guided_upload, export_boq_analysis_pdf, export_boq_analysis_preview_pdf, api_boq_upload_excel, api_boq_summary, generate_quotation_pdf, generate_guidance_pdf, generate_quotation_excel, dashboard_boq_pdf, dashboard_boq_excel

@app.route("/api/notifications/read-all", methods=["POST"])
@dashboard_access_required
def notifications_read_all():
    if current_user.is_authenticated and not session.get("is_guest"):
        Notification.query.filter_by(user_id=current_user.id, is_read=False).update({"is_read": True})
        db.session.commit()
    session["read_notifications"] = [item["id"] for item in build_live_notifications()]
    session.modified = True
    return jsonify({"success": True})


@app.route("/api/notifications")
@dashboard_access_required
def notifications_feed():
    lang = session.get("lang", "en")
    notifications = get_localized_notifications(lang)
    unread_count = sum(1 for item in notifications if not item.get("read"))
    return jsonify({"success": True, "notifications": notifications, "unread_count": unread_count})


# /dashboard/boq/<id>/pdf and /dashboard/boq/<id>/excel removed in BOQ rebuild
# New endpoints: boq_download_pdf and boq_download_excel (see above)


@app.route("/tutorials/faq")
@dashboard_access_required
def tutorials_faq():
    username, company = dashboard_identity()
    return render_template("tutorials/faq.html", username=username, company=company, active_dashboard="tutorials", active_tutorial="faq")


@app.route("/tutorials/dpr-guide")
@dashboard_access_required
def tutorials_dpr():
    username, company = dashboard_identity()
    return render_template("tutorials/dpr_guide.html", username=username, company=company, active_dashboard="tutorials", active_tutorial="dpr")


@app.route("/tutorials/boq-guide")
@dashboard_access_required
def tutorials_boq():
    username, company = dashboard_identity()
    return render_template("tutorials/boq_guide.html", username=username, company=company, active_dashboard="tutorials", active_tutorial="boq")


@app.route("/tutorials/inventory-guide")
@dashboard_access_required
def tutorials_inventory():
    username, company = dashboard_identity()
    return render_template("tutorials/inventory_guide.html", username=username, company=company, active_dashboard="tutorials", active_tutorial="inventory")


@app.route("/tutorials/task-guide")
@dashboard_access_required
def tutorials_task():
    username, company = dashboard_identity()
    return render_template("tutorials/task_guide.html", username=username, company=company, active_dashboard="tutorials", active_tutorial="task")


@app.route("/tutorials/translator-guide")
@dashboard_access_required
def tutorials_translator():
    username, company = dashboard_identity()
    return render_template("tutorials/translator_guide.html", username=username, company=company, active_dashboard="tutorials", active_tutorial="translator")


@app.route("/dashboard/<module>")
@dashboard_access_required
def dashboard_modules(module):
    username, company = dashboard_identity()
    lang = session.get("lang", "en")
    allowed = {
        "translator": ("Translator", "المترجم"),
        "boq": ("Bill of Quantities", "جدول الكميات"),
        "inventory": ("Inventory Tracker", "متابعة المخزون"),
        "tasks": ("Task Management", "إدارة المهام"),
        "chatbot": ("AI Chatbot", "المساعد الذكي"),
    }
    module_titles = allowed.get(module)
    if not module_titles:
        abort(404)
    module_title = module_titles[1] if lang == "ar" else module_titles[0]
    return render_template(
        "dashboard/module_placeholder.html",
        username=username,
        company=company,
        module_title=module_title,
        active_dashboard=module,
        show_demo_data=is_demo_mode(),
    )


@app.route("/blog")
def blog_index():
    posts = [
        {"slug": "vision-2030-ai-construction", "category": "Saudi Market", "title": "Vision 2030 and the Rise of AI-Powered Construction in Saudi Arabia", "excerpt": "How digital delivery, data-led execution, and AI workflows are reshaping giga-project outcomes across the Kingdom."},
        {"slug": "multilingual-construction-teams-best-practices", "category": "Site Operations", "title": "Managing Multilingual Construction Teams: Arabic-English Best Practices", "excerpt": "Methods for reducing rework and safety incidents by making bilingual coordination practical and consistent."},
        {"slug": "predictive-inventory-management-ksa", "category": "Procurement", "title": "How Predictive Inventory Management Saves KSA Mega-Projects Millions", "excerpt": "Why forecast-driven material control is becoming a core discipline for project controls and procurement leaders."},
    ]
    return render_template("blog/index.html", posts=posts)


@app.route("/blog/<slug>")
def blog_article(slug):
    template_name = BLOG_ARTICLES.get(slug)
    if not template_name:
        abort(404)
    return render_template(template_name)


@app.route("/set-lang/<lang>")
def set_lang(lang):
    # Language switching is disabled for now. Keep English only.
    session["lang"] = "en"
    session.modified = True
    referrer = request.referrer
    if referrer:
        return redirect(referrer)
    return redirect(url_for("index"))


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/_internal/sentry-test")
@login_required
def sentry_test():
    """Trigger a test exception to verify Sentry capture. Admin/DEBUG only."""
    if not (getattr(current_user, "is_admin", False) or app.debug):
        abort(404)
    raise ZeroDivisionError("Sentry test — intentional exception from /_internal/sentry-test")


@app.errorhandler(404)
def not_found(error):
    if wants_json_response():
        return jsonify(success=False, error="Not found."), 404
    try:
        return render_template("errors/404.html"), 404
    except Exception:
        return "<h1>Page Not Found</h1><p>The page you requested could not be found.</p>", 404


@app.errorhandler(500)
def internal_server_error(error):
    app.logger.error(f"500 error: {str(error)}")
    if wants_json_response():
        return jsonify(success=False, error="Server error. Please try again."), 500
    try:
        return render_template("errors/500.html"), 500
    except Exception:
        return (
            "<h1>Server Error</h1>"
            "<p>Please try again.</p>"
            "<a href=\"/dashboard\">Go back</a>"
        ), 500


@app.errorhandler(429)
def ratelimit_handler(e):
    retry_after = getattr(e, "retry_after", None) or getattr(e, "description", None) or "a few minutes"
    message_en = f"Too many requests. Please wait {retry_after} and try again."
    message_ar = f"طلبات كثيرة جداً. يرجى الانتظار {retry_after} والمحاولة مرة أخرى."
    if request.path.startswith("/api/") or request.is_json or wants_json_response():
        return jsonify({
            "error": "rate_limited",
            "message": message_en,
            "message_ar": message_ar,
            "retry_after": str(retry_after),
            "success": False,
        }), 429
    try:
        return render_template(
            "errors/429.html",
            message=message_en,
            message_ar=message_ar,
            retry_after=str(retry_after),
        ), 429
    except Exception:
        return message_en, 429


@app.errorhandler(CSRFError)
def csrf_error(e):
    app.logger.warning("CSRF validation failed: %s", e.description)
    if wants_json_response():
        return jsonify(error="Session expired. Please refresh the page.", success=False), 400
    flash("Your session expired. Please refresh the page and try again.", "warning")
    try:
        return render_template("errors/500.html"), 400
    except Exception:
        return "<h1>Session expired</h1><p>Please refresh the page and try again.</p>", 400


@app.after_request
def set_security_headers(response):
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=(self), camera=()"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com "
        "https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "img-src 'self' data: blob:; "
        "connect-src 'self' https://api.openai.com; "
        "media-src 'self' blob:; "
        "object-src 'none'; "
        "base-uri 'self'; "
        "form-action 'self'; "
        "frame-ancestors 'self';"
    )
    if current_user.is_authenticated or session.get("is_guest") or is_protected_path(request.path):
        apply_no_store_headers(response)
    return response


@app.errorhandler(Exception)
def handle_exception(error):
    if isinstance(error, HTTPException):
        if error.code == 404:
            return not_found(error)
        if wants_json_response():
            return jsonify(success=False, error=error.description or error.name), error.code
        return error

    app.logger.error(
        f"Unhandled exception: {str(error)}\n"
        f"{traceback.format_exc()}"
    )
    if wants_json_response():
        return jsonify(success=False, error="Server error. Please try again."), 500
    try:
        return render_template("errors/500.html"), 500
    except Exception:
        return (
            "<h1>Something went wrong</h1>"
            "<a href=\"/dashboard\">Go home</a>"
        ), 500


def _print_startup_status():
    """Print a startup banner showing monitoring + email configuration."""
    app.logger.info("=" * 60)
    app.logger.info("BanaaIQ starting — env=%s", os.getenv("FLASK_ENV", "production"))
    app.logger.info(
        "  Sentry:  %s",
        "enabled" if os.getenv("SENTRY_DSN", "").strip() else "disabled (set SENTRY_DSN in production)",
    )
    app.logger.info("  Email:   provider=%s", app.config.get("EMAIL_PROVIDER", "console"))
    app.logger.info(
        "  Rate limiter: %s",
        "redis" if "redis://" in app.config.get("RATELIMIT_STORAGE_URI", "") else "in-memory (DEV ONLY)",
    )
    app.logger.info("=" * 60)


with app.app_context():
    _print_startup_status()


if __name__ == "__main__":
    app.run(debug=app.config["DEBUG"])
