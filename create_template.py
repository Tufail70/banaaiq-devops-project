from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Worker Attendance"

    navy = PatternFill("solid", fgColor="0A1628")
    headers = ["Worker Name", "Role", "Hours Worked", "Present (Yes/No)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = navy
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    samples = [
        ["Ahmed Al-Otaibi", "Site Engineer", 8, "Yes"],
        ["Khalid Nasser", "Civil Engineer", 8, "Yes"],
        ["Raj Kumar Sharma", "Foreman", 9, "Yes"],
    ]
    for row_idx, row in enumerate(samples, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx).value = value

    widths = [30, 25, 15, 20]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    output_dir = Path("static") / "templates"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "worker_attendance_template.xlsx"
    wb.save(output_path)
    print(f"Template created: {output_path}")


if __name__ == "__main__":
    main()
